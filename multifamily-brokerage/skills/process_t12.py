#!/usr/bin/env python3
"""
process_t12.py — Convert a T12 (Twelve Month Profit & Loss) PDF into the
brokerage's T12 Processor workbook, auto-mapping every GL account to the
brokerage charge codes (Masterkey).

Mapping engine (layered, in priority order):
  1. EXACT-MATCH MEMORY  — mappings.csv: confirmed (account -> code) pairs
     harvested from prior underwritings. Highest confidence.
  2. SECTION RULES       — the statement's own hierarchy (e.g. anything under
     "Payroll and Benefits" is 'pr'; "Utilities" lines split w/tr/e/o by
     keyword). Deterministic.
  3. KEYWORD RULES       — curated patterns per code (vacancy, concession,
     late fee, ...), constrained to the section's allowed codes.
  4. FUZZY MATCH         — similarity vs the corpus, constrained by section;
     >=0.90 auto-accepts, 0.75-0.90 flags for review.
  5. UNMAPPED            — best guess + REVIEW flag. Never silent.

Every run reconciles the coded totals against the statement's own printed
Total Revenue / Total Operating Expense / NOI, so an incomplete or
double-counted mapping cannot pass.

Usage:
    python process_t12.py input.pdf --template "T12 Processor.xlsx"
                          [-o out.xlsx] [--mappings mappings.csv]
    python process_t12.py harvest old1.xlsx old2.xlsx ...   # grow the corpus
"""

import argparse
import calendar
import csv
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime, timedelta

warnings.filterwarnings("ignore")

import pdfplumber
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

from difflib import SequenceMatcher

try:
    from rapidfuzz import fuzz
    def _sim(a, b):
        return fuzz.token_set_ratio(a, b) / 100.0
except ImportError:
    def _sim(a, b):
        return SequenceMatcher(None, a, b).ratio()


def _whole(a, b):
    """Whole-string similarity. token_set_ratio (used by _sim) scores 1.00
    whenever one label's token set is a SUBSET of the other's, so a bare
    generic account name like 'Income' matches 'write off of uncollectible
    other income' perfectly. Any fuzzy hit must also clear this bar or it
    is not trustworthy - found on The Gardens, 7/2026, where the property's
    entire revenue line would otherwise have been auto-coded 'bd'."""
    return SequenceMatcher(None, a, b).ratio()


# An account label that is the whole undifferentiated revenue/expense side of
# a statement (no detail beneath it). Always REVIEW - the split is unknown.
LUMP_INCOME = re.compile(r"^(total\s+)?(income|revenues?|gross (income|"
                         r"revenues?)|collections?)$", re.I)


# ----------------------------------------------------------------------------
# Chart of accounts (Masterkey)
# ----------------------------------------------------------------------------

INCOME_ORDER = ["r", "ll", "v", "nr", "bd", "rw", "ro", "oi"]
EXPENSE_ORDER = ["cs", "rm", "ad", "m", "pr", "w", "tr", "e", "o",
                 "mf", "i", "tx"]
# 'rt' (RUBS trash) lines are listed under the 'ro' category header.
CATEGORY_NAMES = {
    "r": "Rental Income", "ll": "(Loss to Lease) / Gain to Lease",
    "v": "Vacancy", "nr": "Non-Revenue/Concessions", "bd": "Bad Debt",
    "rw": "RUBS - Water/Sewer", "ro": "RUBS - Electric/Gas/Other",
    "oi": "Other Income", "cs": "Contract Services",
    "rm": "Repair & Maintenance", "ad": "Administrative", "m": "Marketing",
    "pr": "Payroll", "w": "Water/Sewer", "tr": "Trash", "e": "Electric",
    "o": "Gas/Other", "mf": "Management Fee", "i": "Insurance",
    "tx": "Real Estate Taxes",
}
INCOME_CODES = set(INCOME_ORDER) | {"rt"}
EXPENSE_CODES = set(EXPENSE_ORDER)

# Which codes are legal within each source-statement section
SECTION_ALLOWED = [
    (re.compile(r"gross potential rent", re.I), {"r", "ll"}),
    (re.compile(r"rental adjustment", re.I), {"v", "nr", "bd"}),
    (re.compile(r"other income", re.I), {"oi", "rw", "rt", "ro"}),
    (re.compile(r"payroll|benefits", re.I), {"pr"}),
    # Yardi-style "FIXED ADMINISTRATIVE" buckets accounting fees together with
    # insurance / management fee / property taxes. Constraining it to {ad}
    # (the plain "admin" rule below) would bury the house-rule mf/i/tx codes,
    # so this more specific rule must be tested first.
    (re.compile(r"fixed\s+admin", re.I), {"ad", "mf", "i", "tx"}),
    (re.compile(r"admin", re.I), {"ad"}),
    (re.compile(r"contract service", re.I), {"cs"}),
    (re.compile(r"marketing|advertis", re.I), {"m"}),
    # "maint" (not just "maintenance") so Yardi's "IN-HOUSE GENERAL
    # MAINT/SUPPLY" is recognised - otherwise its ELECTRICAL / PLUMBING
    # supply lines fall through to the unconstrained utility keywords.
    (re.compile(r"maint|repair|turnover|make.?ready", re.I), {"rm"}),
    # 'ad' allowed here: house rule routes telephone/internet lines that
    # sit under a Utilities section to Administrative
    (re.compile(r"utilit", re.I), {"w", "tr", "e", "o", "ad"}),
    (re.compile(r"management fee", re.I), {"mf"}),
    (re.compile(r"^tax|taxes", re.I), {"tx"}),
    (re.compile(r"insurance", re.I), {"i"}),
]

# Keyword rules: (pattern, code, restrict-to-side). Checked in order.
KEYWORD_RULES = [
    (r"gain.?/?.?loss to lease|loss to lease|gain to lease|ltl", "ll", "inc"),
    (r"vacan", "v", "inc"),
    (r"model|non.?rev|employee unit|office unit|down unit", "v", "inc"),
    (r"concession|write.?off|discount", "nr", "inc"),
    (r"bad debt|delinquen|prepaid|collection|skip|charge.?off", "bd", "inc"),
    (r"water.{0,12}(reimb|income)|sewer.{0,12}(reimb|income)", "rw", "inc"),
    (r"trash.{0,12}(reimb|income)(?!.*valet)", "rt", "inc"),
    (r"(gas|electric).{0,12}(reimb|income)", "ro", "inc"),
    (r"apartment income|rent income|rental income|gross rent|market rent"
     r"|potential rent", "r", "inc"),
    (r"payroll|salar|wage|bonus|incentive|workers comp|leasing comm"
     r"|employee benefit|allowance|health|medical|dental|vision", "pr",
     "exp"),
    (r"management fee", "mf", "exp"),
    # house rules: telephone/answering/uniforms are Administrative
    (r"telephone|\bphones?\b|internet|cable|answering|uniform", "ad", "exp"),
    (r"insurance", "i", "exp"),
    (r"tax", "tx", "exp"),
    (r"valet|trash|garbage|refuse|dumpster", "tr", "exp"),
    (r"water|sewer", "w", "exp"),
    (r"electric", "e", "exp"),
    (r"\bgas\b", "o", "exp"),
    (r"landscap|pest|alarm|security|courtesy patrol|elevator|towing"
     r"|snow|pool service|laundry|washer/dryer expense"
     r"|equipment rental", "cs", "exp"),
    (r"advertis|marketing|resident relation|promotion|signs|banner"
     r"|locator|internet listing|model", "m", "exp"),
    (r"legal|professional|office|postage|bank charge|dues|subscription"
     r"|screening|eviction|computer|software|permit|license|travel"
     r"|training|credit check|answering", "ad", "exp"),
    (r"repair|maint|plumb|carpentr|paint|appliance|hvac|a/?c|electrical"
     r"|supplies|turnover|make.?ready|carpet|flooring|blinds|lock|key"
     r"|janitorial|cleaning|glass|roof|tool", "rm", "exp"),
]


def short_prop(prop):
    """Strip corporate/asset suffixes for file naming: 'Harvest Moon
    Apartments' -> 'Harvest Moon'."""
    return re.sub(r"\s+(apartments?|apts\.?|apartment homes|llc|lp|ltd"
                  r"|inc\.?)$", "", prop.strip(), flags=re.I)


def safe_name(prop):
    """Make a property name usable inside a filename. Portfolio statements
    name several assets on one line ('Wren Rowe/Holly Ln') - the slash would
    otherwise be read as a directory separator."""
    return re.sub(r"\s+", " ", re.sub(r'[\\/:*?"<>|]+', "-", prop)).strip()


def norm(s):
    """Normalize an account label for matching: strip GL number, case,
    punctuation."""
    s = re.sub(r"^\s*\d[\d.\-]*\s*", "", str(s))       # leading GL number
    s = re.sub(r"[^a-z0-9/ ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def gl_number(s):
    m = re.match(r"^\s*(\d[\d.\-]*)\s", str(s))
    return m.group(1) if m else ""


# ----------------------------------------------------------------------------
# Corpus (exact-match memory)
# ----------------------------------------------------------------------------

def load_corpus(path):
    """mappings.csv -> {(gl, norm_name): code}, {norm_name: code}"""
    by_gl, by_name = {}, {}
    if path and os.path.exists(path):
        with open(path, newline="", encoding="utf-8-sig") as f:
            for rec in csv.DictReader(f):
                code = rec["code"].strip()
                nn = norm(rec["account"])
                gl = rec.get("gl", "") or gl_number(rec["account"])
                if gl:
                    by_gl[(gl, nn)] = code
                by_name.setdefault(nn, code)
    return by_gl, by_name


def harvest(paths, out_csv):
    """Extract confirmed (code, account) pairs from prior processor
    workbooks' RawData tabs into the corpus CSV."""
    rows, seen = [], set()
    if os.path.exists(out_csv):
        with open(out_csv, newline="", encoding="utf-8") as f:
            for rec in csv.DictReader(f):
                rows.append(rec)
                seen.add((rec.get("gl", ""), norm(rec["account"])))
    added = 0
    for p in paths:
        wb = load_workbook(p, read_only=True, data_only=True)
        if "RawData" not in wb.sheetnames:
            print(f"  skip {p}: no RawData tab")
            continue
        ws = wb["RawData"]
        for row in ws.iter_rows(min_row=6):
            code, acct = row[0].value, row[1].value
            if not code or not acct:
                continue
            code = str(code).strip().lower()
            if code in ("rev", "exp", "noi") or code not in \
                    (INCOME_CODES | EXPENSE_CODES):
                continue
            key = (gl_number(acct), norm(acct))
            if key in seen or not key[1]:
                continue
            seen.add(key)
            rows.append({"gl": key[0], "account": str(acct).strip(),
                         "code": code, "source": os.path.basename(p)})
            added += 1
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["gl", "account", "code", "source"])
        w.writeheader()
        w.writerows(rows)
    print(f"Corpus: {len(rows)} mappings ({added} new) -> {out_csv}")


# ----------------------------------------------------------------------------
# T12 PDF parser
# ----------------------------------------------------------------------------

@dataclass
class Line:
    kind: str                 # 'section' | 'account' | 'subtotal'
    name: str
    values: list = None       # 12 monthly floats (account/subtotal rows)
    section: str = ""         # innermost section
    side: str = ""            # 'inc' | 'exp'
    code: str = ""            # assigned charge code
    method: str = ""          # exact | section | keyword | fuzzy:0.93 | REVIEW
    review: bool = False
    below: bool = False       # below-the-line (after NOI): excluded
    xside: str = ""           # corpus code rejected for being cross-ledger
    derived: bool = False     # subtotal computed from detail, NOT printed on
                              # the statement (never used as a tie-out target)
    empty: list = None        # per-month mask: True where the SOURCE cell was
                              # blank. A blank is not a zero - the writers
                              # leave those cells empty and Total stays the
                              # sum of the months that carry an amount.
    flag_review: str = ""     # parser-supplied structural reason to REVIEW
                              # (lump revenue, duplicate hand-keyed cells)


# Sections that sit below NOI and are excluded from the operating T12
BELOW_PAT = re.compile(
    r"debt service|capital expenditure|capex|cash flow|distribution"
    r"|special project|reserve|non.?operating expense|interest|mortgage"
    r"|partnership|owner|escrow",
    re.I)


NUM = r"\(?-?[\d,]+\.\d{2}\)?"


def _f(tok):
    tok = tok.replace(",", "")
    neg = tok.startswith("(")
    v = float(tok.strip("()"))
    return -v if neg else v


# ---------------------------------------------------------------------------
# ResMan "Twelve Month Profit and Loss" PDF (validated on Lofts at Taft,
# Cornerstone Residential, 7/2026)
# ---------------------------------------------------------------------------
# Layout: 14 right-aligned numeric columns = 12 monthly "Actual" columns +
# "Adjusted Total" + "Variance" (budget variance - ignored). Labels carry the
# GL number ("4110 Gross Potential Rent") and indentation (char x0) encodes
# the section tree. ResMan emits every word as its own text-showing operation
# with no space characters, so extract_text()/extract_words() glue words
# together and - when a long label physically overlaps the first month column
# - interleave the label with that month's number
# ("4210BadDebt/Write-OffUncollec-t1a,b1l4e4R.0e0nt"). Both are recovered by
# rebuilding tokens from page.chars in content-stream order: a backwards x
# jump starts a new token, an x gap wider than a rendered space separates
# words.

RESMAN_MONTH = re.compile(r"^([A-Z][a-z]{2})[a-z]?\.?\s?(\d{4})$")
RESMAN_NUM = re.compile(r"^\(?-?[\d,]*\d\.\d{2}\)?$")
# grand rows -> the canonical names reconcile()/RawData expect
RESMAN_GRAND = [
    (re.compile(r"^total\s+(operating\s+)?income$", re.I), "Total Revenue"),
    (re.compile(r"^total\s+(operating\s+)?expenses?$", re.I),
     "Total Operating Expenses"),
    (re.compile(r"^net\s+operating\s+income$", re.I),
     "Total Net Operating Income"),
]
RESMAN_SKIP = re.compile(
    r"^(twelve month profit|.*accounting book:|printed\s?\d|account\b"
    r"|actual\b|adjusted\b|variance\b|total$|©|\(c\)\s*resman)", re.I)


def _resman_tokens(chars, gap=1.1):
    """One PDF row's chars -> [(x0, x1, text)] tokens, honouring stream
    order so an overlapping label and month value stay separate."""
    toks, cur, x0, prev = [], "", None, None
    for c in chars:
        t = c["text"]
        if t.strip() == "" and cur:
            toks.append((x0, prev, cur))
            cur, prev = "", None
            continue
        if not cur:
            cur, x0 = t, c["x0"]
        elif c["x0"] < prev - 0.5 or c["x0"] - prev > gap:
            toks.append((x0, prev, cur))
            cur, x0 = t, c["x0"]
        else:
            cur += t
        prev = c["x1"]
    if cur:
        toks.append((x0, prev, cur))
    return toks


def _resman_below(section):
    """Below-the-line test for a ResMan section name. BELOW_PAT is
    deliberately broad; ResMan's operating section '7100 Interest, Insurance
    & Taxes' (property insurance + property taxes, printed above NOI) trips
    it only on the word 'interest'."""
    if not section:
        return False
    if re.search(r"insurance|tax", section, re.I) and not re.search(
            r"debt|capital|capex|mortgage|reserve|distribution",
            section, re.I):
        return False
    return bool(BELOW_PAT.search(section))


def _resman_rows(page, tol=1.6):
    """Group a page's chars into rows (stream order preserved within a
    row), returned top-down."""
    rows = []
    for c in page.chars:
        for r in rows:
            if abs(r[0] - c["top"]) <= tol:
                r[1].append(c)
                break
        else:
            rows.append((c["top"], [c]))
    rows.sort(key=lambda r: r[0])
    return rows


def parse_t12_pdf_resman(path):
    """Return (property_name, [month_end_dates x12], [Line])."""
    lines_out, months, edges = [], [], []
    prop, banner = "", set()         # banner: page-header lines to skip
    side = "inc"
    stack = []                       # [(indent_x0, section name)]
    below = False
    bad_totals = []
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            for top, chars in _resman_rows(page):
                toks = _resman_tokens(chars)
                if not toks:
                    continue
                # ---- month header row: also defines the column edges ----
                # ResMan splits "Jul2025" into two text ops -> pair them up
                mons, k = [], 0
                while k < len(toks):
                    t = toks[k][2]
                    m = RESMAN_MONTH.match(t)
                    if m and m.group(2):
                        mons.append((toks[k][1], m))
                    elif re.fullmatch(r"[A-Z][a-z]{2,3}\.?", t) and \
                            k + 1 < len(toks) and \
                            re.fullmatch(r"\d{4}", toks[k + 1][2]):
                        mons.append((toks[k + 1][1],
                                     RESMAN_MONTH.match(
                                         t.rstrip(".") + toks[k + 1][2])))
                        k += 1
                    k += 1
                mons = [(x1, m) for x1, m in mons if m]
                if len(mons) >= 12:
                    if not months:
                        for x1, m in mons[:12]:
                            d = datetime.strptime(
                                f"{m.group(1)} {m.group(2)}", "%b %Y")
                            last = calendar.monthrange(d.year, d.month)[1]
                            months.append(d.replace(day=last))
                            edges.append(x1)
                    continue
                text = " ".join(t for _, _, t in toks).strip()
                if RESMAN_SKIP.match(text):
                    continue
                nums = [(x1, t) for _, x1, t in toks
                        if RESMAN_NUM.match(t)]
                name = " ".join(t for _, x1, t in toks
                                if not RESMAN_NUM.match(t)).strip()
                indent = toks[0][0]
                if not nums:
                    if pi == 0 and not months:
                        # page banner: property, owner LLC, report title...
                        if not prop:
                            prop = name
                        banner.add(name)
                        continue
                    if not name or name in banner:
                        continue
                    if re.fullmatch(r"INCOME", name):
                        side, stack = "inc", []
                        lines_out.append(Line("section", name, None, "",
                                              side))
                        continue
                    if re.fullmatch(r"EXPENSE", name):
                        side, stack = "exp", []
                        lines_out.append(Line("section", name, None, "",
                                              side))
                        continue
                    # section header
                    while stack and stack[-1][0] >= indent:
                        stack.pop()
                    stack.append((indent, name))
                    ln = Line("section", name, None, name, side)
                    ln.below = below or _resman_below(name)
                    lines_out.append(ln)
                    continue
                if not months or len(nums) < 12:
                    continue
                # right-align each value onto its column
                vals = [0.0] * 12
                printed = None
                for x1, t in nums:
                    j = min(range(len(edges)),
                            key=lambda k: abs(edges[k] - x1))
                    if abs(edges[j] - x1) <= 4.0:
                        vals[j] = _f(t)
                    elif x1 > edges[-1] + 4.0 and printed is None:
                        printed = _f(t)      # "Adjusted Total" column
                if printed is not None and abs(sum(vals) - printed) > 0.05:
                    bad_totals.append((name, sum(vals), printed))
                is_total = re.match(r"^(\d[\d.\-]*\s+)?total\b", name, re.I) \
                    or re.match(r"^net\s+(operating\s+)?income$", name, re.I)
                canon = None
                for pat, nm in RESMAN_GRAND:
                    if pat.fullmatch(re.sub(r"^\d[\d.\-]*\s+", "",
                                            name).strip()):
                        canon = nm
                        break
                if canon:
                    lines_out.append(Line("subtotal", canon, vals, "", side))
                    if canon == "Total Net Operating Income":
                        below = True         # everything after NOI
                    if canon == "Total Revenue":
                        side = "exp"
                    stack = []
                    continue
                if re.fullmatch(r"net income", name, re.I):
                    continue                 # derived bottom line, not used
                sect = stack[-1][1] if stack else ""
                kind = "subtotal" if is_total else "account"
                if is_total:
                    while stack and stack[-1][0] >= indent:
                        stack.pop()
                ln = Line(kind, name, vals, sect, side)
                ln.below = below or _resman_below(sect)
                lines_out.append(ln)
    if bad_totals:
        for nm, got, want in bad_totals:
            print(f"  ROW-TOTAL MISMATCH {nm}: monthly sum {got:,.2f} vs "
                  f"printed total {want:,.2f} (variance {got - want:+,.2f})")
    return prop, months, lines_out


# ---------------------------------------------------------------------------
# QuickBooks Online "Profit and Loss" PDF  (validated on Benbrook Apartments /
# 4639 Williams LLC, Jul 2025 - Jun 2026, 8/2026)
# ---------------------------------------------------------------------------
# Layout (landscape 792pt, 3 pages, page header repeated on every page):
#
#   4639 Williams LLC                 <- property / owner entity
#   Profit and Loss
#   July, 2025-June, 2026
#   Jul 2025  Aug 2025 ... Jun 2026  Total          <- 12 months + Total
#   Income
#     Fee Revenue
#       Convenience Fee   161.95  181.70 ...  1,015.15
#       ...
#     Total for Fee Revenue ...
#   ...
#   Cash Basis  Wednesday, August 05, 2026 ...  1/3 <- page footer
#
# Three structural traits drive a dedicated parser:
#
#  1. **SPARSE ROWS.** Most accounts carry values in only SOME months, and
#     QuickBooks prints nothing at all for the others - it does not print
#     0.00. Reading the tokens in order and zipping them onto months would
#     shift every gap month's money to the left, silently. Values are
#     therefore assigned POSITIONALLY: every numeric token is snapped to the
#     month column whose printed RIGHT EDGE it matches (all values are
#     right-aligned; the header token's right edge sits ~2-3pt left of the
#     data's, well inside the 45pt column pitch). Anything that cannot be
#     snapped inside COL_TOL aborts the parse rather than being dropped.
#     In-period blanks are read as 0.00 (a QBO blank means "no transactions
#     posted", which IS zero) and counted in meta["blank_cells"].
#  2. **The section tree is indentation, and "Total for X" closes a section.**
#     Top-level rows sit at x0~26, sections at ~32, accounts at ~35, but the
#     depth varies (the below-NOI "Other Income" block sits one level up), so
#     an indent STACK is used rather than fixed columns.
#  3. **Cost of Goods Sold sits between Income and Expenses.** QuickBooks
#     prints Income - COGS = Gross Profit, then Gross Profit - Expenses =
#     Net Operating Income. Its printed "Total for Expenses" therefore
#     EXCLUDES COGS. The canonical "Total Operating Expenses" handed to
#     reconcile() is the statement's printed Total for Expenses PLUS its
#     printed Total for Cost of Goods Sold - both printed rows, and the sum
#     is proved by the printed NOI (Total for Income - that sum = printed
#     NOI). It is not a derived/self-referential figure.
#
# Everything after the printed Net Operating Income row (QuickBooks' "Other
# Income"/"Other Expenses" block) is below-the-line -> Capex & Misc.

QBO_MONTH_HDR = re.compile(r"^([A-Z][a-z]{2})\s+(\d{4})$")
QBO_NUM = re.compile(r"^-?\$?\(?-?[\d,]*\d\.\d{2}\)?$")
QBO_TOTAL_FOR = re.compile(r"^Total\s+for\s+(.+)$", re.I)
# top-level computed rows QuickBooks prints with values but which are not
# accounts (and are not "Total for ..." rows)
QBO_COMPUTED = re.compile(
    r"^(gross profit|net operating income|net other income|net income"
    r"|net revenue)$", re.I)
QBO_FOOTER = re.compile(
    r"^(cash|accrual) basis\b|^\d+/\d+$|^page \d+", re.I)
QBO_COL_TOL = 12.0          # pts; column pitch is ~45


def _qbo_f(tok):
    """'1,015.15' / '$23,188.65' / '-$121,349.44' / '(500.00)' -> float."""
    t = tok.replace(",", "").replace("$", "")
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    v = float(t)
    return -v if neg else v


def _qbo_rows(page, tol=2.5):
    """Group a page's words into visual rows, top-down."""
    rows = []
    for w in page.extract_words(use_text_flow=False, keep_blank_chars=False):
        for r in rows:
            if abs(r["top"] - w["top"]) <= tol:
                r["words"].append(w)
                break
        else:
            rows.append({"top": w["top"], "words": [w]})
    for r in rows:
        r["words"].sort(key=lambda x: x["x0"])
    rows.sort(key=lambda r: r["top"])
    return rows


def is_qbo_pdf(path):
    """Sniff: a QuickBooks-Online-style P&L - 'Profit and Loss' title, a
    12-month 'Mon YYYY' header row and QuickBooks' 'Total for X' subtotals."""
    try:
        with pdfplumber.open(path) as pdf:
            txt = pdf.pages[0].extract_text() or ""
    except Exception:
        return False
    if not re.search(r"profit\s+(and|&)\s+loss", txt, re.I):
        return False
    if not re.search(r"^\s*Total\s+for\s+\S", txt, re.I | re.M):
        return False
    return len(re.findall(r"\b[A-Z][a-z]{2}\s+\d{4}\b", txt)) >= 12


def parse_t12_qbo_pdf(path):
    """Return (property, [month-end datetimes], [Line], meta)."""
    meta = {"period": None, "basis": "", "notes": [], "printed_checks": [],
            "blank_cells": [], "dropped_cols": [], "stub_last": False,
            "exempt_full_year": False}
    prop = ""
    months, edges = [], []
    tot_edge = None
    raw = []                       # (indent, name, cells[12], printed_total)
    bad_rows = []

    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            seen_hdr = False
            for row in _qbo_rows(page):
                words = row["words"]
                text = " ".join(w["text"] for w in words).strip()
                if not text or QBO_FOOTER.match(text):
                    continue
                # ---- month header row (also defines the column edges) ----
                mons = []
                k = 0
                while k + 1 < len(words):
                    a, b = words[k], words[k + 1]
                    m = QBO_MONTH_HDR.match(f"{a['text']} {b['text']}")
                    if m:
                        mons.append((b["x1"], m))
                        k += 2
                    else:
                        k += 1
                if len(mons) >= 12:
                    seen_hdr = True
                    if not months:
                        for x1, m in mons[:12]:
                            d = datetime.strptime(
                                f"{m.group(1)} {m.group(2)}", "%b %Y")
                            months.append(_month_end(d.year, d.month))
                            edges.append(x1)
                        tw = [w for w in words
                              if w["text"].lower() == "total"]
                        tot_edge = tw[-1]["x1"] if tw else None
                    continue
                if not seen_hdr:
                    # page banner: entity name / report title / period line
                    if pi == 0 and not prop and \
                            not re.search(r"profit|loss|^\d", text, re.I):
                        prop = text
                    if not meta["period"]:
                        pm = re.match(
                            r"^([A-Z][a-z]+),\s*(\d{4})\s*-\s*"
                            r"([A-Z][a-z]+),\s*(\d{4})$", text)
                        if pm:
                            meta["period"] = (
                                datetime.strptime(
                                    f"{pm.group(1)[:3]} {pm.group(2)}",
                                    "%b %Y"),
                                datetime.strptime(
                                    f"{pm.group(3)[:3]} {pm.group(4)}",
                                    "%b %Y"))
                    continue

                nums = [w for w in words if QBO_NUM.match(w["text"])]
                name = " ".join(w["text"] for w in words
                                if not QBO_NUM.match(w["text"])).strip()
                if not name:
                    continue
                indent = words[0]["x0"]
                if not nums:
                    raw.append((indent, name, None, None))
                    continue
                cells = [None] * 12
                printed = None
                for w in nums:
                    x1 = w["x1"]
                    j = min(range(12), key=lambda k: abs(edges[k] - x1))
                    if abs(edges[j] - x1) <= QBO_COL_TOL:
                        if cells[j] is not None:
                            raise ValueError(
                                f"two values land on {months[j]:%b %Y} for "
                                f"row {name!r}")
                        cells[j] = _qbo_f(w["text"])
                    elif tot_edge is not None and \
                            abs(tot_edge - x1) <= QBO_COL_TOL:
                        printed = _qbo_f(w["text"])
                    else:
                        raise ValueError(
                            f"value {w['text']!r} on row {name!r} (x1="
                            f"{x1:.1f}) does not line up with any month "
                            f"column - refusing to guess its month")
                for j, v in enumerate(cells):
                    if v is None:
                        meta["blank_cells"].append((name, f"{months[j]:%b %Y}"))
                vals = [0.0 if v is None else v for v in cells]
                if printed is not None and abs(sum(vals) - printed) > 0.005:
                    bad_rows.append((name, sum(vals), printed))
                raw.append((indent, name, vals, printed))

    if not months:
        raise ValueError("no 12-month header row found in the QuickBooks PDF")
    if re.search(r"cash basis", " ".join(
            (p.extract_text() or "") for p in pdfplumber.open(path).pages),
            re.I):
        meta["basis"] = "Cash"

    # ---- build the Line list from the indent stack ----------------------
    lines_out = []
    stack = []                    # [(indent, section name)]
    side = "inc"
    below = False
    grands = {}                   # canonical name -> monthly values
    printed_rows = {}             # "Total for X" -> values (for verification)

    for indent, name, vals, printed in raw:
        low = name.strip().lower()
        if vals is None:                       # section header
            while stack and stack[-1][0] >= indent:
                stack.pop()
            stack.append((indent, name))
            ln = Line("section", name, None, name, side)
            ln.below = below
            lines_out.append(ln)
            continue

        tf = QBO_TOTAL_FOR.match(name)
        if tf:
            sect = tf.group(1).strip()
            printed_rows[sect.lower()] = vals
            while stack and stack[-1][1].strip().lower() != sect.lower():
                stack.pop()
            parent = stack[-1][1] if stack else ""
            if stack:
                stack.pop()
            if low == "total for income":
                grands["rev"] = vals
                side = "exp"           # everything after income is expense
                lines_out.append(Line("subtotal", "Total Revenue", vals,
                                      "", "inc"))
                continue
            if low == "total for expenses":
                grands["exp"] = vals
                lines_out.append(Line("subtotal", name, vals, parent, side))
                continue
            if low == "total for cost of goods sold":
                grands["cogs"] = vals
            ln = Line("subtotal", name, vals, parent, side)
            ln.below = below
            lines_out.append(ln)
            continue

        if QBO_COMPUTED.match(low):
            if low == "net operating income":
                grands["noi"] = vals
                lines_out.append(Line("subtotal",
                                      "Total Net Operating Income", vals,
                                      "", ""))
                below = True
                stack = []
                continue
            grands[low.replace(" ", "_")] = vals
            ln = Line("subtotal", name, vals, "", side)
            ln.below = below
            lines_out.append(ln)
            continue

        sect = stack[-1][1] if stack else ""
        ln = Line("account", name, vals, sect, side)
        ln.below = below or bool(BELOW_PAT.search(sect or ""))
        lines_out.append(ln)

    # ---- canonical Total Operating Expenses = Expenses + COGS -----------
    if "exp" in grands:
        opex = list(grands["exp"])
        if "cogs" in grands:
            opex = [a + b for a, b in zip(opex, grands["cogs"])]
            meta["notes"].append(
                "QuickBooks prints Cost of Goods Sold as its own block above "
                "Gross Profit, so its printed 'Total for Expenses' "
                f"({sum(grands['exp']):,.2f}) EXCLUDES COGS "
                f"({sum(grands['cogs']):,.2f}). The canonical Total "
                f"Operating Expenses is the sum of those two printed rows, "
                f"{sum(opex):,.2f} - proved by the printed NOI.")
        # insert right after the printed "Total for Expenses" row
        idx = next((i for i, l in enumerate(lines_out)
                    if l.kind == "subtotal"
                    and l.name.strip().lower() == "total for expenses"),
                   len(lines_out))
        lines_out.insert(idx + 1,
                         Line("subtotal", "Total Operating Expenses", opex,
                              "", "exp"))
        grands["opex"] = opex

    # ---- verification ----------------------------------------------------
    n = 12
    chk = meta["printed_checks"]

    def _cmp(label, got, want):
        d = got - want
        chk.append(f"  {'OK ' if abs(d) <= 0.005 else 'VARIANCE'} {label}: "
                   f"printed {want:,.2f} vs detail {got:,.2f} "
                   f"(diff {d:+,.2f})")
        return abs(d) <= 0.005

    # (a) every "Total for X" row vs the account detail inside section X
    for i, ln in enumerate(lines_out):
        if ln.kind != "subtotal":
            continue
        tf = QBO_TOTAL_FOR.match(ln.name)
        if not tf:
            continue
        sect = tf.group(1).strip().lower()
        kids = [l for l in lines_out
                if l.kind == "account"
                and (l.section or "").strip().lower() == sect]
        if not kids:
            continue
        for k in range(n):
            got = sum(l.values[k] for l in kids)
            if abs(got - ln.values[k]) > 0.005:
                _cmp(f"{ln.name} {months[k]:%b %Y}", got, ln.values[k])
                ln.values[k] = got             # monthly detail wins
        _cmp(f"{ln.name} [{len(kids)} accounts] ANNUAL",
             sum(sum(l.values[k] for l in kids) for k in range(n)),
             sum(ln.values))

    # (b) grand rows vs the parsed detail
    def _side_sum(k, want_side, want_below=False):
        return sum(l.values[k] for l in lines_out
                   if l.kind == "account" and l.side == want_side
                   and bool(l.below) == want_below)

    if "rev" in grands:
        _cmp("Total for Income ANNUAL",
             sum(_side_sum(k, "inc") for k in range(n)), sum(grands["rev"]))
    if "opex" in grands:
        _cmp("Total Operating Expenses (Expenses + COGS) ANNUAL",
             sum(_side_sum(k, "exp") for k in range(n)), sum(grands["opex"]))
    if "gross_profit" in grands and "rev" in grands and "cogs" in grands:
        _cmp("Gross Profit ANNUAL",
             sum(grands["rev"]) - sum(grands["cogs"]),
             sum(grands["gross_profit"]))
    if "noi" in grands and "rev" in grands and "opex" in grands:
        _cmp("Net Operating Income ANNUAL",
             sum(grands["rev"]) - sum(grands["opex"]), sum(grands["noi"]))
    if "net_income" in grands and "noi" in grands and \
            "net_other_income" in grands:
        _cmp("Net Income ANNUAL",
             sum(grands["noi"]) + sum(grands["net_other_income"]),
             sum(grands["net_income"]))
    if "net_other_income" in grands:
        oi = printed_rows.get("other income")
        oe = printed_rows.get("other expenses")
        if oi is not None and oe is not None:
            _cmp("Net Other Income ANNUAL",
                 sum(oi) - sum(oe), sum(grands["net_other_income"]))

    if bad_rows:
        for nm, got, want in bad_rows:
            chk.append(f"  ROW-TOTAL MISMATCH {nm}: monthly sum {got:,.2f} "
                       f"vs printed Total {want:,.2f} "
                       f"(variance {got - want:+,.2f})")
    else:
        chk.insert(0, f"  OK  every printed row Total ties to its own "
                      f"monthly cells ({sum(1 for r in raw if r[3] is not None)}"
                      f" rows checked)")
    if meta["blank_cells"]:
        # one entry per ROW (not per cell) - a QBO statement is sparse by
        # nature and a per-cell list would be hundreds of lines of noise
        n_cells = len(meta["blank_cells"])
        agg = {}
        for nm, mo in meta["blank_cells"]:
            agg.setdefault(nm, []).append(mo)
        meta["blank_cells"] = [(nm, f"{len(v)} mo") for nm, v in agg.items()]
        meta["blank_label"] = (
            f"{n_cells} month cell(s) across {len(agg)} row(s) print blank "
            "in QuickBooks (no transactions posted) and are read as $0.00; "
            "every affected row's printed Total ties to its monthly cells")
    return prop, months, lines_out, meta


def parse_t12_pdf(path):
    """Return (property_name, [month_end_dates x12], [Line])."""
    with pdfplumber.open(path) as pdf:
        head = pdf.pages[0].extract_text() or ""
    if is_qbo_pdf(path):
        p, m, l, _meta = parse_t12_qbo_pdf(path)
        return p, m, l
    if re.search(r"ResMan", head, re.I) or \
            len(re.findall(r"\b[A-Z][a-z]{2}\d{4}\b", head)) >= 12:
        return parse_t12_pdf_resman(path)

    lines_out = []
    prop, months = "", []
    side = "inc"
    section_stack = []
    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            txt = page.extract_text() or ""
            for raw in txt.split("\n"):
                raw = raw.strip()
                if not raw:
                    continue
                m = re.findall(r"([A-Z][a-z]{2}) (\d{1,2}), (\d{4})", raw)
                if len(m) >= 12:
                    if not months:
                        months = [datetime.strptime(f"{a} {b} {c}",
                                                    "%b %d %Y")
                                  for a, b, c in m[:12]]
                    continue
                # page headers
                if raw.startswith(("Twelve Month Profit", "For Year",
                                   "Period End", "Account ", "Page ")):
                    continue
                if pi == 0 and not prop and not months and \
                        not re.search(NUM, raw) and \
                        raw not in ("Revenue",):
                    if "Profit" not in raw:
                        prop = raw
                        continue
                nums = re.findall(NUM, raw)
                name = re.sub(rf"({NUM}\s*)+$", "", raw).strip()
                if len(nums) >= 12:
                    vals = [_f(t) for t in nums[:12]]   # drop YTD col
                    if name.lower().startswith("total"):
                        lines_out.append(Line("subtotal", name, vals,
                                              section_stack[-1] if
                                              section_stack else "", side))
                    else:
                        lines_out.append(Line("account", name, vals,
                                              section_stack[-1] if
                                              section_stack else "", side))
                elif not nums:
                    if re.fullmatch(r"Operating Expenses?", name):
                        side = "exp"
                        section_stack = []
                        lines_out.append(Line("section", name, None, "",
                                              side))
                        continue
                    if name == "Revenue":
                        side = "inc"
                        lines_out.append(Line("section", name, None, "",
                                              side))
                        continue
                    if name.lower().startswith("total"):
                        # subtotal of an empty section
                        lines_out.append(Line("subtotal", name,
                                              [0.0] * 12,
                                              section_stack[-1] if
                                              section_stack else "", side))
                        if section_stack:
                            section_stack.pop()
                        continue
                    if name == prop:      # property name repeated as header
                        continue
                    if re.search(r"(curr\.?|current|reporting)\s+period",
                                 name, re.I):
                        continue          # report footnote, not a section
                    section_stack = [name]
                    lines_out.append(Line("section", name, None, name, side))

    # Everything after the printed NOI row — and anything in a below-the-line
    # section — is excluded from the operating T12.
    seen_noi = False
    for ln in lines_out:
        if seen_noi or BELOW_PAT.search(ln.section or ""):
            ln.below = True
        if ln.kind == "subtotal" and re.fullmatch(
                r"Total Net Operating Income", ln.name, re.I):
            seen_noi = True
    return prop, months, lines_out


# ----------------------------------------------------------------------------
# T12 XLSX parser — Yardi "Statement (N months)" export
# ----------------------------------------------------------------------------
# Layout (validated on The Meadows (tmtx), 7/2026):
#   row 1: property name + code, e.g. "The Meadows (tmtx)"
#   row 2: "Statement (12 months)"      row 3: "Period = Dec 2025-Jun 2026"
#   row 4: "Book = Accrual"
#   header row: month cells ("Jan 2026", "Feb 2026", ...) starting in col C.
#   body: col A = GL account number, col B = account name whose LEADING
#   SPACES encode the hierarchy. Section headers carry a GL + name and no
#   values; aggregates are "TOTAL x" / "SUB-TOTAL x" / "NET OTHER INCOME" /
#   "NET OPERATING INCOME".
#
# Two structural differences from the PM-prepared layout above drive the
# separate parser:
#   1. THERE IS NO TOTALS COLUMN. Yardi prints monthly columns only, so there
#      are no printed annual figures to check rows against. Instead every
#      aggregate row is verified against the monthly detail it aggregates,
#      column by column (stronger than the annual-only check), by the greedy
#      consumer in _yardi_verify().
#   2. The number of month columns is whatever the export contains and does
#      NOT have to agree with the "Statement (12 months)" caption or the
#      "Period =" line - both are Yardi report parameters, not data. The
#      COLUMNS are the truth. A short export is refused rather than padded
#      (see _require_full_year); never fabricate months.
#
# The account name handed downstream is "<GL> <Name>" so the corpus matcher's
# gl_number()/norm() can key on the GL number, matching the PDF parser.

def _is_yardi_xlsx(path):
    """True when the sheet looks like a Yardi statement export: a month
    header row with the months starting at col C or later, and a body whose
    col A is mostly bare GL numbers with the label in col B."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(path, read_only=True, data_only=True)
    except Exception:
        return False
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True, max_row=400)]
    hdr = None
    for i, r in enumerate(rows[:40]):
        hits = [j for j, v in enumerate(r) if _yardi_month(v)]
        if len(hits) >= 2 and min(hits) >= 2:
            hdr = i
            break
    if hdr is None:
        return False
    gl, named = 0, 0
    for r in rows[hdr + 1:]:
        a = str(r[0] or "").strip()
        b = str(r[1] or "").strip() if len(r) > 1 else ""
        if not a and not b:
            continue
        named += 1
        if re.fullmatch(r"\d[\d.\-]*", a) and b:
            gl += 1
    return named >= 5 and gl >= 0.6 * named


def _yardi_month(tok):
    """'Jan 2026' / 'Jan-2026' / 'January 2026' -> datetime, else None."""
    tok = str(tok or "").strip()
    m = re.match(r"^([A-Za-z]{3,9})\.?[-/. ]*(\d{4})$", tok)
    if not m:
        return None
    for cand in (m.group(1), m.group(1)[:3]):
        for fmt in ("%B", "%b"):
            try:
                return datetime.strptime(f"{cand.title()} {m.group(2)}",
                                         f"{fmt} %Y")
            except ValueError:
                pass
    return None


# Aggregate (subtotal) row labels. Yardi mixes "TOTAL x" / "SUB-TOTAL x" with
# unprefixed "NET OTHER INCOME" / "NET OPERATING INCOME" roll-ups; the plain
# section header "NET GROSS POTENTIAL RENT" carries no values so it can never
# be confused with these.
YARDI_AGG = re.compile(
    r"^(sub[-\s]*)?total\b"
    r"|^net\s+(other\s+income|operating\s+income)\b", re.I)
YARDI_NOI = re.compile(r"^net\s+operating\s+income\b", re.I)


def _require_full_year(months, path, allow_partial=False):
    """A T-12 must be twelve months. Report captions and period lines lie
    (Yardi happily titles a six-month export "Statement (12 months)"); the
    month columns do not. Refuse a short statement instead of padding it out.

    Exempt: sources that declare a reporting period of their own and label
    every output PARTIAL PERIOD (the Google-Sheets statement parser) - there
    the short period is the documented subject of the file, not a surprise.
    """
    if len(months) >= 12 or allow_partial:
        return
    sys.exit(
        "ERROR: this statement carries only %d month column%s (%s-%s) - "
        "a T-12 needs 12.\n"
        "       Source: %s\n"
        "       The report caption/period line is a report parameter, not "
        "data; the month COLUMNS are the truth and they are short.\n"
        "       Request a full twelve-month export from the PM rather than "
        "annualising or padding. Re-run with --allow-partial only if a "
        "partial-period workbook is genuinely wanted."
        % (len(months), "" if len(months) == 1 else "s",
           months[0].strftime("%b %Y") if months else "?",
           months[-1].strftime("%b %Y") if months else "?",
           os.path.basename(path)))


def _yardi_verify(valued):
    """Verify each aggregate against the detail it aggregates.

    `valued` is the ordered list of (name, kind, values) for every row that
    carries numbers. Yardi nests aggregates (TOTAL RENTAL INCOME rolls up
    four subtotals, which in turn roll up the account rows), so this walks
    forward keeping a list of not-yet-consumed vectors: for each aggregate
    the LONGEST contiguous run of unconsumed predecessors that sums to it is
    consumed, and the aggregate itself stays unconsumed so a parent can roll
    it up. Longest-first matters where a group sums to zero (TOTAL RENTS) -
    a shortest-first match would under-consume.

    Returns a list of (name, months_sum, aggregate_sum) for aggregates that
    could not be tied to the detail beneath them.
    """
    fails, pending = [], []          # pending: [(idx, values)]
    for idx, (name, kind, vals) in enumerate(valued):
        if kind != "subtotal":
            pending.append((idx, vals))
            continue
        if YARDI_NOI.search(name):
            continue                 # NOI is income - expense, not a sum
        hit = None
        for take in range(len(pending), 0, -1):
            run = pending[-take:]
            if all(abs(sum(v[i] for _, v in run) - vals[i]) <= 0.05
                   for i in range(len(vals))):
                hit = take
                break
        if hit is None:
            fails.append((name, sum(v for _, vv in pending for v in vv),
                          sum(vals)))
            pending = [(idx, vals)]
            continue
        pending = pending[:-hit] + [(idx, vals)]
    return fails


def parse_t12_yardi_xlsx(path, trust_monthly=False, allow_partial=False):
    """Yardi 'Statement (N months)' export -> (property, months, [Line])."""
    from openpyxl import load_workbook as _lw

    def _num(v):
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        try:
            return float(str(v).replace(",", "").replace("$", ""))
        except ValueError:
            return None

    wb = _lw(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    month_cols, months, hdr_i = [], [], None
    for i, r in enumerate(rows):
        got = [(j, _yardi_month(v)) for j, v in enumerate(r)]
        got = [(j, d) for j, d in got if d]
        if len(got) >= 2 and min(j for j, _ in got) >= 2:
            month_cols = [j for j, _ in got]
            months = [d for _, d in got]
            hdr_i = i
            break
    if hdr_i is None:
        sys.exit("ERROR: no month header row found in Yardi xlsx T12.")
    print(f"  Yardi layout: {len(months)} month column(s) "
          f"{months[0]:%b %Y}-{months[-1]:%b %Y}")
    for r in rows[:hdr_i]:
        cap = str(r[0] or "").strip()
        if re.match(r"^(period|statement)\b", cap, re.I):
            print(f"  report caption: {cap!r} (parameter, not data)")
    _require_full_year(months, path, allow_partial)

    # property: first text line, minus Yardi's "(code)" suffix
    prop = ""
    for r in rows[:hdr_i]:
        a = str(r[0] or "").strip()
        if a and not re.match(r"^(statement|period|book)\b", a, re.I):
            prop = a
            break
    prop = re.sub(r"\s*\(.*$", "", prop).strip()

    lines_out, side, section = [], "inc", ""
    valued = []
    for r in rows[hdr_i + 1:]:
        gl = str(r[0] or "").strip()
        raw = str(r[1] or "") if len(r) > 1 else ""
        name = re.sub(r"\s+", " ", raw).strip()
        if not name:
            continue
        vals = [_num(r[j]) if j < len(r) else None for j in month_cols]
        has_vals = any(v is not None for v in vals)
        mvals = [v or 0.0 for v in vals]
        is_agg = bool(YARDI_AGG.search(name))

        if not has_vals:
            section = name
            lines_out.append(Line("section", name, None, section, side))
            continue
        if is_agg:
            valued.append((name, "subtotal", mvals))
            lines_out.append(Line("subtotal", name, mvals, section, side))
            # everything after the printed revenue grand total is expense
            if re.match(r"^total\s+(operating\s+)?(income|revenue)s?$",
                        name, re.I):
                side = "exp"
            continue
        valued.append((name, "account", mvals))
        disp = name.title() if name.isupper() else name
        lines_out.append(Line("account", f"{gl} {disp}".strip(),
                              mvals, section, side))

    # below-the-line: anything after the printed NOI row (debt service,
    # capex, reserves, distributions) is excluded from operations
    seen_noi = False
    for ln in lines_out:
        if seen_noi or BELOW_PAT.search(ln.section or ""):
            ln.below = True
        if ln.kind == "subtotal" and YARDI_NOI.search(ln.name):
            seen_noi = True

    # structural check: every aggregate vs the detail beneath it
    for nm, got, want in _yardi_verify(valued):
        print(f"  SUBTOTAL-ROW MISMATCH {nm}: detail {got:,.2f} "
              f"vs printed {want:,.2f} (variance {got - want:+,.2f})")

    def grand(pat):
        for ln in lines_out:
            if ln.kind == "subtotal" and re.search(pat, ln.name, re.I):
                return sum(ln.values)
        return None

    rev = grand(r"^total\s+income$|^total\s+revenue$")
    exp = grand(r"^total\s+operating\s+expenses?$")
    noi = grand(r"^net\s+operating\s+income")

    # coded detail vs the statement's own grand rows, per side
    det = {"inc": 0.0, "exp": 0.0}
    for ln in lines_out:
        if ln.kind == "account" and not ln.below:
            det[ln.side] = det.get(ln.side, 0.0) + sum(ln.values)
    inc_var = det["inc"] - rev if rev is not None else 0.0
    exp_var = det["exp"] - exp if exp is not None else 0.0
    bad = False
    for label, resid, printed, detail in (
            ("Total Income", inc_var, rev, det["inc"]),
            ("Total Operating Expenses", exp_var, exp, det["exp"])):
        if abs(resid) > 0.05:
            bad = True
            print(f"  GRAND-ROW MISMATCH {label}: monthly detail "
                  f"{detail:,.2f} vs printed grand row {printed:,.2f} "
                  f"(variance {resid:+,.2f})")
    if bad:
        if not trust_monthly:
            sys.exit("ERROR: Yardi grand rows do not tie to the monthly "
                     "detail - aborting. (re-run with --trust-monthly to "
                     "treat the monthly detail as source of truth)")
        print("  --trust-monthly: using monthly detail; grand totals "
              f"adjusted by income {inc_var:+,.2f} / expense "
              f"{exp_var:+,.2f} for reconciliation.")
        rev, exp = det["inc"], det["exp"]
        if noi is not None:
            noi += inc_var - exp_var

    canon = []
    if rev is not None:
        canon.append(("Total Revenue", rev))
    if exp is not None:
        canon.append(("Total Operating Expenses", exp))
    if noi is not None:
        canon.append(("Total Net Operating Income", noi))
    lines_out = [l for l in lines_out if not (
        l.kind == "subtotal" and re.search(
            r"^total\s+income$|^total\s+operating\s+expenses?$"
            r"|^total\s+expense$|^net\s+operating\s+income", l.name, re.I))]
    for nm, a in canon:
        lines_out.append(
            Line("subtotal", nm, [a] + [0.0] * (len(months) - 1), "", ""))
    return prop, months, lines_out


# ----------------------------------------------------------------------------
# Google-Sheets-printed income statement PDF (partial-period capable)
# ----------------------------------------------------------------------------
# Layout (validated on The Gardens Apartments, 7/2026):
#   row 1  "<Property> Income Statement"
#   row 2  "01/01/2026 - 7/25/2026, By Month, Cash Basis"
#   row 3  "<Property>"
#   row 4  "Account"  +  month headers "01-2026" ... "12-2026"
#   body   "Income" (section head that ALSO carries the revenue values),
#          "Expense" (bare section head), account rows, "Total Expenses"
# There is no Total column, no printed Total Revenue, and no printed NOI.
# Columns are right-aligned to the month header, so values are assigned
# positionally (nearest right edge) - a row with a gap month (e.g. no
# management fee in July) keeps its remaining values in the correct columns
# instead of shifting left, which is what a naive text parse would do.
#
# PARTIAL PERIODS: the reporting period in the subtitle is authoritative.
# Month columns outside it are dropped (after asserting they are empty/zero)
# and `months` is returned with the ACTUAL number of reported months - never
# padded to 12 and never annualized.

GS_MONTH_HDR = re.compile(r"^(0[1-9]|1[0-2])-(\d{4})$")
GS_PERIOD = re.compile(
    r"(\d{1,2})/(\d{1,2})/(\d{4})\s*[-–]\s*(\d{1,2})/(\d{1,2})/(\d{4})")


def _month_end(y, m):
    return datetime(y + (m == 12), (m % 12) + 1, 1) - timedelta(days=1)


def _gs_rows(page, tol=2.0):
    """Group a page's words into visual rows keyed by baseline top."""
    rows = []
    for w in page.extract_words(use_text_flow=False, keep_blank_chars=False):
        for r in rows:
            if abs(r["top"] - w["top"]) <= tol:
                r["words"].append(w)
                break
        else:
            rows.append({"top": w["top"], "words": [w]})
    for r in rows:
        r["words"].sort(key=lambda x: x["x0"])
    rows.sort(key=lambda r: r["top"])
    return rows


def is_gsheet_pdf(path):
    """Cheap sniff: a header row of >=2 'MM-YYYY' month tokens."""
    try:
        with pdfplumber.open(path) as pdf:
            txt = pdf.pages[0].extract_text() or ""
    except Exception:
        return False
    for raw in txt.split("\n"):
        toks = raw.split()
        if sum(1 for t in toks if GS_MONTH_HDR.match(t)) >= 2:
            return True
    return False


def parse_t12_gsheet_pdf(path):
    """Return (property, [month datetimes], [Line], meta).

    meta = {"period": (start, end), "n_months": N, "stub_last": bool,
            "basis": "Cash"|"Accrual"|"", "notes": [...],
            "dropped_cols": [...], "blank_cells": [(account, month), ...]}
    """
    meta = {"period": None, "stub_last": False, "basis": "", "notes": [],
            "dropped_cols": [], "blank_cells": [], "printed_checks": []}
    prop = ""
    hdr_months, hdr_x = [], []
    body = []

    with pdfplumber.open(path) as pdf:
        for pi, page in enumerate(pdf.pages):
            for row in _gs_rows(page):
                words = row["words"]
                text = " ".join(w["text"] for w in words)
                mtoks = [w for w in words if GS_MONTH_HDR.match(w["text"])]
                if len(mtoks) >= 2:
                    if not hdr_months:
                        for w in mtoks:
                            mm, yy = GS_MONTH_HDR.match(w["text"]).groups()
                            hdr_months.append(_month_end(int(yy), int(mm)))
                            hdr_x.append(w["x1"])
                    continue
                if not hdr_months:
                    # pre-header: title / period line / property name
                    pm = GS_PERIOD.search(text)
                    if pm:
                        a = datetime(int(pm.group(3)), int(pm.group(1)),
                                     int(pm.group(2)))
                        b = datetime(int(pm.group(6)), int(pm.group(4)),
                                     int(pm.group(5)))
                        meta["period"] = (a, b)
                        if re.search(r"cash basis", text, re.I):
                            meta["basis"] = "Cash"
                        elif re.search(r"accrual", text, re.I):
                            meta["basis"] = "Accrual"
                        continue
                    if re.search(r"income statement|profit|loss", text, re.I):
                        if not prop:
                            prop = re.sub(
                                r"\s*(income statement|profit (and|&) loss"
                                r"|statement of operations).*$", "", text,
                                flags=re.I).strip()
                        continue
                    if text.strip().lower() == "account":
                        continue
                    if not re.search(NUM, text) and text.strip():
                        prop = text.strip()          # repeated property row
                    continue
                body.append((pi, words))

    if not hdr_months:
        raise ValueError("no 'MM-YYYY' month header row found")

    ncols = len(hdr_months)

    def col_of(w):
        """Nearest month column by right edge (values are right-aligned)."""
        d = [abs(w["x1"] - x) for x in hdr_x]
        j = d.index(min(d))
        return j if d[j] <= 14 else None

    # ---- read body rows into a raw grid -------------------------------
    label_cut = min(hdr_x) - 45          # anything left of col 1 is the label
    raw = []
    for pi, words in body:
        name = " ".join(w["text"] for w in words
                        if w["x0"] < label_cut
                        and not re.fullmatch(NUM, w["text"])).strip()
        cells = [None] * ncols
        for w in words:
            if not re.fullmatch(NUM, w["text"].replace("$", "")):
                continue
            j = col_of(w)
            if j is None:
                continue
            cells[j] = _f(w["text"].replace("$", ""))
        if not name and all(c is None for c in cells):
            continue
        raw.append((name, cells))

    # ---- decide the reported month window ------------------------------
    if meta["period"]:
        a, b = meta["period"]
        keep = [j for j, d in enumerate(hdr_months)
                if (d.year, d.month) >= (a.year, a.month)
                and (d.year, d.month) <= (b.year, b.month)]
        meta["stub_last"] = (b.date() != _month_end(b.year, b.month).date())
    else:
        keep = [j for j in range(ncols)
                if any(c[j] is not None for _, c in raw)]
    if not keep:
        raise ValueError("no month columns fall inside the reporting period")
    keep = list(range(min(keep), max(keep) + 1))

    # columns outside the period must be empty or zero - never silently drop
    for j in range(ncols):
        if j in keep:
            continue
        junk = [(nm, c[j]) for nm, c in raw if c[j] not in (None, 0.0)]
        vals = [(nm, c[j]) for nm, c in raw if c[j] is not None]
        if junk:
            raise ValueError(
                f"non-zero values in {hdr_months[j]:%m-%Y}, outside the "
                f"reported period: {junk}")
        if vals:
            meta["dropped_cols"].append(
                f"{hdr_months[j]:%m-%Y}: " +
                ", ".join(f"{nm} $0.00" for nm, _ in vals))

    months = [hdr_months[j] for j in keep]

    # ---- build Line objects --------------------------------------------
    lines_out, side, section = [], "inc", ""
    for name, cells in raw:
        vals = [cells[j] for j in keep]
        has = any(v is not None for v in vals)
        low = name.strip().lower()

        if re.fullmatch(r"(total\s+)?(income|revenues?|revenue)", low):
            side, section = "inc", name.strip()
        elif re.fullmatch(r"(operating\s+)?expenses?", low):
            side, section = "exp", name.strip()

        if not has:
            lines_out.append(Line("section", name, None, section, side))
            continue
        for j, v in enumerate(vals):
            if v is None:
                meta["blank_cells"].append((name, f"{months[j]:%m-%Y}"))
        vals = [0.0 if v is None else v for v in vals]
        # A row like "Income" that is both the section head AND the only
        # line carrying that side's values is emitted once, as an account -
        # a duplicate section marker would just read as a phantom row.
        kind = "subtotal" if low.startswith("total") else "account"
        lines_out.append(Line(kind, name, vals, section, side))

    # ---- printed subtotal rows vs the monthly detail --------------------
    # Dmytro's standing rule: monthly detail wins; variances are recorded.
    n = len(months)
    for ln in lines_out:
        if ln.kind != "subtotal":
            continue
        want = None
        if re.fullmatch(r"total\s+(operating\s+)?expenses?", ln.name,
                        re.I):
            want = [sum(l.values[i] for l in lines_out
                        if l.kind == "account" and l.side == "exp")
                    for i in range(n)]
        elif re.fullmatch(r"total\s+(income|revenues?)", ln.name, re.I):
            want = [sum(l.values[i] for l in lines_out
                        if l.kind == "account" and l.side == "inc")
                    for i in range(n)]
        if want is None:
            continue
        for i in range(n):
            d = ln.values[i] - want[i]
            tag = "OK " if abs(d) <= 0.005 else "VARIANCE"
            meta["printed_checks"].append(
                f"  {tag} {ln.name} {months[i]:%b %Y}: printed "
                f"{ln.values[i]:,.2f} vs detail {want[i]:,.2f} "
                f"(diff {d:+,.2f})")
            if abs(d) > 0.005:
                ln.values[i] = want[i]          # monthly detail wins
        dt = sum(ln.values) - sum(want)
        meta["printed_checks"].append(
            f"  {'OK ' if abs(dt) <= 0.005 else 'VARIANCE'} {ln.name} "
            f"PERIOD TOTAL: printed {sum(ln.values):,.2f} vs detail "
            f"{sum(want):,.2f} (diff {dt:+,.2f})")

    # ---- derive the control subtotals the statement does not print ------
    have = {re.sub(r"\s+", " ", l.name.strip().lower())
            for l in lines_out if l.kind == "subtotal"}
    inc = [sum(l.values[i] for l in lines_out
               if l.kind == "account" and l.side == "inc") for i in range(n)]
    exp = [sum(l.values[i] for l in lines_out
               if l.kind == "account" and l.side == "exp") for i in range(n)]
    if not any(re.fullmatch(r"total\s+(income|revenues?)", h) for h in have):
        lines_out.insert(
            [i for i, l in enumerate(lines_out)
             if l.kind == "section" and l.side == "exp"][0]
            if any(l.kind == "section" and l.side == "exp"
                   for l in lines_out) else len(lines_out),
            Line("subtotal", "Total Revenue", inc, "", "inc", derived=True))
        meta["notes"].append(
            "Statement prints no Total Revenue row - 'Total Revenue' is "
            "derived from the parsed revenue detail (not a tie-out target).")
    if not any(re.fullmatch(r"total\s+(operating\s+)?expenses?", h)
               for h in have):
        lines_out.append(Line("subtotal", "Total Operating Expenses", exp,
                              "", "exp", derived=True))
        meta["notes"].append(
            "Statement prints no Total Operating Expenses row - derived "
            "from detail.")
    if not any("net operating income" in h for h in have):
        lines_out.append(Line("subtotal", "Total Net Operating Income",
                              [a - b for a, b in zip(inc, exp)], "", "",
                              derived=True))
        meta["notes"].append(
            "Statement prints no NOI row - NOI is derived as "
            "Total Revenue less Total Operating Expenses.")

    meta["n_months"] = n
    return prop, months, lines_out, meta


# ----------------------------------------------------------------------------
# T12 XLSX parser (owner/PM-prepared income statement workbooks)
# ----------------------------------------------------------------------------
# Layout (validated on Clark Duplexes & Pecan Townhomes, 5/2026):
#   row 1: "Income Statement - T12..."   row 2: owner LLC
#   row 3: property name                 row 4: "Period Range: ..."
#   months header row: 12 cells like "June-2025", "Sept-2025", "April-2026"
#   a TOTALS column holds each row's printed annual total
#   body: ALL-CAPS section headers (no values), indented account rows,
#   "Total ..." subtotal rows (annual-only for the grand rows), single-line
#   ALL-CAPS accounts (INSURANCE / TAXES), "NOI - Net Operating Income".
# Every printed annual total is checked against the row's monthly sum, and
# canonical Total Revenue / Total Operating Expenses / NOI subtotal lines are
# emitted from the statement's own printed grand totals so reconcile() and
# the RawData sum-check work unchanged.

# Grand-total finalization shared by the xlsx statement parsers.
#
# `row_fails` is [(name, months_sum, printed_annual, kind, side)] for every row
# whose printed annual total disagreed with its own monthly cells. The account
# variances are folded into the printed grand totals so reconcile() compares
# the coded output against a CORRECTED statement total instead of a widened
# tolerance; the grand rows themselves are then re-checked against the monthly
# detail (a grand row can be hardcoded / double-counted even when every account
# row ties) and the residual reported. Monthly detail always wins - house rule.
#
# `pats` names the grand rows in the dialect being parsed:
#   rev/exp/noi  regexes that FIND the printed grand row (first match wins)
#   drop         regex for the raw grand rows to remove before the canonical
#                "Total Revenue"/"Total Operating Expenses"/"Total Net
#                Operating Income" lines reconcile() and RawData expect are
#                appended (leaving both would double-count in RawData).

_GRAND_PATS_DEFAULT = {
    "rev": r"total operating income|total revenue",
    "exp": r"total operating expense",
    "noi": r"\bNOI\b|net operating income",
    "drop": r"total operating income|total operating expense"
            r"|net operating income",
}

# RealPage/OneSite prints "Total Income" / "Total Operating Expenses" /
# "Net Operating Income (Loss)", and nests same-suffix rows above them
# ("Total Rental Income", "Total Other Income", "Total Other Non-Operating
# Expenses"), so every pattern is anchored.
_GRAND_PATS_ONESITE = {
    "rev": r"^total\s+(operating\s+)?(income|revenue)s?$",
    "exp": r"^total\s+(operating\s+)?expenses?$",
    "noi": r"^net\s+operating\s+income",
    "drop": r"^total\s+(operating\s+)?(income|revenue)s?$"
            r"|^total\s+(operating\s+)?expenses?$"
            r"|^net\s+operating\s+income",
}


def _xlsx_grand_finalize(lines_out, row_fails, trust_monthly, pats):
    """Report row-total variances, adjust and verify the printed grand rows,
    then swap them for the canonical Total Revenue / Total Operating Expenses
    / Total Net Operating Income lines. Returns the finished line list."""
    inc_var = exp_var = 0.0
    for n, got, want, kind, sd in row_fails:
        tag = "ROW-CHECK MISMATCH" if kind == "account" else \
              "SUBTOTAL-ROW MISMATCH"
        print(f"  {tag} {n}: months sum {got:,.2f} "
              f"vs printed TOTALS {want:,.2f} "
              f"(variance {got - want:+,.2f})")
        if kind != "account":
            continue          # printed subtotals are never used downstream
        if sd == "inc":
            inc_var += got - want
        else:
            exp_var += got - want

    # canonical grand subtotals from the statement's own printed numbers
    def grand_annual(pat):
        for ln in lines_out:
            if ln.kind == "subtotal" and re.search(pat, ln.name, re.I):
                a = getattr(ln, "values_annual", None)
                return a if a is not None else sum(ln.values)
        return None

    rev = grand_annual(pats["rev"])
    exp = grand_annual(pats["exp"])
    noi = grand_annual(pats["noi"])
    if rev is not None:
        rev += inc_var
    if exp is not None:
        exp += exp_var

    # A grand row can itself be internally wrong (hardcoded / double-counting
    # monthly cells) even when every account row ties. Compare the adjusted
    # printed grand against the monthly detail and report the residual.
    det = {"inc": 0.0, "exp": 0.0}
    for ln in lines_out:
        if ln.kind == "account" and not ln.below:
            det[ln.side] = det.get(ln.side, 0.0) + sum(ln.values)
    resid_inc = det["inc"] - rev if rev is not None else 0.0
    resid_exp = det["exp"] - exp if exp is not None else 0.0
    for label, resid, printed, detail in (
            ("Total Operating Income", resid_inc, rev, det["inc"]),
            ("Total Operating Expense", resid_exp, exp, det["exp"])):
        if abs(resid) > 0.05:
            print(f"  GRAND-ROW MISMATCH {label}: monthly detail "
                  f"{detail:,.2f} vs printed grand row "
                  f"{printed:,.2f} (after row-check adjustments) "
                  f"(variance {resid:+,.2f})")
    rev = det["inc"] if rev is not None else None
    exp = det["exp"] if exp is not None else None
    inc_var += resid_inc
    exp_var += resid_exp
    if noi is not None:
        noi += inc_var - exp_var

    if row_fails or abs(resid_inc) > 0.05 or abs(resid_exp) > 0.05:
        if not trust_monthly:
            sys.exit("ERROR: xlsx row totals do not tie out - aborting. "
                     "(re-run with --trust-monthly to treat the monthly "
                     "detail as source of truth)")
        print("  --trust-monthly: using monthly detail; grand totals "
              f"adjusted by income {inc_var:+,.2f} / expense "
              f"{exp_var:+,.2f} for reconciliation.")

    canon = []
    if rev is not None:
        canon.append(("Total Revenue", rev))
    if exp is not None:
        canon.append(("Total Operating Expenses", exp))
    if noi is not None:
        canon.append(("Total Net Operating Income", noi))
    # drop the raw grand rows (names don't match reconcile patterns anyway)
    lines_out = [l for l in lines_out if not (
        l.kind == "subtotal" and re.search(pats["drop"], l.name, re.I))]
    for nm, a in canon:
        lines_out.append(Line("subtotal", nm, [a] + [0.0] * 11, "", ""))
    return lines_out


def parse_t12_xlsx(path, trust_monthly=False, allow_partial=False):
    """Return (property_name, [month datetimes x12], [Line]).

    trust_monthly: when a row's printed annual TOTALS disagrees with the sum
    of its monthly cells, treat the monthly detail as source of truth - warn
    loudly, keep going, and adjust the printed grand totals by the known
    variance so reconciliation checks the coded output against the corrected
    statement totals (never silently widened tolerances).
    """
    from openpyxl import load_workbook as _lw

    # Other xlsx dialects have a different geometry entirely (Yardi: GL number
    # in col A, indented name in col B, no TOTALS column; RealPage/OneSite:
    # date column headers and an indentation tree in col A). Dispatch to their
    # own parsers rather than bending this one out of shape; this generic
    # owner/PM-prepared layout stays the fallback.
    for detect, parser in T12_XLSX_PARSERS:
        if detect(path):
            return parser(path, trust_monthly=trust_monthly,
                          allow_partial=allow_partial)

    def _month(tok):
        tok = str(tok or "").strip()
        # accepts "June-2025", "Sept-2025", "May 2025", "Apr. 2025",
        # "Sept. 2025" (abbreviations with a trailing period + space)
        m = re.match(r"^([A-Za-z]{3,9})\.?[-/. ]*(\d{4})$", tok)
        if not m:
            return None
        for cand in (m.group(1), m.group(1)[:3]):
            for fmt in ("%B", "%b"):
                try:
                    return datetime.strptime(
                        f"{cand.title()} {m.group(2)}", f"{fmt} %Y")
                except ValueError:
                    pass
        return None

    def _num(v):
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return float(str(v).replace(",", "").replace("$", ""))

    wb = _lw(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # months header + TOTALS column
    month_cols, months, hdr_i = [], [], None
    for i, r in enumerate(rows):
        parsed = [(j, _month(v)) for j, v in enumerate(r)]
        got = [(j, d) for j, d in parsed if d]
        if len(got) >= 12:
            month_cols = [j for j, _ in got[:12]]
            months = [d for _, d in got[:12]]
            hdr_i = i
            break
    if hdr_i is None:
        sys.exit("ERROR: no 12-month header row found in xlsx T12.")
    tot_col = None
    for r in rows[:hdr_i + 3]:
        for j, v in enumerate(r):
            # column A holds "Total ..." row labels - only look right of it
            if j and str(v or "").strip().upper() in ("TOTALS", "TOTAL"):
                tot_col = j
    prop = ""
    for r in rows[:hdr_i]:
        a = str(r[0] or "").strip()
        if a.lower().startswith("period range"):
            break
        if a and "income statement" not in a.lower():
            prop = a          # last text line before Period Range wins
    prop = re.sub(r"\s*\(.*$", "", prop).strip()

    # Names that appear as "Total <X>" rows: a row whose own label is <X> is a
    # section header even if the sheet leaves stray zeros on it (seen on
    # MWhite "CONTRACT EXPENSES", 4/2026).
    total_tails = set()
    for r in rows[hdr_i + 1:]:
        nm = re.sub(r"\s+", " ", str(r[0] or "")).strip()
        m = re.match(r"^total\s+(.+)$", nm, re.I)
        if m:
            total_tails.add(m.group(1).strip().lower())

    lines_out, side, section = [], "inc", ""
    row_fails = []
    for r in rows[hdr_i + 1:]:
        name = re.sub(r"\s+", " ", str(r[0] or "")).strip()
        if not name or name.lower().startswith("account name"):
            continue
        vals = [_num(r[j]) for j in month_cols]
        annual = _num(r[tot_col]) if tot_col is not None else None
        has_vals = any(v is not None for v in vals)
        is_total = bool(re.match(r"^total\b", name, re.I))
        is_noi = bool(re.search(r"\bNOI\b|net operating income", name, re.I))
        # a header carrying only zeros still owns a "Total <name>" row
        if has_vals and not is_total and not is_noi and \
                name.lower() in total_tails and \
                not any(v for v in vals) and not annual:
            has_vals = False

        if is_total or is_noi:
            mvals = [v or 0.0 for v in vals]
            if has_vals and annual is not None and \
                    abs(sum(mvals) - annual) > 0.05:
                row_fails.append((name, sum(mvals), annual, "subtotal", side))
            lines_out.append(Line("subtotal", name, mvals, section, side))
            lines_out[-1].values_annual = annual
            # after the printed revenue grand total everything is expense
            if re.match(r"^total\s+(operating\s+)?(income|revenue)\b",
                        name, re.I):
                side = "exp"
            continue
        if not has_vals:
            # section header. "Operating Income & Expense" style wrappers name
            # both sides - only flip on an unambiguous one.
            section = name
            if re.fullmatch(r"(operating\s+)?(income|revenue)s?", name, re.I):
                side = "inc"
            elif re.search(r"expense", name, re.I) and \
                    not re.search(r"income|revenue", name, re.I):
                side = "exp"
            lines_out.append(Line("section", name, None, section, side))
            continue
        # account row (incl. single-line ALL-CAPS accounts: their own section)
        sec = name if name.isupper() else section
        if side == "inc" and re.search(r"expense|insurance|tax", sec, re.I):
            side = "exp"
        if name.isupper() and re.search(r"insurance|tax", name, re.I):
            side = "exp"
        mvals = [v or 0.0 for v in vals]
        if annual is not None and abs(sum(mvals) - annual) > 0.05:
            row_fails.append((name, sum(mvals), annual, "account", side))
        lines_out.append(Line("account", name.title() if name.isupper()
                              else name, mvals, sec, side))

    # below-the-line: anything after the printed NOI row, or in a
    # debt-service/capex/reserve section, is excluded from operations
    seen_noi = False
    for ln in lines_out:
        if seen_noi or BELOW_PAT.search(ln.section or ""):
            ln.below = True
        if ln.kind == "subtotal" and re.search(
                r"\bNOI\b|net operating income", ln.name, re.I):
            seen_noi = True

    return prop, months, _xlsx_grand_finalize(
        lines_out, row_fails, trust_monthly, _GRAND_PATS_DEFAULT)


# ----------------------------------------------------------------------------
# RealPage / OneSite "Twelve Month Trailing Income Statement" XLSX
# ----------------------------------------------------------------------------
# Validated on Synott Square (location 04-356, ACCRUAL book, 6/2026).
#
#   row 1   property name              row 2  report title
#   row 3   "June 30, 2026"            rows 4-6  Reporting Book / As of Date /
#                                      Location ("04-356--Synott Square")
#   header  twelve period-END dates ("07/31/2025" ... "06/30/2026") in cols
#           B..M plus a "Twelve Month" / "Total" column (col N). The row under
#           the dates repeats "Actual" per month.
#   body    the whole statement tree lives in COLUMN A as leading-space
#           indentation (2 spaces per level); cols B..N of a section-header row
#           carry that same whitespace padding instead of numbers, which is how
#           headers are told from data rows. Accounts are "<GL> - <Name>"
#           ("6000 - Rental Income"). Roll-ups are usually "Total <section>",
#           but OneSite also repeats the header's own name for the roll-up
#           ("Net Rental Income", "Non-Operating Expenses") - so any valued
#           non-account row is a subtotal whatever it is called.
#   footer  "Created on: ..." at indent 0.
#
# Because the indentation is explicit there is no need to guess run lengths
# (the Yardi parser's problem): every roll-up is verified against exactly the
# not-yet-consumed rows indented deeper than it. Two rows are DIFFERENCES, not
# sums - "Net Operating Income (Loss)" (income less expenses) and "Total
# Current Net Income" (NOI less non-operating) - so a net-named roll-up that
# fails the sum test is retried as first-child-less-the-rest.
#
# Below the printed NOI: Owner Expense, Debt Services, Capital General and
# Other Non-Operating Expenses -> Capex & Misc workbook, excluded from ops.

ONESITE_ACCT = re.compile(r"^\d{3,6}\s*[-–]\s*\S")
ONESITE_NET = re.compile(r"net\s+operating\s+income|net\s+income", re.I)
ONESITE_SKIP = re.compile(r"^(created on|printed on|page \d)", re.I)


def _onesite_month(v):
    """A OneSite column header: a period-END date, either a real date cell or
    'MM/DD/YYYY' text -> datetime at the first of that month."""
    if hasattr(v, "year") and hasattr(v, "month") and not isinstance(v, str):
        try:
            return datetime(v.year, v.month, 1)
        except (TypeError, ValueError):
            return None
    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", str(v or "").strip())
    if not m:
        return None
    mo, dd, yy = (int(x) for x in m.groups())
    if not (1 <= mo <= 12 and 1 <= dd <= 31 and 1900 <= yy <= 2999):
        return None
    return datetime(yy, mo, 1)


def _is_onesite_xlsx(path):
    """True when the sheet looks like a OneSite trailing income statement: a
    header row of MM/DD/YYYY period-end dates right of column A, and a body of
    INDENTED "<GL> - <Name>" account rows in column A."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(path, read_only=True, data_only=True)
    except Exception:
        return False
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True, max_row=400)]
    hdr = None
    for i, r in enumerate(rows[:40]):
        if len([j for j, v in enumerate(r) if j and _onesite_month(v)]) >= 2:
            hdr = i
            break
    if hdr is None:
        return False
    gl = 0
    for r in rows[hdr + 1:]:
        a = str(r[0] or "")
        if a.startswith(" ") and ONESITE_ACCT.match(a.strip()):
            gl += 1
    return gl >= 5


def parse_t12_onesite_xlsx(path, trust_monthly=False, allow_partial=False):
    """RealPage/OneSite trailing income statement -> (property, months, [Line])."""
    from openpyxl import load_workbook as _lw

    def _num(v):
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip().replace(",", "").replace("$", "")
            if not s:
                return None
            neg = s.startswith("(") and s.endswith(")")
            try:
                f = float(s.strip("()"))
            except ValueError:
                return None
            return -f if neg else f
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    wb = _lw(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    hdr_i, month_cols, months = None, [], []
    for i, r in enumerate(rows[:40]):
        got = [(j, _onesite_month(v)) for j, v in enumerate(r) if j]
        got = [(j, d) for j, d in got if d]
        if len(got) >= 2:
            hdr_i, month_cols, months = i, [j for j, _ in got], \
                [d for _, d in got]
            break
    if hdr_i is None:
        sys.exit("ERROR: no dated month header row found in OneSite xlsx T12.")

    # the Total column sits right of the last month; its caption is split over
    # the header rows ("Twelve Month" / "Total")
    tot_col, width = None, max(len(r) for r in rows[hdr_i:hdr_i + 3])
    for j in range(max(month_cols) + 1, width):
        cap = " ".join(str((rows[k][j] if j < len(rows[k]) else "") or "")
                       for k in range(hdr_i, min(hdr_i + 3, len(rows))))
        if "total" in cap.lower():
            tot_col = j
            break

    prop = str(rows[0][0] or "").strip() if rows else ""
    if not prop:
        for r in rows[:hdr_i]:
            if str(r[0] or "").strip().lower().startswith("location"):
                loc = str((r[1] if len(r) > 1 else "") or "").strip()
                prop = re.sub(r"^[\d\-]+--", "", loc).strip()
    prop = re.sub(r"\s*\(.*$", "", prop).strip()

    lines_out, stack, pending = [], [], []
    row_fails, agg_fails, n_checked = [], [], 0
    side = "inc"
    for r in rows[hdr_i + 1:]:
        raw = r[0] if r else None
        if raw is None or not str(raw).strip():
            continue
        s = str(raw)
        indent = len(s) - len(s.lstrip(" \t"))
        name = re.sub(r"\s+", " ", s.strip())
        if ONESITE_SKIP.match(name):
            continue
        vals = [_num(r[j]) if j < len(r) else None for j in month_cols]
        annual = (_num(r[tot_col])
                  if tot_col is not None and tot_col < len(r) else None)

        if all(v is None for v in vals):
            # section header (its month cells hold OneSite's indent padding)
            while stack and stack[-1][0] >= indent:
                stack.pop()
            stack.append((indent, name))
            lines_out.append(Line("section", name, None, name, side))
            continue

        mvals = [v or 0.0 for v in vals]
        is_acct = bool(ONESITE_ACCT.match(name))
        if annual is not None:
            n_checked += 1
            if abs(sum(mvals) - annual) > 0.05:
                row_fails.append((name, sum(mvals), annual,
                                  "account" if is_acct else "subtotal", side))
        if is_acct:
            lines_out.append(Line("account", name,
                                  mvals, stack[-1][1] if stack else "", side))
            lines_out[-1].values_annual = annual
            pending.append((indent, name, mvals))
            continue

        # roll-up row: consumes every not-yet-consumed row indented deeper
        while stack and stack[-1][0] >= indent:
            stack.pop()
        k = len(pending)
        while k and pending[k - 1][0] > indent:
            k -= 1
        take = pending[k:]
        if take:
            ssum = [sum(p[2][x] for p in take) for x in range(len(mvals))]
            ok = all(abs(a - b) <= 0.05 for a, b in zip(ssum, mvals))
            if not ok and ONESITE_NET.search(name):
                dif = [take[0][2][x] - sum(p[2][x] for p in take[1:])
                       for x in range(len(mvals))]
                ok = all(abs(a - b) <= 0.05 for a, b in zip(dif, mvals))
            if not ok:
                agg_fails.append((name, sum(ssum), sum(mvals),
                                  [p[1] for p in take]))
        lines_out.append(Line("subtotal", name, mvals,
                              stack[-1][1] if stack else "", side))
        lines_out[-1].values_annual = annual
        pending = pending[:k] + [(indent, name, mvals)]
        if re.match(r"^total\s+(operating\s+)?(income|revenue)s?$",
                    name, re.I):
            side = "exp"          # everything after the revenue grand row

    # below-the-line: anything after the printed NOI row, or in a
    # debt-service/capex/reserve section, is excluded from operations
    seen_noi = False
    for ln in lines_out:
        if seen_noi or BELOW_PAT.search(ln.section or ""):
            ln.below = True
        if ln.kind == "subtotal" and re.search(
                r"\bNOI\b|net operating income", ln.name, re.I):
            seen_noi = True

    n_agg = sum(1 for l in lines_out if l.kind == "subtotal")
    print(f"  OneSite row check: {n_checked} row(s) vs the printed "
          f"'Twelve Month Total' column"
          + (" - all tie." if not row_fails else
             f" - {len(row_fails)} MISMATCH."))
    print(f"  OneSite structure check: {n_agg} roll-up row(s) vs the detail "
          f"beneath them (indentation tree)"
          + (" - all tie." if not agg_fails else
             f" - {len(agg_fails)} MISMATCH."))
    for nm, got, want, kids in agg_fails:
        print(f"  SECTION-SUBTOTAL MISMATCH {nm}: detail beneath it sums to "
              f"{got:,.2f} vs printed {want:,.2f} "
              f"(variance {got - want:+,.2f}); children: {', '.join(kids)}")
    if agg_fails and not trust_monthly:
        sys.exit("ERROR: OneSite roll-up rows do not tie to the detail "
                 "beneath them - aborting. (re-run with --trust-monthly to "
                 "treat the monthly detail as source of truth)")

    return prop, months, _xlsx_grand_finalize(
        lines_out, row_fails, trust_monthly, _GRAND_PATS_ONESITE)


def apply_exclusions(lines, months, exclude, reason=""):
    """Drop named account lines from operations entirely (--exclude-account).

    For statements where ownership has CONFIRMED a line is not a real cost -
    typically a hand-keyed duplicate of another line - the honest fix is to
    remove it and say so, loudly, rather than leave a known-bad number in the
    T-12 or quietly patch a cell. This is deliberately a general flag and not
    a special case: it takes account names, matches them exactly (case- and
    whitespace-insensitive), refuses to no-op silently if a name matches
    nothing, and returns a note per removed line carrying the account, the
    side, every non-zero monthly value removed, the removed total and the
    stated reason. Nothing about an exclusion is ever silent - it prints to
    the console, lands in the delivery notes, and is written as a red note on
    the Trailing Financials tab.

    Returns (kept_lines, notes).
    """
    if not exclude:
        return lines, []
    want = {re.sub(r"\s+", " ", n).strip().lower() for n in exclude}
    kept, removed, notes = [], [], []
    for ln in lines:
        key = re.sub(r"\s+", " ", ln.name or "").strip().lower()
        if ln.kind == "account" and key in want:
            removed.append(ln)
        else:
            kept.append(ln)
    hit = {re.sub(r"\s+", " ", l.name or "").strip().lower() for l in removed}
    missing = sorted(want - hit)
    if missing:
        sys.exit("ERROR: --exclude-account matched no account line for %s. "
                 "Account names must match the statement exactly (they are "
                 "listed in the run output); refusing to silently exclude "
                 "nothing." % ", ".join(repr(m) for m in missing))
    for ln in removed:
        detail = "; ".join(
            "%s = %s" % (months[k].strftime("%b %Y"), f"{v:,.2f}")
            for k, v in enumerate(ln.values) if v)
        notes.append(
            "EXCLUDED FROM OPERATIONS (--exclude-account, never silent): "
            "account %r [%s side] removed in full. Reason: %s. Monthly values "
            "removed: %s (all other months 0.00). Removed total: %s. This "
            "line is gone from Total Operating Expenses, NOI, the Trailing "
            "Financials tab, the Final T-12 model-import tab and RawData."
            % (ln.name, "expense" if ln.side == "exp" else "income",
               reason or "not stated",
               detail or "none - the line was zero across the window",
               f"{sum(ln.values):,.2f}"))
    return kept, notes


# ----------------------------------------------------------------------------
# Owner-made "Rental schedule" operating statement (.xls / .xlsx)
# ----------------------------------------------------------------------------
# Validated on Heritage Ridge Apartments, 7550 NW 10th Street, Oklahoma City OK
# (owner/bookkeeper-prepared, .xls BIFF8, 8/2026).
#
#   rows 1-14  free-form header block: "Rental Property Statement for ...",
#              "Address of Rental Property:" + address, "Operating Statement
#              as of:" + a real date cell, acquisition/sale dates, percentage
#              owned, weeks rented.  There is NO property name anywhere in the
#              file - only the street address - so the parser falls back to the
#              address and says so; pass --property to name the deal.
#   header     ONE row of real date cells, one per month, running left to
#              right with NO totals column and NO year subtotals.  The dates
#              are day-of-month 22 formatted `mmm-yy`; only the month matters.
#              Heritage Ridge carries 52 of them (Apr-2022 .. Jul-2026), i.e.
#              four-plus fiscal years side by side - so this dialect is
#              inherently WIDER than a T-12 and a trailing window has to be
#              cut out of it (see `window` / `window_end`).
#   body       one row per account, label in col A, no sections, no GL
#              numbers, no indentation.  A revenue grand row ("Total Gross
#              Rent") sits at the top; every row below it is an expense.
#              "Total Expenses" / "Net Profit/(Loss)" / "NOI" are printed
#              roll-up rows - on Heritage Ridge only NOI carries any numbers.
#   footer     free text keyed to `**` markers used inside the grid.
#
# Owner spreadsheets are hand-maintained, so this parser assumes nothing:
#
#   * ORPHAN ROWS.  Heritage Ridge's "Rental Income" / "Other Income" /
#     "Other" rows (directly under the month header, exactly where the revenue
#     detail belongs) hold *date serials* in their month cells, plus - on the
#     Rental Income row - seven stray numbers that no printed total includes.
#     Any body row with a DATE cell in a month column is structurally corrupt:
#     it is dropped from the statement and reported verbatim, never guessed at.
#   * `**`-MARKED TEXT AMOUNTS.  Footnoted cells are stored as text
#     ('**23610.82', '**$4387.66').  They are real amounts: parsed to floats
#     and listed in the notes with the footnote text they point at.
#   * LUMP REVENUE.  When the revenue grand row has no surviving detail beneath
#     it (as here, the detail rows being the orphans above) the grand row is
#     PROMOTED to an account so the money is actually in the T-12, and the
#     Total Revenue subtotal becomes derived.  The mapping engine's
#     LUMP_INCOME rule then REVIEW-flags it - the rent/other-income split is a
#     judgement call, never silent.
#   * DUPLICATE-CELL GUARD.  Two different accounts carrying byte-identical
#     non-zero values in the same month is the classic hand-keyed copy/paste
#     double-count (Heritage Ridge: "Other" repeats "Water" in Aug-2025 and
#     Apr-2026, and ownership's own NOI excludes exactly those two cells).
#     Reported per occurrence with the amount; the value is KEPT, because
#     monthly detail wins and deleting a line ownership printed is not the
#     parser's call.
#   * PER-MONTH ROLL-UP CHECK.  There is no totals column, so every printed
#     roll-up row is checked column by column against the detail it rolls up
#     (revenue less expenses for NOI, sum of expenses for Total Expenses).
#     A printed roll-up cell that is blank is NOT read as zero - it is a month
#     ownership did not compute, and it disqualifies that roll-up as a
#     tie-out target for the window total (which is then derived, `~`).

OWNER_RS_MARK = re.compile(
    r"rental property statement|operating statement as of"
    r"|address of rental property", re.I)
OWNER_RS_REV_GRAND = re.compile(r"^total\s+(gross\s+)?(rent|income|revenue)s?", re.I)
OWNER_RS_EXP_GRAND = re.compile(r"^total\s+(operating\s+)?expenses?$", re.I)
OWNER_RS_NOI = re.compile(r"^\s*(noi|net\s+operating\s+income)\b", re.I)
OWNER_RS_NET = re.compile(r"^net\s+(profit|income|loss)", re.I)
_OWNER_RS_DATE = object()          # sentinel: a date cell where a number belongs


def _owner_rs_grid(path):
    """Read every sheet of an .xls/.xlsx into plain python grids.

    `.xls` (BIFF) is read with **xlrd** directly - deterministic, no external
    process, and it preserves the cell TYPE, which this dialect needs (a date
    serial sitting in a money column is how the orphan rows are detected; once
    converted to xlsx by a spreadsheet app that distinction can survive as a
    format only).  If xlrd is not installed the file is converted once with

        libreoffice --headless --convert-to xlsx --outdir <tmp> <file>

    and read with openpyxl - the same grid either way, so results are
    reproducible on a machine with only one of the two available.

    Returns [(sheet_name, rows)]; each cell is None, float, str, or datetime.
    """
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook as _lw
        wb = _lw(path, read_only=True, data_only=True)
        return [(ws.title, [list(r) for r in ws.iter_rows(values_only=True)])
                for ws in wb.worksheets]

    try:
        import xlrd
    except ImportError:
        xlrd = None
    if xlrd is not None:
        bk = xlrd.open_workbook(path)
        out = []
        for sh in bk.sheets():
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cv = sh.cell(r, c)
                    if cv.ctype == xlrd.XL_CELL_DATE:
                        try:
                            row.append(xlrd.xldate_as_datetime(cv.value,
                                                               bk.datemode))
                        except Exception:
                            row.append(None)
                    elif cv.ctype == xlrd.XL_CELL_NUMBER:
                        row.append(float(cv.value))
                    elif cv.ctype == xlrd.XL_CELL_TEXT:
                        row.append(cv.value)
                    elif cv.ctype == xlrd.XL_CELL_ERROR:
                        row.append("#ERROR!")
                    else:
                        row.append(None)
                rows.append(row)
            out.append((sh.name, rows))
        return out

    import subprocess
    import tempfile
    td = tempfile.mkdtemp(prefix="ownerrs_")
    subprocess.run(["libreoffice", "--headless", "--convert-to", "xlsx",
                    "--outdir", td, path], check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    conv = os.path.join(
        td, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
    return _owner_rs_grid(conv)


def _owner_rs_header(rows):
    """Locate the month header row: >=12 real date cells, one month apart.

    Returns (row_index, [column indexes], [datetime first-of-month]) or None.
    """
    for i, r in enumerate(rows):
        got = [(j, v) for j, v in enumerate(r)
               if hasattr(v, "year") and hasattr(v, "month")
               and not isinstance(v, str)]
        if len(got) < 12:
            continue
        cols = [j for j, _ in got]
        months = [datetime(v.year, v.month, 1) for _, v in got]
        ok = all(0 < (months[k + 1] - months[k]).days <= 31
                 for k in range(len(months) - 1))
        if ok:
            return i, cols, months
    return None


def _is_owner_rental_schedule(path):
    """A hand-made owner statement: a month-per-column grid with no totals
    column, carrying this dialect's header wording (or living on a sheet
    literally called "Rental schedule")."""
    if not path.lower().endswith((".xls", ".xlsx", ".xlsm")):
        return False
    try:
        sheets = _owner_rs_grid(path)
    except Exception:
        return False
    for name, rows in sheets:
        if _owner_rs_header(rows) is None:
            continue
        if re.search(r"rental\s+schedule", name or "", re.I):
            return True
        for r in rows[:20]:
            for v in r:
                if isinstance(v, str) and OWNER_RS_MARK.search(v):
                    return True
    return False


def _owner_rs_num(v):
    """Coerce a body cell. Returns (value, starred) where value is None
    (empty), a float, or the _OWNER_RS_DATE sentinel."""
    if v is None:
        return None, False
    if hasattr(v, "year") and hasattr(v, "month") and not isinstance(v, str):
        return _OWNER_RS_DATE, False
    if isinstance(v, (int, float)):
        return float(v), False
    s = str(v).strip()
    if not s:
        return None, False
    starred = "*" in s
    t = s.replace("*", "").replace("$", "").replace(",", "").strip()
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()")
    try:
        f = float(t)
    except ValueError:
        return None, False
    return (-f if neg else f), starred


def parse_t12_owner_rental_schedule(path, trust_monthly=False,
                                    allow_partial=False, window=12,
                                    window_end=None, exclude=None,
                                    exclude_reason=""):
    """Owner-made month-per-column operating statement -> (prop, months,
    lines, meta).

    The source is normally WIDER than twelve months (Heritage Ridge: 52), so a
    trailing window is cut out of it:

        window      number of trailing months to keep (default 12; 0 = all)
        window_end  'YYYY-MM' to end the window on a specific month; default
                    is the last month that carries ANY account data.

    Months are never padded, reordered or annualised - the window is a
    contiguous slice of the columns ownership actually printed.
    """
    sheets = _owner_rs_grid(path)
    pick = None
    for name, rows in sheets:
        h = _owner_rs_header(rows)
        if h is not None:
            pick = (name, rows, h)
            break
    if pick is None:
        sys.exit("ERROR: no month-per-column header row found in %s"
                 % os.path.basename(path))
    sheet_name, rows, (hdr_i, mcols, all_months) = pick
    notes, checks = [], []

    # ---- header block: address, as-of date, footnotes ---------------------
    addr, asof, prop = "", None, ""
    for r in rows[:hdr_i]:
        for j, v in enumerate(r):
            if not isinstance(v, str):
                continue
            lab = v.strip().lower()
            if lab.startswith("address of rental property"):
                for w in r[j + 1:]:
                    if isinstance(w, str) and w.strip():
                        addr = w.strip()
                        break
            elif lab.startswith("operating statement as of"):
                for w in r[j + 1:]:
                    if hasattr(w, "year") and not isinstance(w, str):
                        asof = w
                        break
    if addr:
        prop = addr.split(",")[0].strip()
        notes.append("this statement carries NO property name - only the "
                     "address %r. Property defaulted to the street address; "
                     "pass --property to name the deal." % addr)
    if asof:
        notes.append("statement 'as of' date printed in the header: %s"
                     % asof.strftime("%m/%d/%Y"))

    # ---- body -------------------------------------------------------------
    other_sheets = [n for n, _ in sheets if n != sheet_name]
    n_all = len(all_months)
    lines_all, side = [], "inc"
    orphans, starred, dup_notes = [], [], []
    printed_rollups = []          # (name, kind, [value|None per month])

    for r in rows[hdr_i + 1:]:
        name = re.sub(r"\s+", " ", str(r[0] or "")).strip()
        if not name:
            continue
        raw = [(_owner_rs_num(r[j]) if j < len(r) else (None, False))
               for j in mcols]
        vals = [v for v, _ in raw]
        if any(v is _OWNER_RS_DATE for v in vals):
            n_date = sum(1 for v in vals if v is _OWNER_RS_DATE)
            n_num = sum(1 for v in vals if isinstance(v, float))
            stray = ", ".join(
                "%s=%s" % (all_months[k].strftime("%b-%y"), f"{vals[k]:,.2f}")
                for k in range(n_all) if isinstance(vals[k], float))
            orphans.append((name, n_date, n_num, stray))
            continue
        for k, (v, st) in enumerate(raw):
            if st and v is not None:
                starred.append((name, all_months[k].strftime("%b-%y"), v))

        is_rev_grand = bool(OWNER_RS_REV_GRAND.match(name))
        is_exp_grand = bool(OWNER_RS_EXP_GRAND.match(name))
        is_noi = bool(OWNER_RS_NOI.match(name))
        is_net = bool(OWNER_RS_NET.match(name))
        if is_rev_grand or is_exp_grand or is_noi or is_net:
            printed_rollups.append((name, ("rev" if is_rev_grand else
                                           "exp" if is_exp_grand else
                                           "noi" if is_noi else "net"),
                                    list(vals)))
            ln = Line("subtotal", name, [v or 0.0 for v in vals], "", side)
            ln.printed = list(vals)
            lines_all.append(ln)
            if is_rev_grand:
                side = "exp"          # everything under the revenue grand row
            continue
        if all(v is None for v in vals):
            continue                  # label-only / spacer row
        ln = Line("account", name, [v or 0.0 for v in vals], "", side)
        ln.empty = [v is None for v in vals]   # blank != zero; kept per ROW,
        lines_all.append(ln)                   # two rows can share a label

    if orphans:
        for nm, nd, nn, stray in orphans:
            notes.append(
                "ORPHAN ROW DROPPED %r: %d of its %d month cells hold DATE "
                "serials, not amounts (structurally corrupt)%s. Not parsed, "
                "not guessed at."
                % (nm, nd, n_all, ("; stray numbers on the same row, included "
                                   "in no printed total: " + stray)
                   if nn else ""))

    # ---- window ------------------------------------------------------------
    acct_idx = [k for k in range(n_all)
                if any(l.values[k] for l in lines_all if l.kind == "account")]
    if not acct_idx:
        sys.exit("ERROR: no month column carries any account data in %s"
                 % os.path.basename(path))
    end = acct_idx[-1]
    if window_end:
        want = datetime.strptime(window_end + "-01", "%Y-%m-%d")
        if want not in all_months:
            sys.exit("ERROR: --months-ending %s is not a column on this "
                     "statement (columns run %s..%s)."
                     % (window_end, all_months[0].strftime("%Y-%m"),
                        all_months[-1].strftime("%Y-%m")))
        end = all_months.index(want)
    n = int(window or 0) or (end + 1)
    start = max(0, end + 1 - n)
    sl = slice(start, end + 1)
    months = all_months[sl]
    notes.append(
        "source spans %d month columns (%s - %s) with NO totals column and no "
        "year subtotals; trailing window cut to %s - %s (%d months)."
        % (n_all, all_months[0].strftime("%b %Y"),
           all_months[-1].strftime("%b %Y"), months[0].strftime("%b %Y"),
           months[-1].strftime("%b %Y"), len(months)))
    if end != n_all - 1:
        notes.append("columns %s - %s carry no account data and were left out "
                     "of the window."
                     % (all_months[end + 1].strftime("%b %Y"),
                        all_months[-1].strftime("%b %Y")))
    for ln in lines_all:
        ln.values = ln.values[sl]
        if hasattr(ln, "printed"):
            ln.printed = ln.printed[sl]
        if ln.empty is not None:      # dataclass field; None = never set
            ln.empty = ln.empty[sl]
    lines = [l for l in lines_all
             if l.kind != "account" or any(l.values)]
    dropped = [l.name for l in lines_all
               if l.kind == "account" and not any(l.values)]
    if dropped:
        notes.append("accounts with no activity inside the window (dropped): "
                     + ", ".join(sorted(set(dropped))))

    # Confirmed exclusions are applied BEFORE every downstream check, so the
    # printed roll-up rows are compared against the corrected detail (an
    # excluded duplicate is exactly what made ownership's own NOI disagree).
    lines, excl_notes = apply_exclusions(lines, months, exclude,
                                         exclude_reason)
    notes.extend(excl_notes)

    # ---- blank cells inside the window ------------------------------------
    # Only GAPS are worth reporting: a month with no amount on an account that
    # is otherwise billed every month (Heritage Ridge: Water in Jul 2026, the
    # month the statement was cut). An account that is genuinely occasional
    # ("Other": 2 hits in 12 months) is not a gap and would only be noise.
    blanks = []
    for l in lines:
        if l.kind != "account" or not hasattr(l, "empty"):
            continue
        e = l.empty
        if not any(e) or sum(1 for x in e if not x) < len(months) * 0.75:
            continue
        for k, is_empty in enumerate(e):
            if is_empty:
                blanks.append((l.name, months[k].strftime("%b %Y")))

    # ---- lump revenue ------------------------------------------------------
    rev_accts = [l for l in lines if l.kind == "account" and l.side == "inc"]
    rev_grand = next((l for l in lines if l.kind == "subtotal"
                      and OWNER_RS_REV_GRAND.match(l.name)), None)
    if rev_grand is not None and not rev_accts:
        notes.append(
            "LUMP REVENUE: the revenue grand row %r has no surviving detail "
            "rows beneath it, so it IS the revenue side. Promoted to an "
            "account (whole undifferentiated revenue, no rent / other-income "
            "split) and REVIEW-flagged; Total Revenue is derived from it."
            % rev_grand.name)
        rev_grand.kind = "account"
        rev_grand.side = "inc"
        rev_grand.flag_review = ("LUMP_INCOME - whole undifferentiated "
                                 "revenue side, no rent / other-income split")

    # ---- duplicate-cell guard ---------------------------------------------
    accts = [l for l in lines if l.kind == "account"]
    for k in range(len(months)):
        seen = {}
        for l in accts:
            v = l.values[k]
            if not v:
                continue
            if round(v, 2) in seen:
                l.flag_review = ("duplicate hand-keyed cell(s) shared with "
                                 "another account - possible double-count")
                dup_notes.append(
                    "DUPLICATE CELL %s: %r and %r both carry %s - identical "
                    "hand-keyed values in one month are the classic "
                    "double-count. Value KEPT (monthly detail wins); confirm "
                    "with ownership before underwriting."
                    % (months[k].strftime("%b %Y"), seen[round(v, 2)],
                       l.name, f"{v:,.2f}"))
            else:
                seen[round(v, 2)] = l.name
    notes.extend(dup_notes)

    # ---- per-month roll-up checks -----------------------------------------
    inc_v = [sum(l.values[k] for l in accts if l.side == "inc")
             for k in range(len(months))]
    exp_v = [sum(l.values[k] for l in accts if l.side == "exp")
             for k in range(len(months))]
    usable = {}
    for ln in lines:
        if ln.kind != "subtotal" or not hasattr(ln, "printed"):
            continue
        want = {"exp": exp_v,
                "noi": [i - e for i, e in zip(inc_v, exp_v)],
                "rev": inc_v}.get(
            "rev" if OWNER_RS_REV_GRAND.match(ln.name) else
            "exp" if OWNER_RS_EXP_GRAND.match(ln.name) else
            "noi" if OWNER_RS_NOI.match(ln.name) else "net")
        if want is None:
            checks.append("  ~ %s: printed on %d of %d months; this dialect's "
                          "net-income rows roll up nothing checkable - "
                          "ignored." % (ln.name,
                                        sum(1 for v in ln.printed
                                            if v is not None), len(months)))
            continue
        have = [k for k in range(len(months)) if ln.printed[k] is not None]
        bad = [k for k in have if abs(ln.printed[k] - want[k]) > 0.05]
        if not have:
            checks.append("  ~ %s: ownership printed NO values for this row "
                          "inside the window - derived from detail instead."
                          % ln.name)
        else:
            checks.append(
                "  %s %s: printed on %d of %d window months; %d tie to the "
                "monthly detail exactly%s"
                % ("OK " if not bad else "MISMATCH", ln.name, len(have),
                   len(months), len(have) - len(bad),
                   "." if not bad else ":"))
            for k in bad:
                checks.append(
                    "      %s printed %s vs monthly detail %s "
                    "(detail - printed = %s)"
                    % (months[k].strftime("%b %Y"),
                       f"{ln.printed[k]:,.2f}", f"{want[k]:,.2f}",
                       f"{want[k] - ln.printed[k]:+,.2f}"))
            if len(have) < len(months):
                gaps = [months[k].strftime("%b %Y") for k in range(len(months))
                        if ln.printed[k] is None]
                checks.append(
                    "      not printed for %s - a blank roll-up cell is a "
                    "month ownership did not compute, NOT a zero, so this row "
                    "cannot tie out the window total." % ", ".join(gaps))
        usable[ln.name] = (not bad) and len(have) == len(months)

    if any(v is False for v in usable.values()) and not trust_monthly:
        print("Printed roll-up rows vs parsed monthly detail:")
        for c in checks:
            print(c)
        sys.exit("ERROR: printed roll-up rows do not tie to the monthly "
                 "detail - aborting. (re-run with --trust-monthly to treat "
                 "the monthly detail as source of truth; every variance is "
                 "printed above)")

    # ---- canonical grand rows ---------------------------------------------
    lines = [l for l in lines if not (l.kind == "subtotal")]
    for nm, tot, src in (("Total Revenue", sum(inc_v), "rev"),
                         ("Total Operating Expenses", sum(exp_v), "exp"),
                         ("Total Net Operating Income",
                          sum(inc_v) - sum(exp_v), "noi")):
        ln = Line("subtotal", nm, [tot] + [0.0] * (len(months) - 1), "", "")
        ln.derived = True
        lines.append(ln)
    notes.append(
        "no Total Revenue / Total Operating Expenses row is printed anywhere "
        "on this statement and the printed NOI row is incomplete inside the "
        "window, so all three grand totals are DERIVED from the monthly "
        "detail and reported with '~' - they are not independent tie-out "
        "targets.")

    # a red note under the month headers naming what ownership left blank /
    # keyed twice inside the window - the two things that move NOI here
    bits = []
    excl_line = ""
    if excl_notes:
        excl_line = "; ".join(
            "%r %s line excluded - %s"
            % (n, "expense", exclude_reason or "reason not stated")
            for n in exclude) + ". "
        excl_line = excl_line.replace('"', "'")
    if blanks:
        bits.append("no amount reported by ownership for "
                    + "; ".join("%s [%s]" % (a, m) for a, m in blanks)
                    + " - cell left blank, not zero")
    if dup_notes:
        bits.append("%d duplicate hand-keyed cell(s) inside the window "
                    "(see delivery notes) are included as printed"
                    % len(dup_notes))
    note_line = (excl_line + ("Data quality: " + "; ".join(bits) + "."
                              if bits else "")).strip() or None

    meta = {
        "period": (months[0], months[-1].replace(
            day=_month_end(months[-1].year, months[-1].month).day)),
        "note_line": note_line,
        "blank_label": ("In-window months an otherwise-monthly account "
                        "reports NO amount (blank, NOT zero - the Total "
                        "column is a sum of 11 real months for these)"),
        "basis": "",
        "notes": notes,
        "dropped_cols": [],
        "blank_cells": blanks,
        "printed_checks": checks,
        "sheet": sheet_name,
        "other_sheets": other_sheets,
        "exempt_full_year": False,
        "exclusions_applied": True,
    }
    return prop, months, lines, meta


# ---------------------------------------------------------------------------
# AppFolio "Cash Flow - N Month" / "Income Statement" XLSX export
# (validated on Vista Lago, Tyler TX, Crosspointe Management, Jul 2025-Jun 2026)
#
# Geometry, and how it differs from every other xlsx dialect we handle:
#   * A parameter preamble (Exported On / Properties / Period Range /
#     Accounting Basis / Level of Detail) sits above the header row.
#   * Header row: col A is literally "Account Name", the month columns start
#     immediately at col B as "Mon YYYY" text, and the totals column is
#     captioned "Total". Yardi's detector requires months at col C or later,
#     so it does not fire here; OneSite's requires MM/DD/YYYY period-end
#     headers, so it does not fire either. Both were confirmed False on this
#     file - but the GENERIC parser did fire as the fallback and produced
#     garbage (it double-counted subtotals into the expense grand row and
#     summed the cash-balance rows), which is why this parser exists.
#   * The statement tree is leading-space indentation in col A, 4 per level.
#     A section header's month cells are genuinely EMPTY (None) - unlike
#     OneSite, which pads them with whitespace.
#   * Leaves and section headers share indent levels. "Advertising" and
#     "SUPPLIES" are single-line ACCOUNTS sitting at the same indent as the
#     section header "CLEANING AND MAINTENANCE". They cannot be told apart by
#     indent or by case, so classification is structural: a valued row is a
#     roll-up only if deeper-indented rows are still unconsumed beneath it
#     (or it carries an aggregate caption); otherwise it is an account.
#
# Two AppFolio-specific traps this parser handles explicitly:
#
#   1. RESTATEMENT ROWS. Below NOI the report restates the same numbers under
#      new captions - "Total Income" / "Total Expense" / "Net Income" repeat
#      Total Operating Income / Total Operating Expense / NOI, and "Cash Flow"
#      = Net Income + Net Other Items. They carry no new information. They are
#      VERIFIED against their operating counterparts month by month and then
#      dropped, rather than trusted or silently ignored - left in, they would
#      have been swept below-the-line and shown up in the Capex & Misc
#      workbook as an "expense" spanning both sides of the ledger.
#
#   2. CASH-BALANCE ROWS. "Beginning Cash", "Beginning Cash + Cash Flow" and
#      "Actual Ending Cash" are point-in-time BALANCES, not flows: their Total
#      column is a balance (Jul's opening / Jun's closing), NOT the sum of the
#      twelve monthly cells. Every generic row check therefore "fails" on them
#      by ~$381k. They are excluded from the statement entirely (they are not
#      below-the-line activity either - putting them in Capex & Misc would
#      fabricate ~$400k of transactions). The exclusion is reported, never
#      silent, and the roll-forward is checked for information.

APPFOLIO_INDENT = 4

# balance (not flow) rows: excluded from the statement, reported not silent
APPFOLIO_BALANCE = re.compile(
    r"^(beginning cash( \+ cash flow)?|actual ending cash|ending cash"
    r"|beginning balance|ending balance)$", re.I)

# rows that are aggregates by caption even if they consumed no children
APPFOLIO_AGG = re.compile(
    r"^(total\b|net\b|noi\b|cash flow$|gross\s+(profit|income)$)", re.I)

# CROSS-LEDGER summary rows. Expenses are carried positive, so a row that
# mixes the revenue and expense sides (or restates rows from both) can never
# be verified by summing the detail beneath it - the indentation consumer
# would happily "sum" 785k of income and 189k of expense to 974k. These are
# instead verified by IDENTITY against their operating counterparts, month by
# month (see the identity-check block below), and then dropped as duplicates.
APPFOLIO_RESTATE = re.compile(
    r"^(total income|total expenses?|net income|cash flow)$", re.I)

# AppFolio's grand rows are exact captions; anchor them so the restatements
# ("Total Income"/"Total Expense") can never be picked up as the grand row.
_GRAND_PATS_APPFOLIO = {
    "rev": r"^total operating income$",
    "exp": r"^total operating expenses?$",
    "noi": r"^noi\b|^net operating income",
    "drop": r"^total operating income$|^total operating expenses?$"
            r"|^noi\b|^net operating income",
}


def _appfolio_month(tok):
    """'Jul 2025' / 'July 2025' / 'Jul-2025' -> datetime, else None."""
    tok = str(tok or "").strip()
    m = re.match(r"^([A-Za-z]{3,9})\.?[-/. ]+(\d{4})$", tok)
    if not m:
        return None
    for cand in (m.group(1), m.group(1)[:3]):
        for fmt in ("%B", "%b"):
            try:
                return datetime.strptime(
                    f"{cand.title()} {m.group(2)}", f"{fmt} %Y")
            except ValueError:
                pass
    return None


def _is_appfolio_xlsx(path):
    """True for an AppFolio financial export: a report-parameter preamble
    above a header row whose col A is 'Account Name' with "Mon YYYY" month
    columns starting at col B, over an indented account tree.

    Deliberately narrow - it keys on AppFolio's own parameter lines AND the
    'Account Name' caption AND months at col B, so it cannot collide with the
    Yardi (months at col C+, GL in col A) or OneSite (MM/DD/YYYY headers)
    detectors, nor with the generic owner layout (no parameter preamble, and
    its totals column is captioned TOTAL/TOTALS beside ALL-CAPS sections)."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(path, read_only=True, data_only=True)
    except Exception:
        return False
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True, max_row=400)]

    preamble = 0
    for r in rows[:20]:
        a = str(r[0] or "").strip() if r else ""
        if re.match(r"^(exported on|period range|accounting basis|properties|"
                    r"level of detail|additional cash gl accounts|"
                    r"include zero balance)\s*:", a, re.I):
            preamble += 1
    if preamble < 2:
        return False

    hdr = None
    for i, r in enumerate(rows[:30]):
        if not r or str(r[0] or "").strip().lower() != "account name":
            continue
        hits = [j for j, v in enumerate(r) if j and _appfolio_month(v)]
        if len(hits) >= 2 and min(hits) == 1:
            hdr = i
            break
    if hdr is None:
        return False

    # an indented tree must follow
    return sum(1 for r in rows[hdr + 1:]
               if r and isinstance(r[0], str) and r[0].startswith(" ")) >= 5


def parse_t12_appfolio_xlsx(path, trust_monthly=False, allow_partial=False):
    """AppFolio Cash Flow / Income Statement xlsx -> (property, months, [Line])."""
    from openpyxl import load_workbook as _lw

    def _num(v):
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip().replace(",", "").replace("$", "")
            if not s:
                return None
            neg = s.startswith("(") and s.endswith(")")
            try:
                f = float(s.strip("()"))
            except ValueError:
                return None
            return -f if neg else f
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    wb = _lw(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    hdr_i, month_cols, months = None, [], []
    for i, r in enumerate(rows[:30]):
        if not r or str(r[0] or "").strip().lower() != "account name":
            continue
        got = [(j, _appfolio_month(v)) for j, v in enumerate(r) if j]
        got = [(j, d) for j, d in got if d]
        if len(got) >= 2:
            hdr_i = i
            month_cols = [j for j, _ in got]
            months = [d for _, d in got]
            break
    if hdr_i is None:
        sys.exit("ERROR: no 'Account Name' + month header row found in the "
                 "AppFolio xlsx export.")

    tot_col = None
    for j in range(max(month_cols) + 1, len(rows[hdr_i])):
        if str(rows[hdr_i][j] or "").strip().lower() in ("total", "totals"):
            tot_col = j
            break

    # ---- property + reporting parameters from the preamble --------------
    prop, basis, period, caption = "", "", "", ""
    for r in rows[:hdr_i]:
        a = str(r[0] or "").strip() if r else ""
        if not a:
            continue
        if re.match(r"^properties?\s*:", a, re.I):
            val = a.split(":", 1)[1].strip()
            # "Vista Lago - 14200 CR 1134 Tyler, TX 75709" -> "Vista Lago"
            # split on the dash that introduces the street address
            prop = re.split(r"\s+-\s+(?=\d)", val, 1)[0].strip()
        elif re.match(r"^accounting basis\s*:", a, re.I):
            basis = a.split(":", 1)[1].strip()
        elif re.match(r"^period range\s*:", a, re.I):
            period = a.split(":", 1)[1].strip()
        elif not caption:
            caption = a
    if not prop:
        sys.exit("ERROR: AppFolio export has no 'Properties:' line - pass "
                 "--property to name the deliverable.")
    if "," in prop or re.match(r"^all\b", prop, re.I):
        print(f"  NOTE: multi-property AppFolio export ({prop}) - pass "
              f"--property for the combined name.")

    print(f"  AppFolio export: {caption or 'financial statement'}")
    if period:
        print(f"  Period Range (report parameter): {period}")
    if basis:
        print(f"  Accounting Basis: {basis}"
              + ("  <- CASH BASIS: revenue/expense land when money moves, "
                 "not when earned/incurred." if basis.lower().startswith("cash")
                 else ""))

    # ---- body ------------------------------------------------------------
    lines_out, stack, pending = [], [], []
    row_fails, agg_fails, skipped = [], [], []
    n_checked = 0
    side = "inc"

    for r in rows[hdr_i + 1:]:
        raw = r[0] if r else None
        if raw is None or not str(raw).strip():
            continue
        s = str(raw)
        indent = len(s) - len(s.lstrip(" \t"))
        name = re.sub(r"\s+", " ", s.strip())

        vals = [_num(r[j]) if j < len(r) else None for j in month_cols]
        annual = (_num(r[tot_col])
                  if tot_col is not None and tot_col < len(r) else None)

        # cash-balance rows: their Total column is a BALANCE, not a sum
        if APPFOLIO_BALANCE.match(name):
            skipped.append((name, [v or 0.0 for v in vals], annual))
            continue

        if all(v is None for v in vals):
            while stack and stack[-1][0] >= indent:
                stack.pop()
            stack.append((indent, name))
            lines_out.append(Line("section", name, None, name, side))
            continue

        mvals = [v or 0.0 for v in vals]

        # how many not-yet-consumed rows sit deeper than this one?
        k = len(pending)
        while k and pending[k - 1][0] > indent:
            k -= 1
        take = pending[k:]
        is_agg = bool(take) or bool(APPFOLIO_AGG.match(name))

        if annual is not None:
            n_checked += 1
            if abs(sum(mvals) - annual) > 0.05:
                row_fails.append((name, sum(mvals), annual,
                                  "subtotal" if is_agg else "account", side))

        if not is_agg:
            lines_out.append(Line("account", name, mvals,
                                  stack[-1][1] if stack else "", side))
            lines_out[-1].values_annual = annual
            pending.append((indent, name, mvals))
            continue

        # roll-up row
        while stack and stack[-1][0] >= indent:
            stack.pop()
        cross = bool(APPFOLIO_RESTATE.match(name))
        if take and not cross:
            # must equal exactly the not-yet-consumed detail beneath it
            ssum = [sum(p[2][x] for p in take) for x in range(len(mvals))]
            if not all(abs(a - b) <= 0.05 for a, b in zip(ssum, mvals)):
                agg_fails.append((name, sum(ssum), sum(mvals),
                                  [p[1] for p in take]))
        lines_out.append(Line("subtotal", name, mvals,
                              stack[-1][1] if stack else "", side))
        lines_out[-1].values_annual = annual
        # A cross-ledger row consumes nothing (it is checked by identity), and
        # an aggregate that consumed nothing restates rows already in pending -
        # neither introduces new detail, so neither may enter pending or a
        # later roll-up would double-count it.
        if cross:
            pass
        elif take:
            pending = pending[:k] + [(indent, name, mvals)]
        if re.match(_GRAND_PATS_APPFOLIO["rev"], name, re.I):
            side = "exp"          # everything after the revenue grand row

    # ---- restatement rows: verify, then drop -----------------------------
    by_name = {}
    for ln in lines_out:
        if ln.kind == "subtotal":
            by_name.setdefault(ln.name.lower(), ln)

    def _vals(nm):
        ln = by_name.get(nm)
        return ln.values if ln else None

    restate_fails, restated, ident_ok = [], [], []
    noi_ln = next((l for l in lines_out if l.kind == "subtotal"
                   and re.search(_GRAND_PATS_APPFOLIO["noi"], l.name, re.I)),
                  None)
    toi, toe = _vals("total operating income"), _vals("total operating expense")

    def _identity(label, got, want, drop):
        """Verify a cross-ledger row month by month against the operating
        rows it restates. `drop` marks it as a duplicate to be removed."""
        if got is None or want is None:
            return
        (restated if drop else ident_ok).append(label)
        if not all(abs(a - b) <= 0.05 for a, b in zip(got, want)):
            restate_fails.append((label, sum(got), sum(want)))

    # the grand NOI row itself: income less expense (kept, not dropped)
    if noi_ln is not None and toi and toe:
        _identity("NOI = Total Operating Income - Total Operating Expense",
                  noi_ln.values, [a - b for a, b in zip(toi, toe)], False)
    _identity("total income = Total Operating Income",
              _vals("total income"), toi, True)
    _identity("total expense = Total Operating Expense",
              _vals("total expense"), toe, True)
    _identity("net income = NOI", _vals("net income"),
              noi_ln.values if noi_ln else None, True)
    ni, noth = _vals("net income"), _vals("net other items")
    _identity("cash flow = Net Income + Net Other Items", _vals("cash flow"),
              [a + b for a, b in zip(ni, noth)] if ni and noth else None, True)

    drop_names = {lbl.split(" = ")[0] for lbl in restated}
    lines_out = [l for l in lines_out
                 if not (l.kind == "subtotal" and l.name.lower() in drop_names)]

    # ---- below-the-line: anything after the printed NOI -------------------
    seen_noi = False
    for ln in lines_out:
        if seen_noi or BELOW_PAT.search(ln.section or ""):
            ln.below = True
        if ln.kind == "subtotal" and re.search(
                _GRAND_PATS_APPFOLIO["noi"], ln.name, re.I):
            seen_noi = True

    # ---- report ----------------------------------------------------------
    n_agg = sum(1 for l in lines_out if l.kind == "subtotal")
    print(f"  AppFolio row check: {n_checked} row(s) vs the printed 'Total' "
          f"column" + (" - all tie." if not row_fails else
                       f" - {len(row_fails)} MISMATCH."))
    print(f"  AppFolio structure check: {n_agg} roll-up row(s) vs the detail "
          f"beneath them (indentation tree)"
          + (" - all tie." if not agg_fails else
             f" - {len(agg_fails)} MISMATCH."))
    for nm, got, want, kids in agg_fails:
        print(f"  SECTION-SUBTOTAL MISMATCH {nm}: detail beneath it sums to "
              f"{got:,.2f} vs printed {want:,.2f} "
              f"(variance {got - want:+,.2f}); children: {', '.join(kids)}")
    if restated or ident_ok:
        print(f"  AppFolio cross-ledger identity check: "
              f"{len(restated) + len(ident_ok)} row(s) verified month by "
              f"month against their operating counterparts"
              + (" - all tie." if not restate_fails else
                 f" - {len(restate_fails)} MISMATCH."))
        for lbl in ident_ok:
            print(f"    OK  {lbl}  (kept)")
        for lbl in restated:
            print(f"    OK  {lbl}  (duplicate - dropped)"
                  if not any(f[0] == lbl for f in restate_fails) else "")
    for lbl, got, want in restate_fails:
        print(f"  IDENTITY MISMATCH {lbl}: {got:,.2f} vs {want:,.2f} "
              f"(variance {got - want:+,.2f})")
    if skipped:
        print(f"  AppFolio cash roll-forward: {len(skipped)} balance row(s) "
              f"excluded from the statement (point-in-time balances, not "
              f"flows): " + ", ".join(n for n, _, _ in skipped))
        beg = next((v for n, v, _ in skipped
                    if re.match(r"^beginning cash$", n, re.I)), None)
        end = next((v for n, v, _ in skipped
                    if re.match(r"^actual ending cash$", n, re.I)), None)
        if beg and end:
            brk = [i for i in range(len(end) - 1)
                   if abs(end[i] - beg[i + 1]) > 0.05]
            print("    roll-forward (ending cash -> next beginning cash): "
                  + ("continuous." if not brk else
                     f"BREAKS at {len(brk)} month boundary(ies)."))
            cfv = _vals("cash flow")
            if cfv:
                drift = [(months[i], end[i] - beg[i] - cfv[i])
                         for i in range(len(end))
                         if abs(end[i] - beg[i] - cfv[i]) > 0.05]
                if drift:
                    print(f"    NOTE: actual bank movement differs from the "
                          f"GL Cash Flow line in {len(drift)} month(s) "
                          f"(total {sum(d for _, d in drift):+,.2f}) - "
                          "non-GL cash activity; informational only.")

    if agg_fails or restate_fails:
        if not trust_monthly:
            sys.exit("ERROR: AppFolio roll-up/restatement rows do not tie to "
                     "the detail beneath them - aborting. (re-run with "
                     "--trust-monthly to treat the monthly detail as source "
                     "of truth)")
        print("  --trust-monthly: monthly detail wins over the printed "
              "roll-up rows.")

    return prop, months, _xlsx_grand_finalize(
        lines_out, row_fails, trust_monthly, _GRAND_PATS_APPFOLIO)


# ----------------------------------------------------------------------------
# ResMan "Trailing Profit And Loss Detail" XLSX
# ----------------------------------------------------------------------------
# Added for Eclipse of White Rock 8/5/2026 (Ci Mgmt, Jan-Dec 2025, Accrual,
# Accounting Book: Default). This is ResMan's trailing P&L exported to Excel -
# a different animal from the ResMan PDF (`parse_t12_pdf_resman`), which has no
# column tree at all.
#
#   row 1   property name
#   row 3   "Trailing Profit And Loss Detail"
#   row 4   "<Month Year>- <basis> - Accounting Book: <book>"
#   row 6   header: "Account" in col A, twelve "<Mon> <YYYY> Actual" month
#           columns (E..P) and an "Adjusted Total" column (Q)
#
# The statement tree is encoded by WHICH COLUMN the label sits in - there is no
# leading whitespace to measure and no "Total <X>" caption convention that can
# be trusted on its own ("4000 Total RENTAL INCOME" carries the GL number
# first, so an unanchored `^total` test never fires):
#
#   col A   grand NOI row                ("NET OPERATING INCOME")
#   col B   ledger side + grand totals    ("INCOME", "TOTAL INCOME",
#                                          "EXPENSE", "TOTAL EXPENSE")
#   col C   section header + its roll-up  ("4000 RENTAL INCOME",
#                                          "4000 Total RENTAL INCOME")
#   col D   accounts                      ("4103 RENT INCOME")
#
# A section header carries no month values; its roll-up does. Level is
# therefore unambiguous without any indent guessing, and the roll-up's GL
# number is asserted against its section header's so a mis-nested export is
# caught rather than silently rolled up into the wrong section.
#
# Watch for: reimbursement accounts booked INSIDE an expense section (Eclipse
# carries 4301 UTILITY REIMBURSEMENT / 4302 ELECTRICITY REIMBURSEMENT as
# negative expense inside "6200 UTILITIES"). They stay on the expense side -
# moving them to revenue would push money across the ledger, which the corpus
# cross-ledger guard exists to prevent - but the run prints them explicitly so
# the netting is never invisible.

_GRAND_PATS_RESMAN_TR = {
    "rev": r"^total\s+income$",
    "exp": r"^total\s+(operating\s+)?expenses?$",
    "noi": r"^net\s+operating\s+income$",
    "drop": r"^total\s+income$|^total\s+(operating\s+)?expenses?$"
            r"|^net\s+operating\s+income$",
}

RESMAN_TR_MONTH = re.compile(r"^([A-Za-z]{3,9})\.?\s+(\d{4})\s+Actual$", re.I)


def _resman_tr_month(v):
    """'Jan 2025 Actual' / 'May 2025\\nActual' -> datetime, else None."""
    tok = re.sub(r"\s+", " ", str(v or "")).strip()
    m = RESMAN_TR_MONTH.match(tok)
    if not m:
        return None
    for cand in (m.group(1), m.group(1)[:3]):
        for fmt in ("%B", "%b"):
            try:
                return datetime.strptime(
                    f"{cand.title()} {m.group(2)}", f"{fmt} %Y")
            except ValueError:
                pass
    return None


def _resman_tr_header(rows):
    """-> (header_row_index, [month cols], [months], total col) or Nones."""
    for i, r in enumerate(rows[:15]):
        if not r or str(r[0] or "").strip().lower() != "account":
            continue
        got = [(j, _resman_tr_month(v)) for j, v in enumerate(r) if j >= 2]
        got = [(j, d) for j, d in got if d]
        if len(got) < 2:
            continue
        mcols = [j for j, _ in got]
        tot = None
        for j, v in enumerate(r):
            if j > max(mcols) and re.fullmatch(
                    r"adjusted total|total", re.sub(r"\s+", " ",
                                                    str(v or "")).strip(),
                    re.I):
                tot = j
                break
        return i, mcols, [d for _, d in got], tot
    return None, [], [], None


def _is_resman_trailing_xlsx(path):
    """True for a ResMan 'Trailing Profit And Loss Detail' xlsx export.

    Keyed on the report title AND an 'Account' header row carrying
    '<Mon> <YYYY> Actual' month captions at col C or right of it - narrow
    enough that it cannot collide with AppFolio ('Account Name' in col A with
    bare 'Mon YYYY' captions), OneSite (MM/DD/YYYY period-end headers), Yardi
    (GL number in col A, no header caption row) or the generic owner layout
    (plain 'Mon-YYYY' captions and a TOTAL/TOTALS column)."""
    from openpyxl import load_workbook as _lw
    try:
        wb = _lw(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True, max_row=40)]
    except Exception:
        return False
    if not any(re.search(r"trailing\s+profit\s+and\s+loss", str(r[0] or ""),
                         re.I)
               for r in rows[:8] if r):
        return False
    hdr_i, _, _, tot = _resman_tr_header(rows)
    return hdr_i is not None and tot is not None


def parse_t12_resman_trailing_xlsx(path, trust_monthly=False,
                                   allow_partial=False):
    """ResMan Trailing P&L Detail xlsx -> (property, months, [Line])."""
    from openpyxl import load_workbook as _lw

    def _num(v):
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip().replace(",", "").replace("$", "")
            if not s:
                return None
            neg = s.startswith("(") and s.endswith(")")
            try:
                f = float(s.strip("()"))
            except ValueError:
                return None
            return -f if neg else f
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    wb = _lw(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    hdr_i, month_cols, months, tot_col = _resman_tr_header(rows)
    if hdr_i is None:
        sys.exit("ERROR: no 'Account' + '<Mon> <YYYY> Actual' header row "
                 "found in the ResMan trailing P&L xlsx.")
    if tot_col is None:
        sys.exit("ERROR: ResMan trailing P&L xlsx has no 'Adjusted Total' "
                 "column - every row total check would be unverifiable.")

    label_cols = list(range(0, min(month_cols)))     # A..D
    acct_col, sec_col = max(label_cols), max(label_cols) - 1

    # ---- header block ----------------------------------------------------
    prop = re.sub(r"\s+", " ", str(rows[0][0] or "")).strip() if rows else ""
    caption, basis, book = "", "", ""
    for r in rows[1:hdr_i]:
        a = re.sub(r"\s+", " ", str(r[0] or "")).strip() if r else ""
        if not a:
            continue
        if re.search(r"profit\s+and\s+loss", a, re.I):
            caption = a
            continue
        m = re.search(r"\b(accrual|cash)\b", a, re.I)
        if m:
            basis = m.group(1).title()
        m = re.search(r"accounting book\s*:\s*(.+)$", a, re.I)
        if m:
            book = m.group(1).strip()
    print(f"  ResMan export: {caption or 'Trailing Profit And Loss Detail'}"
          + (f" | {basis} basis" if basis else "")
          + (f" | Accounting Book: {book}" if book else ""))

    # ---- body ------------------------------------------------------------
    lines_out, row_fails = [], []
    sec_subs = []            # (section, name, mvals, side) roll-ups
    sub_fails, nest_fails = [], []
    acct_by_section = {}
    grand = {}
    cur_section, n_checked = "", 0
    side = "inc"
    reimb = []               # revenue-coded accounts inside expense sections

    for r in rows[hdr_i + 1:]:
        lab_j = next((j for j in label_cols
                      if j < len(r) and str(r[j] or "").strip()), None)
        if lab_j is None:
            continue
        name = re.sub(r"\s+", " ", str(r[lab_j])).strip()
        if not name:
            continue
        vals = [_num(r[j]) if j < len(r) else None for j in month_cols]
        annual = (_num(r[tot_col]) if tot_col < len(r) else None)
        has_vals = any(v is not None for v in vals)
        mvals = [v or 0.0 for v in vals]

        # ---- section header (no month values) ----------------------------
        if not has_vals:
            if lab_j >= sec_col:
                cur_section = name
            if re.fullmatch(r"(operating\s+)?(income|revenue)s?", name, re.I):
                side = "inc"
            elif re.fullmatch(r"(operating\s+)?expenses?", name, re.I):
                side = "exp"
            lines_out.append(Line("section", name, None, name, side))
            continue

        kind = ("account" if lab_j == acct_col else
                "subtotal" if lab_j == sec_col else "grand")

        if annual is not None:
            n_checked += 1
            if abs(sum(mvals) - annual) > 0.05:
                row_fails.append((name, sum(mvals), annual,
                                  "account" if kind == "account"
                                  else "subtotal", side))

        if kind == "account":
            lines_out.append(Line("account", name, mvals, cur_section, side))
            lines_out[-1].values_annual = annual
            acct_by_section.setdefault(cur_section, []).append((name, mvals))
            if side == "exp" and gl_number(name).startswith("4") and \
                    sum(mvals) < 0:
                reimb.append((name, cur_section, sum(mvals)))
            continue

        if kind == "subtotal":
            # "4000 Total RENTAL INCOME" rolls up section "4000 RENTAL INCOME":
            # assert the GL numbers agree before trusting the nesting.
            g_sub, g_sec = gl_number(name), gl_number(cur_section)
            if g_sub and g_sec and g_sub != g_sec:
                nest_fails.append((name, cur_section))
            kids = acct_by_section.get(cur_section, [])
            if kids:
                ssum = [sum(k[1][x] for k in kids) for x in range(len(mvals))]
                if not all(abs(a - b) <= 0.05 for a, b in zip(ssum, mvals)):
                    sub_fails.append((name, sum(ssum), sum(mvals),
                                      [k[0] for k in kids]))
            sec_subs.append((cur_section, name, mvals, side))
            lines_out.append(Line("subtotal", name, mvals, cur_section, side))
            lines_out[-1].values_annual = annual
            continue

        # ---- grand row ---------------------------------------------------
        key = ("rev" if re.fullmatch(_GRAND_PATS_RESMAN_TR["rev"], name, re.I)
               else "exp" if re.fullmatch(_GRAND_PATS_RESMAN_TR["exp"], name,
                                          re.I)
               else "noi" if re.fullmatch(_GRAND_PATS_RESMAN_TR["noi"], name,
                                          re.I) else None)
        lines_out.append(Line("subtotal", name, mvals, "", side))
        lines_out[-1].values_annual = annual
        if key:
            grand[key] = mvals
        if key == "rev":
            side = "exp"          # everything after the revenue grand row

    # ---- grand rows vs the section roll-ups beneath them ------------------
    grand_fails = []

    def _side_sum(sd):
        picked = [s for s in sec_subs if s[3] == sd]
        if not picked:
            return None
        return [sum(s[2][x] for s in picked) for x in range(len(months))]

    for key, label, want in (("rev", "TOTAL INCOME", _side_sum("inc")),
                             ("exp", "TOTAL EXPENSE", _side_sum("exp"))):
        got = grand.get(key)
        if got is None or want is None:
            continue
        if not all(abs(a - b) <= 0.05 for a, b in zip(got, want)):
            grand_fails.append((label, sum(want), sum(got)))
    if grand.get("noi") and grand.get("rev") and grand.get("exp"):
        want = [a - b for a, b in zip(grand["rev"], grand["exp"])]
        if not all(abs(a - b) <= 0.05
                   for a, b in zip(grand["noi"], want)):
            grand_fails.append(("NET OPERATING INCOME = TOTAL INCOME - "
                                "TOTAL EXPENSE", sum(want),
                                sum(grand["noi"])))

    # ---- below-the-line: anything after the printed NOI --------------------
    seen_noi = False
    for ln in lines_out:
        if seen_noi or BELOW_PAT.search(ln.section or ""):
            ln.below = True
        if ln.kind == "subtotal" and re.fullmatch(
                _GRAND_PATS_RESMAN_TR["noi"], ln.name, re.I):
            seen_noi = True

    # ---- report -----------------------------------------------------------
    print(f"  ResMan row check: {n_checked} row(s) vs the printed 'Adjusted "
          f"Total' column" + (" - all tie." if not row_fails else
                              f" - {len(row_fails)} MISMATCH."))
    print(f"  ResMan section check: {len(sec_subs)} section roll-up row(s) vs "
          f"their account detail, month by month"
          + (" - all tie." if not sub_fails else
             f" - {len(sub_fails)} MISMATCH."))
    for nm, got, want, kids in sub_fails:
        print(f"  SECTION-SUBTOTAL MISMATCH {nm}: detail beneath it sums to "
              f"{got:,.2f} vs printed {want:,.2f} "
              f"(variance {got - want:+,.2f}); children: {', '.join(kids)}")
    for nm, sec in nest_fails:
        print(f"  STRUCTURE MISMATCH: roll-up '{nm}' closes section '{sec}' "
              f"but their GL numbers disagree.")
    print(f"  ResMan grand check: TOTAL INCOME / TOTAL EXPENSE vs their "
          f"section roll-ups and NOI vs income less expense, month by month"
          + (" - all tie." if not grand_fails else
             f" - {len(grand_fails)} MISMATCH."))
    for lbl, want, got in grand_fails:
        print(f"  GRAND MISMATCH {lbl}: detail {want:,.2f} vs printed "
              f"{got:,.2f} (variance {want - got:+,.2f})")
    if reimb:
        print(f"  NOTE: {len(reimb)} revenue-numbered account(s) are booked "
              f"inside an EXPENSE section as negative expense (the statement "
              f"nets them against the cost). They are kept on the expense "
              f"side - reclassifying them to Other Income would move money "
              f"across the ledger:")
        for nm, sec, tot in reimb:
            print(f"    {nm}  [{sec}]  {tot:,.2f}")

    if sub_fails or grand_fails or nest_fails:
        if not trust_monthly:
            sys.exit("ERROR: ResMan roll-up rows do not tie to the detail "
                     "beneath them - aborting. (re-run with --trust-monthly "
                     "to treat the monthly detail as source of truth)")
        print("  --trust-monthly: monthly detail wins over the printed "
              "roll-up rows.")

    return prop, months, _xlsx_grand_finalize(
        lines_out, row_fails, trust_monthly, _GRAND_PATS_RESMAN_TR)


# Detection-based dispatch for xlsx T-12s (mirrors the rent-roll XLSX_PARSERS
# convention). parse_t12_xlsx itself is the fallback owner/PM-prepared layout.
T12_XLSX_PARSERS = [
    (_is_resman_trailing_xlsx, parse_t12_resman_trailing_xlsx),
    (_is_appfolio_xlsx, parse_t12_appfolio_xlsx),
    (_is_onesite_xlsx, parse_t12_onesite_xlsx),
    (_is_yardi_xlsx, parse_t12_yardi_xlsx),
]


# ----------------------------------------------------------------------------
# Mapping engine
# ----------------------------------------------------------------------------

def section_codes(section, side):
    for pat, codes in SECTION_ALLOWED:
        if pat.search(section or ""):
            return codes
    return INCOME_CODES if side == "inc" else EXPENSE_CODES


def map_codes(lines, by_gl, by_name):
    corpus_names = list(by_name.items())
    for ln in lines:
        if ln.kind != "account":
            continue
        if ln.below:
            ln.code, ln.method = "", "below-the-line (excluded)"
            continue
        allowed = section_codes(ln.section, ln.side)
        gl, nn = gl_number(ln.name), norm(ln.name)
        side_codes = INCOME_CODES if ln.side == "inc" else EXPENSE_CODES

        # 0) undifferentiated lump-sum revenue line: the statement gives no
        #    rent/other-income split, so booking it anywhere is a judgement
        #    call. Park it in Rental Income (the dominant component of gross
        #    collections) and ALWAYS flag it - never let it pass silently.
        if ln.side == "inc" and LUMP_INCOME.match(ln.name.strip()) and \
                "r" in allowed:
            ln.code, ln.review = "r", True
            ln.method = ("REVIEW (undifferentiated lump-sum revenue line - "
                         "no rent / other-income detail on the statement; "
                         "parked in Rental Income)")
            continue

        # 1) exact memory
        code = by_gl.get((gl, nn)) or by_name.get(nn)
        # An exact hit may come from a statement that carried the same label
        # on the OTHER side of the ledger - "LEGAL FEES" as an expense vs
        # Yardi's resident-charged legal-fee income, "WASHER/DRYER RENTAL" as
        # laundry income vs the equipment-lease expense. Honouring it would
        # move money across the revenue/expense line and break the
        # reconciliation, so reject it, fall through to the rules layers and
        # always flag the line. (The Meadows, 7/2026.)
        if code and code not in side_codes:
            ln.xside, code = code, None
        if code:
            ln.code, ln.method = code, "exact"
            if code not in allowed:
                ln.review = True
                ln.method = "exact (conflicts with section - REVIEW)"
            continue
        # 2) single-code section
        if len(allowed) == 1:
            ln.code, ln.method = next(iter(allowed)), "section"
            continue
        # 3) keyword rules within allowed set
        hit = None
        for pat, code, restrict in KEYWORD_RULES:
            if restrict != ln.side:
                continue
            if code in allowed and re.search(pat, nn, re.I):
                hit = code
                break
        if hit:
            ln.code, ln.method = hit, "keyword"
            continue
        # 4) fuzzy vs corpus, constrained to allowed.
        #    _sim is rapidfuzz's token_set_ratio, which scores 1.00 whenever
        #    one label's token set is a SUBSET of the other's. Two guards:
        #      (a) ties are broken by whole-string similarity, so the most
        #          literally similar corpus entry wins rather than whichever
        #          one the corpus happened to list first;
        #      (b) if EVERY top-scoring candidate is a strict superset of the
        #          account name - i.e. the name contributes no token that
        #          distinguishes them - and those candidates disagree about
        #          the code, the match is a subset artifact, not evidence.
        #          A bare "Income" scores 1.00 against 'fee income' (oi),
        #          'gas income' (ro) and 'hap income' (r) alike; taking the
        #          first would have auto-coded The Gardens' entire revenue
        #          line (7/2026). Such lines are defaulted and flagged.
        best, best_s, best_w, top = None, 0.0, 0.0, []
        for cand_name, cand_code in corpus_names:
            if cand_code not in allowed:
                continue
            s = _sim(nn, cand_name)
            if s > best_s:
                best_s, top = s, [(cand_name, cand_code)]
                best, best_w = cand_code, _whole(nn, cand_name)
            elif s == best_s:
                top.append((cand_name, cand_code))
                w = _whole(nn, cand_name)
                if w > best_w:
                    best, best_w = cand_code, w
        if best and best_s >= 0.90:
            inp = set(nn.split())
            frag_only = bool(top) and all(
                inp < set(c.split()) for c, _ in top)
            if frag_only and len({cc for _, cc in top}) > 1:
                default = "oi" if ln.side == "inc" else \
                    ("rm" if "rm" in allowed else next(iter(allowed)))
                ln.code, ln.review = default, True
                ln.method = (
                    f"REVIEW (fuzzy {best_s:.2f} is a token-subset artifact: "
                    f"{len(top)} corpus labels all contain '{nn}' plus extra "
                    f"qualifiers and disagree "
                    f"({'/'.join(sorted({cc for _, cc in top}))}); "
                    f"defaulted to '{default}')")
                continue
            ln.code, ln.method = best, f"fuzzy:{best_s:.2f}"
            continue
        # 5) unmapped -> default bucket + REVIEW
        default = "oi" if ln.side == "inc" else \
            ("rm" if "rm" in allowed else next(iter(allowed)))
        if best and best_s >= 0.75:
            default = best
        ln.code, ln.method, ln.review = default, \
            f"REVIEW (best guess, fuzzy={best_s:.2f})", True

    # A parser can KNOW a line is suspect for a structural reason the mapping
    # layers cannot see (an undifferentiated lump-revenue row that only looks
    # like a clean rent total; two accounts carrying byte-identical hand-keyed
    # cells). `Line.flag_review` carries that reason through to REVIEW.
    for ln in lines:
        if ln.kind == "account" and getattr(ln, "flag_review", ""):
            ln.review = True
            ln.method = f"{ln.method} [REVIEW: {ln.flag_review}]"

    # cross-ledger corpus hits are never silent, whichever layer resolved them
    for ln in lines:
        if ln.kind == "account" and ln.xside:
            ln.review = True
            ln.method = (f"{ln.method} [corpus said '{ln.xside}' - wrong "
                         f"side of ledger, rejected - REVIEW]")
    return lines


# ----------------------------------------------------------------------------
# Reconciliation
# ----------------------------------------------------------------------------

def reconcile(lines):
    """Compare coded sums vs the statement's own grand totals.

    Subtotals flagged `derived` were computed by the parser because the
    statement never printed them - they are reported for information but are
    NEVER used as a tie-out target (that would be checking the parse against
    itself)."""
    def grand(name_pat):
        for ln in lines:
            if ln.kind == "subtotal" and not ln.derived and \
                    re.fullmatch(name_pat, ln.name, re.I):
                return sum(ln.values)
        return None

    inc = sum(sum(l.values) for l in lines
              if l.kind == "account" and l.code in INCOME_CODES)
    exp = sum(sum(l.values) for l in lines
              if l.kind == "account" and l.code in EXPENSE_CODES)
    checks, ok = [], True

    for label, got, want in [
            ("Total Revenue", inc, grand(r"Total (Revenue|Income)s?")),
            ("Total Operating Expense", exp,
             grand(r"Total (Operating )?Expenses?")),
            ("NOI", inc - exp, grand(r"Total Net Operating Income"))]:
        if want is None:
            checks.append(f"  ~ {label}: {got:,.2f} (no printed total on the "
                          f"statement - derived from parsed detail)")
            continue
        good = abs(got - want) <= 0.05
        ok &= good
        checks.append(f"  {'OK ' if good else 'MISMATCH'} {label}: "
                      f"coded {got:,.2f} vs statement {want:,.2f}")
    return ok, "\n".join(checks)


# ----------------------------------------------------------------------------
# Workbook writer
# ----------------------------------------------------------------------------

def _purge_broken_names(wb):
    """Remove defined names that reference deleted sheets or #REF! — they
    crash JS-based Excel loaders (Power BI web, import tools)."""
    sheets = set(wb.sheetnames)

    def is_bad(name, ref):
        if not ref or "#REF!" in ref or name.startswith("ExternalData"):
            return True
        for q, u in re.findall(r"'([^']+)'!|([A-Za-z0-9_. ]+)!", ref):
            if (q or u) not in sheets:
                return True
        return False

    for holder in [wb] + list(wb.worksheets):
        dn = getattr(holder, "defined_names", None)
        if dn is None:
            continue
        for n in list(dn):
            try:
                ref = dn[n].value or ""
            except Exception:
                ref = ""
            if is_bad(n, ref):
                del dn[n]


def _normalize_xlsx(path):
    """Rewrite an openpyxl-saved package the way Excel writes it: move
    inline strings into xl/sharedStrings.xml and untype empty cells.
    DOM-based loaders fetch sharedStrings unconditionally and crash
    ('getElementsByTagName of null') when it is absent."""
    import shutil
    import zipfile

    SST_T = ("application/vnd.openxmlformats-officedocument.spreadsheetml"
             ".sharedStrings+xml")
    SST_R = ("http://schemas.openxmlformats.org/officeDocument/2006/"
             "relationships/sharedStrings")
    strings, index = [], {}
    total_refs = 0

    def fix_sheet(xml):
        nonlocal total_refs

        def repl(m):
            nonlocal total_refs
            head, tail, inner = m.group(1), m.group(2), m.group(3) or ""
            t = re.search(r"<t[^>]*>.*?</t>", inner, re.S)
            t_elem = t.group(0) if t else "<t/>"
            if t_elem not in index:
                index[t_elem] = len(strings)
                strings.append(t_elem)
            total_refs += 1
            return f'<c{head} t="s"{tail}><v>{index[t_elem]}</v></c>'

        xml = re.sub(
            r'<c([^>]*?) t="inlineStr"([^>]*)>(?:<is>(.*?)</is>)?</c>',
            repl, xml, flags=re.S)
        # empty cells typed as numbers -> untyped empty cells
        xml = re.sub(r'<c([^>]*?) t="n"([^>]*)></c>', r"<c\1\2/>", xml)
        xml = re.sub(r'<c([^>]*?) t="n"([^>]*)/>', r"<c\1\2/>", xml)
        return xml

    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        parts = {n: zin.read(n) for n in zin.namelist()}
        for name in list(parts):
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                parts[name] = fix_sheet(parts[name].decode("utf-8")
                                        ).encode("utf-8")
            elif re.match(r"xl/tables/table\d+\.xml$", name):
                # the template table was query-backed; its connection parts
                # are gone, so strip the external-data markers (Excel
                # otherwise repairs with 'Removed Part: External data range')
                t = parts[name].decode("utf-8")
                t = t.replace(' tableType="queryTable"', "")
                t = re.sub(r' queryTableFieldId="\d+"', "", t)
                m = re.search(r'<table[^>]* ref="([^"]+)"', t)
                if m:      # sync the autoFilter range to the table range
                    t = re.sub(r'<autoFilter ref="[^"]+"/>',
                               f'<autoFilter ref="{m.group(1)}"/>', t)
                parts[name] = t.encode("utf-8")
            elif name == "xl/workbook.xml":
                w = parts[name].decode("utf-8")
                w = re.sub(r"<connections[^/>]*/>", "", w)
                parts[name] = w.encode("utf-8")
        if strings:
            sst = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<sst xmlns="http://schemas.openxmlformats.org/'
                   'spreadsheetml/2006/main" '
                   f'count="{total_refs}" uniqueCount="{len(strings)}">'
                   + "".join(f"<si>{t}</si>" for t in strings) + "</sst>")
            parts["xl/sharedStrings.xml"] = sst.encode("utf-8")
            ct = parts["[Content_Types].xml"].decode("utf-8")
            if "sharedStrings" not in ct:
                ct = ct.replace("</Types>",
                                f'<Override PartName="/xl/sharedStrings.xml"'
                                f' ContentType="{SST_T}"/></Types>')
                parts["[Content_Types].xml"] = ct.encode("utf-8")
            rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
            if "sharedStrings" not in rels:
                rels = rels.replace(
                    "</Relationships>",
                    f'<Relationship Id="rIdSST" Type="{SST_R}" '
                    f'Target="sharedStrings.xml"/></Relationships>')
                parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
        for name, data in parts.items():
            zout.writestr(name, data)
    shutil.move(tmp, path)


def write_workbook(template, out_path, prop, months, lines,
                   raw_only=False, keep_raw=False, header_note=None,
                   real_idx=None, note_line=None):
    """Fill RawData with the coded statement. Unless raw_only, also build the
    send-out 'Trailing Financials' tab as formatted VALUES (replicating the
    Power BI grouping) and delete all helper tabs.

    `months` may be SHORTER than 12 (partial-period statement). The unused
    template month columns are left genuinely blank - never zero-filled and
    never annualized - and the Total column stays in its template position
    (RawData/Final T-12 col O, Trailing Financials col N) so every downstream
    formula and the model import still line up. `Total` therefore means
    "sum of the months actually reported".

    `real_idx` (--pad-to-12) is the list of display-column indexes that carry
    real data, e.g. [6,7,8,9,10,11] when a six-month statement is shown on a
    full trailing-twelve axis. `months` then holds all twelve dated headers
    while every `Line.values` list still holds only the real months: padded
    columns are simply never written, so they stay genuinely EMPTY (not zero)
    and every sum, sum-check and reconciliation continues to run on real
    months only.

    `note_line` is an extra red note written under the Trailing Financials
    header (used to say which months ownership did not provide)."""
    n = len(months)
    if n > 12:
        raise ValueError(f"{n} month columns - template holds 12")
    slots = list(real_idx) if real_idx is not None else list(range(n))
    wb = load_workbook(template)

    # ---------------- RawData ----------------
    ws = wb["RawData"]
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        for c in row:
            c.value = None
    for i, d in enumerate(months):
        ws.cell(row=5, column=3 + i, value=d.replace(day=1))
    r = 6
    for ln in lines:
        if ln.kind == "section":
            ws.cell(row=r, column=2, value=ln.name)
        else:
            code = ln.code
            if ln.kind == "subtotal":
                nm = re.sub(r"\s+", " ", ln.name.strip().lower())
                code = ""
                # match by pattern - statements print these rows under many
                # names ("Total Expenses", "Total Operating Expenses", ...)
                if re.fullmatch(r"total (revenues?|incomes?)", nm):
                    code = "rev"
                elif re.fullmatch(r"total (operating )?expenses?", nm):
                    code = "exp"
                elif re.fullmatch(r"(total )?(net operating income|noi)", nm):
                    code = "noi"
            if code:
                ws.cell(row=r, column=1, value=code)
                if ln.review:
                    ws.cell(row=r, column=17, value="REVIEW: " + ln.method)
            elif ln.kind == "account" and ln.below:
                ws.cell(row=r, column=17, value="below-the-line: excluded")
            ws.cell(row=r, column=2, value=ln.name)
            # Line.empty marks month cells the source left BLANK. A blank is
            # not a zero (Heritage Ridge: no water bill posted for the month
            # the statement was cut) - write nothing there. The Total column
            # stays the sum of the months that do carry an amount.
            emask = getattr(ln, "empty", None) or [False] * len(ln.values)
            for i, v in enumerate(ln.values):
                if emask[i]:
                    continue
                ws.cell(row=r, column=3 + slots[i], value=round(v, 2))
            ws.cell(row=r, column=15, value=round(sum(ln.values), 2))
        r += 1

    # ---- RawData sum-check (rows 1-3 equivalent, cols C..O) ----
    # Recompute the sheet's own Revenue/Expense/NOI check rows from the
    # values just written; each row-total must be within +/- 10.
    inc_s = exp_s = rev_s = expx_s = noi_s = 0.0
    for rr in ws.iter_rows(min_row=6, max_row=r - 1):
        code = str(rr[0].value or "").strip().lower()
        vals = sum(c.value or 0 for c in rr[2:15])   # C..O
        if code in INCOME_CODES:
            inc_s += vals
        elif code in EXPENSE_CODES:
            exp_s += vals
        elif code == "rev":
            rev_s += vals
        elif code == "exp":
            expx_s += vals
        elif code == "noi":
            noi_s += vals
    checks = {"Total Revenue Check": inc_s - rev_s,
              "Expense Check": exp_s - expx_s,
              "NOI Check": (inc_s - exp_s) - noi_s}
    sumcheck_ok = all(abs(v) <= 10 for v in checks.values())
    for label, v in checks.items():
        print(f"  RawData {label}: {v:+,.2f} "
              f"({'OK' if abs(v) <= 10 else 'FAIL (tolerance +/-10)'})")
    if not sumcheck_ok:
        print("  !! Sum-check failed - keeping the RawData tab in the "
              "output for inspection.")

    if raw_only:
        _purge_broken_names(wb)
        wb.save(out_path)
        _normalize_xlsx(out_path)
        return sumcheck_ok

    # ------------- Trailing Financials as formatted values -------------
    from openpyxl.styles import Font, Border, Side

    accounts = [l for l in lines if l.kind == "account" and l.code]
    tf = wb["Trailing Financials"]
    acct_fmt = tf["B7"].number_format          # accounting format
    base = tf["A7"].font
    f_norm = Font(name=base.name, size=base.size)
    f_bold = Font(name=base.name, size=base.size, bold=True)
    thin_bottom = Border(bottom=Side(style="thin"))
    top_border = Border(top=Side(style="thin"))

    for row in tf.iter_rows(min_row=3, max_row=tf.max_row):
        for c in row:
            c.value = None
            c.font = f_norm
            c.border = Border()
            c.number_format = "General"

    tf["A1"] = prop
    tf["A2"] = header_note or (
        "Trailing 12 Month Income/Expense" if n == 12 else
        f"PARTIAL PERIOD - {n} MONTHS ONLY ({months[0]:%b %Y} - "
        f"{months[-1]:%b %Y}); NOT a T-12. Total column = sum of these "
        f"{n} months only.")
    tf["A1"].font = f_bold
    tf["A2"].font = f_bold
    for i, d in enumerate(months):
        c = tf.cell(row=3, column=2 + i, value=d.replace(day=1))
        c.number_format = "mmm-yy"
        c.font = f_bold
        c.border = thin_bottom
    c = tf.cell(row=3, column=14, value="Total")
    c.font = f_bold
    c.border = thin_bottom
    tf["A3"].border = thin_bottom
    if note_line:
        # dedicated red note line, directly under the month header row and
        # above the first category (row 4 is blank in the template)
        nc = tf.cell(row=4, column=1, value=note_line)
        nc.font = Font(name=base.name, size=base.size, bold=True,
                       color="9C0006")

    def members(code):
        if code == "ro":
            return [l for l in accounts if l.code in ("ro", "rt")]
        return [l for l in accounts if l.code == code]

    def put(r, name, vals=None, bold=False, border=False, mask=None):
        c = tf.cell(row=r, column=1, value=name)
        c.font = f_bold if bold else f_norm
        if vals is not None:
            mask = mask or [False] * len(vals)
            for i, v in enumerate(vals):
                # blank in the source stays blank on the send-out tab
                cc = tf.cell(row=r, column=2 + slots[i],
                             value=None if mask[i] else round(v, 2))
                cc.number_format = acct_fmt
                cc.font = f_bold if bold else f_norm
                if border:
                    cc.border = top_border
            cc = tf.cell(row=r, column=14, value=round(sum(vals), 2))
            cc.number_format = acct_fmt
            cc.font = f_bold          # Total column always bold (house rule)
            if border:
                cc.border = top_border
        if border:
            c.border = top_border

    r = 5
    for code in INCOME_ORDER:
        mem = members(code)
        if not mem:          # trim category heads with no accounts under them
            continue
        put(r, CATEGORY_NAMES[code], bold=True)      # category head: no zeros
        r += 1
        for l in mem:
            put(r, l.name, l.values, mask=getattr(l, "empty", None))
            r += 1
    # month count comes from the statement, not a hardcoded 12: a --allow-
    # partial run has fewer columns and must not index past its own data
    inc_vals = [sum(l.values[i] for l in accounts
                    if l.code in INCOME_CODES) for i in range(len(slots))]
    put(r, "Total Revenue", inc_vals, bold=True, border=True)
    r += 2
    put(r, "Expense", bold=True)
    r += 1
    for code in EXPENSE_ORDER:
        mem = members(code)
        if not mem:          # trim category heads with no accounts under them
            continue
        put(r, CATEGORY_NAMES[code], bold=True)
        r += 1
        for l in mem:
            put(r, l.name, l.values, mask=getattr(l, "empty", None))
            r += 1
    exp_vals = [sum(l.values[i] for l in accounts
                    if l.code in EXPENSE_CODES) for i in range(len(slots))]
    put(r, "Total Expense", exp_vals, bold=True, border=True)
    r += 2
    noi = [a - b for a, b in zip(inc_vals, exp_vals)]
    put(r, "Net Operating Income", noi, bold=True, border=True)

    # ------------- Final T-12 (hidden, kept for model import) -------------
    ft = wb["Final T-12"]
    for row in ft.iter_rows(min_row=2, max_row=ft.max_row):
        for c in row:
            c.value = None
    ft["A2"] = "date"
    ft["B2"] = "Account"
    for i, d in enumerate(months):
        ft.cell(row=2, column=3 + i, value=d.replace(day=1))
    ft.cell(row=2, column=15, value="Total")

    def ft_put(r, code, name, vals=None):
        ft.cell(row=r, column=1, value=code)
        ft.cell(row=r, column=2, value=name)
        if vals is not None:
            for i, v in enumerate(vals):
                ft.cell(row=r, column=3 + slots[i], value=round(v, 2))
            ft.cell(row=r, column=15, value=round(sum(vals), 2))

    fr = 4
    for code in INCOME_ORDER:
        ft_put(fr, code, CATEGORY_NAMES[code])
        fr += 1
        for l in members(code):
            ft_put(fr, l.code, l.name, l.values)
            fr += 1
    ft_put(fr, "rev", "Total Revenue", inc_vals)
    fr += 2
    ft.cell(row=fr, column=2, value="Expense")
    fr += 1
    for code in EXPENSE_ORDER:
        ft_put(fr, code, CATEGORY_NAMES[code])
        fr += 1
        for l in members(code):
            ft_put(fr, l.code, l.name, l.values)
            fr += 1
    ft_put(fr, "exp", "Total Expense", exp_vals)
    fr += 2
    ft_put(fr, "noi", "Net Operating Income", noi)
    if "Final_T_12" in ft.tables:
        ft.tables["Final_T_12"].ref = f"A1:O{fr}"
    ft.sheet_state = "hidden"

    # ------------- delete helper tabs -------------
    keep = ("Trailing Financials", "Final T-12", "RawData") \
        if (keep_raw or not sumcheck_ok) \
        else ("Trailing Financials", "Final T-12")
    for name in list(wb.sheetnames):
        if name not in keep:
            del wb[name]
    wb.move_sheet("Trailing Financials", -wb.sheetnames.index(
        "Trailing Financials"))

    _purge_broken_names(wb)
    wb.calculation.calcId = 191029
    wb.calculation.fullCalcOnLoad = True
    wb.save(out_path)
    _normalize_xlsx(out_path)
    return sumcheck_ok


def write_capex(out_path, prop, months, lines, real_idx=None):
    """Second output: all below-the-line items (Debt Service, CapEx, Cash
    Flow Adjustments, ...) in the same formatted-values layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side

    below = [l for l in lines if l.below]
    if not any(l.kind == "account" for l in below):
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Capex & Misc"
    ws.sheet_view.showGridLines = False
    f_norm = Font(name="Calibri", size=11)
    f_bold = Font(name="Calibri", size=11, bold=True)
    thin_bottom = Border(bottom=Side(style="thin"))
    top_border = Border(top=Side(style="thin"))
    ACCT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"_);_(@_)'

    ws["A1"] = prop
    n = len(months)
    slots = list(real_idx) if real_idx is not None else list(range(n))
    ws["A2"] = ("Capex & Miscellaneous (Below the Line) - "
                f"{months[0]:%b %Y} to {months[-1]:%b %Y}"
                + ("" if n == 12 else
                   f"  [PARTIAL PERIOD: {n} months only]"))
    ws["A1"].font = f_bold
    ws["A2"].font = f_bold
    for i, d in enumerate(months):
        c = ws.cell(row=3, column=2 + i, value=d.replace(day=1))
        c.number_format = "mmm-yy"
        c.font = f_bold
        c.border = thin_bottom
    c = ws.cell(row=3, column=14, value="Total")
    c.font = f_bold
    c.border = thin_bottom
    ws["A3"].border = thin_bottom
    ws.column_dimensions["A"].width = 42
    for col in "BCDEFGHIJKLMN":
        ws.column_dimensions[col].width = 11

    def put(r, name, vals=None, bold=False, border=False):
        c = ws.cell(row=r, column=1, value=name)
        c.font = f_bold if bold else f_norm
        if border:
            c.border = top_border
        if vals is not None:
            # months in B.., Total pinned to col N regardless of month count
            for i, v in enumerate(vals):
                cc = ws.cell(row=r, column=2 + slots[i], value=round(v, 2))
                cc.number_format = ACCT
                cc.font = f_bold if bold else f_norm
                if border:
                    cc.border = top_border
            cc = ws.cell(row=r, column=14, value=round(sum(vals), 2))
            cc.number_format = ACCT
            cc.font = f_bold
            if border:
                cc.border = top_border

    # keep source order: section head (bold, no values), accounts, subtotals
    r = 5
    grand = [0.0] * len(slots)
    pending_section = None
    for ln in below:
        if ln.kind == "section":
            pending_section = ln.name
            continue
        if pending_section:
            put(r, pending_section, bold=True)
            pending_section = None
            r += 1
        if ln.kind == "account":
            put(r, ln.name, ln.values)
            grand = [g + v for g, v in zip(grand, ln.values)]
        else:                      # printed subtotal row
            put(r, ln.name, ln.values, bold=True, border=True)
            r += 1                 # blank line after each section total
        r += 1
    put(r, "Total Capex & Misc", grand, bold=True, border=True)

    wb.save(out_path)
    _normalize_xlsx(out_path)
    return True


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main(argv=None):
    if argv is None:
        argv = sys.argv[1:]
    if argv and argv[0] == "harvest":
        ap = argparse.ArgumentParser()
        ap.add_argument("cmd")
        ap.add_argument("files", nargs="+")
        ap.add_argument("--mappings", default="t12_mappings.csv")
        a = ap.parse_args(argv)
        harvest(a.files, a.mappings)
        return 0

    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("pdf")
    ap.add_argument("--template", required=True,
                    help="T12 Processor workbook to fill")
    ap.add_argument("-o", "--output")
    ap.add_argument("--mappings", default=None,
                    help="corpus CSV (default: t12_mappings.csv next to "
                         "this script)")
    ap.add_argument("--property", help="override property name")
    ap.add_argument("--raw-only", action="store_true",
                    help="only fill the RawData tab (keep all template tabs "
                         "untouched) for the Power BI flow")
    ap.add_argument("--keep-raw", action="store_true",
                    help="keep the coded RawData tab in the T-12 output")
    ap.add_argument("--header-note",
                    help="override the Trailing Financials A2 header line")
    ap.add_argument("--trust-monthly", action="store_true",
                    help="xlsx T12s: monthly detail wins over hardcoded "
                         "printed row totals (variance reported, grand "
                         "totals adjusted for reconciliation)")
    ap.add_argument("--allow-partial", action="store_true",
                    help="accept a statement with fewer than 12 month "
                         "columns (produces a PARTIAL-period workbook, not a "
                         "T-12; months are never padded or annualised)")
    ap.add_argument("--trailing", type=int, default=12,
                    help="owner month-per-column statements that are WIDER "
                         "than a T-12 (multi-year grids): number of trailing "
                         "months to cut out (default 12; 0 = keep all)")
    ap.add_argument("--months-ending", metavar="YYYY-MM",
                    help="end the trailing window on this month instead of "
                         "the last month carrying data")
    ap.add_argument("--exclude-account", action="append", metavar="NAME",
                    default=None,
                    help="drop a named account line from operations "
                         "entirely (repeatable). For lines OWNERSHIP HAS "
                         "CONFIRMED are not real costs - typically a "
                         "hand-keyed duplicate of another line. The removed "
                         "account, side, every monthly value removed and the "
                         "reason are printed, added to the delivery notes "
                         "and written as a red note on the Trailing "
                         "Financials tab: an exclusion is never silent. A "
                         "name that matches nothing aborts the run.")
    ap.add_argument("--exclude-reason", default="",
                    help="why the --exclude-account lines were removed; "
                         "quoted verbatim in the run output, the notes and "
                         "the workbook")
    ap.add_argument("--pad-to-12", action="store_true",
                    help="show a short statement on a full trailing-12 axis "
                         "ending at its last real month: the missing months "
                         "get dated column headers but genuinely EMPTY data "
                         "cells (never zeros), the Total column stays the sum "
                         "of the real months, and a red note names the months "
                         "ownership did not provide. Implies --allow-partial.")
    args = ap.parse_args(argv)
    if args.pad_to_12:
        args.allow_partial = True

    mpath = args.mappings or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "t12_mappings.csv")
    by_gl, by_name = load_corpus(mpath)
    print(f"Corpus: {len(by_name)} account mappings loaded"
          if by_name else "Corpus: none found - rules/fuzzy layers only")

    meta = None
    if _is_owner_rental_schedule(args.pdf):
        # owner-made month-per-column grid (.xls/.xlsx), usually multi-year
        prop, months, lines, meta = parse_t12_owner_rental_schedule(
            args.pdf, trust_monthly=args.trust_monthly,
            allow_partial=args.allow_partial, window=args.trailing,
            window_end=args.months_ending, exclude=args.exclude_account,
            exclude_reason=args.exclude_reason)
    elif args.pdf.lower().endswith((".xlsx", ".xlsm")):
        prop, months, lines = parse_t12_xlsx(
            args.pdf, trust_monthly=args.trust_monthly,
            allow_partial=args.allow_partial)
    elif is_gsheet_pdf(args.pdf):
        prop, months, lines, meta = parse_t12_gsheet_pdf(args.pdf)
    elif is_qbo_pdf(args.pdf):
        prop, months, lines, meta = parse_t12_qbo_pdf(args.pdf)
    else:
        prop, months, lines = parse_t12_pdf(args.pdf)
    if args.property:
        prop = args.property
    # A T-12 must be twelve months, whatever the source. The Google-Sheets
    # statement parser (the only one that returns `meta`) reads an explicit
    # reporting period out of the document and labels every output PARTIAL
    # PERIOD, so it is exempt; everything else must be full or opted in.
    if meta is None or not meta.get("exempt_full_year", True):
        _require_full_year(months, args.pdf, allow_partial=args.allow_partial)
    # Parsers that verify printed roll-ups internally apply exclusions
    # themselves (before those checks run); for every other source the same
    # flag is honoured here, right after the parse.
    if args.exclude_account and not (meta or {}).get("exclusions_applied"):
        lines, _excl = apply_exclusions(lines, months, args.exclude_account,
                                        args.exclude_reason)
        for n in _excl:
            print("   " + n)
    n_acct = sum(1 for l in lines if l.kind == "account")
    n_mo = len(months)
    print(f"Parsed: {prop} | {months[0]:%b %Y}-{months[-1]:%b %Y} | "
          f"{n_mo} month(s) | {n_acct} account lines")

    # ---- --pad-to-12: full trailing-12 axis, padded months left EMPTY ----
    real_idx, pad_note = None, None
    if args.pad_to_12 and n_mo < 12:
        last = months[-1]
        axis = []
        for k in range(11, -1, -1):
            y, mo = last.year, last.month - k
            while mo < 1:
                mo += 12
                y -= 1
            axis.append(datetime(y, mo, 1))
        real = axis[12 - n_mo:]
        if [(d.year, d.month) for d in real] != \
                [(d.year, d.month) for d in months]:
            sys.exit("ERROR: --pad-to-12 needs contiguous months ending at "
                     "the last reported month; got "
                     f"{[f'{d:%b %Y}' for d in months]}.")
        pad = axis[:12 - n_mo]
        real_idx = list(range(12 - n_mo, 12))
        months = axis[:12 - n_mo] + list(months)   # keep the real objects
        n_mo = 12
        pad_note = (f"{pad[0]:%b %Y} - {pad[-1]:%b %Y} not provided by "
                    f"ownership - columns intentionally blank.")
        print(f"!! --pad-to-12: trailing-12 axis {axis[0]:%b %Y}-"
              f"{axis[-1]:%b %Y}; {len(pad)} month(s) "
              f"({pad[0]:%b %Y}-{pad[-1]:%b %Y}) have NO data and are written "
              f"as EMPTY cells (never zeros). Total column and every "
              f"sum-check/reconciliation still run on the "
              f"{len(real_idx)} real months only.")

    # ---- partial-period banner -------------------------------------------
    stub = bool(meta and meta.get("stub_last"))
    if n_mo < 12:
        print(f"!! PARTIAL PERIOD: {n_mo} of 12 months. NOT a T-12. "
              f"Missing months are left BLANK (never zero-filled, never "
              f"annualized); Total = sum of the {n_mo} reported months.")
    if meta:
        if meta.get("period"):
            a, b = meta["period"]
            print(f"   Reporting period: {a:%m/%d/%Y} - {b:%m/%d/%Y}"
                  + (f"  [{meta['basis']} basis]" if meta["basis"] else ""))
        if stub:
            b = meta["period"][1]
            print(f"!! STUB MONTH: {b:%b %Y} covers only "
                  f"{b.replace(day=1):%m/%d}-{b:%m/%d/%Y} "
                  f"({b.day} of {_month_end(b.year, b.month).day} days) - "
                  f"do not treat it as a full month.")
        for note in meta.get("notes", []):
            print(f"   NOTE: {note}")
        for d in meta.get("dropped_cols", []):
            print(f"   DROPPED (outside period, value was zero): {d}")
        if meta.get("blank_cells"):
            print("   " + meta.get(
                "blank_label",
                "In-period blank cells read as $0.00 (validated against "
                "the printed section total)") + ": "
                  + "; ".join(f"{a} [{m}]"
                              for a, m in meta["blank_cells"]))
        if meta.get("printed_checks"):
            print("Printed row/section totals vs parsed monthly detail "
                  "(monthly detail wins - house rule):")
            for c in meta["printed_checks"]:
                print(c)

    map_codes(lines, by_gl, by_name)
    ok, report = reconcile(lines)

    stats = {}
    flagged = []
    for l in lines:
        if l.kind != "account":
            continue
        key = l.method.split(":")[0].split(" ")[0]
        stats[key] = stats.get(key, 0) + 1
        if l.review:
            flagged.append(l)
    n_below = stats.pop("below-the-line", 0)

    sprop = short_prop(prop)
    out = args.output
    if not out:
        if n_mo == 12:
            out = f"T-12 - {safe_name(sprop)} - {months[-1]:%B %Y}.xlsx"
        else:
            span = (f"{months[0]:%b}-{months[-1]:%b %Y}"
                    if months[0].year == months[-1].year
                    else f"{months[0]:%b %Y}-{months[-1]:%b %Y}")
            out = (f"T-{n_mo} ({span}) - {safe_name(sprop)}.xlsx")
    header_note = args.header_note
    if not header_note and n_mo < 12:
        span = (f"{months[0]:%b}-{months[-1]:%b %Y}"
                if months[0].year == months[-1].year
                else f"{months[0]:%b %Y} - {months[-1]:%b %Y}")
        header_note = (
            f"*** PARTIAL PERIOD - T-{n_mo} ({span}), NOT A T-12: only "
            f"{n_mo} of 12 months reported. Blank month columns = no data "
            f"(not zero); Total column = sum of these {n_mo} months only; "
            f"nothing annualized.")
        if meta and meta.get("period"):
            a, b = meta["period"]
            header_note += (f" Period {a:%m/%d/%Y}-{b:%m/%d/%Y}"
                            + (f", {meta['basis']} basis." if meta["basis"]
                               else "."))
        if stub:
            b = meta["period"][1]
            header_note += (f" FINAL MONTH IS A STUB: {b:%b %Y} covers "
                            f"{b.replace(day=1):%m/%d}-{b:%m/%d/%Y} only "
                            f"({b.day} of {_month_end(b.year, b.month).day} "
                            f"days).")
    print("RawData sum-check (written values, tolerance +/- 10):")
    sumcheck_ok = write_workbook(args.template, out, sprop, months, lines,
                                 raw_only=args.raw_only,
                                 keep_raw=args.keep_raw,
                                 header_note=header_note,
                                 real_idx=real_idx,
                                 note_line=pad_note or (
                                     meta.get("note_line") if meta else None))
    capex_out = f"Capex & Misc - {months[-1]:%B %Y}.xlsx"
    wrote_capex = write_capex(capex_out, sprop, months, lines,
                              real_idx=real_idx)

    print(f"Mapping: {stats}" +
          (f" | {n_below} below-the-line lines excluded (kept in RawData, "
           f"uncoded)" if n_below else ""))
    print("Reconciliation vs statement totals:")
    print(report)
    if flagged:
        print(f"\n{len(flagged)} line(s) flagged for review "
              f"(also marked in RawData col Q):")
        for l in flagged:
            print(f"  - {l.name}  [{l.section}] -> '{l.code}' ({l.method})")
    else:
        print("No lines flagged for review.")
    print(f"Output: {out}")
    if wrote_capex:
        print(f"Output: {capex_out} (below-the-line items)")
    return 0 if (ok and sumcheck_ok) else 1


if __name__ == "__main__":
    sys.exit(main())
