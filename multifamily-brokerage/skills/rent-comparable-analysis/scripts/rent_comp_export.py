#!/usr/bin/env python3
"""
TMG Rent Comparable Analysis exporter
=====================================
Consumes Dwellsy Comps API flat CSVs (the --flat output of
dwellsy_comps_lookup.py) plus a subject-property JSON, and produces:

  1. An Excel export ("<Subject> - Rent Comps.xlsx") with:
       - "Rent Comparison Grid"      subject + property-level comp grid
                                     (Property Name | Address | City | State |
                                      Zip | Year Built | # of Units | Avg. Size |
                                      Avg. Rent/Unit | Avg. $/SF | Amenities |
                                      Interior Quality)
       - "Granular Rent Comparison"  per-subject-floor-plan blocks: comps whose
                                     unit SF is within --sf-variance of the plan
       - "Comp Listings"             the cleaned unit-level listing data
  2. A client-facing, TMG-branded PDF ("<Subject> - Rent Comparable
     Analysis.pdf"): cover, rent comparison grid, comps by unit type, subject
     positioning, and rent trending behavior (trend lines, leasing activity,
     rent-vs-size scatter).

Inputs
------
  --csv       glob(s) of Dwellsy flat CSVs (repeatable)
  --subject   JSON file:
     {"name": "...", "address": "...", "city": "...", "state": "TX",
      "zip": "...", "year_built": 1970, "units": 20,
      "lat": 30.14891, "lon": -96.39483,
      "plans": [{"label": "1x1", "bed": 1, "sf": 500,
                 "market": 850, "effective": 900, "units": 6}, ...]}
  --names     optional CSV mapping Dwellsy addresses to verified identities and
              merging multi-address communities:
              address_1,name,units,year_built,community
  --out       output directory (default .)

House rules baked in:
- No sub-5-unit comp properties (--min-units, default 5).
- URBAN MODE (auto when the qualifying pool is plentiful — more than
  --urban-threshold listings — or forced with --urban / --no-urban):
    * comps must be within +/-10 years of the subject's vintage
      (--vintage-window);
    * when the subject is under 100 units, complexes of 150+ units are
      excluded (amenity-package mismatch);
    * rent outliers are pruned per bed type by standard deviation
      (+/-2 sigma, --sd-cut) INSTEAD of the fixed sane-rent band.
  Filters apply as a trim ladder — vintage first keeps unknown-year comps
  only if needed, and each step is relaxed if it would leave fewer than
  --min-comps properties.
- Query side: pull a MINIMUM 3-mile radius per API request, then trim here.
- Medians/averages are PROPERTY-weighted wherever quoted; one community split
  across several addresses is merged via the names CSV before counting;
  Dwellsy unit counts / vintages are labelled "per Dwellsy" unless the names
  CSV marks them verified.
"""

import argparse
import base64
import csv
import glob
import io
import json
import math
import os
import statistics as st
import sys
from collections import defaultdict
from datetime import datetime

# ---------------------------------------------------------------------------
# Brand (chrome only -- data series use the validated categorical palette)
# ---------------------------------------------------------------------------
NAVY = "#1B3E6F"        # TMG brand navy   (headers, rules, table chrome)
GOLD = "#FDB714"        # TMG brand gold   (accents, subject highlight)
MID_NAVY = "#345279"
PALE_NAVY = "#DCE6F2"   # subject-row shading
GRAY_TXT = "#52514E"
LIGHT_GRID = "#D9D9D9"

# Validated categorical palette (dataviz reference instance, light mode).
# Series = unit types, assigned in fixed order Studio,1BR,2BR,3BR+.
SERIES = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100"]
BED_ORDER = ["0", "1", "2", "3"]
BED_LABEL = {"0": "Studio", "1": "1 Bed", "2": "2 Bed", "3": "3 Bed"}

SANE_RENT = (300.0, 10000.0)
EARTH_MI = 3958.8


def as_f(v, default=None):
    try:
        x = float(str(v).replace("$", "").replace(",", "").strip())
        return x
    except (TypeError, ValueError):
        return default


def as_i(v):
    t = str(v).strip()
    return int(float(t)) if t.replace(".", "", 1).isdigit() else None


def haversine(a, b, c, d):
    r = math.radians
    dlat, dlon = r(c - a), r(d - b)
    h = math.sin(dlat / 2) ** 2 + math.cos(r(a)) * math.cos(r(c)) * math.sin(dlon / 2) ** 2
    return 2 * EARTH_MI * math.asin(math.sqrt(h))


def med(v):
    return st.median(v) if v else None


def money(v, dec=0):
    return "" if v in (None, "") else f"${v:,.{dec}f}"


# ---------------------------------------------------------------------------
# Load & clean
# ---------------------------------------------------------------------------

def load_listings(csv_globs, subject, names_map):
    rows, seen = [], set()
    for g in csv_globs:
        for path in sorted(glob.glob(g)):
            with open(path, newline="") as f:
                for r in csv.DictReader(f):
                    key = (r.get("id") or
                           f'{r.get("address_1")}|{r.get("address_2")}|{r.get("last_listing_creation_time")}')
                    if key in seen:
                        continue
                    seen.add(key)
                    r["rent"] = as_f(r.get("listing_amount"))
                    r["sf"] = as_f(r.get("square_feet"))
                    r["beds"] = str(r.get("bedrooms", "")).strip()
                    r["units_dw"] = as_i(r.get("community_unit_count"))
                    r["year_dw"] = as_i(r.get("year_built"))
                    r["active"] = r.get("property_listing_status") == "active"
                    r["created"] = (r.get("last_listing_creation_time") or "")[:10]
                    la, lo = as_f(r.get("latitude")), as_f(r.get("longitude"))
                    r["mi"] = (haversine(subject["lat"], subject["lon"], la, lo)
                               if la and lo and subject.get("lat") else as_f(r.get("distance_miles")))
                    a1 = (r.get("address_1") or "").strip()
                    nm = names_map.get(a1, {})
                    r["prop_key"] = nm.get("community") or a1
                    r["sane"] = (r["rent"] is not None and
                                 SANE_RENT[0] <= r["rent"] <= SANE_RENT[1])
                    rows.append(r)
    return rows


def load_names(path):
    m = {}
    if path and os.path.exists(path):
        with open(path, newline="") as f:
            for r in csv.DictReader(f):
                m[r["address_1"].strip()] = {
                    "name": r.get("name", "").strip(),
                    "units": as_i(r.get("units")),
                    "year_built": as_i(r.get("year_built")),
                    "community": (r.get("community") or "").strip() or None,
                }
    return m


def build_properties(rows, names_map):
    """Collapse unit listings to property level (community-merged)."""
    groups = defaultdict(list)
    for r in rows:
        groups[r["prop_key"]].append(r)
    props = []
    for key, lst in groups.items():
        sane = [r for r in lst if r["sane"]]
        if not sane:
            continue
        a1s = sorted({r["address_1"] for r in lst})
        nm = next((names_map[a] for a in a1s if a in names_map and names_map[a].get("name")), {})
        by_bed = defaultdict(list)
        for r in sane:
            by_bed[r["beds"]].append(r)
        rents = [r["rent"] for r in sane]
        sfs = [r["sf"] for r in sane if r["sf"]]
        psfs = [r["rent"] / r["sf"] for r in sane if r["sf"]]
        props.append({
            "key": key,
            "name": nm.get("name") or key,
            "verified": bool(nm.get("name")),
            "address": (a1s[0] if len(a1s) == 1 else
                        " / ".join(a1s) if len(a1s) <= 3 else
                        f"{a1s[0]} et al. ({len(a1s)} addresses)"),
            "city": lst[0].get("address_city", ""),
            "state": lst[0].get("address_state", ""),
            "zip": lst[0].get("address_zip", ""),
            "units": nm.get("units") or lst[0]["units_dw"],
            "units_src": "verified" if nm.get("units") else "Dwellsy",
            "year": nm.get("year_built") or next((r["year_dw"] for r in lst if r["year_dw"]), None),
            "year_src": "verified" if nm.get("year_built") else "Dwellsy",
            "mi": med([r["mi"] for r in sane if r["mi"] is not None]),
            "n": len(sane),
            "n_active": sum(1 for r in sane if r["active"]),
            "avg_rent": st.mean(rents),
            "avg_sf": st.mean(sfs) if sfs else None,
            "avg_psf": st.mean(psfs) if psfs else None,
            "by_bed": {b: {
                "n": len(v),
                "avg_rent": st.mean([r["rent"] for r in v]),
                "avg_sf": st.mean([r["sf"] for r in v if r["sf"]]) if any(r["sf"] for r in v) else None,
                "min_rent": min(r["rent"] for r in v),
                "max_rent": max(r["rent"] for r in v),
            } for b, v in by_bed.items()},
            "rows": sane,
        })
    props.sort(key=lambda p: (p["mi"] if p["mi"] is not None else 99))
    return props


def bed_benchmarks(props):
    """Property-weighted per-bed medians across the comp set."""
    out = {}
    for b in BED_ORDER:
        prop_avgs, prop_psf, all_rows = [], [], []
        for p in props:
            v = p["by_bed"].get(b)
            if not v:
                continue
            prop_avgs.append(v["avg_rent"])
            if v["avg_sf"]:
                prop_psf.append(v["avg_rent"] / v["avg_sf"])
            all_rows += [r for r in p["rows"] if r["beds"] == b]
        if not prop_avgs:
            continue
        out[b] = {
            "props": len(prop_avgs),
            "n": len(all_rows),
            "prop_med": med(prop_avgs),
            "list_med": med([r["rent"] for r in all_rows]),
            "psf_med": med(prop_psf),
            "sf_med": med([r["sf"] for r in all_rows if r["sf"]]),
        }
    return out


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def write_xlsx(path, subject, props, rows, sf_variance, urban=False):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    navy_fill = PatternFill("solid", fgColor=NAVY.lstrip("#"))
    pale_fill = PatternFill("solid", fgColor=PALE_NAVY.lstrip("#"))
    gold_side = Side(style="thick", color=GOLD.lstrip("#"))
    thin = Side(style="thin", color=LIGHT_GRID.lstrip("#"))
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    sub_font = Font(name="Calibri", bold=True, color=NAVY.lstrip("#"), size=11)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Rent Comparison Grid --------------------------------
    ws = wb.active
    ws.title = "Rent Comparison Grid"
    headers = ["Property Name", "Address", "City", "State", "Zip", "Year Built",
               "# of Units", "Avg. Size", "Avg. Rent/Unit", "Avg. $/SF",
               "Amenities", "Interior Quality"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=gold_side)

    def grid_row(ws, r, vals, subject_row=False):
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = Border(bottom=thin)
            if subject_row:
                cell.fill = pale_fill
                cell.font = sub_font
            if c in (8,):
                cell.number_format = "#,##0"
            if c == 9:
                cell.number_format = '"$"#,##0'
            if c == 10:
                cell.number_format = '"$"0.00'

    # subject first (asking/market rents)
    s_units = sum(p["units"] for p in subject["plans"])
    s_sf = sum(p["sf"] * p["units"] for p in subject["plans"]) / s_units
    s_rent = sum(p["market"] * p["units"] for p in subject["plans"]) / s_units
    grid_row(ws, 2, [subject["name"], subject["address"], subject["city"],
                     subject["state"], subject["zip"], subject.get("year_built") or "",
                     subject["units"], round(s_sf), round(s_rent),
                     round(s_rent / s_sf, 2), "", ""], subject_row=True)
    r = 3
    for p in props:
        grid_row(ws, r, [p["name"], p["address"], p["city"], p["state"], p["zip"],
                         p["year"] or "", p["units"] or "",
                         round(p["avg_sf"]) if p["avg_sf"] else "",
                         round(p["avg_rent"]),
                         round(p["avg_psf"], 2) if p["avg_psf"] else "", "", ""])
        r += 1
    prune_txt = ("rent outliers pruned by standard deviation per bed type"
                 if urban else
                 f"rents outside ${SANE_RENT[0]:,.0f}-${SANE_RENT[1]:,.0f} excluded")
    urban_txt = (" Urban comp rules applied: vintage within 10 yrs of subject; "
                 "no 150+ unit complexes for a sub-100-unit subject."
                 if urban else "")
    note = (f"Comp data: Dwellsy Comps API, {len(rows)} unit listings collapsed to "
            f"{len(props)} properties; {prune_txt}. Subject Avg. Rent/Unit is the "
            f"unit-weighted market (asking) rent. Unverified unit counts / vintages "
            f"are per Dwellsy. House rule: comp properties under 5 units are "
            f"excluded from this comp set." + urban_txt)
    ws.cell(row=r + 1, column=1, value=note).font = Font(size=8, italic=True,
                                                         color="808080")
    widths = [26, 30, 12, 7, 8, 10, 9, 9, 13, 9, 22, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ---- Sheet 2: Granular per-floor-plan blocks ----------------------
    ws2 = wb.create_sheet("Granular Rent Comparison")
    ws2.cell(row=1, column=1, value="Granular - Rent Comparison Grid").font = \
        Font(bold=True, size=13, color=NAVY.lstrip("#"))
    ws2.cell(row=2, column=1, value="Max Variance in floor plan SF")
    ws2.cell(row=2, column=4, value=sf_variance)
    r = 4
    for plan in subject["plans"]:
        bed = str(plan["bed"])
        lo, hi = plan["sf"] * (1 - sf_variance), plan["sf"] * (1 + sf_variance)
        ws2.cell(row=r, column=1, value=plan["label"]).font = sub_font
        ws2.cell(row=r, column=2, value=BED_LABEL.get(bed, bed).upper()).font = sub_font
        r += 1
        for c, h in enumerate(["PROPERTY", "SIZE", "RENT", "$/SF"], start=1):
            cell = ws2.cell(row=r, column=c, value=h)
            cell.font = hdr_font
            cell.fill = navy_fill
        r += 1
        for p in props:
            match = [x for x in p["rows"] if x["beds"] == bed and x["sf"] and lo <= x["sf"] <= hi]
            if not match:
                continue
            sf = st.mean([x["sf"] for x in match])
            rent = st.mean([x["rent"] for x in match])
            ws2.cell(row=r, column=1, value=p["name"])
            ws2.cell(row=r, column=2, value=round(sf)).number_format = "#,##0"
            ws2.cell(row=r, column=3, value=round(rent)).number_format = '"$"#,##0'
            ws2.cell(row=r, column=4, value=round(rent / sf, 2)).number_format = '"$"0.00'
            r += 1
        # subject row last, shaded (mirrors the example workbook layout)
        ws2.cell(row=r, column=1, value=subject["name"]).font = sub_font
        ws2.cell(row=r, column=2, value=plan["sf"]).number_format = "#,##0"
        ws2.cell(row=r, column=3, value=plan["market"]).number_format = '"$"#,##0'
        ws2.cell(row=r, column=4, value=round(plan["market"] / plan["sf"], 2)
                 ).number_format = '"$"0.00'
        for c in range(1, 5):
            ws2.cell(row=r, column=c).fill = pale_fill
        r += 3
    for col, w in zip("ABCD", [28, 9, 10, 8]):
        ws2.column_dimensions[col].width = w

    # ---- Sheet 3: cleaned listings ------------------------------------
    ws3 = wb.create_sheet("Comp Listings")
    cols = ["prop_key", "address_1", "address_2", "address_city", "beds",
            "bathrooms", "sf", "year_dw", "rent", "created",
            "property_listing_status", "mi", "units_dw"]
    heads = ["Property", "Address", "Unit", "City", "Beds", "Baths", "SF",
             "Year Built*", "Rent", "Listed", "Status", "Miles", "# Units*"]
    ws3.append(heads)
    for c in range(1, len(heads) + 1):
        ws3.cell(row=1, column=c).font = hdr_font
        ws3.cell(row=1, column=c).fill = navy_fill
    for i, x in enumerate(sorted(rows, key=lambda x: (x["mi"] or 99)), start=2):
        for c, k in enumerate(cols, start=1):
            v = x.get(k)
            if k == "mi" and v is not None:
                v = round(v, 2)
            ws3.cell(row=i, column=c, value=v)
    ws3.cell(row=len(rows) + 3, column=1,
             value="* per Dwellsy; verify before citing").font = \
        Font(size=8, italic=True, color="808080")
    ws3.freeze_panes = "A2"
    for col, w in zip("ABCDEFGHIJKLM", [24, 22, 8, 12, 6, 6, 7, 10, 9, 11, 9, 7, 8]):
        ws3.column_dimensions[col].width = w

    wb.save(path)


# ---------------------------------------------------------------------------
# Charts (matplotlib -> PNG buffers for the PDF)
# ---------------------------------------------------------------------------

def _style_ax(ax):
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines[["left", "bottom"]].set_color(LIGHT_GRID)
    ax.tick_params(colors=GRAY_TXT, labelsize=8)
    ax.grid(axis="y", color=LIGHT_GRID, linewidth=0.6, alpha=0.6)
    ax.set_axisbelow(True)


def chart_rankings(props, subject, bench):
    """One horizontal-bar panel per bed type: property avg rents, subject
    highlighted in gold, comp median as a navy dashed line."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    MAX_BARS = 12   # client-facing: nearest properties only, keep it readable
    beds = [b for b in BED_ORDER if b in bench]
    fig, axes = plt.subplots(1, len(beds), figsize=(10.4, 3.4))
    if len(beds) == 1:
        axes = [axes]
    for ax, b in zip(axes, beds):
        rows = []
        for p in props:                       # props arrive distance-sorted
            v = p["by_bed"].get(b)
            if v:
                rows.append((p["name"], v["avg_rent"], False))
        dropped = max(0, len(rows) - MAX_BARS)
        rows = rows[:MAX_BARS]
        splan = [p for p in subject["plans"] if str(p["bed"]) == b]
        if splan:
            s_rent = sum(p["market"] * p["units"] for p in splan) / sum(p["units"] for p in splan)
            rows.append((subject["name"], s_rent, True))
        rows.sort(key=lambda x: x[1])
        names = [n if len(n) <= 24 else n[:22] + "…" for n, _, _ in rows]
        vals = [v for _, v, _ in rows]
        colors = [GOLD if s else "#B9C6DA" for _, _, s in rows]
        ax.barh(range(len(rows)), vals, color=colors, height=0.62, zorder=3)
        for i, (n, v, s) in enumerate(rows):
            ax.text(v + max(vals) * 0.015, i, money(v), va="center", fontsize=7.5,
                    color="#0b0b0b", fontweight="bold" if s else "normal")
        m = bench[b]["prop_med"]
        ax.axvline(m, color=NAVY, linestyle=(0, (4, 3)), linewidth=1.2, zorder=4)
        ax.text(m, -0.62, f"comp median {money(m)} ", fontsize=7, color=NAVY,
                va="top", ha="right")
        ax.set_ylim(-1.1, len(rows) - 0.4)
        ax.set_yticks(range(len(rows)))
        ax.set_yticklabels(names, fontsize=7.5)
        for tl, (_, _, s) in zip(ax.get_yticklabels(), rows):
            if s:
                tl.set_fontweight("bold")
                tl.set_color(NAVY)
        title = f"{BED_LABEL.get(b, b)} — avg asking rent"
        if dropped:
            title += f" (nearest {MAX_BARS})"
        ax.set_title(title, fontsize=9.5, color=NAVY, loc="left",
                     fontweight="bold")
        ax.set_xlim(0, max(vals) * 1.22)
        _style_ax(ax)
        ax.grid(axis="y", visible=False)
        ax.grid(axis="x", color=LIGHT_GRID, linewidth=0.6, alpha=0.6)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200)
    plt.close(fig)
    return buf


def chart_trend(rows):
    """Quarterly median asking rent by bed type over the lookback window.
    Quarterly (not monthly) so thin small-market months don't read as noise;
    x is a shared sorted index so multi-series ordering can never scramble."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    quarterly = defaultdict(lambda: defaultdict(list))
    for r in rows:
        if not (r["sane"] and r["created"]):
            continue
        qu = f'{r["created"][:4]} Q{(int(r["created"][5:7]) - 1) // 3 + 1}'
        quarterly[r["beds"]][qu].append(r["rent"])
    quarters = sorted({q for d in quarterly.values() for q in d})
    idx = {q: i for i, q in enumerate(quarters)}
    fig, ax = plt.subplots(figsize=(6.4, 3.2))
    plotted = 0
    for b in BED_ORDER:
        if b not in quarterly:
            continue
        pts = [(idx[q], med(v), len(v)) for q, v in sorted(quarterly[b].items())]
        if len(pts) < 2:
            continue
        color = SERIES[BED_ORDER.index(b)]
        xs, ys = [p[0] for p in pts], [p[1] for p in pts]
        ax.plot(xs, ys, color=color, linewidth=2, marker="o", markersize=3.5,
                label=BED_LABEL.get(b, b), zorder=3)
        ax.annotate(BED_LABEL.get(b, b), (xs[-1], ys[-1]), textcoords="offset points",
                    xytext=(6, -3), fontsize=8, color="#0b0b0b", fontweight="bold")
        plotted += 1
    if not plotted:
        plt.close(fig)
        return None
    ax.set_title("Median asking rent by quarter listed", fontsize=10, color=NAVY,
                 loc="left", fontweight="bold")
    ax.set_xticks(range(len(quarters)))
    ax.set_xticklabels(quarters)
    ax.set_xlim(-0.4, len(quarters) - 0.2)
    ax.tick_params(axis="x", rotation=45)
    ax.yaxis.set_major_formatter(lambda v, _: f"${v:,.0f}")
    ax.legend(frameon=False, fontsize=8, loc="upper left")
    _style_ax(ax)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200)
    plt.close(fig)
    return buf


def chart_activity(rows):
    """Listings placed on market per quarter, stacked by bed type."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    q = defaultdict(lambda: defaultdict(int))
    for r in rows:
        if not (r["sane"] and r["created"]):
            continue
        y, m = r["created"][:4], int(r["created"][5:7])
        qu = f"{y} Q{(m - 1) // 3 + 1}"
        q[qu][r["beds"]] += 1
    quarters = sorted(q)
    fig, ax = plt.subplots(figsize=(4.4, 3.2))
    bottom = [0] * len(quarters)
    for b in BED_ORDER:
        vals = [q[x].get(b, 0) for x in quarters]
        if not any(vals):
            continue
        color = SERIES[BED_ORDER.index(b)]
        ax.bar(quarters, vals, bottom=bottom, color=color, width=0.62,
               label=BED_LABEL.get(b, b), edgecolor="white", linewidth=1, zorder=3)
        bottom = [a + v for a, v in zip(bottom, vals)]
    ax.set_title("Listings placed on market per quarter", fontsize=10,
                 color=NAVY, loc="left", fontweight="bold")
    ax.tick_params(axis="x", rotation=45)
    ax.legend(frameon=False, fontsize=8)
    _style_ax(ax)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200)
    plt.close(fig)
    return buf


def chart_scatter(rows, subject):
    """Rent vs SF, colored by bed type; subject plans as navy diamonds with a
    gold rim (distinct from every series color). Legend sits BELOW the plot,
    outside the axes, so it can never read as data points."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(6.4, 3.7))
    for b in BED_ORDER:
        pts = [(r["sf"], r["rent"]) for r in rows
               if r["sane"] and r["beds"] == b and r["sf"]]
        if not pts:
            continue
        color = SERIES[BED_ORDER.index(b)]
        ax.scatter([p[0] for p in pts], [p[1] for p in pts], s=26, color=color,
                   alpha=0.75, edgecolors="white", linewidths=1,
                   label=BED_LABEL.get(b, b), zorder=3)
        # dashed least-squares trend for this bed type (needs SF spread)
        if len(pts) >= 3 and len({p[0] for p in pts}) > 1:
            slope, intercept = st.linear_regression([p[0] for p in pts],
                                                    [p[1] for p in pts])
            x0, x1 = min(p[0] for p in pts), max(p[0] for p in pts)
            ax.plot([x0, x1], [intercept + slope * x0, intercept + slope * x1],
                    color=color, linestyle=(0, (5, 4)), linewidth=1.3,
                    alpha=0.9, zorder=2)
    for p in subject["plans"]:
        ax.scatter([p["sf"]], [p["market"]], marker="D", s=95, color=NAVY,
                   edgecolors=GOLD, linewidths=1.6, zorder=5)
    ax.scatter([], [], marker="D", s=75, color=NAVY, edgecolors=GOLD,
               linewidths=1.4, label=f'{subject["name"]} (asking)')
    ax.set_xlabel("Unit square feet", fontsize=8.5, color=GRAY_TXT)
    ax.set_ylabel("Asking rent", fontsize=8.5, color=GRAY_TXT)
    ax.yaxis.set_major_formatter(lambda v, _: f"${v:,.0f}")
    ax.set_title("Every comp listing — rent vs unit size", fontsize=10,
                 color=NAVY, loc="left", fontweight="bold")
    ax.legend(frameon=False, fontsize=8.5, loc="upper center",
              bbox_to_anchor=(0.5, -0.22), ncol=4, columnspacing=1.4,
              handletextpad=0.4)
    _style_ax(ax)
    fig.tight_layout(rect=(0, 0.06, 1, 1))
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=200)
    plt.close(fig)
    return buf


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def write_pdf(path, subject, props, rows, bench, logo_path, report_date,
              urban=False, trend_rows=None, comp_window=None):
    trend_rows = trend_rows if trend_rows is not None else rows
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (BaseDocTemplate, Frame, PageTemplate,
                                    Paragraph, Spacer, Table, TableStyle,
                                    Image, PageBreak, KeepTogether)

    W, H = landscape(letter)
    navy, gold = colors.HexColor(NAVY), colors.HexColor(GOLD)
    pale = colors.HexColor(PALE_NAVY)
    gray = colors.HexColor(GRAY_TXT)

    st_h1 = ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=26,
                           textColor=navy, leading=30)
    st_h2 = ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=15,
                           textColor=navy, spaceAfter=2)
    st_sub = ParagraphStyle("sub", fontName="Helvetica", fontSize=10,
                            textColor=gray, spaceAfter=10)
    st_body = ParagraphStyle("body", fontName="Helvetica", fontSize=9,
                             textColor=colors.HexColor("#222222"), leading=12)
    st_note = ParagraphStyle("note", fontName="Helvetica-Oblique", fontSize=7,
                             textColor=colors.HexColor("#808080"), leading=9)
    st_cell = ParagraphStyle("cell", fontName="Helvetica", fontSize=8, leading=10)
    st_cellb = ParagraphStyle("cellb", fontName="Helvetica-Bold", fontSize=8,
                              leading=10, textColor=navy)

    def header_footer(canv, doc):
        canv.saveState()
        # header rule + logo
        if os.path.exists(logo_path):
            canv.drawImage(logo_path, W - 2.15 * inch, H - 0.62 * inch,
                           width=1.65 * inch, height=0.252 * inch, mask="auto")
        canv.setFillColor(navy)
        canv.setFont("Helvetica-Bold", 9)
        canv.drawString(0.55 * inch, H - 0.5 * inch,
                        f'{subject["name"]} — Rent Comparable Analysis')
        canv.setFillColor(gray)
        canv.setFont("Helvetica", 8)
        canv.drawString(0.55 * inch, H - 0.63 * inch, f"Report date: {report_date}")
        canv.setStrokeColor(gold)
        canv.setLineWidth(2)
        canv.line(0.55 * inch, H - 0.72 * inch, W - 0.55 * inch, H - 0.72 * inch)
        # footer band
        canv.setFillColor(navy)
        canv.rect(0, 0, W, 0.42 * inch, stroke=0, fill=1)
        canv.setFillColor(colors.white)
        canv.setFont("Helvetica", 7.5)
        canv.drawString(0.55 * inch, 0.17 * inch,
                        "The Multifamily Group  |  Confidential — prepared for client review")
        canv.drawRightString(W - 0.55 * inch, 0.17 * inch, f"Page {doc.page}")
        canv.restoreState()

    def cover(canv, doc):
        canv.saveState()
        canv.setFillColor(navy)
        canv.rect(0, 0, W, 1.05 * inch, stroke=0, fill=1)
        canv.setFillColor(gold)
        canv.rect(0, 1.05 * inch, W, 0.07 * inch, stroke=0, fill=1)
        canv.setFillColor(colors.white)
        canv.setFont("Helvetica", 8.5)
        canv.drawString(0.7 * inch, 0.62 * inch, "The Multifamily Group")
        canv.drawString(0.7 * inch, 0.42 * inch,
                        "Rent comp data: Dwellsy Comps API  |  Community identities "
                        "verified via public listings")
        if os.path.exists(logo_path):
            canv.drawImage(logo_path, W - 3.25 * inch, H - 1.25 * inch,
                           width=2.6 * inch, height=0.397 * inch, mask="auto")
        canv.restoreState()

    doc = BaseDocTemplate(path, pagesize=(W, H),
                          leftMargin=0.55 * inch, rightMargin=0.55 * inch,
                          topMargin=0.95 * inch, bottomMargin=0.6 * inch)
    frame = Frame(doc.leftMargin, doc.bottomMargin, W - 1.1 * inch,
                  H - 1.55 * inch, id="f")
    cover_frame = Frame(0.7 * inch, 1.3 * inch, W - 1.4 * inch, H - 2.6 * inch,
                        id="c")
    doc.addPageTemplates([PageTemplate(id="cover", frames=[cover_frame],
                                       onPage=cover),
                          PageTemplate(id="page", frames=[frame],
                                       onPage=header_footer)])
    story = []

    # ----- cover ---------------------------------------------------------
    s_units = sum(p["units"] for p in subject["plans"])
    s_sf = sum(p["sf"] * p["units"] for p in subject["plans"]) / s_units
    s_mkt = sum(p["market"] * p["units"] for p in subject["plans"]) / s_units
    s_eff = sum((p.get("effective") or p["market"]) * p["units"]
                for p in subject["plans"]) / s_units
    story.append(Spacer(1, 0.9 * inch))
    story.append(Paragraph("Rent Comparable Analysis",
                           ParagraphStyle("pre", fontName="Helvetica-Bold",
                                          fontSize=12, textColor=gold)))
    story.append(Spacer(1, 4))
    story.append(Paragraph(subject["name"], st_h1))
    story.append(Paragraph(
        f'{subject["address"]}, {subject["city"]}, {subject["state"]} '
        f'{subject["zip"]}', st_sub))
    story.append(Spacer(1, 14))
    stats = [
        ("Units", f'{subject["units"]}'),
        ("Avg Unit Size", f"{s_sf:,.0f} SF"),
        ("Avg Market Rent", money(s_mkt)),
        ("Avg In-Place Rent", money(s_eff)),
        ("Comp Properties", f"{len(props)}"),
        ("Comp Listings", f"{len([r for r in rows if r['sane']])}"),
    ]
    tile_rows = [[Paragraph(v, ParagraphStyle("tv", fontName="Helvetica-Bold",
                                              fontSize=16, textColor=navy,
                                              alignment=1)) for _, v in stats],
                 [Paragraph(k, ParagraphStyle("tk", fontName="Helvetica",
                                              fontSize=8, textColor=gray,
                                              alignment=1)) for k, _ in stats]]
    t = Table(tile_rows, colWidths=[(W - 1.4 * inch) / 6] * 6)
    t.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 1.2, gold),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 1), (-1, 1), 8),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))
    story.append(Paragraph(f"Report date: {report_date}", st_body))
    story.append(PageBreak())

    # ----- page 2: Rent Comparison Grid ----------------------------------
    story.append(Paragraph("Rent Comparison Grid", st_h2))
    story.append(Paragraph("Property-level comparison — subject shown first; "
                           "comp figures are averages of cleaned Dwellsy "
                           "listings.", st_sub))
    heads = ["Property Name", "Address", "City", "St", "Zip", "Year\nBuilt",
             "# of\nUnits", "Avg.\nSize", "Avg.\nRent/Unit", "Avg.\n$/SF",
             "Dist.\n(mi)", "Listings"]
    data = [heads]
    data.append([Paragraph(subject["name"], st_cellb), subject["address"],
                 subject["city"], subject["state"], subject["zip"],
                 str(subject.get("year_built") or "—"), str(subject["units"]),
                 f"{s_sf:,.0f}", money(s_mkt), f"${s_mkt / s_sf:,.2f}", "—", "—"])
    for p in props:
        yr = str(p["year"]) + ("" if p["year_src"] == "verified" else "*") if p["year"] else "—"
        un = str(p["units"]) + ("" if p["units_src"] == "verified" else "*") if p["units"] else "—"
        data.append([Paragraph(p["name"], st_cell), Paragraph(p["address"], st_cell),
                     p["city"], p["state"], p["zip"], yr, un,
                     f'{p["avg_sf"]:,.0f}' if p["avg_sf"] else "—",
                     money(p["avg_rent"]),
                     f'${p["avg_psf"]:,.2f}' if p["avg_psf"] else "—",
                     f'{p["mi"]:.1f}' if p["mi"] is not None else "—",
                     str(p["n"])])
    cw = [1.55, 1.75, 0.75, 0.32, 0.5, 0.55, 0.5, 0.55, 0.75, 0.55, 0.5, 0.6]
    t = Table(data, colWidths=[c * inch for c in cw], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.5),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 8),
        ("BACKGROUND", (0, 1), (-1, 1), pale),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, gold),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 2), (-1, -1),
         [colors.white, colors.HexColor("#F4F7FB")]),
        ("LINEBELOW", (0, 1), (-1, -1), 0.4, colors.HexColor(LIGHT_GRID)),
        ("TOPPADDING", (0, 0), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5),
    ]))
    story.append(t)
    story.append(Spacer(1, 6))
    grid_note = ("* Unit count / year built per Dwellsy — not independently "
                 "verified. Multi-address communities merged before "
                 "averaging. House rule: comp properties under 5 units are "
                 "excluded.")
    if urban:
        grid_note += (" Urban comp rules applied: vintage within 10 years of "
                      "the subject; no 150+ unit complexes for a sub-100-unit "
                      "subject; rent outliers pruned by standard deviation "
                      "per bed type.")
    if comp_window:
        grid_note += (f" Comp figures reflect listings from the last "
                      f"{comp_window} months; the trending page uses the "
                      f"full lookback.")
    story.append(Paragraph(grid_note, st_note))
    story.append(PageBreak())

    # ----- page 3: comps by unit type ------------------------------------
    story.append(Paragraph("Comparable Rents by Unit Type", st_h2))
    story.append(Paragraph("Property-weighted benchmarks; subject plans "
                           "positioned against the comp median.", st_sub))
    bh = ["Unit Type", "Comp\nProps", "Comp\nListings", "Median\nSF",
          "Property-Wtd\nMedian Rent", "Median\n$/SF"]
    bd = [bh]
    for b in BED_ORDER:
        if b not in bench:
            continue
        v = bench[b]
        bd.append([BED_LABEL.get(b, b), str(v["props"]), str(v["n"]),
                   f'{v["sf_med"]:,.0f}' if v["sf_med"] else "—",
                   money(v["prop_med"]),
                   f'${v["psf_med"]:,.2f}' if v["psf_med"] else "—"])
    t = Table(bd, colWidths=[1.3 * inch] * 6)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 8),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, gold),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F4F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 12))
    buf = chart_rankings(props, subject, bench)
    story.append(Image(buf, width=9.55 * inch, height=3.12 * inch))
    story.append(Spacer(1, 4))
    story.append(Paragraph("Gold = subject (asking). Bars are each property's "
                           "average asking rent for the unit type; dashed line "
                           "is the property-weighted comp median.", st_note))
    story.append(PageBreak())

    # ----- page 4: subject positioning -----------------------------------
    story.append(Paragraph("Subject Positioning & Rent Upside", st_h2))
    story.append(Paragraph("Subject floor plans vs the comp set — the gap "
                           "between in-place rent and comp-supported rent is "
                           "the value-add story.", st_sub))
    ph = ["Floor Plan", "Bed", "SF", "Units", "Market\nRent", "In-Place\nRent",
          "Comp Median\n(same bed)", "PSF-Implied\nRent", "In-Place vs\nComp",
          "Upside\n/Unit/Mo"]
    pd_ = [ph]
    tot_up = 0.0
    for p in subject["plans"]:
        b = str(p["bed"])
        v = bench.get(b)
        eff = p.get("effective") or p["market"]
        comp_med = v["prop_med"] if v else None
        implied = v["psf_med"] * p["sf"] if v and v["psf_med"] else None
        target = min(x for x in (comp_med, implied) if x) if (comp_med or implied) else None
        upside = max(0.0, (target - eff)) if target else None
        if upside:
            tot_up += upside * p["units"]
        pd_.append([p["label"], b, f'{p["sf"]:,}', str(p["units"]),
                    money(p["market"]), money(eff),
                    money(comp_med) if comp_med else "—",
                    money(implied) if implied else "—",
                    f"{eff / comp_med * 100:,.0f}%" if comp_med else "—",
                    money(upside) if upside is not None else "—"])
    pd_.append(["Total demonstrated upside", "", "", "", "", "", "", "", "",
                f"{money(tot_up)}/mo"])
    t = Table(pd_, colWidths=[1.05 * inch, 0.4 * inch, 0.55 * inch, 0.5 * inch,
                              0.8 * inch, 0.8 * inch, 1.05 * inch, 1.0 * inch,
                              0.9 * inch, 0.95 * inch])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.5),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 8.5),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, gold),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
         [colors.white, colors.HexColor("#F4F7FB")]),
        ("BACKGROUND", (0, -1), (-1, -1), pale),
        ("FONT", (0, -1), (-1, -1), "Helvetica-Bold", 9),
        ("TEXTCOLOR", (0, -1), (-1, -1), navy),
        ("SPAN", (0, -1), (8, -1)),
        ("ALIGN", (0, -1), (0, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 6))
    story.append(Paragraph("Comp-supported target = the lesser of the same-bed "
                           "property-weighted median and the PSF-implied rent "
                           "for the plan's square footage (conservative). "
                           "Upside floors at $0.", st_note))
    story.append(Spacer(1, 10))
    buf = chart_scatter(rows, subject)
    story.append(Image(buf, width=6.9 * inch, height=3.67 * inch))
    story.append(PageBreak())

    # ----- page 5: trending behavior --------------------------------------
    story.append(Paragraph("Market Trending Behavior", st_h2))
    win_txt = (f"the comp pages use listings from the last {comp_window} "
               f"months; trends below use the full pulled lookback."
               if comp_window else
               "trends use the full pulled lookback window.")
    story.append(Paragraph("Listing-derived trends across the comp "
                           f"properties — {win_txt}", st_sub))
    imgs = []
    b1 = chart_trend(trend_rows)
    if b1:
        imgs.append(Image(b1, width=5.55 * inch, height=2.78 * inch))
    b2 = chart_activity(trend_rows)
    if b2:
        imgs.append(Image(b2, width=3.8 * inch, height=2.78 * inch))
    if imgs:
        t = Table([imgs], colWidths=[5.7 * inch, 3.95 * inch][:len(imgs)])
        t.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(t)
    # trailing-vs-prior movement table
    dated = sorted([r for r in trend_rows if r["sane"] and r["created"]],
                   key=lambda r: r["created"])
    if dated:
        cut = dated[-1]["created"][:10]
        cut_dt = datetime.strptime(cut, "%Y-%m-%d")
        recent, prior = defaultdict(list), defaultdict(list)
        for r in dated:
            d = (cut_dt - datetime.strptime(r["created"], "%Y-%m-%d")).days
            (recent if d <= 182 else prior)[r["beds"]].append(r["rent"])
        mh = ["Unit Type", "Prior Median\n(6-24 mo ago)", "Recent Median\n(last 6 mo)",
              "Movement", "Recent\nListings"]
        md = [mh]
        for b in BED_ORDER:
            if not (recent.get(b) and prior.get(b)):
                continue
            pm, rm = med(prior[b]), med(recent[b])
            md.append([BED_LABEL.get(b, b), money(pm), money(rm),
                       f"{(rm - pm) / pm * 100:+.1f}%", str(len(recent[b]))])
        if len(md) > 1:
            story.append(Spacer(1, 10))
            t = Table(md, colWidths=[1.3 * inch] * 5)
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), navy),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 7.5),
                ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
                ("LINEBELOW", (0, 0), (-1, 0), 1.5, gold),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(t)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        "Trend medians are computed from the quarter each comp listing was "
        "placed on market (asking rents). Thin quarters reflect small-market listing "
        "volume — read direction, not precision. Data: Dwellsy Comps API; "
        "property identities verified via public listing sites.", st_note))

    # switch template after cover
    story.insert(1, _NextTemplate("page"))
    doc.build(story)


class _NextTemplate:
    """Tiny flowable that switches the active page template."""
    def __init__(self, name):
        from reportlab.platypus.doctemplate import NextPageTemplate
        self._f = NextPageTemplate(name)

    def __new__(cls, name):
        from reportlab.platypus.doctemplate import NextPageTemplate
        return NextPageTemplate(name)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("--csv", action="append", required=True,
                    help="glob of Dwellsy flat CSVs (repeatable)")
    ap.add_argument("--subject", required=True, help="subject JSON file")
    ap.add_argument("--names", help="verified names CSV "
                                    "(address_1,name,units,year_built,community)")
    ap.add_argument("--logo", default=os.path.join(os.path.dirname(
        os.path.abspath(__file__)), "tmg_logo.png"))
    ap.add_argument("--out", default=".")
    ap.add_argument("--sf-variance", type=float, default=0.25,
                    help="granular grid SF match band (default 0.25)")
    ap.add_argument("--min-units", type=int, default=5,
                    help="exclude comp properties below this unit count "
                         "(house rule: no sub-5-unit complexes; default 5). "
                         "Unit count is the verified value from --names when "
                         "given, else Dwellsy's. Set 1 to disable.")
    ap.add_argument("--urban", dest="urban", action="store_true", default=None,
                    help="force urban-mode filters on")
    ap.add_argument("--no-urban", dest="urban", action="store_false",
                    help="force urban-mode filters off (sparse market)")
    ap.add_argument("--urban-threshold", type=int, default=15,
                    help="auto-enable urban mode when the qualifying pool "
                         "exceeds this many listings (default 15)")
    ap.add_argument("--vintage-window", type=int, default=10,
                    help="urban mode: max |comp year built - subject year "
                         "built| (default 10)")
    ap.add_argument("--sd-cut", type=float, default=3.0,
                    help="urban mode: prune listings whose rent is beyond "
                         "this many standard deviations from the per-bed "
                         "mean (default 2.0)")
    ap.add_argument("--min-comps", type=int, default=5,
                    help="relax urban filters rather than fall below this "
                         "many comp properties (default 5)")
    ap.add_argument("--comp-months", type=int, default=6,
                    help="URBAN MODE ONLY: trim the COMP SET to listings "
                         "from the last N months (falling back to 12, then "
                         "the full window, if fewer than --min-comps "
                         "properties survive); trends always use the full "
                         "pulled lookback (default 6; 0 disables)")
    ap.add_argument("--report-date",
                    default=datetime.now().strftime("%-m/%-d/%Y"))
    args = ap.parse_args(argv)

    subject = json.load(open(args.subject))
    subject.setdefault("units", sum(p["units"] for p in subject["plans"]))
    names_map = load_names(args.names)
    rows = load_listings(args.csv, subject, names_map)
    if not rows:
        sys.exit("ERROR: no comp listings loaded")
    props = build_properties(rows, names_map)

    def keep(props_kept, rows_in):
        keys = {p["key"] for p in props_kept}
        return [r for r in rows_in if r["prop_key"] in keys]

    # House rule 1: multifamily comps only -- drop sub-`min_units` complexes
    # (and unknown unit counts) from the comp set entirely.
    if args.min_units > 1:
        before = len(props)
        props = [p for p in props
                 if p["units"] and p["units"] >= args.min_units]
        rows = keep(props, rows)
        print(f"min-units {args.min_units}: kept {len(props)} comp "
              f"properties, excluded {before - len(props)}")
    if not props:
        sys.exit(f"ERROR: no comp properties with {args.min_units}+ units -- "
                 f"widen the search (submarket sweep) before exporting")

    # Urban mode: on when forced, else when the qualifying pool is plentiful.
    n_sane = sum(1 for r in rows if r["sane"])
    urban = args.urban if args.urban is not None else n_sane > args.urban_threshold
    print(f"urban mode: {'ON' if urban else 'OFF'} "
          f"({n_sane} qualifying listings, threshold {args.urban_threshold}"
          f"{', forced' if args.urban is not None else ''})")

    if urban:
        # House rule 2: amenity-package parity -- a sub-100-unit subject is
        # not comped against 150+ unit complexes.
        if subject["units"] < 100:
            cand = [p for p in props if p["units"] < 150]
            if len(cand) >= args.min_comps:
                if len(cand) < len(props):
                    print(f"unit cap: excluded "
                          f"{len(props) - len(cand)} complexes of 150+ units")
                props, rows = cand, keep(cand, rows)
            else:
                print(f"unit cap SKIPPED: would leave {len(cand)} < "
                      f"{args.min_comps} comps")

        # House rule 3: vintage within +/-window years of the subject.
        # Trim ladder: strict -> keep unknown vintages -> skip.
        sy = subject.get("year_built")
        if sy and args.vintage_window > 0:
            lo, hi = sy - args.vintage_window, sy + args.vintage_window
            strict = [p for p in props if p["year"] and lo <= p["year"] <= hi]
            loose = strict + [p for p in props if not p["year"]]
            if len(strict) >= args.min_comps:
                print(f"vintage {lo}-{hi}: kept {len(strict)} of "
                      f"{len(props)} (unknown vintages dropped)")
                props, rows = strict, keep(strict, rows)
            elif len(loose) >= args.min_comps:
                print(f"vintage {lo}-{hi}: kept {len(loose)} of {len(props)} "
                      f"({len(loose) - len(strict)} unknown vintages KEPT to "
                      f"hold {args.min_comps}+ comps -- verify them)")
                props, rows = loose, keep(loose, rows)
            else:
                print(f"vintage filter SKIPPED: would leave {len(loose)} < "
                      f"{args.min_comps} comps")

        # House rule 4: outlier pruning by standard deviation (per bed type,
        # two passes) INSTEAD of the fixed sane-rent band.
        if sum(1 for r in rows if r["rent"] is not None) > args.urban_threshold:
            for r in rows:
                r["sane"] = r["rent"] is not None
            for _ in range(2):
                for b in {r["beds"] for r in rows}:
                    sample = [r for r in rows if r["beds"] == b and r["sane"]]
                    if len(sample) < 4:
                        continue
                    mu = st.mean(r["rent"] for r in sample)
                    sd = st.pstdev(r["rent"] for r in sample)
                    if sd == 0:
                        continue
                    for r in sample:
                        if abs(r["rent"] - mu) > args.sd_cut * sd:
                            r["sane"] = False
            pruned = sum(1 for r in rows if r["rent"] is not None
                         and not r["sane"])
            print(f"SD pruning (+/-{args.sd_cut} sigma per bed, 2 passes): "
                  f"removed {pruned} outlier listings")

        # Rebuild property aggregates from the trimmed listing set.
        rebuilt = build_properties(rows, names_map)
        order = {p["key"]: i for i, p in enumerate(props)}
        props = [p for p in rebuilt if p["key"] in order]

    if not props:
        sys.exit("ERROR: no comp properties survived the urban trim ladder")

    # House rule: pull deep (24 months) so trends have a full lookback, then
    # trim the COMP SET to the freshest window that still holds min-comps
    # properties (try --comp-months, then 12, then keep the full window).
    # trend_rows keeps every filtered/pruned listing for the trending page.
    trend_rows = rows
    comp_rows, comp_window = rows, None
    dated = [r for r in rows if r["sane"] and r["created"]]
    # Trim only in urban mode: in a sparse market the deep lookback IS the
    # comp set — recency-trimming it would delete real comps.
    if urban and dated and args.comp_months > 0:
        newest = max(datetime.strptime(r["created"], "%Y-%m-%d")
                     for r in dated)
        for win in (args.comp_months, 12):
            cand = [r for r in rows if r["created"] and
                    (newest - datetime.strptime(r["created"], "%Y-%m-%d")
                     ).days <= win * 30.44]
            cand_props = [p for p in build_properties(cand, names_map)
                          if p["key"] in {q["key"] for q in props}]
            if len(cand_props) >= args.min_comps:
                comp_rows, comp_window = cand, win
                props = cand_props
                break
        print(f"comp recency window: "
              f"{'last %d months' % comp_window if comp_window else 'full lookback'}"
              f" ({len(comp_rows)} listings, {len(props)} properties); "
              f"trends use the full {len(trend_rows)}-listing dataset")
    bench = bed_benchmarks(props)

    os.makedirs(args.out, exist_ok=True)
    xlsx = os.path.join(args.out, f'{subject["name"]} - Rent Comps.xlsx')
    pdf = os.path.join(args.out,
                       f'{subject["name"]} - Rent Comparable Analysis.pdf')
    write_xlsx(xlsx, subject, props, comp_rows, args.sf_variance, urban=urban)
    write_pdf(pdf, subject, props, comp_rows, bench, args.logo,
              args.report_date, urban=urban, trend_rows=trend_rows,
              comp_window=comp_window)

    n_sane = len([r for r in comp_rows if r["sane"]])
    print(f"Comp set: {len(comp_rows)} listings ({n_sane} within rent band) "
          f"-> {len(props)} properties")
    for b in BED_ORDER:
        if b in bench:
            v = bench[b]
            print(f'  {BED_LABEL[b]:7s} props={v["props"]:2d} n={v["n"]:3d} '
                  f'prop-med {money(v["prop_med"])}  psf-med '
                  f'{"${:.2f}".format(v["psf_med"]) if v["psf_med"] else "-"}')
    print(f"Wrote:\n  {xlsx}\n  {pdf}")


if __name__ == "__main__":
    main()
