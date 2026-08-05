#!/usr/bin/env python3
"""
harvest_t12_corpus.py — Bulk-build the T12 charge-code mapping corpus from
prior underwritings.

Input: an Excel file whose FIRST COLUMN lists deal folders, e.g.
    C:\\Users\\dmytr\\TMG Dropbox\\- Underwritings\\Bryce\\Bryce - (Springtown,
    TX) - North Park Townhomes\\- Info for Buyers - Financials\\Templates
(one per row; a full path to the .xlsx itself also works). In each folder the
script opens the T12 processor workbook (same filename everywhere), reads the
"Final T-12" tab, and extracts every (GL account -> charge code) pair.

Output: t12_mappings.csv — the corpus process_t12.py uses for exact-match
mapping. Re-running is safe: results are merged and de-duplicated. When two
deals coded the same account differently, the MAJORITY code wins and the
conflict is reported so you can adjudicate.

Usage (Windows):
    pip install openpyxl
    python harvest_t12_corpus.py directories.xlsx
    python harvest_t12_corpus.py directories.xlsx --filename "T12  Property Name  Month Year  Processor.xlsx"
    python harvest_t12_corpus.py directories.xlsx --out t12_mappings.csv
"""

import argparse
import csv
import os
import re
import sys
import warnings
from collections import Counter, defaultdict

warnings.filterwarnings("ignore")
from openpyxl import load_workbook

# Default template filename to look for inside each listed folder.
DEFAULT_TEMPLATE_NAME = "T12  Property Name  Month Year  Processor.xlsx"

VALID_CODES = {"r", "ll", "v", "nr", "bd", "rw", "rt", "ro", "oi",
               "cs", "rm", "ad", "m", "pr", "w", "tr", "e", "o",
               "mf", "i", "tx"}

# Category-header labels that appear in col B with a code in col A but are
# NOT accounts — never harvest these.
CATEGORY_LABELS = {
    "rental income", "(loss to lease) / gain to lease", "vacancy",
    "non-revenue/concessions", "bad debt", "rubs - water/sewer",
    "rubs - electric/gas/other", "other income", "contract services",
    "repair & maintenance", "administrative", "marketing", "payroll",
    "water/sewer", "trash", "electric", "gas/other", "management fee",
    "insurance", "real estate taxes", "total revenue", "total income",
    "total expense", "net operating income", "income", "expense",
    "revenue", "account", "date",
}


def norm(s):
    s = re.sub(r"^\s*\d[\d.\-]*\s*", "", str(s))
    s = re.sub(r"[^a-z0-9/ ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


# normalize the header labels through the same pipeline used for matching
CATEGORY_LABELS = {norm(x) for x in CATEGORY_LABELS}


def gl_number(s):
    m = re.match(r"^\s*(\d[\d.\-]*)\s", str(s))
    return m.group(1) if m else ""


def find_workbook(entry, template_name):
    """Resolve a spreadsheet row to a workbook path."""
    entry = str(entry).strip().strip('"')
    if not entry:
        return None, "empty row"
    if entry.lower().endswith((".xlsx", ".xlsm")):
        return (entry, None) if os.path.exists(entry) else \
            (None, "file not found")
    if not os.path.isdir(entry):
        return None, "folder not found"
    exact = os.path.join(entry, template_name)
    if os.path.exists(exact):
        return exact, None
    # fallbacks: any workbook that looks like the processor
    cands = [f for f in os.listdir(entry)
             if f.lower().endswith((".xlsx", ".xlsm"))
             and not f.startswith("~$")]
    for f in cands:
        if "processor" in f.lower() and "t12" in f.lower().replace("-", ""):
            return os.path.join(entry, f), None
    for f in cands:
        if "processor" in f.lower():
            return os.path.join(entry, f), None
    return None, (f"no '{template_name}' (or *Processor*.xlsx) here; "
                  f"found: {', '.join(cands[:4]) or 'no xlsx files'}")


def extract_pairs(xlsx_path):
    """(code, account) pairs from the 'Final T-12' tab (falls back to
    RawData if Final T-12 is absent)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = next((s for s in ("Final T-12", "RawData") if s in wb.sheetnames),
                 None)
    if sheet is None:
        return None, "no 'Final T-12' or 'RawData' tab"
    ws = wb[sheet]
    pairs = []
    for row in ws.iter_rows(min_row=2, max_col=2):
        code = row[0].value
        acct = row[1].value if len(row) > 1 else None
        if code is None or acct is None:
            continue
        code = str(code).strip().lower()
        acct = str(acct).strip()
        if code not in VALID_CODES:
            continue
        if norm(acct) in CATEGORY_LABELS or not norm(acct):
            continue                      # category header row, not an account
        pairs.append((code, acct))
    wb.close()
    return pairs, sheet


def main(argv=None):
    ap = argparse.ArgumentParser(
        description="Build t12_mappings.csv from prior underwritings")
    ap.add_argument("directories_xlsx",
                    help="Excel file listing deal folders in column A")
    ap.add_argument("--filename", default=DEFAULT_TEMPLATE_NAME,
                    help=f"template filename inside each folder "
                         f"(default: '{DEFAULT_TEMPLATE_NAME}')")
    ap.add_argument("--out", default="t12_mappings.csv",
                    help="corpus CSV to create/merge into")
    ap.add_argument("--sheet", default=None,
                    help="sheet of the directories file to read "
                         "(default: first)")
    args = ap.parse_args(argv)

    # ---- read directory list ----
    wb = load_workbook(args.directories_xlsx, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    entries = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v and str(v).strip() and "\\" in str(v) or (v and "/" in str(v)):
            entries.append(str(v).strip())
    wb.close()
    print(f"{len(entries)} folder(s) listed in {args.directories_xlsx}\n")

    # ---- existing corpus (merge) ----
    votes = defaultdict(Counter)          # key -> Counter(code)
    meta = {}                             # key -> (gl, account, source)
    if os.path.exists(args.out):
        with open(args.out, newline="", encoding="utf-8-sig") as f:
            for rec in csv.DictReader(f):
                key = (rec.get("gl", ""), norm(rec["account"]))
                n = int(rec.get("count", 1) or 1)
                votes[key][rec["code"].strip()] += n
                meta.setdefault(key, (rec.get("gl", ""), rec["account"],
                                      rec.get("source", "")))
        print(f"Merging into existing corpus ({len(votes)} accounts)\n")

    # ---- harvest ----
    ok = failed = 0
    for entry in entries:
        path, err = find_workbook(entry, args.filename)
        label = os.path.basename(os.path.dirname(os.path.dirname(entry))) \
            if os.path.isdir(entry) else os.path.basename(entry)
        if err:
            print(f"  SKIP  {entry}\n        -> {err}")
            failed += 1
            continue
        try:
            pairs, src_sheet = extract_pairs(path)
        except Exception as e:
            print(f"  SKIP  {path}\n        -> {type(e).__name__}: {e}")
            failed += 1
            continue
        if pairs is None:
            print(f"  SKIP  {path}\n        -> {src_sheet}")
            failed += 1
            continue
        for code, acct in pairs:
            key = (gl_number(acct), norm(acct))
            votes[key][code] += 1
            meta.setdefault(key, (gl_number(acct), acct, label))
        print(f"  OK    {label}: {len(pairs)} coded lines ({src_sheet})")
        ok += 1

    # ---- resolve majorities, report conflicts ----
    out_rows, conflicts = [], []
    for key, counter in sorted(votes.items(), key=lambda kv: kv[0]):
        gl, acct, src = meta[key]
        winner, n = counter.most_common(1)[0]
        row = {"gl": gl, "account": acct, "code": winner,
               "count": sum(counter.values()), "source": src,
               "conflicts": ""}
        if len(counter) > 1:
            others = "; ".join(f"{c}x{k}" for k, c in
                               counter.most_common()[1:])
            row["conflicts"] = others
            conflicts.append((acct, dict(counter)))
        out_rows.append(row)

    with open(args.out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["gl", "account", "code", "count",
                                          "source", "conflicts"])
        w.writeheader()
        w.writerows(out_rows)

    print(f"\nDone: {ok} workbook(s) harvested, {failed} skipped.")
    print(f"Corpus: {len(out_rows)} unique account mappings -> {args.out}")
    if conflicts:
        print(f"\n{len(conflicts)} account(s) coded differently across deals "
              f"(majority kept - review these):")
        for acct, counter in conflicts:
            print(f"  - {acct}: {counter}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
