#!/usr/bin/env python3
"""
process_rent_roll.py — Convert a property-management rent roll PDF into a
standardized "Rent Roll" Excel workbook (rediQ-style layout, matching
'RR - Harvest Moon - 7-1-2026.xlsx').

Currently supported source formats:
  * SSI410 "Rent Roll Report" (Simply Software / SSI) — the Harvest Moon layout.
  * AppFolio rent roll XLSX export (Unit / BD/BA / Tenant / Status ...).
  * Yardi/ResMan-style "Rent Roll" XLSX export (Unit / Unit Type / Unit Sq Ft
    / Resident / Name / Market Rent / Actual Rent ..., with
    Current/Notice/Vacant + Future Residents sections and a Summary Groups
    block) — the Meadows (tmtx) layout.

The code is structured so additional PDF layouts can be added later:
subclass RentRollParser, implement .detect() and .parse(), and register it
in PARSERS. The writer (write_workbook) is format-agnostic: it consumes a
list of UnitRecord objects.

Usage:
    python process_rent_roll.py input.pdf [-o output.xlsx]

Output columns (one row per unit):
    Unit No. | Floor Plan | Net Sf | Bed | Bath | Lease Type | Renovation
    Status | Occupancy Status | Market Rent | Contractual Rent | Recurring
    Concessions | Net Effective Rent | Supplemental Rent | Upfront
    Concessions | Emp./Other Discounts | Other Income | Lease Start Date |
    Lease Expiration | Lease Term (months) | MTM | Move In Date | Move Out
    Date | Vac. Notice

Mapping rules (derived from the 7-1-2026 example workbook):
  * Contractual Rent  = the RENT lease-charge for the unit's primary resident.
  * Recurring Concessions / Upfront Concessions / Emp. Discounts = charges
    whose codes match the concession/discount patterns at the top of this
    file (CONC, EMPL, DISC, FREE RENT, ...). Unrecognized negative charges
    are routed to Upfront Concessions when magnitude > $200, otherwise to
    Recurring Concessions. None exist in the Harvest Moon SSI410 roll, so
    those columns are blank there.
  * Other Income      = all remaining non-RENT charges (TRASH, WATER, ...).
  * Net Effective Rent= Contractual Rent + recurring concessions +
                        discounts (negative amounts). Upfront concessions
                        are reported in col N but deliberately EXCLUDED
                        from NER so one-time specials don't drag down
                        average in-place rents. 0 for vacant units.
  * Primary resident  = highest-priority resident status: C (current) >
                        N (notice) > L (leased/future) > P (previous) > X.
  * Occupancy Status  = "Occupied" for OC/NA/NU/NR apt statuses,
                        "Vacant" for V* statuses or units with no residents.
  * Vac. Notice       = "Yes" when the apt is on notice (NA/NR) or the
                        primary resident status is N.
  * MTM               = "Yes" when the lease term-type indicates month-to-month.
"""

import argparse
import math
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime, date

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula


# ----------------------------------------------------------------------------
# Charge-code classification
# ----------------------------------------------------------------------------
# Every lease charge is routed to exactly one output column based on its code.
# Codes not matched below fall through to Other Income (col P).
# Extend these patterns as new PM systems / charge codes are encountered.

RENT_CODES = re.compile(r"^RENT$|^BASE\s*RENT$|^RESIDENT\s+RENT$", re.I)

# Housing-authority / voucher subsidy paid on the resident's behalf (HAP).
# House ruling (Dmytro, 7/31/2026): Contractual Rent is the FULL contract
# rent, i.e. the tenant portion PLUS the subsidy, and NER includes both — a
# voucher unit is not a discounted unit, the rent is simply paid by two
# payors. These codes therefore classify as "rent"; Resident.subsidy_charge /
# .tenant_rent keep the split available for reporting.
SUBSIDY_CODES = re.compile(
    r"^SUBSIDY$|^HAP$|^HAP\s*(RENT|INCOME|SUBSIDY)$|^HOUSING\s*ASSIST\w*$"
    r"|^SEC(?:TION)?\s*8$|^S8$|^VOUCHER$", re.I)

# Recurring (monthly) concessions -> col K. Typically negative amounts.
RECURRING_CONC_CODES = re.compile(
    r"CONC|^RCONC|LOSS.?TO.?LEASE|^LTL$", re.I)

# One-time / upfront concessions -> col N (rarely appear as recurring
# charges on a rent roll, but some systems list them).
UPFRONT_CONC_CODES = re.compile(
    r"UPFRONT|^UCONC|MOVE.?IN.?(CONC|SPECIAL)|FREE.?RENT|^SPCL|SPECIAL", re.I)

# Employee / model / other discounts -> col O.
DISCOUNT_CODES = re.compile(
    r"^EMP|EMPL|EMPLOYEE|^DISC|DISCOUNT|COURTESY|^OFFDISC"
    # OneSite: recurring credit that zeroes out an office/admin unit's
    # rent. It offsets the rent every month, so it belongs in col O
    # (inside NER), never in Upfront Concessions (which NER excludes).
    r"|^ADMIN\s*UNIT$", re.I)


def classify_charge(code, amount):
    """Return output bucket for a charge: rent|recurring_conc|upfront_conc|
    discount|other."""
    c = code.strip()
    if RENT_CODES.search(c) or SUBSIDY_CODES.search(c):
        return "rent"
    if RECURRING_CONC_CODES.search(c):
        return "recurring_conc"
    if UPFRONT_CONC_CODES.search(c):
        return "upfront_conc"
    if DISCOUNT_CODES.search(c):
        return "discount"
    # A negative non-RENT charge with no recognized code is almost always a
    # concession/discount of some kind — safer here than inflating
    # (deflating) Other Income. Magnitude over $200 is treated as a one-time
    # (upfront) concession; smaller amounts as recurring.
    if amount is not None and amount < 0:
        return "upfront_conc" if abs(amount) > 200 else "recurring_conc"
    return "other"


# ----------------------------------------------------------------------------
# Data model
# ----------------------------------------------------------------------------

@dataclass
class Resident:
    resident_id: str = ""
    name: str = ""
    status: str = ""             # C / N / L / P / X
    charges: list = field(default_factory=list)   # [(code, amount, is_future)]
    gross_possible: float = None
    actual_charges: float = None
    move_in: date = None
    move_out: date = None
    lease_start: date = None
    lease_expires: date = None
    term_type: str = ""          # e.g. "12-R", "6", "MTM"
    deposit: float = None
    other_deposit: float = None
    surety_bond: float = None
    ending_balance: float = None

    def _bucket(self, bucket):
        vals = [a for c, a, _ in self.charges
                if classify_charge(c, a) == bucket]
        return sum(vals) if vals else None

    @property
    def rent_charge(self):
        """Full contract rent: tenant portion + any housing subsidy."""
        return self._bucket("rent")

    @property
    def subsidy_charge(self):
        """Housing-authority (HAP) portion of the contract rent, if broken
        out by the source. Included in rent_charge — this is for reporting
        the split, not a separate revenue line."""
        vals = [a for c, a, _ in self.charges if SUBSIDY_CODES.search(c.strip())]
        return sum(vals) if vals else None

    @property
    def tenant_rent(self):
        """Resident-paid portion of the contract rent (rent_charge less
        subsidy)."""
        r = self.rent_charge
        if r is None:
            return None
        return r - (self.subsidy_charge or 0)

    @property
    def recurring_concessions(self):
        return self._bucket("recurring_conc")

    @property
    def upfront_concessions(self):
        return self._bucket("upfront_conc")

    @property
    def discounts(self):
        return self._bucket("discount")

    @property
    def other_income(self):
        return self._bucket("other")

    @property
    def total_charges(self):
        return sum(a for _, a, _ in self.charges)


@dataclass
class UnitRecord:
    unit: str
    floor_plan: str = ""
    sqft: float = None
    apt_status: str = ""         # OC / NA / VU / ...
    market_rent: float = None
    residents: list = field(default_factory=list)
    # Lease Type (output col F). Filled only when the source says something
    # concrete about the lease itself -- e.g. "Section 8 Voucher" for units
    # whose charge detail carries a housing-authority subsidy. The floor-plan
    # code is left untouched so plan rollups/sqft still work.
    lease_type: str = ""
    # Explicit bed/bath. When bed_bath_explicit is True these are returned
    # verbatim by .bed_bath -- including None, which renders blank. Parsers
    # set this for layouts whose floor-plan code does NOT encode bed/bath
    # (guessing from the code would be inventing numbers).
    bed_explicit: float = None
    bath_explicit: float = None
    bed_bath_explicit: bool = False
    # Fields filled from --sqft-est / --bedbath-est: best estimates, NOT from
    # the rent roll. Members of {"fp", "sqft", "bed", "bath"}; the writers
    # highlight exactly these cells red and print the estimate note.
    estimated: set = field(default_factory=set)

    STATUS_PRIORITY = {"C": 0, "N": 1, "L": 2, "P": 3, "X": 4}

    @property
    def primary(self):
        """Primary resident: Current > Notice > Leased/future > Previous."""
        if not self.residents:
            return None
        return min(self.residents,
                   key=lambda r: self.STATUS_PRIORITY.get(r.status.upper(), 9))

    @property
    def is_vacant(self):
        s = self.apt_status.upper()
        if s.startswith("V"):
            return True
        # No current/notice resident at all -> vacant
        return not any(r.status.upper() in ("C", "N") for r in self.residents)

    @property
    def on_notice(self):
        if self.apt_status.upper() in ("NA", "NR", "NU"):
            return True
        p = self.primary
        return bool(p and p.status.upper() == "N")

    @property
    def bed_bath(self):
        if self.bed_bath_explicit:
            return self.bed_explicit, self.bath_explicit
        fp = self.floor_plan.strip()
        if re.fullmatch(r"eff\w*|studio|s0|e", fp, re.I):
            return 0, 1
        m = re.match(r"^(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", fp)
        if m:
            b = float(m.group(1)); ba = float(m.group(2))
            return (int(b) if b == int(b) else b,
                    int(ba) if ba == int(ba) else ba)
        # Yardi/ResMan-style unit-type codes carry bed x bath in the code
        # itself: "tm1x1a" -> 1/1, "tm2x1b" -> 2/1, "A2X2" -> 2/2. Must run
        # before the "A1"/"B2" fallback below, which would read only the
        # first digit and guess the bath count.
        m = re.search(r"(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)", fp)
        if m:
            b = float(m.group(1)); ba = float(m.group(2))
            return (int(b) if b == int(b) else b,
                    int(ba) if ba == int(ba) else ba)
        m = re.match(r"^[A-Za-z]*(\d)", fp)   # e.g. "A1", "B2"
        if m:
            n = int(m.group(1))
            return n, max(1, n - (0 if n <= 2 else 1))
        return None, None


# ----------------------------------------------------------------------------
# SSI410 parser
# ----------------------------------------------------------------------------

class RentRollParser:
    """Base class for source-format parsers."""
    name = "base"

    @staticmethod
    def detect(pdf) -> bool:
        raise NotImplementedError

    def parse(self, pdf):
        """Return (property_name, as_of_date, [UnitRecord], report_checks)."""
        raise NotImplementedError

    # "PDF" or "xlsx export"; main() sets it from the input's extension so
    # the provenance line does not call an xlsx export a PDF.
    source_kind = "PDF"

    def source_note(self, asof):
        """Provenance line written under the last Rent Roll data row."""
        return (f"Generated from {self.name} rent roll {self.source_kind} "
                f"({asof.strftime('%m/%d/%Y') if asof else 'unknown date'}). "
                "Contractual Rent = RENT charge; Other Income = sum of "
                "non-RENT charges for the primary (current/notice) resident.")


class SSI410Parser(RentRollParser):
    """Parser for the SSI410 'Rent Roll Report' layout."""
    name = "SSI410"

    # x-coordinate windows for each column (from measured word positions)
    W_APT     = (0,   66)
    W_ID      = (66,  82)
    W_TYPE    = (82,  112)
    W_STATUS  = (112, 130)
    W_NAME    = (130, 266)
    W_RS      = (266, 281)
    W_SQFT    = (281, 328)
    W_MARKET  = (328, 368)
    W_CODE    = (368, 404)
    W_CHARGE  = (404, 448)
    W_STAR    = (448, 456)
    W_GROSS   = (456, 505)
    W_ACTUAL  = (505, 555)
    W_DATE    = (555, 585)   # M/I (line 1) and M/O (line 2)
    W_LEASE   = (585, 648)   # lease expires (line 1), trm-type (line 2)
    W_DEPOSIT = (648, 706)   # sec deposit (line 1), other deposit (line 2)
    W_BALANCE = (706, 9999)

    STOP_MARKERS = ("Total by Occupancy Status", "Grand Total",
                    "Applicants/Leases", "Summary of Actual Charges")

    @staticmethod
    def detect(pdf) -> bool:
        txt = pdf.pages[0].extract_text() or ""
        return "SSI410" in txt or "Rent Roll Report" in txt

    # -- helpers -------------------------------------------------------------

    @staticmethod
    def _num(s):
        if s is None:
            return None
        s = s.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            v = float(s)
        except ValueError:
            return None
        return -v if neg else v

    @staticmethod
    def _date(s):
        if not s:
            return None
        for fmt in ("%m/%d/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s.strip(), fmt).date()
            except ValueError:
                pass
        return None

    def _in(self, words, window):
        lo, hi = window
        return [w for w in words if lo <= w["x0"] < hi]

    def _txt(self, words, window):
        return " ".join(w["text"] for w in self._in(words, window)).strip()

    # -- header --------------------------------------------------------------

    def _parse_header(self, page):
        """Property name & as-of date from the centered header block."""
        words = page.extract_words()
        lines = {}
        for w in words:
            lines.setdefault(round(w["top"]), []).append(w)
        centered = []
        for top in sorted(lines):
            ws = sorted(lines[top], key=lambda w: w["x0"])
            x0 = ws[0]["x0"]; x1 = ws[-1]["x1"]
            mid = (x0 + x1) / 2
            if abs(mid - page.width / 2) < 60 and x0 > 150:
                centered.append(" ".join(w["text"] for w in ws))
        prop, asof = "", None
        # centered block: [report title, mgmt co, property, date]
        for ln in centered:
            d = None
            for fmt in ("%B %d, %Y", "%b %d, %Y"):
                try:
                    d = datetime.strptime(ln.strip(), fmt).date()
                except ValueError:
                    pass
            if d:
                asof = d
        if len(centered) >= 3:
            prop = centered[2]
        if asof is None:
            m = re.search(r"Select:\s*(\d{2}/\d{2}/\d{2,4})",
                          page.extract_text() or "")
            if m:
                asof = self._date(m.group(1))
        return prop, asof

    def _parse_checks(self, pdf):
        """Pull the report's own totals for reconciliation."""
        checks = {}
        for page in pdf.pages:
            txt = page.extract_text() or ""
            m = re.search(r"(\d[\d,]*)\s+Apts?,\s+([\d,]+)\s+Sq\.\s*Ft", txt)
            if m and "unit_count" not in checks:
                checks["unit_count"] = int(m.group(1).replace(",", ""))
                checks["total_sqft"] = float(m.group(2).replace(",", ""))
            m = re.search(r"Occupied\s+(\d+)\s+([\d,]+)\s+([\d,.]+)", txt)
            if m and "occupied_count" not in checks and "Occupancy Status" in txt:
                checks["occupied_count"] = int(m.group(1))
                checks["occupied_market_rent"] = self._num(m.group(3))
            if "Current/On-Notice" in txt and "current_lease_charges" not in checks:
                m = re.search(r"Current/On-Notice\s+([\d,.()]+)", txt)
                if m:
                    checks["current_lease_charges"] = self._num(m.group(1))
        return checks

    # -- body ----------------------------------------------------------------

    def parse(self, pdf):
        prop, asof = self._parse_header(pdf.pages[0])
        checks = self._parse_checks(pdf)

        units = {}          # unit -> UnitRecord
        cur_unit = None
        cur_res = None
        res_line = 0        # line index within current resident block
        stopped = False

        for page in pdf.pages:
            if stopped:
                break
            words = page.extract_words()
            lines = {}
            for w in words:
                lines.setdefault(round(w["top"]), []).append(w)

            page_text = page.extract_text() or ""
            page_stop_tops = []
            for marker in self.STOP_MARKERS:
                if marker in page_text:
                    # find the vertical position of the marker
                    for top, ws in lines.items():
                        joined = " ".join(w["text"] for w in
                                          sorted(ws, key=lambda w: w["x0"]))
                        if marker in joined:
                            page_stop_tops.append(top)
            stop_top = min(page_stop_tops) if page_stop_tops else None

            for top in sorted(lines):
                if stop_top is not None and top >= stop_top:
                    break
                ws = sorted(lines[top], key=lambda w: w["x0"])
                apt_txt = self._txt(ws, self.W_APT)
                rs_txt = self._txt(ws, self.W_RS)
                mkt_txt = self._txt(ws, self.W_MARKET)

                is_new_block = bool(
                    re.match(r"^\d[\w\s.-]*$", apt_txt)
                    and (rs_txt or mkt_txt or self._txt(ws, self.W_STATUS))
                )

                if is_new_block:
                    unit_id = re.sub(r"\s+", " ", apt_txt)
                    if unit_id not in units:
                        units[unit_id] = UnitRecord(unit=unit_id)
                    cur_unit = units[unit_id]
                    fp = self._txt(ws, self.W_TYPE)
                    if fp:
                        cur_unit.floor_plan = fp
                    st = self._txt(ws, self.W_STATUS)
                    if st:
                        cur_unit.apt_status = st
                    sq = self._num(self._txt(ws, self.W_SQFT))
                    if sq:
                        cur_unit.sqft = sq
                    mk = self._num(mkt_txt)
                    if mk is not None:
                        cur_unit.market_rent = mk

                    cur_res = Resident(
                        name=self._txt(ws, self.W_NAME),
                        status=rs_txt,
                        gross_possible=self._num(self._txt(ws, self.W_GROSS)),
                        actual_charges=self._num(self._txt(ws, self.W_ACTUAL)),
                        move_in=self._date(self._txt(ws, self.W_DATE)),
                        lease_expires=self._date(self._txt(ws, self.W_LEASE)),
                        deposit=self._num(self._txt(ws, self.W_DEPOSIT)),
                        ending_balance=self._num(self._txt(ws, self.W_BALANCE)),
                    )
                    cur_unit.residents.append(cur_res)
                    res_line = 1
                elif cur_res is not None:
                    res_line += 1
                    if res_line == 2:
                        rid = self._txt(ws, self.W_ID)
                        if rid:
                            cur_res.resident_id = rid
                        mo = self._date(self._txt(ws, self.W_DATE))
                        if mo:
                            cur_res.move_out = mo
                        tt = self._txt(ws, self.W_LEASE)
                        if tt:
                            cur_res.term_type = tt
                        od = self._num(self._txt(ws, self.W_DEPOSIT))
                        if od is not None:
                            cur_res.other_deposit = od

                # charge line (applies to first line of block too)
                if cur_res is not None:
                    code = self._txt(ws, self.W_CODE)
                    amt = self._num(self._txt(ws, self.W_CHARGE))
                    star = self._txt(ws, self.W_STAR) == "*"
                    if code and code != "Total:" and amt is not None:
                        cur_res.charges.append((code, amt, star))

            if stop_top is not None:
                stopped = True

        return prop, asof, list(units.values()), checks


class AppFolioXlsxParser(RentRollParser):
    """Parser for AppFolio-style rent roll XLSX exports.

    Columns: Unit | BD/BA | Tenant | Status | Market Rent | Rent | Deposit |
    Lease From | Lease To | Move-in | Move-out | Past Due | NSF | Late.
    """
    name = "AppFolio-xlsx"

    STATUS_MAP = {          # AppFolio status -> (apt_status, resident status)
        "current": ("OC", "C"),
        "notice-rented": ("NA", "N"),
        "notice-unrented": ("NA", "N"),
        "evict": ("OC", "C"),
        "vacant-rented": ("VR", ""),
        "vacant-unrented": ("VU", ""),
    }

    @staticmethod
    def detect_xlsx(path):
        try:
            from openpyxl import load_workbook as _lw
            wb = _lw(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1, max_row=15):
                vals = [str(c.value or "").strip().lower() for c in row]
                if "unit" in vals and "bd/ba" in vals and "status" in vals:
                    return True
        except Exception:
            pass
        return False

    @staticmethod
    def _d(v):
        """Cell -> date; AppFolio uses 1970-01-01 as an empty placeholder."""
        if isinstance(v, datetime):
            d = v.date()
        elif isinstance(v, date):
            d = v
        else:
            return None
        return None if d.year <= 1971 else d

    @staticmethod
    def _fp(bdba):
        """'2/1.00' -> '2/1', '2/1.50' -> '2/1.5'."""
        m = re.match(r"^\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)",
                     str(bdba or ""))
        if not m:
            return str(bdba or "").strip()

        def trim(x):
            f = float(x)
            return str(int(f)) if f == int(f) else str(f)
        return f"{trim(m.group(1))}/{trim(m.group(2))}"

    def parse(self, path):
        from openpyxl import load_workbook as _lw
        wb = _lw(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[c.value for c in r] for r in ws.iter_rows()]

        asof, prop, header_i = None, "", None
        checks = {}
        for i, r in enumerate(rows):
            a = str(r[0] or "")
            m = re.match(r"As of:\s*(\d{2}/\d{2}/\d{4})", a)
            if m:
                asof = datetime.strptime(m.group(1), "%m/%d/%Y").date()
            vals = [str(v or "").strip().lower() for v in r]
            if "unit" in vals and "bd/ba" in vals:
                header_i = i
                cols = {v: j for j, v in enumerate(vals)}
                break
        if header_i is None:
            sys.exit("ERROR: AppFolio header row not found.")

        def col(name):
            return cols.get(name)

        units = []
        adv_count, adv_market = 0, 0.0
        group = ""          # current property group label (multi-property)
        for r in rows[header_i + 1:]:
            unit = str(r[0] or "").strip()
            # Single-unit properties (SFR/duplex portfolios) print the
            # property group header row and then a detail row with an EMPTY
            # Unit cell - the property name IS the unit. Adopt the group's
            # short name so the unit is not silently dropped.
            if not unit and r[col("bd/ba")] and group:
                unit = group
            if not unit:
                continue
            # "Include Advertised Rentals" adds one placeholder listing row
            # per floor plan (marketing name like '2Bed/2.5Bath Luxury
            # Townhomes...', no tenant) - not a physical unit. Skip it and
            # back it out of the report's own totals.
            if re.match(r"^\d+\s*Bed\s*/", unit, re.I) and \
                    not str(r[col("tenant")] or "").strip():
                adv_count += 1
                adv_market += float(r[col("market rent")] or 0)
                continue
            status = str(r[col("status")] or "").strip()
            # totals row(s): '65 Units' / 'Total 65 Units' / '89.2% Occupied'
            mt = re.match(r"^(?:total\s+)?(\d+)\s+units?$", unit, re.I)
            if mt or "% occupied" in status.lower():
                if mt:
                    checks["unit_count"] = int(mt.group(1))
                mo = re.match(r"([\d.]+)%\s*occupied", status, re.I)
                if mo and "unit_count" in checks:
                    checks["occupied_count"] = round(
                        float(mo.group(1)) / 100 * checks["unit_count"])
                if r[col("market rent")] is not None:
                    checks["total_market_rent"] = float(
                        r[col("market rent")])
                if r[col("rent")] is not None:
                    checks["total_contract_rent"] = float(r[col("rent")])
                continue
            # repeated column-header row (multi-property exports repeat the
            # header before each property group)
            if unit.lower() == "unit" and \
                    str(r[col("bd/ba")] or "").strip().lower() == "bd/ba":
                continue
            if not r[col("bd/ba")]:          # property group header row
                name = unit.split(" - ")[0].strip()
                if name and not re.match(r"^(?:total\s+)?\d+\s+units?$",
                                         name, re.I):
                    group = name
                    prop = f"{prop} & {name}" if prop else name
                continue

            apt_st, res_st = self.STATUS_MAP.get(
                status.lower(), ("OC", "C") if r[col("tenant")] else
                ("VU", ""))
            u = UnitRecord(unit=unit, floor_plan=self._fp(r[col("bd/ba")]),
                           sqft=None, apt_status=apt_st,
                           market_rent=float(r[col("market rent")] or 0)
                           or None)
            rent = r[col("rent")]
            if res_st:
                res = Resident(
                    name=str(r[col("tenant")] or "").strip(),
                    status=res_st,
                    charges=[("RENT", float(rent), False)]
                    if rent not in (None, "") else [],
                    lease_start=self._d(r[col("lease from")]),
                    lease_expires=self._d(r[col("lease to")]),
                    move_in=self._d(r[col("move-in")]),
                    move_out=self._d(r[col("move-out")]),
                )
                u.residents.append(res)
            units.append(u)
        if adv_count:
            print(f"Excluded {adv_count} advertised-rental placeholder "
                  f"row(s) (${adv_market:,.0f} market rent backed out of "
                  "report totals).")
            if "unit_count" in checks:
                checks["unit_count"] -= adv_count
            if "total_market_rent" in checks:
                checks["total_market_rent"] -= adv_market
        return prop, asof, units, checks


class YardiRentRollXlsxParser(RentRollParser):
    """Parser for Yardi/ResMan-style 'Rent Roll' XLSX exports.

    Layout (validated on The Meadows (tmtx), 7/30/2026):

        Rent Roll
        <Property Name> (<code>)
        As Of = 07/30/2026
        Month Year = 07/2026
        Unit | Unit Type | Unit    | Resident | Name | Market | Actual | ...
             |           | Sq Ft   |          |      | Rent   | Rent   |
        Current/Notice/Vacant Residents          <- section banner (col A)
        2101 | tm1x1a | 650 | t0014478 | Jane Doe GHA | 1048 | 324 | ...
        ...
        Future Residents/Applicants              <- section banner
        4101 | tm2x2a | 865 | t0019597 | New Tenant | 1386 | 0 | ...
        ...
                     | Total | <Property> | <market> | <actual> | <deposits>
        Summary Groups | ... | Square Footage | Market Rent | Actual Rent |
                             Security Deposit | Other Deposits | # Of Units |
                             % Unit Occupancy | % Sqft Occupied
        Current/Notice/Vacant Residents | ...   <- the report's own totals

    Notes on the mapping:
      * The column header spans two physical rows; they are joined per
        column ("Unit"+"Sq Ft" -> "unit sq ft").
      * One row per unit in the Current/Notice/Vacant section — that section
        IS the rent roll (152 rows = 152 doors). Rows in the Future
        Residents/Applicants section are *applicants for units that already
        appear above*, so they are attached to the existing unit as a
        status-"L" resident and never counted as extra doors.
      * "VACANT" in the Resident/Name columns marks a vacant unit; a vacant
        unit with a future applicant becomes VR (vacant-rented), otherwise
        VU. An occupied unit with a Move Out date is on notice (NA / resident
        status "N") — it still counts as occupied, exactly as the report's
        own Occupied Units total does.
      * Floor plan = the unit-type code (tm1x1a / tm1x1c / ...). The codes
        are NOT interchangeable: sqft differs per code (650 / 653 / 659 for
        the three 1x1s), so collapsing them to "1/1" would lose real data.
        Bed/bath comes from the NxN inside the code (UnitRecord.bed_bath).
      * Market Rent / Actual Rent are per-unit printed values; Actual Rent
        is mapped to a single RENT charge (Contractual Rent). There is no
        charge detail in this export, so Other Income / concession columns
        stay blank.
      * There is no lease-start and no lease-type/MTM column in this export.
        Move In feeds col U, Lease Expiration col R; Lease Start (col Q) and
        MTM (col T) are left blank — expired lease dates are NOT read as
        month-to-month (same house rule as the AppFolio parser).

    VARIANT: "Rent Roll with Lease Charges" (The Meadows, 7/31/2026)
    ---------------------------------------------------------------
    Same title block, sections and Summary Groups block, but the single
    "Actual Rent" column is replaced by "Charge Code" + "Amount", and each
    unit's charges continue on unnamed rows below it, closed by a per-unit
    `Total` row:

        2101 | tm1x1a | 650 | t0014478 | Jane Doe GHA | 1048 | rent    | 324
             |        |     |          |              |      | petrent | 10
             |        |     |          |              |      | subsidy | 1056
             |        |     |          |              |      | Total   | 1390

      * Every printed per-unit `Total` is checked against the sum of that
        unit's charge rows; a single mismatch aborts the run.
      * Charges route through classify_charge(): `rent` AND `subsidy` are
        both Contractual Rent (house ruling 7/31/2026 — a voucher unit's
        contract rent is the tenant portion plus the HAP payment, and NER
        includes both); everything else falls in the usual buckets, so
        petrent / washdry / cam / garage / move-in land in Other Income.
      * A unit whose charges include a subsidy code is labelled
        "Section 8 Voucher" in Lease Type (col F). The floor-plan code is
        left alone so plan rollups and sqft are unaffected.
      * The report's own "Summary of Charges by Charge Code" block is tied
        out code by code, and the Balance column totals are tied out too.
      * The Summary Groups block sits one column further right here, so its
        columns are located from its own two-row header instead of by
        position, and its rent column is "Lease Charges" (ALL charges) —
        that ties to total lease charges, NOT to Contractual Rent, which is
        checked separately against rent + subsidy from the charge-code block.
    """
    name = "Yardi-xlsx"

    # charge codes that never appear on a rent-roll charge block as revenue
    SUBSIDY_LABEL = "Section 8 Voucher"

    # Anchored on purpose: the lease-charges variant prints a qualifier line
    # "(Current/Notice Residents Only)" under the charge-code summary, which
    # an unanchored "current.*residents" would read as a new detail section.
    SECTION_CURRENT = re.compile(r"^current[\w/\s]*residents", re.I)
    SECTION_FUTURE = re.compile(r"^future\s+residents|^applicants", re.I)

    # summary-block row labels -> key prefix
    SUMMARY_ROWS = {
        "current/notice/vacant residents": "total",
        "occupied units": "occupied",
        "total vacant units": "vacant",
        "total non rev units": "nonrev",
        "future residents/applicants": "future",
        "totals:": "grand",
    }

    @staticmethod
    def _num(v):
        """'123,784.00' / 1048 / None -> float or None."""
        if v is None or isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "").replace("$", "")
        if s in ("", "-"):
            return None
        try:
            return float(s)
        except ValueError:
            return None

    @staticmethod
    def _d(v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v or "").strip()
        for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, f).date()
            except ValueError:
                continue
        return None

    @classmethod
    def _rows(cls, path):
        from openpyxl import load_workbook as _lw
        wb = _lw(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [[c.value for c in r] for r in ws.iter_rows()]

    @staticmethod
    def _join_header(rows, i):
        """Join the 2-row header at rows[i]/rows[i+1] into one name per
        column: 'Unit' over 'Sq Ft' -> 'unit sq ft'."""
        top = [str(v or "").strip() for v in rows[i]]
        nxt = ([str(v or "").strip() for v in rows[i + 1]]
               if i + 1 < len(rows) else [""] * len(top))
        nxt += [""] * (len(top) - len(nxt))
        return [" ".join(x for x in (a, b) if x).lower()
                for a, b in zip(top, nxt)]

    @classmethod
    def _find_header(cls, rows):
        """Return (header_row_index, {name: col}) for the 2-row header.

        Both variants are accepted: the plain roll carries an "Actual Rent"
        column, the "with Lease Charges" roll carries "Charge Code"+"Amount"
        instead.
        """
        for i, r in enumerate(rows[:20]):
            joined = cls._join_header(rows, i)
            if "unit" in joined and "unit type" in joined and \
                    "market rent" in joined and \
                    ("actual rent" in joined
                     or ("charge code" in joined and "amount" in joined)):
                return i, {n: j for j, n in enumerate(joined) if n}
        return None, {}

    @classmethod
    def detect_xlsx(cls, path):
        try:
            rows = cls._rows(path)
        except Exception:
            return False
        i, _ = cls._find_header(rows)
        return i is not None

    def parse(self, path):
        rows = self._rows(path)
        hdr_i, cols = self._find_header(rows)
        if hdr_i is None:
            sys.exit("ERROR: Yardi rent roll header row not found.")

        def cell(r, name):
            j = cols.get(name)
            return r[j] if j is not None and j < len(r) else None

        # ---- title block: property name + as-of date ---------------------
        prop, asof = "", None
        for r in rows[:hdr_i]:
            s = str(r[0] or "").strip()
            if not s:
                continue
            m = re.match(r"as\s*of\s*[:=]\s*(\d{1,2}/\d{1,2}/\d{4})", s, re.I)
            if m:
                asof = datetime.strptime(m.group(1), "%m/%d/%Y").date()
                continue
            # report title: "Rent Roll", "Rent Roll with Lease Charges", ...
            if re.match(r"rent\s*roll\b", s, re.I) or \
                    re.match(r"month\s*year\s*[:=]", s, re.I):
                continue
            if not prop:
                # "The Meadows (tmtx)" -> "The Meadows"
                prop = re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()

        # "with Lease Charges" variant: charge code + amount instead of a
        # single Actual Rent column, with per-unit charge blocks below.
        charges_mode = "charge code" in cols and "actual rent" not in cols
        if charges_mode:
            self.name = "Yardi-xlsx (with lease charges)"

        units, order = {}, []
        checks, section = {}, None
        printed_total = {}
        future_market_conflicts = []
        code_totals = {}             # printed Summary of Charges by code
        code_grand = None
        unit_totals = {}             # unit -> printed per-unit Total
        cur_res = None               # resident receiving continuation charges
        cur_unit = None
        sum_cols = None              # Summary Groups col map (by header name)

        def scell(r, name):
            j = sum_cols.get(name) if sum_cols else None
            return r[j] if j is not None and j < len(r) else None

        i = hdr_i + 2
        while i < len(rows):
            r = rows[i]
            i += 1
            vals = [str(v or "").strip() for v in r]
            a = vals[0] if vals else ""
            rest = [v for v in vals[1:] if v]

            # ---- section banners (label alone on the row) ----------------
            if a and not rest:
                if self.SECTION_CURRENT.search(a):
                    section = "current"
                elif self.SECTION_FUTURE.search(a):
                    section = "future"
                elif re.match(r"summary\s+groups", a, re.I):
                    section = "summary"
                elif re.match(r"summary\s+of\s+charges", a, re.I):
                    section = "chargecodes"
                continue

            # ---- per-unit charge continuation rows ----------------------
            # (no unit identity; the charge code column carries the payload)
            if charges_mode and section in ("current", "future") and \
                    not str(cell(r, "unit") or "").strip() and \
                    str(cell(r, "charge code") or "").strip():
                code = str(cell(r, "charge code") or "").strip()
                amt = self._num(cell(r, "amount"))
                if code.lower() == "total":
                    if cur_unit is not None:
                        unit_totals[cur_unit] = amt or 0.0
                    continue
                if cur_res is not None and amt is not None:
                    cur_res.charges.append((code, amt, False))
                continue

            # ---- report "Total" line under the detail -------------------
            if any(re.fullmatch(r"total", v, re.I) for v in vals[:6]) and \
                    section in ("current", "future"):
                printed_total = {
                    "market": self._num(cell(r, "market rent")),
                    "actual": self._num(cell(r, "amount" if charges_mode
                                              else "actual rent")),
                    "deposit": self._num(cell(r, "resident deposit")),
                    "balance": self._num(cell(r, "balance")),
                }
                section = "summary"
                continue

            # ---- summary block ------------------------------------------
            if section == "summary":
                if re.match(r"summary\s+groups", a, re.I):
                    # locate the summary columns from its own 2-row header;
                    # the block shifts right in the lease-charges variant
                    hdr = self._join_header(rows, i - 1)
                    sum_cols = {n: j for j, n in enumerate(hdr) if n}
                    i += 1                        # consume the 2nd header row
                    continue
                key = self.SUMMARY_ROWS.get(a.lower())
                if key and sum_cols:
                    got = {
                        "sqft": scell(r, "square footage"),
                        "market": scell(r, "market rent"),
                        # plain roll prints "Actual Rent"; the lease-charges
                        # roll prints "Lease Charges" = ALL charges
                        "actual": scell(r, "actual rent") if not charges_mode
                        else scell(r, "lease charges"),
                        "deposit": scell(r, "security deposit"),
                        "other_deposit": scell(r, "other deposits"),
                        "count": scell(r, "# of units"),
                        "pct_units": scell(r, "% unit occupancy"),
                        "pct_sqft": scell(r, "% sqft occupied"),
                        "balance": scell(r, "balance"),
                    }
                    for k, v in got.items():
                        v = self._num(v)
                        if v is not None:
                            checks[f"{key}_{k}"] = v
                continue

            # ---- "Summary of Charges by Charge Code" block --------------
            if section == "chargecodes":
                if not a or re.match(r"charge\s*code$", a, re.I):
                    continue
                amt = next((self._num(v) for v in r[1:]
                            if self._num(v) is not None), None)
                if amt is None:
                    continue
                if re.fullmatch(r"total", a, re.I):
                    code_grand = amt
                else:
                    code_totals[a] = code_totals.get(a, 0.0) + amt
                continue

            # ---- detail rows --------------------------------------------
            unit = str(cell(r, "unit") or "").strip()
            utype = str(cell(r, "unit type") or "").strip()
            if not unit or not utype or section is None:
                continue
            if unit.lower() == "unit":            # repeated header
                continue

            sqft = self._num(cell(r, "unit sq ft"))
            market = self._num(cell(r, "market rent"))
            name = str(cell(r, "name") or "").strip()
            rid = str(cell(r, "resident") or "").strip()
            move_in = self._d(cell(r, "move in"))
            move_out = self._d(cell(r, "move out"))
            expires = self._d(cell(r, "lease expiration"))
            dep = self._num(cell(r, "resident deposit"))
            odep = self._num(cell(r, "other deposit"))
            bal = self._num(cell(r, "balance"))
            vacant = name.upper() == "VACANT" or rid.upper() == "VACANT"
            # first charge of the block sits on the unit row itself
            if charges_mode:
                first_code = str(cell(r, "charge code") or "").strip()
                first_amt = self._num(cell(r, "amount"))
                charges = ([(first_code, first_amt, False)]
                           if first_code and first_amt is not None else [])
            else:
                actual = self._num(cell(r, "actual rent"))
                charges = [("RENT", actual, False)] if actual is not None \
                    else []
            cur_unit, cur_res = unit, None

            if section == "future":
                # applicant for a unit that already exists above
                u = units.get(unit)
                if u is None:                     # defensive: unseen unit
                    u = UnitRecord(unit=unit, floor_plan=utype, sqft=sqft,
                                   apt_status="VU", market_rent=market)
                    units[unit] = u
                    order.append(unit)
                elif market is not None and u.market_rent is not None and \
                        abs(market - u.market_rent) > 0.01:
                    future_market_conflicts.append(
                        (unit, u.market_rent, market))
                cur_res = Resident(
                    resident_id=rid, name=name, status="L",
                    charges=list(charges),
                    move_in=move_in, lease_expires=expires,
                    ending_balance=bal)
                u.residents.append(cur_res)
                if u.apt_status.upper().startswith("V"):
                    u.apt_status = "VR"           # vacant, pre-leased
                # a per-unit Total row for an applicant belongs to that
                # applicant's charges, keyed separately from the unit itself
                cur_unit = f"{unit} (future)"
                continue

            # current / notice / vacant section
            if unit in units:
                print(f"WARNING: duplicate unit row for {unit} — keeping "
                      "the first occurrence.")
                continue
            if vacant:
                apt_status = "VU"
                res_status = ""
            elif move_out is not None:
                apt_status = "NA"                 # notice
                res_status = "N"
            else:
                apt_status = "OC"
                res_status = "C"
            u = UnitRecord(unit=unit, floor_plan=utype, sqft=sqft,
                           apt_status=apt_status, market_rent=market)
            if res_status:
                cur_res = Resident(
                    resident_id=rid, name=name, status=res_status,
                    charges=list(charges),
                    move_in=move_in, move_out=move_out,
                    lease_expires=expires, deposit=dep, other_deposit=odep,
                    ending_balance=bal)
                u.residents.append(cur_res)
            units[unit] = u
            order.append(unit)

        # ---- per-unit printed Total rows must equal the charge detail ----
        if charges_mode:
            bad = []
            for key, want in unit_totals.items():
                un = key[:-9] if key.endswith(" (future)") else key
                u = units.get(un)
                if u is None:
                    bad.append((key, None, want))
                    continue
                if key.endswith(" (future)"):
                    got = sum(r.total_charges for r in u.residents
                              if r.status.upper() == "L")
                else:
                    got = sum(r.total_charges for r in u.residents
                              if r.status.upper() in ("C", "N"))
                if abs(got - want) > 0.01:
                    bad.append((key, got, want))
            if bad:
                for key, got, want in bad[:20]:
                    print(f"  unit {key}: charges sum "
                          f"{'n/a' if got is None else format(got, ',.2f')} "
                          f"vs printed Total {want:,.2f}")
                sys.exit(f"ERROR: {len(bad)} unit(s) whose printed Total row "
                         "disagrees with their charge detail — parse is "
                         "wrong, refusing to write the workbook.")
            missing = [u for u in order if u not in unit_totals
                       and f"{u} (future)" not in unit_totals]
            if missing:
                sys.exit("ERROR: no printed Total row found for unit(s): "
                         + ", ".join(missing[:20]))

            # subsidised units get the Lease Type label (floor plan intact)
            for u in units.values():
                if any((r.subsidy_charge or 0) > 0 for r in u.residents):
                    u.lease_type = self.SUBSIDY_LABEL

        for unit, have, got in future_market_conflicts:
            print(f"NOTE: unit {unit} future-resident row prints market rent "
                  f"{got:,.0f} vs {have:,.0f} on the unit row — kept the "
                  "unit row's value.")

        # ---- normalize the summary block into reconcile()'s check keys ---
        alias = {
            "unit_count": "total_count",
            "total_sqft": "total_sqft",
            "total_market_rent": "total_market",
            "total_contract_rent": "total_actual",
            "total_deposits": "total_deposit",
            "total_other_deposits": "total_other_deposit",
            "occupied_count": "occupied_count",
            "occupied_sqft": "occupied_sqft",
            "occupied_market_rent": "occupied_market",
            "vacant_count": "vacant_count",
            "vacant_sqft": "vacant_sqft",
            "vacant_market_rent": "vacant_market",
            "nonrev_count": "nonrev_count",
            "future_count": "future_count",
            "future_sqft": "future_sqft",
            "future_market_rent": "future_market",
            "pct_unit_occupancy": "total_pct_units",
            "pct_sqft_occupied": "total_pct_sqft",
        }
        if charges_mode:
            # "Lease Charges" is every charge, not contract rent
            alias.pop("total_contract_rent")
            alias["current_lease_charges"] = "total_actual"
        out = {k: checks[v] for k, v in alias.items() if v in checks}
        for k in ("unit_count", "occupied_count", "vacant_count",
                  "nonrev_count", "future_count"):
            if k in out:
                out[k] = int(round(out[k]))

        if charges_mode:
            # Contract rent comes from the report's own charge-code block:
            # every code that classifies as rent (tenant rent + subsidy).
            if code_totals:
                out["charge_totals"] = dict(code_totals)
                rent_codes = {c: a for c, a in code_totals.items()
                              if classify_charge(c, a) == "rent"}
                if rent_codes:
                    out["total_contract_rent"] = sum(rent_codes.values())
                    print("Contract rent = " + " + ".join(
                        f"{c} {a:,.2f}" for c, a in sorted(rent_codes.items()))
                        + f" = {sum(rent_codes.values()):,.2f}")
                if code_grand is not None:
                    got = sum(code_totals.values())
                    if abs(got - code_grand) > 0.01:
                        sys.exit("ERROR: charge-code block sums to "
                                 f"{got:,.2f} vs its printed Total "
                                 f"{code_grand:,.2f}.")
                    if "current_lease_charges" in out and \
                            abs(code_grand - out["current_lease_charges"]) \
                            > 0.01:
                        sys.exit("ERROR: charge-code block total "
                                 f"{code_grand:,.2f} disagrees with the "
                                 "summary block's Lease Charges "
                                 f"{out['current_lease_charges']:,.2f}.")
            out["unit_charge_totals"] = {
                (k[:-9] if k.endswith(" (future)") else k): v
                for k, v in unit_totals.items()
                if not k.endswith(" (future)")}
            # resident balances (current/notice and applicants)
            if "total_balance" in checks:
                out["report_total_balance"] = checks["total_balance"]
                out["parsed_balance"] = sum(
                    r.ending_balance or 0 for u in units.values()
                    for r in u.residents if r.status.upper() in ("C", "N"))
            if "future_balance" in checks:
                out["report_future_balance"] = checks["future_balance"]
                out["parsed_future_balance"] = sum(
                    r.ending_balance or 0 for u in units.values()
                    for r in u.residents if r.status.upper() == "L")

        # the detail "Total" line must agree with the summary block
        if printed_total.get("market") is not None:
            for key, ck in (("market", "total_market_rent"),
                            ("actual", "current_lease_charges" if charges_mode
                             else "total_contract_rent"),
                            ("deposit", "total_deposits")):
                v = printed_total.get(key)
                if v is not None and ck in out and abs(v - out[ck]) > 0.01:
                    sys.exit(f"ERROR: report's Total line {key} {v:,.2f} "
                             f"disagrees with its summary block "
                             f"{out[ck]:,.2f}.")
            # the grand Total line's balance covers current + applicants
            gb = printed_total.get("balance")
            if gb is not None and "total_balance" in checks:
                want = checks["total_balance"] + checks.get("future_balance", 0)
                if abs(gb - want) > 0.01:
                    sys.exit(f"ERROR: report's Total line balance {gb:,.2f} "
                             f"disagrees with its summary block {want:,.2f}.")
        return prop, asof, [units[u] for u in order], out


# ----------------------------------------------------------------------------
# ResMan parser
# ----------------------------------------------------------------------------

class ResManParser(RentRollParser):
    """Parser for the ResMan 'Rent Roll' PDF (landscape, 792x612).

    Layout (validated on Lofts at Taft / Cornerstone Residential, 6/29/2026):

      Unit | Type | Sq. Feet | Residents | Status | Market Rent |
      Description | Amount | Move In | Lease Start | Lease End | Move Out |
      Surety Bonds | Deposits | Balance

    * Units are grouped under section headers ("Current"); each unit row is
      followed by one indented line per recurring lease charge and a printed
      "Total" line for the unit.
    * A long floor-plan name wraps around the unit row: the first token sits
      on the line ABOVE and the remainder on the line BELOW (e.g. "F2" /
      unit row / "Large" -> "F2 Large"). Those continuation lines carry
      nothing outside the Type column, which is how they are recognised.
    * Vacant units carry the literal resident name "Vacant Unit", no status
      letter, and no lease dates.
    * The Description column is HARD-TRUNCATED at ~16 characters
      ("Renters Legal Lia", "Storage Space Re"). The report's own
      "Total Charges" / "Total Credits" summary carries the full
      descriptions, so charge names are resolved against it (see
      _resolve_descriptions) before classification -- otherwise a $300
      concession printed as "Fetick and Grego" would be unclassifiable.
    * Month-to-month is taken from the explicit "Month to Month Fee" charge,
      never inferred from an expired lease end date (house rule).
    """
    name = "ResMan"

    # x-coordinate windows, from measured word positions on the 792pt page
    W_UNIT    = (0,   58)
    W_TYPE    = (58,  90)
    W_SQFT    = (90,  118)
    W_NAME    = (118, 186)
    W_STATUS  = (186, 220)
    W_MARKET  = (220, 300)
    W_DESC    = (300, 400)
    W_AMOUNT  = (400, 438)
    W_MOVEIN  = (438, 484)
    W_LSTART  = (484, 531)
    W_LEND    = (531, 578)
    W_MOVEOUT = (578, 630)
    W_SURETY  = (630, 676)
    W_DEPOSIT = (676, 726)
    W_BALANCE = (726, 9999)

    # A description whose last glyph reaches this x has hit the column's
    # clip width and may be truncated; shorter ones are complete.
    TRUNC_X = 384.0

    ROW_TOL = 2.0        # pt; must stay < the ~4.5pt wrapped-type line gap

    # Lines that start in the Unit column but are not units
    NON_UNIT_TOKENS = ("Printed", "Unit", "Current", "Future", "Notice",
                       "Vacant", "Applicant", "Total", "Description",
                       "Type", "Account", "Property", "Collections", "*")

    STOP_MARKERS = ("Total Charges", "Property Occupancy", "Collections")

    _NUM_RE = re.compile(r"^-?\(?\$?[\d,]+\.\d{2}\)?$|^-?\(?\$?[\d,]+\)?$")

    # -- helpers -------------------------------------------------------------

    @staticmethod
    def _num(s):
        if s is None:
            return None
        s = s.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            v = float(s)
        except ValueError:
            return None
        return -v if neg else v

    @staticmethod
    def _date(s):
        if not s:
            return None
        for fmt in ("%m/%d/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s.strip(), fmt).date()
            except ValueError:
                pass
        return None

    def _in(self, words, window):
        lo, hi = window
        return [w for w in words if lo <= w["x0"] < hi]

    def _txt(self, words, window):
        return " ".join(w["text"] for w in self._in(words, window)).strip()

    def _numcol(self, words, window):
        """First numeric token in a column (ignores stray label text)."""
        for w in self._in(words, window):
            if self._NUM_RE.match(w["text"]):
                return self._num(w["text"])
        return None

    def _rows(self, page):
        """Cluster words into visual rows, keeping wrapped-type continuation
        lines (~4.5pt off the unit row) as rows of their own."""
        words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))
        rows, cur, cur_top = [], [], None
        for w in words:
            if cur_top is None or abs(w["top"] - cur_top) <= self.ROW_TOL:
                cur.append(w)
                cur_top = w["top"] if cur_top is None else cur_top
            else:
                rows.append((cur_top, sorted(cur, key=lambda x: x["x0"])))
                cur, cur_top = [w], w["top"]
        if cur:
            rows.append((cur_top, sorted(cur, key=lambda x: x["x0"])))
        return rows

    @staticmethod
    def detect(pdf) -> bool:
        txt = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
        return "ResMan" in txt and "Rent Roll" in txt

    # -- header --------------------------------------------------------------

    def _parse_header(self, page):
        prop, asof = "", None
        centered = []
        for top, ws in self._rows(page):
            if top > 110:
                break
            x0, x1 = ws[0]["x0"], ws[-1]["x1"]
            if abs((x0 + x1) / 2 - page.width / 2) < 60 and x0 > 150:
                centered.append(" ".join(w["text"] for w in ws))
        for ln in centered:
            d = self._date(ln)
            if d:
                asof = d
        if centered:
            prop = centered[0]
        return prop, asof

    # -- the report's own summary blocks -------------------------------------

    def _parse_summary(self, pdf):
        """Charge/credit summary ("Total Charges" / "Total Credits") and the
        Property / Unit Type occupancy tables on the trailing pages."""
        charges, credits = {}, {}
        charge_total = credit_total = None
        checks = {}
        for page in pdf.pages:
            txt = page.extract_text() or ""
            if "Total Charges" in txt:
                mode = None
                for top, ws in self._rows(page):
                    joined = " ".join(w["text"] for w in ws)
                    if "Total Charges" in joined:
                        mode = "table"
                        continue
                    if mode != "table" or "Description" in joined:
                        continue
                    # left half = charges, right half = credits
                    for side, lo, hi, dhi, bag in (
                            ("c", 0, 300, 190, charges),
                            ("r", 300, 9999, 560, credits)):
                        cell = [w for w in ws if lo <= w["x0"] < hi]
                        if not cell:
                            continue
                        desc = " ".join(w["text"] for w in cell
                                        if w["x0"] < dhi).strip()
                        nums = [w for w in cell if w["x0"] >= dhi
                                and self._NUM_RE.match(w["text"])]
                        if not nums:
                            continue
                        val = self._num(nums[-1]["text"])
                        if desc:
                            bag[desc] = bag.get(desc, 0.0) + val
                        elif side == "c":
                            charge_total = val
                        else:
                            credit_total = val
            if "Property Occupancy" in txt:
                m = re.search(r"Total\s+Occupied\s+([\d,.]+)\s+[\d.]+%\s+"
                              r"(\d+)\s+[\d.]+%\s+([\d,.]+)", txt)
                if m:
                    checks["occupied_market_rent"] = self._num(m.group(1))
                    checks["occupied_count"] = int(m.group(2))
                    checks["occupied_sqft"] = self._num(m.group(3))
                m = re.search(r"Total\s+Vacant\s+([\d,.]+)\s+[\d.]+%\s+"
                              r"(\d+)\s+[\d.]+%\s+([\d,.]+)", txt)
                if m:
                    checks["vacant_market_rent"] = self._num(m.group(1))
                    checks["vacant_count"] = int(m.group(2))
                    checks["vacant_sqft"] = self._num(m.group(3))
                # Unit Type occupancy: "<type> Occupied|Vacant mkt % n % sf %"
                fps = {}
                seen_unit_type = False
                for top, ws in self._rows(page):
                    joined = " ".join(w["text"] for w in ws)
                    if "Unit Type" in joined:
                        seen_unit_type = True
                        continue
                    if not seen_unit_type:
                        continue
                    m = re.match(r"^(.+?)\s+(Occupied|Vacant)\s+([\d,.]+)\s+"
                                 r"[\d.]+%\s+(\d+)\s+[\d.]+%\s+([\d,.]+)\s+"
                                 r"[\d.]+%$", joined)
                    if not m:
                        continue
                    fp = m.group(1).strip()
                    e = fps.setdefault(fp, {"units": 0, "market": 0.0,
                                            "sqft": 0.0})
                    e["units"] += int(m.group(4))
                    e["market"] += self._num(m.group(3))
                    e["sqft"] += self._num(m.group(5))
                if fps:
                    checks["floor_plan_totals"] = fps
                    checks["unit_count"] = sum(v["units"] for v in fps.values())
                    checks["total_market_rent"] = round(
                        sum(v["market"] for v in fps.values()), 2)
                    checks["total_sqft"] = sum(v["sqft"] for v in fps.values())
        if charges:
            checks["charge_totals"] = charges
        if credits:
            checks["credit_totals"] = credits
        if charge_total is not None:
            checks["total_charges"] = charge_total
        if credit_total is not None:
            checks["total_credits"] = credit_total
        for k in ("Resident Rent", "Rent", "Base Rent"):
            if k in charges:
                checks["total_contract_rent"] = charges[k]
                break
        return checks

    # -- description resolution ---------------------------------------------

    def _resolve_descriptions(self, units, checks):
        """Map truncated detail descriptions to the full names printed in the
        report's charge/credit summary.

        1. exact match;
        2. if the text hit the column clip width, unambiguous prefix match in
           either direction ("Renters Legal Lia" -> "Renters Legal Liability",
           "Credit Builder - M" -> "Credit Builder");
        3. residual match -- an unresolved group whose total equals exactly
           one summary line's still-unallocated remainder is that line
           (this is what identifies ResMan's per-lease memo descriptions,
           e.g. "Fetick and Grego" = the single $300
           "Concession - Resident Rent" credit, and "Rent" = the remainder of
           "Resident Rent").
        Anything still unresolved is left verbatim and reported; the
        by-description reconciliation will then fail loudly.
        """
        charges = checks.get("charge_totals") or {}
        credits = checks.get("credit_totals") or {}
        groups = {}          # (desc, is_credit) -> [total, hit_clip]
        for u in units:
            for r in u.residents:
                for desc, amt, clipped in r.charges:
                    key = (desc, amt < 0)
                    g = groups.setdefault(key, [0.0, False])
                    g[0] += amt
                    g[1] = g[1] or clipped

        resolved, notes = {}, []
        alloc = {}

        def cands(is_credit):
            return credits if is_credit else charges

        # passes 1 & 2
        for (desc, is_credit), (total, clipped) in sorted(groups.items()):
            pool = cands(is_credit)
            low = desc.lower()
            hit = [c for c in pool if c.lower() == low]
            if not hit and clipped:
                hit = [c for c in pool
                       if c.lower().startswith(low) or low.startswith(c.lower())]
            if len(hit) == 1:
                resolved[(desc, is_credit)] = hit[0]
                alloc[(hit[0], is_credit)] = \
                    alloc.get((hit[0], is_credit), 0.0) + abs(total)

        # pass 3: residual match
        for (desc, is_credit), (total, clipped) in sorted(groups.items()):
            if (desc, is_credit) in resolved:
                continue
            pool = cands(is_credit)
            hit = [c for c in pool
                   if abs((pool[c] - alloc.get((c, is_credit), 0.0))
                          - abs(total)) <= 0.005]
            if len(hit) == 1:
                resolved[(desc, is_credit)] = hit[0]
                alloc[(hit[0], is_credit)] = \
                    alloc.get((hit[0], is_credit), 0.0) + abs(total)
                notes.append(f"    {desc!r} -> {hit[0]!r} "
                             f"(matched by residual {abs(total):,.2f})")
            else:
                notes.append(f"    {desc!r} UNRESOLVED "
                             f"(total {total:,.2f}) - left verbatim")

        for u in units:
            for r in u.residents:
                r.charges = [(resolved.get((d, a < 0), d), a, c)
                             for d, a, c in r.charges]
        return notes

    # -- body ----------------------------------------------------------------

    def parse(self, pdf):
        prop, asof = self._parse_header(pdf.pages[0])
        checks = self._parse_summary(pdf)

        units, order = {}, []
        cur_unit = cur_res = None
        pending_type = ""          # wrapped floor-plan token seen above a row
        just_opened = False        # last row was a unit row (accept a suffix)
        unit_printed_totals = {}
        surety = deposits = balance = 0.0

        for page in pdf.pages:
            rows = self._rows(page)
            for idx, (top, ws) in enumerate(rows):
                joined = " ".join(w["text"] for w in ws)
                if any(m in joined for m in self.STOP_MARKERS):
                    cur_unit = cur_res = None
                    just_opened = False
                    break

                unit_txt = self._txt(ws, self.W_UNIT)
                type_txt = self._txt(ws, self.W_TYPE)
                desc_txt = self._txt(ws, self.W_DESC)
                mkt = self._numcol(ws, self.W_MARKET)

                # wrapped floor-plan continuation: nothing outside the Type col
                if type_txt and not unit_txt and not desc_txt and \
                        all(self.W_TYPE[0] <= w["x0"] < self.W_TYPE[1]
                            for w in ws):
                    if just_opened and cur_unit is not None:
                        cur_unit.floor_plan = \
                            (cur_unit.floor_plan + " " + type_txt).strip()
                    else:
                        pending_type = type_txt
                    continue

                # grand-total strip under the last unit: market / surety /
                # deposits / balance, with no unit and no description
                if (not unit_txt and not desc_txt and mkt is not None
                        and self._numcol(ws, self.W_SURETY) is not None
                        and self._numcol(ws, self.W_BALANCE) is not None):
                    checks["report_total_market_rent"] = mkt
                    checks["report_total_surety_bonds"] = \
                        self._numcol(ws, self.W_SURETY)
                    checks["report_total_deposits"] = \
                        self._numcol(ws, self.W_DEPOSIT)
                    checks["report_total_balance"] = \
                        self._numcol(ws, self.W_BALANCE)
                    continue

                is_unit = (unit_txt and mkt is not None
                           and unit_txt.split()[0] not in self.NON_UNIT_TOKENS
                           and re.match(r"^[A-Za-z]*\d", unit_txt))
                just_opened = False

                if is_unit:
                    uid = re.sub(r"\s+", " ", unit_txt)
                    cur_unit = UnitRecord(unit=uid)
                    units[uid] = cur_unit
                    order.append(uid)
                    cur_unit.floor_plan = (pending_type + " " + type_txt).strip()
                    pending_type = ""
                    cur_unit.sqft = self._numcol(ws, self.W_SQFT)
                    cur_unit.market_rent = mkt

                    name = self._txt(ws, self.W_NAME)
                    status = self._txt(ws, self.W_STATUS)
                    vacant = (not status) and re.match(r"^vacant\b", name, re.I)
                    cur_unit.apt_status = "VU" if vacant else "OC"

                    cur_res = Resident(
                        name=name,
                        status="" if vacant else status,
                        move_in=self._date(self._txt(ws, self.W_MOVEIN)),
                        lease_start=self._date(self._txt(ws, self.W_LSTART)),
                        lease_expires=self._date(self._txt(ws, self.W_LEND)),
                        move_out=self._date(self._txt(ws, self.W_MOVEOUT)),
                        surety_bond=self._numcol(ws, self.W_SURETY),
                        deposit=self._numcol(ws, self.W_DEPOSIT),
                        ending_balance=self._numcol(ws, self.W_BALANCE),
                    )
                    cur_unit.residents.append(cur_res)
                    surety += cur_res.surety_bond or 0
                    deposits += cur_res.deposit or 0
                    balance += cur_res.ending_balance or 0
                    just_opened = True
                    # fall through: the unit row itself carries the 1st charge

                if cur_res is None or not desc_txt:
                    continue

                amt = self._numcol(ws, self.W_AMOUNT)
                if amt is None:
                    continue
                if desc_txt.lower() == "total":
                    if cur_unit is not None:
                        unit_printed_totals[cur_unit.unit] = amt
                    continue
                dwords = self._in(ws, self.W_DESC)
                clipped = bool(dwords) and dwords[-1]["x1"] >= self.TRUNC_X
                cur_res.charges.append((desc_txt, amt, clipped))

        unit_list = [units[u] for u in order]

        # MTM comes from the explicit Month-to-Month fee, not from an expired
        # lease end date (house rule: never infer MTM from dates).
        for u in unit_list:
            for r in u.residents:
                if any(re.match(r"month\s*to\s*month", d, re.I)
                       for d, _, _ in r.charges):
                    r.term_type = "MTM"

        notes = self._resolve_descriptions(unit_list, checks)
        if notes:
            print("Charge descriptions resolved against the report's own "
                  "charge/credit summary (the detail column is truncated):")
            print("\n".join(notes))

        # Bed count is encoded in the floor-plan code (F1/G1 = 1BR,
        # F2/F2.2/G2 = 2BR); bath count is NOT in a ResMan rent roll, so it
        # is left blank rather than guessed.
        for u in unit_list:
            m = re.match(r"^[A-Za-z]+\s*(\d)", u.floor_plan.strip())
            u.bed_explicit = int(m.group(1)) if m else None
            u.bath_explicit = None
            u.bed_bath_explicit = True

        checks["unit_charge_totals"] = unit_printed_totals
        checks["parsed_surety_bonds"] = round(surety, 2)
        checks["parsed_deposits"] = round(deposits, 2)
        checks["parsed_balance"] = round(balance, 2)
        return prop, asof, unit_list, checks


class ResManSummaryParser(RentRollParser):
    """Parser for the ResMan **'Rent Roll Summary'** PDF - the one-row-per-
    unit condensation of the full ResMan rent roll (no per-charge blocks).

    Validated on McNeil Star Apartments (Touchstone Property Management,
    Dallas TX, 8/3/2026, 32 units, 2 pages).

    Layout (792 x 612 landscape):

      Unit | Type | Sq. Feet | Residents | Status | Market Rent | Rent |
      Other Charges | Credits | Total | Move In | Start | Lease End |
      Move Out | Surety Bonds | Deposits | Balance

    * One row per unit. Vacant units print the literal resident name
      "Vacant Unit" and a blank Status; occupied statuses seen: C (current),
      UE (under eviction - still an occupied door, resident status C; the
      unit is FLAGged for the operator).
    * Numeric columns are right-aligned; values are keyed to the printed
      right edges (NUM_RIGHT), dates to their own (DATE_RIGHT), so a row
      with blanks (vacant rows print no dates) keeps every value in the
      correct column.
    * Rent -> a RENT charge; Other Charges -> one OTHER CHARGES lump (the
      summary does not break out codes) -> Other Income; Credits (printed
      positive) -> a negative CREDIT charge routed via classify_charge.
      Every unit's printed Total is checked against Rent + Other Charges -
      Credits; a single mismatch aborts the run.
    * The report's own tie-outs, all parsed and reconciled: the grand strip
      under the last unit (market / rent / other / credits / total / surety /
      deposits / balance), the Property Occupancy table and the Unit Type
      Occupancy table (per-plan market rent / units / sqft).
    * Bed/bath is NOT encoded in the Type codes (A2 at 597 sf is a 1-bed at
      McNeil Star - the code digit would lie) - left blank for --bedbath /
      --bedbath-est.
    * No MTM or lease-term columns exist; MTM is never inferred from an
      expired lease (house rule).
    """

    name = "ResMan-Summary"

    W_UNIT   = (0,   46)
    W_TYPE   = (46,  74)
    W_SQFT   = (74,  100)
    W_NAME   = (100, 183)
    W_STATUS = (183, 213)

    NUM_RIGHT = (("market", 257.8), ("rent", 314.6), ("other", 357.8),
                 ("credits", 401.0), ("total", 450.7), ("surety", 673.2),
                 ("deposit", 726.5), ("balance", 779.8))
    DATE_RIGHT = (("movein", 496.1), ("lstart", 539.3), ("lend", 582.5),
                  ("moveout", 625.0))
    NUM_TOL = 3.5
    ROW_TOL = 2.5

    _NUM_RE = re.compile(r"^\(?-?[\d,]+\.\d{2}\)?$")
    _DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{2,4}$")

    SKIP_RE = re.compile(
        r"^(Printed\b|Page \d|Current$|Unit Type|Sq\. Other"
        r"|\* denotes|©|\(c\) ?ResMan)", re.I)

    STATUS_MAP = {          # printed status -> (apt_status, resident status)
        "C": ("OC", "C"),
        "UE": ("OC", "C"),          # under eviction: occupied door
        "N": ("NA", "N"),
        "NU": ("NA", "N"),
    }

    @staticmethod
    def _num(s):
        if s is None:
            return None
        s = s.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            v = float(s)
        except ValueError:
            return None
        return -v if neg else v

    @staticmethod
    def _date(s):
        if not s:
            return None
        for fmt in ("%m/%d/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(s.strip(), fmt).date()
            except ValueError:
                pass
        return None

    def _rows(self, page):
        rows = []
        for w in sorted(page.extract_words(), key=lambda z: (z["top"],
                                                             z["x0"])):
            if rows and abs(w["top"] - rows[-1][0]) <= self.ROW_TOL:
                rows[-1][1].append(w)
            else:
                rows.append((w["top"], [w]))
        return [(t, sorted(ws, key=lambda z: z["x0"])) for t, ws in rows]

    def _txt(self, ws, window):
        lo, hi = window
        return " ".join(w["text"] for w in ws if lo <= w["x0"] < hi).strip()

    @staticmethod
    def detect(pdf) -> bool:
        txt = pdf.pages[0].extract_text() or ""
        return "Rent Roll Summary" in txt

    def source_note(self, asof):
        return (f"Generated from the ResMan 'Rent Roll Summary' PDF "
                f"({asof.strftime('%m/%d/%Y') if asof else 'unknown date'}). "
                "Contractual Rent = the Rent column; Other Income = the "
                "'Other Charges' column (the summary prints one lump per "
                "unit, no charge codes). The report carries no lease-term, "
                "MTM or concession detail.")

    def parse(self, pdf):
        self.flags = []
        prop, asof = "", None
        units, order = {}, []
        checks = {}
        grand = None

        for pi, page in enumerate(pdf.pages):
            page_txt = page.extract_text() or ""
            in_summary = "Property Occupancy" in page_txt
            for top, ws in self._rows(page):
                joined = " ".join(w["text"] for w in ws)
                if self.SKIP_RE.match(joined):
                    continue
                # ---- title band -------------------------------------
                if pi == 0 and not units and top < 115:
                    d = self._date(joined)
                    if d:
                        asof = d
                    elif not re.search(r"Rent Roll|Management", joined) \
                            and not prop:
                        prop = joined
                    continue
                nums = {}
                for w in ws:
                    if not self._NUM_RE.match(w["text"]):
                        continue
                    for key, rx in self.NUM_RIGHT:
                        if abs(w["x1"] - rx) <= self.NUM_TOL:
                            nums[key] = self._num(w["text"])
                            break
                dates = {}
                for w in ws:
                    if not self._DATE_RE.match(w["text"]):
                        continue
                    for key, rx in self.DATE_RIGHT:
                        if abs(w["x1"] - rx) <= 4.5:
                            dates[key] = self._date(w["text"])
                            break

                unit_txt = self._txt(ws, self.W_UNIT)
                is_unit = bool(re.match(r"^\d", unit_txt)) and \
                    "market" in nums

                # ---- grand strip: numbers, no unit, no name ----------
                if not is_unit and "market" in nums and "balance" in nums \
                        and not self._txt(ws, self.W_NAME):
                    grand = nums
                    continue
                if not is_unit:
                    continue

                name = self._txt(ws, self.W_NAME)
                status = self._txt(ws, self.W_STATUS)
                vacant = bool(re.match(r"^vacant\b", name, re.I))
                u = UnitRecord(unit=unit_txt,
                               floor_plan=self._txt(ws, self.W_TYPE),
                               sqft=self._num(self._txt(ws, self.W_SQFT)))
                # bed/bath not encoded in the Type code - explicit blanks
                u.bed_explicit = u.bath_explicit = None
                u.bed_bath_explicit = True
                u.market_rent = nums.get("market")
                if vacant:
                    u.apt_status = "VU"
                else:
                    apt_st, res_st = self.STATUS_MAP.get(
                        status.upper(), (None, None))
                    if apt_st is None:
                        apt_st, res_st = "OC", "C"
                        self.flags.append(
                            f"unit {unit_txt}: unrecognised status "
                            f"'{status}' - treated as occupied/current")
                    if status.upper() == "UE":
                        self.flags.append(
                            f"unit {unit_txt}: status UE (under eviction) - "
                            "counted occupied; balance "
                            f"{nums.get('balance', 0):,.2f}")
                    u.apt_status = apt_st
                    charges = []
                    if nums.get("rent") is not None:
                        charges.append(("RENT", nums["rent"], False))
                    if nums.get("other"):
                        charges.append(("OTHER CHARGES", nums["other"],
                                        False))
                    if nums.get("credits"):
                        charges.append(("CREDIT", -abs(nums["credits"]),
                                        False))
                    r = Resident(
                        name=name, status=res_st, charges=charges,
                        move_in=dates.get("movein"),
                        lease_start=dates.get("lstart"),
                        lease_expires=dates.get("lend"),
                        move_out=dates.get("moveout"),
                        surety_bond=nums.get("surety"),
                        deposit=nums.get("deposit"),
                        ending_balance=nums.get("balance"))
                    u.residents.append(r)
                    # printed per-unit Total vs Rent + Other - Credits
                    want = nums.get("total")
                    got = (nums.get("rent") or 0) + (nums.get("other") or 0) \
                        - abs(nums.get("credits") or 0)
                    if want is not None and abs(got - want) > 0.01:
                        self.flags.append(
                            f"unit {unit_txt}: printed Total {want:,.2f} != "
                            f"Rent+Other-Credits {got:,.2f}")
                        checks["_bad_totals"] = \
                            checks.get("_bad_totals", 0) + 1
                if unit_txt in units:
                    self.flags.append(f"duplicate unit row {unit_txt} - "
                                      "keeping the first")
                    continue
                units[unit_txt] = u
                order.append(unit_txt)

            # ---- the report's own occupancy tables -------------------
            if in_summary:
                m = re.search(r"Total\s+Occupied\s+([\d,.]+)\s+[\d.]+%\s+"
                              r"(\d+)\s+[\d.]+%\s+([\d,.]+)", page_txt)
                if m:
                    checks["occupied_market_rent"] = self._num(m.group(1))
                    checks["occupied_count"] = int(m.group(2))
                    checks["occupied_sqft"] = self._num(m.group(3))
                m = re.search(r"Total\s+Vacant\s+([\d,.]+)\s+[\d.]+%\s+"
                              r"(\d+)\s+[\d.]+%\s+([\d,.]+)", page_txt)
                if m:
                    checks["vacant_market_rent"] = self._num(m.group(1))
                    checks["vacant_count"] = int(m.group(2))
                    checks["vacant_sqft"] = self._num(m.group(3))
                fps = {}
                seen_ut = False
                for _, ws2 in self._rows(page):
                    joined = " ".join(w["text"] for w in ws2)
                    if "Unit Type Occupancy" in joined:
                        seen_ut = True
                        continue
                    if not seen_ut:
                        continue
                    m = re.match(r"^(.+?)\s+(Occupied|Vacant)\s+([\d,.]+)\s+"
                                 r"[\d.]+%\s+(\d+)\s+[\d.]+%\s+([\d,.]+)\s+"
                                 r"[\d.]+%$", joined)
                    if not m:
                        continue
                    fp = m.group(1).strip()
                    e = fps.setdefault(fp, {"units": 0, "market": 0.0,
                                            "sqft": 0.0})
                    e["units"] += int(m.group(4))
                    e["market"] += self._num(m.group(3))
                    e["sqft"] += self._num(m.group(5))
                if fps:
                    checks["floor_plan_totals"] = fps
                    checks["unit_count"] = sum(v["units"]
                                               for v in fps.values())
                    checks["total_market_rent"] = round(
                        sum(v["market"] for v in fps.values()), 2)
                    checks["total_sqft"] = sum(v["sqft"]
                                               for v in fps.values())

        unit_list = [units[k] for k in order]
        if grand:
            checks["total_contract_rent"] = grand.get("rent")
            checks["current_lease_charges"] = grand.get("total")
            checks["total_deposits"] = grand.get("deposit")
            checks["report_total_market_rent"] = grand.get("market")
            checks["report_total_surety_bonds"] = grand.get("surety")
            checks["report_total_balance"] = grand.get("balance")
            checks["parsed_surety_bonds"] = round(
                sum(r.surety_bond or 0 for u in unit_list
                    for r in u.residents), 2)
            checks["parsed_balance"] = round(
                sum(r.ending_balance or 0 for u in unit_list
                    for r in u.residents), 2)
            # grand strip Other Charges / Credits vs the parsed charges
            for key in ("other", "credits"):
                want = grand.get(key)
                if want is None:
                    continue
                if key == "other":
                    got = sum(a for u in unit_list for r in u.residents
                              for c, a, _ in r.charges
                              if c == "OTHER CHARGES")
                else:
                    got = -sum(a for u in unit_list for r in u.residents
                               for c, a, _ in r.charges if c == "CREDIT")
                if abs(got - abs(want)) > 0.01:
                    self.flags.append(
                        f"grand strip {key} {want:,.2f} != parsed "
                        f"{got:,.2f}")
        if checks.get("_bad_totals"):
            sys.exit(f"ERROR: {checks['_bad_totals']} unit(s) whose printed "
                     "Total disagrees with Rent + Other Charges - Credits - "
                     "parse is wrong, refusing to write the workbook.")
        return prop, asof, unit_list, checks


class OwnerSheetPdfParser(RentRollParser):
    """Owner-maintained rent roll spreadsheet printed to PDF.

    Validated on Gardens Apartments (Paris, TX, 7/2026) - a Google Sheets
    print with four columns:

        Unit #  |  Tenant  |  Rent  |  Unit Type
        3-101   |  David Villanueva  |  $750.00  |  1 BR/ 1 BA
        3-109   |  Vacant            |           |  1 BR/ 1 BA

    Characteristics of the owner-made genre (as opposed to a PM system
    export) that this parser is built around:

    * **No as-of date.** Owner sheets almost never print one. `parse`
      searches for one anyway (an "as of" line, or any bare date in the
      header band) and, failing that, returns ``asof=None`` with
      ``asof_found=False`` so `main` can require an explicit ``--asof``
      rather than silently inventing a date. Never guess a precise date
      into the filename - that date is what Dmytro reads the file by.
    * **No market rent, no lease dates, no charge codes.** Only the actual
      contract rent is given, so Market Rent / lease dates / Other Income
      stay blank; the single rent figure maps to one RENT charge on a
      synthetic "current" resident. Do not back-fill market rent from the
      contract rent - blank is honest, a copied number is not.
    * **No sqft.** Fill via ``--sqft`` from a cited public source (see the
      missing-sqft protocol in CLAUDE.md).
    * **Occupancy by tenant name.** A blank tenant cell, or a marker word
      ("Vacant", "VAC", "Empty"), means vacant. There is no status column.
    * **Rarely a totals row.** If a trailing currency figure appears under
      the Rent column with no unit, it is captured as the report's printed
      contract-rent total. Because owner sheets print no unit/occupancy
      counts, `parse` *always* additionally re-derives unit count, occupied
      count and rent sum through a second, independent extraction path
      (pdfplumber `extract_tables`, i.e. the ruling-line grid) and files
      those as checks labelled "re-extract". Reconciliation therefore has
      something real to tie to even when the sheet prints no totals at all.
    """

    name = "OwnerSheet-PDF"

    # x-coordinate windows (word centres) for each column. Measured from the
    # Gardens print; generous enough to absorb column drift between sheets.
    W_UNIT   = (0,   170)
    W_TENANT = (170, 320)
    W_RENT   = (320, 400)
    W_TYPE   = (400, 9999)

    HEADER_TOKENS = ("unit", "tenant", "rent")
    VACANT_WORDS = {"vacant", "vac", "vacnt", "empty", "vacante", "vaccant"}
    # cells that look like a unit id but are structural, not a unit
    NON_UNIT = {"total", "totals", "grand total", "unit", "unit #", "sum",
                "subtotal", "vacant"}

    ROW_TOL = 3.0        # pts; words within this vertical distance = one row

    # -- helpers -------------------------------------------------------------

    @staticmethod
    def _money(s):
        """'$1,050.00' / '(750)' -> float; '' -> None."""
        s = (s or "").strip()
        if not s:
            return None
        neg = s.startswith("(") and s.endswith(")")
        s = re.sub(r"[^0-9.\-]", "", s)
        if not s or s in ("-", "."):
            return None
        try:
            v = float(s)
        except ValueError:
            return None
        return -v if neg else v

    @staticmethod
    def _is_unit(s):
        s = (s or "").strip()
        if not s or s.lower() in OwnerSheetPdfParser.NON_UNIT:
            return False
        # 3-101, 101, 12B, A-4, "Bldg 3 101"
        return bool(re.fullmatch(r"[A-Za-z0-9#]+(?:[-/ ][A-Za-z0-9#]+)*", s))

    @staticmethod
    def _fp(unit_type):
        """'1 BR/ 1 BA' -> '1/1'; '2BR/1.5BA' -> '2/1.5'; 'Studio' -> '0/1'.

        Returns "" when the cell is blank or unrecognised - an empty floor
        plan is dropped from the Floor Plan tabs by `_floor_plans`, which is
        the correct outcome: better a visibly missing plan than a guessed one.
        """
        s = (unit_type or "").strip()
        if not s:
            return ""
        if re.match(r"^(studio|eff)", s, re.I):
            return "0/1"
        m = re.search(r"(\d+(?:\.\d+)?)\s*(?:BR|BD|BED)S?\b\s*[/\-]?\s*"
                      r"(\d+(?:\.\d+)?)\s*(?:BA|BTH|BATH)", s, re.I)
        if not m:
            m = re.search(r"(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)", s)
        if not m:
            return ""

        def trim(x):
            f = float(x)
            return str(int(f)) if f == int(f) else str(f)
        return f"{trim(m.group(1))}/{trim(m.group(2))}"

    @classmethod
    def _cells(cls, words):
        """Bucket one row's words into the four column windows."""
        out = {"unit": [], "tenant": [], "rent": [], "type": []}
        for w in sorted(words, key=lambda z: z["x0"]):
            c = (w["x0"] + w["x1"]) / 2
            for key, (lo, hi) in (("unit", cls.W_UNIT),
                                  ("tenant", cls.W_TENANT),
                                  ("rent", cls.W_RENT),
                                  ("type", cls.W_TYPE)):
                if lo <= c < hi:
                    out[key].append(w["text"])
                    break
        return {k: " ".join(v).strip() for k, v in out.items()}

    @classmethod
    def _rows(cls, page):
        """[(top, {cells})] for one page, grouped by vertical position."""
        groups = []
        for w in sorted(page.extract_words(), key=lambda z: z["top"]):
            if groups and abs(w["top"] - groups[-1][0]) <= cls.ROW_TOL:
                groups[-1][1].append(w)
            else:
                groups.append((w["top"], [w]))
        return [(top, cls._cells(ws)) for top, ws in groups]

    @staticmethod
    def detect(pdf) -> bool:
        txt = (pdf.pages[0].extract_text() or "")
        for line in txt.splitlines()[:15]:
            low = line.lower()
            if all(t in low for t in ("unit", "tenant", "rent")) and \
                    "type" in low:
                return True
        return False

    # -- independent cross-check --------------------------------------------

    @classmethod
    def _reextract(cls, pdf):
        """Second, independent pass over the PDF via pdfplumber's
        table/ruling-line extractor (the positional word parser above never
        touches this code path). Returns (units, occupied, rent_sum).

        Owner sheets print no unit or occupancy counts, so this is what the
        reconciliation block ties the parse to.
        """
        n = occ = 0
        total = 0.0
        header_seen = False
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table:
                    cells = [(c or "").strip() for c in row]
                    # The title band above the column header ("Gardens
                    # Apartments") is a single left-aligned cell that passes
                    # _is_unit; anchor on the header row exactly as the
                    # positional pass does so it is not counted as a door.
                    if not header_seen:
                        low = " ".join(cells).lower()
                        if all(t in low for t in cls.HEADER_TOKENS):
                            header_seen = True
                        continue
                    if len(cells) < 3 or not cls._is_unit(cells[0]):
                        continue
                    n += 1
                    tenant = cells[1]
                    rent = cls._money(cells[2])
                    if tenant and tenant.lower() not in cls.VACANT_WORDS:
                        occ += 1
                        total += rent or 0.0
        return n, occ, total

    # -- main parse ----------------------------------------------------------

    def parse(self, pdf):
        self.asof_found = False
        self.flags = []
        prop, asof = "", None
        units, checks = [], {}
        printed_total = None
        header_seen = False

        for pi, page in enumerate(pdf.pages):
            for top, c in self._rows(page):
                joined = " ".join(v for v in c.values() if v).strip()
                if not joined:
                    continue
                low = joined.lower()

                # ---- header / title band (page 1, above the column header)
                if not header_seen:
                    if all(t in low for t in self.HEADER_TOKENS):
                        header_seen = True
                        continue
                    m = re.search(r"as[- ]of[:\s]*"
                                  r"(\d{1,2}/\d{1,2}/\d{2,4})", joined, re.I)
                    if not m:
                        m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{4})\b", joined)
                    if m:
                        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                            try:
                                asof = datetime.strptime(
                                    m.group(1), fmt).date()
                                self.asof_found = True
                                break
                            except ValueError:
                                pass
                    elif not prop and pi == 0:
                        prop = joined
                    continue

                unit = c["unit"]

                # ---- totals row: money under Rent, no unit id
                if not self._is_unit(unit):
                    v = self._money(c["rent"])
                    if v is not None and not c["tenant"]:
                        printed_total = v
                    continue

                tenant = c["tenant"]
                rent = self._money(c["rent"])
                vacant = (not tenant) or tenant.lower() in self.VACANT_WORDS

                u = UnitRecord(unit=unit,
                               floor_plan=self._fp(c["type"]),
                               sqft=None,
                               apt_status="VU" if vacant else "OC",
                               market_rent=None)
                if not vacant:
                    u.residents.append(Resident(
                        name=tenant, status="C",
                        charges=[("RENT", rent, False)]
                        if rent is not None else []))
                    if rent is None:
                        self.flags.append(
                            f"unit {unit}: occupied by {tenant} but no rent "
                            "printed")
                elif rent is not None:
                    self.flags.append(
                        f"unit {unit}: marked vacant but shows rent "
                        f"${rent:,.2f} - rent excluded")
                if not u.floor_plan:
                    self.flags.append(
                        f"unit {unit}: no Unit Type printed - floor plan, "
                        "bed/bath and sqft left blank")
                units.append(u)

        dupes = {u.unit for u in units
                 if [x.unit for x in units].count(u.unit) > 1}
        if dupes:
            self.flags.append("duplicate unit id(s): " + ", ".join(sorted(dupes)))

        # ---- checks: printed totals row (if any) + independent re-extraction
        src = {}
        if printed_total is not None:
            checks["total_contract_rent"] = printed_total
            src["total_contract_rent"] = "printed totals row"
        rx_n, rx_occ, rx_rent = self._reextract(pdf)
        checks["unit_count"] = rx_n
        checks["occupied_count"] = rx_occ
        src["unit_count"] = "re-extract"
        src["occupied_count"] = "re-extract"
        if printed_total is None:
            checks["total_contract_rent"] = rx_rent
            src["total_contract_rent"] = "re-extract"
        else:
            checks["reextract_contract_rent"] = rx_rent
            src["reextract_contract_rent"] = "re-extract"
        checks["_src"] = src

        if not self.asof_found:
            self.flags.append(
                "the sheet prints NO as-of date - the date in the filename "
                "and cell B2 came from --asof, not from the source")
        return prop, asof, units, checks

    def source_note(self, asof):
        d = asof.strftime("%m/%d/%Y") if asof else "unknown date"
        stamp = ("as-of date supplied externally (--asof); the sheet prints "
                 "none" if not getattr(self, "asof_found", False)
                 else "as-of date read from the sheet")
        return (f"Generated from an owner-prepared rent roll spreadsheet "
                f"printed to PDF ({d}; {stamp}). Contractual Rent = the "
                "sheet's Rent column. The sheet provides no market rent, "
                "lease dates, concessions, other income or square footage - "
                "those columns are intentionally blank. Net Sf, where "
                "present, is web-sourced, not from the rent roll.")


class OneSiteRentsParser(RentRollParser):
    """Parser for the RealPage **OneSite Rents v3.0 'RENT ROLL DETAIL'** PDF
    (report id mgt-521-003), report type "Details + Summary".

    Validated on Capstone Real Estate Services Inc. - Synott Square
    (Houston TX, as of 06/29/2026, 108 units, 30 pages).

    Layout (792 x 612 landscape, Arial 6.2pt, fixed column x-positions):

      Unit | Floorplan | Unit Designation | SQFT | Unit/Lease Status | Name |
      Move-In / Move-Out | Lease Start | Lease End | Market Rent |
      Sub Journal + Trans Code | Rent | Lease Charges/Credits |
      Total Billing | Dep On Hand | Other balance

    Structure
    ---------
    * One **lease block** per unit row: unit identity + the lease's first
      charge line.  Every further charge of that lease is its own row
      carrying only Name + Sub Journal/Trans Code + Rent / Lease Charges.
    * The resident **name is re-printed (wrapped) on every charge row**, so
      the full name is assembled from the opening row plus the name-only
      continuation rows that precede the lease's *second* charge row.
    * A row with no unit number but a status in the Unit/Lease Status column
      ("Pending", "Pending renewal", "Applicant") opens an **additional lease
      block on the same unit** - a future/renewal lease.  Its amounts are all
      flagged with `*` ("indicates amounts not included in detail totals"),
      so those residents are status `L` and never inflate the door count or
      the in-place rent.
    * Move-In is on the opening row of a lease; **Move-Out prints on the next
      row in the same column**.
    * Sub Journal and Trans Code are rendered without a gap
      ("RESIDENTPESTCONTROLREIMB"), and a long trans code overflows into the
      Rent column ("...REIMB0.00"). Codes are therefore split using the
      report's own "summary billing by sub journal" list, and every parsed
      code is asserted to exist in the "summary billing by transaction code"
      list.

    Charge routing
    --------------
    * The report prints the unit rent in its own **Rent** column and every
      other charge/credit in **Lease Charges/Credits**; both column totals
      are tied out separately.
    * ``RENT`` -> Contractual Rent.  ``ADMINUNIT`` (office/admin unit credit)
      and ``EMPLCRED`` (employee credit) are recurring credits that offset
      the unit's rent, so they land in Emp./Other Discounts (col O) and are
      inside NER - never in Upfront Concessions, which NER excludes.
    * ``MTOM`` is the explicit month-to-month fee and is the ONLY source of
      the MTM flag (house rule: never infer MTM from an expired lease).
    * HAP/voucher codes are handled by the shared SUBSIDY_CODES machinery
      (none present at Synott Square).

    Summary tie-out
    ---------------
    Everything the "Details + Summary" report prints is reconciled: the
    detail grand-total row (6 columns), the "Amt / SQFT" market/leased
    square footage, the floor-plan table (9 measures per plan + the
    totals/averages row), the "occupancy and rents summary" status buckets
    (market rent / # units / potential rent), the sub-journal billing
    summary and the transaction-code billing summary, plus every lease's
    printed Total Billing.
    """

    name = "RealPage OneSite Rents"

    # --- column geometry (measured word positions, 792pt landscape) --------
    W_UNIT   = (14, 50)
    W_FP     = (50, 140)
    W_DESIG  = (140, 190)
    W_SQFT   = (190, 215)
    W_STATUS = (215, 266)
    W_NAME   = (266, 325)
    W_MIO    = (325, 363)      # Move-In (row 1) / Move-Out (row 2)
    W_LSTART = (363, 402)
    W_LEND   = (402, 445)
    W_CODE   = (490, 592)      # sub journal + trans code, glued together

    # Right-aligned numeric columns, keyed by their printed right edge.
    NUM_RIGHT = (("market", 484.8), ("rent", 602.1), ("charges", 643.2),
                 ("billing", 689.2), ("dep", 737.6), ("balance", 771.6))
    NUM_TOL = 3.0
    ROW_TOL = 3.5              # < the 10pt minimum line gap

    _NUM_RE = re.compile(r"^\(?-?[\d,]+\.\d{2}\)?$|^\(?-?[\d,]+\)?$")
    _NUM_TAIL_RE = re.compile(r"^(.*?)(\(?-?[\d,]*\.\d{2}\)?)$")

    # Rows belonging to the repeated page header / footer.
    SKIP_RE = re.compile(
        r"OneSite Rents|^Parameters:|RENT ROLL DETAIL|^As of \d|^details$"
        r"|indicates amounts not included|^Market Other$|^Credits$"
        r"|Unit/Lease|Floorplan Designation SQFT", re.I)

    # Unit/Lease Status vocabulary
    FUTURE_STATUS_RE = re.compile(
        r"^(pending|applicant|approved|future|reserved)\b", re.I)
    NONREV_STATUS_RE = re.compile(r"^(admin|down|model|non.?rev)\b", re.I)

    # Charge codes that mark a unit's lease type for col F.
    LEASE_TYPE_CODES = (
        (re.compile(r"^ADMIN\s*UNIT$", re.I), "Non-Revenue (Admin/Office)"),
        (re.compile(r"^EMPL?\s*CRED(IT)?$", re.I), "Employee Unit"),
    )

    MTM_CODE_RE = re.compile(r"^MTOM$|^MTM$|^M2M$|MONTH.?TO.?MONTH", re.I)

    # ------------------------------------------------------------------ #
    # helpers
    # ------------------------------------------------------------------ #

    @staticmethod
    def _num(s):
        if s is None:
            return None
        s = s.replace(",", "").replace("$", "").strip()
        neg = s.startswith("(") and s.endswith(")")
        s = s.strip("()")
        try:
            v = float(s)
        except ValueError:
            return None
        return -v if neg else v

    @staticmethod
    def _date(s):
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(s.strip(), fmt).date()
            except ValueError:
                pass
        return None

    @staticmethod
    def _r2(x, places=2):
        """Round half-up to `places` decimals - the report's convention
        (1081.1166 -> 1081.12, 989.375 -> 989.38, 760.259 -> 760)."""
        f = 10 ** places
        return math.floor(abs(x) * f + 0.5) / f * (-1 if x < 0 else 1)

    def _rows(self, page):
        """Cluster words into visual rows on their vertical centre (fields on
        one printed line sit on baselines up to ~2.5pt apart)."""
        words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))
        out, cur, cy = [], [], None
        for w in words:
            c = (w["top"] + w["bottom"]) / 2.0
            if cy is None or abs(c - cy) <= self.ROW_TOL:
                cur.append(w)
                cy = c if cy is None else (cy * (len(cur) - 1) + c) / len(cur)
            else:
                out.append(sorted(cur, key=lambda x: x["x0"]))
                cur, cy = [w], c
        if cur:
            out.append(sorted(cur, key=lambda x: x["x0"]))
        return out

    def _split(self, ws):
        """Split one visual row into its columns.

        Returns (text_cells, num_cells, starred) where text_cells maps a
        column name to its joined text, num_cells maps a numeric column name
        to its float value, and starred is the set of numeric columns whose
        value carried the '*' (excluded-from-totals) marker.
        """
        text = {}
        nums = {}
        starred = set()
        leftovers = []
        for w in ws:
            t, x0, x1 = w["text"], w["x0"], w["x1"]
            if t == "*":
                for key, rx in self.NUM_RIGHT:
                    if rx + 0.5 < x0 < rx + 12:
                        starred.add(key)
                        break
                continue
            # numeric (possibly with an overflowed trans-code prefix glued on)
            m = self._NUM_TAIL_RE.match(t)
            if m and m.group(2):
                hit = next((k for k, rx in self.NUM_RIGHT
                            if abs(x1 - rx) <= self.NUM_TOL), None)
                if hit:
                    nums[hit] = self._num(m.group(2))
                    if m.group(1):
                        text.setdefault("code", []).append(m.group(1))
                    continue
            for key, (lo, hi) in (("unit", self.W_UNIT), ("fp", self.W_FP),
                                  ("desig", self.W_DESIG),
                                  ("sqft", self.W_SQFT),
                                  ("status", self.W_STATUS),
                                  ("name", self.W_NAME), ("mio", self.W_MIO),
                                  ("lstart", self.W_LSTART),
                                  ("lend", self.W_LEND)):
                if lo <= x0 < hi:
                    text.setdefault(key, []).append(t)
                    break
            else:
                lo, hi = self.W_CODE
                if lo <= x0 < hi:
                    text.setdefault("code", []).append(t)
                else:
                    leftovers.append(w)
        joined = {k: (" ".join(v) if k != "code" else "".join(v))
                  for k, v in text.items()}
        return joined, nums, starred, leftovers

    # ------------------------------------------------------------------ #

    @staticmethod
    def detect(pdf) -> bool:
        txt = pdf.pages[0].extract_text() or ""
        return ("OneSite" in txt and
                re.search(r"RENT\s+ROLL\s+DETAIL", txt, re.I) is not None)

    def source_note(self, asof):
        return (f"Generated from RealPage OneSite Rents v3.0 'Rent Roll "
                f"Detail' PDF (as of "
                f"{asof.strftime('%m/%d/%Y') if asof else 'unknown date'}). "
                "Contractual Rent = the lease's RENT charge (Rent column); "
                "Other Income = the remaining Lease Charges/Credits for the "
                "current/on-notice lease; pending / renewal / applicant "
                "leases (amounts flagged '*' by the report and excluded from "
                "its detail totals) are carried as future residents only.")

    # ------------------------------------------------------------------ #

    def parse(self, pdf):
        self.flags = []
        prop, asof = self._header(pdf)

        rows = []
        for page in pdf.pages:
            for ws in self._rows(page):
                txt = " ".join(w["text"] for w in ws).strip()
                if not txt or self.SKIP_RE.search(txt):
                    continue
                rows.append(ws)

        summary = self._parse_summary(rows)
        detail = self._parse_detail(rows, summary)
        return prop, asof, detail[0], self._checks(detail, summary)

    # -- report header -------------------------------------------------- #

    def _header(self, pdf):
        txt = pdf.pages[0].extract_text() or ""
        prop = None
        for line in txt.splitlines()[:3]:
            m = re.search(r"OneSite\s+Rents\s+v[\d.]+\s+(.*?)\s+Page\s+\d+\s+of",
                          line)
            if m:
                who = m.group(1).strip()
                prop = who.split(" - ")[-1].strip() if " - " in who else who
                break
        m = re.search(r"As of\s+(\d{2}/\d{2}/\d{4})", txt)
        asof = self._date(m.group(1)) if m else None
        return prop, asof

    # -- summary section ------------------------------------------------ #

    def _parse_summary(self, rows):
        s = {"floor_plans": {}, "fp_totals": None, "occupancy": {},
             "occ_totals": None, "sub_journals": {}, "sub_journal_total": None,
             "trans_codes": {}, "trans_code_total": None,
             "market_sqft": None, "leased_sqft": None, "detail_totals": None,
             "start_row": None}
        state = None
        for i, ws in enumerate(rows):
            txt = re.sub(r"\s+", " ", " ".join(w["text"] for w in ws)).strip()
            low = txt.lower()

            if low.startswith("totals:") and s["start_row"] is None:
                # grand-total row of the detail section (bold, so its right
                # edges are shifted a couple of points - read positionally)
                vals = [self._num(w["text"]) for w in ws
                        if self._NUM_RE.match(w["text"])]
                if len(vals) == 6:
                    s["detail_totals"] = dict(zip(
                        ("market", "rent", "charges", "billing", "dep",
                         "balance"), vals))
                continue

            m = re.search(r"Amt\s*/\s*SQFT:\s*Market\s*=\s*([\d,]+)\s*SQFT;\s*"
                          r"Leased\s*=\s*([\d,]+)\s*SQFT", txt, re.I)
            if m:
                s["market_sqft"] = self._num(m.group(1))
                s["leased_sqft"] = self._num(m.group(2))
                s["start_row"] = i
                state = None
                continue
            if s["start_row"] is None:
                continue

            if low.startswith("floorplan") and "# units" in low:
                state = "fp"
                continue
            if low.startswith("occupancy and rents summary"):
                state = None
                continue
            if low.startswith("unit status") and "potential rent" in low:
                state = "occ"
                continue
            if low.startswith("summary billing by sub journal"):
                state = None
                continue
            if low.startswith("sub journal amount"):
                state = "sj"
                continue
            if low.startswith("summary billing by transaction code"):
                state = None
                continue
            if low.startswith("code amount"):
                state = "tc"
                continue

            if state == "fp":
                if low.startswith("totals / averages"):
                    s["fp_totals"] = self._fp_row(ws)
                    state = None
                    continue
                label = " ".join(w["text"] for w in ws if w["x0"] < 140)
                vals = self._fp_row(ws)
                if label and vals:
                    s["floor_plans"][label] = vals
                continue

            if state == "occ":
                if low.startswith("totals:"):
                    s["occ_totals"] = self._occ_row(ws)[1]
                    state = None
                    continue
                label, vals = self._occ_row(ws)
                if label:
                    s["occupancy"][self._occ_key(label)] = vals
                continue

            if state in ("sj", "tc"):
                nums = [w for w in ws if self._NUM_RE.match(w["text"])]
                label = " ".join(w["text"] for w in ws
                                 if w["x0"] < 200 and w not in nums).strip()
                amt = self._num(nums[-1]["text"]) if nums else None
                if low.startswith("total:"):
                    s["sub_journal_total" if state == "sj"
                      else "trans_code_total"] = amt
                    state = None
                elif label and amt is not None:
                    (s["sub_journals"] if state == "sj"
                     else s["trans_codes"])[label] = amt
                continue
        return s

    FP_KEYS = ("units", "avg_sqft", "avg_market", "market_psf", "avg_leased",
               "leased_psf", "occupied", "occupancy_pct", "available")

    def _fp_row(self, ws):
        vals = [self._num(w["text"]) for w in ws
                if self._NUM_RE.match(w["text"])]
        if len(vals) != len(self.FP_KEYS):
            return None
        return dict(zip(self.FP_KEYS, vals))

    OCC_RIGHT = (("market", 238.5), ("units", 309.0), ("potential", 416.5))

    def _occ_row(self, ws):
        vals = {}
        label = []
        for w in ws:
            if self._NUM_RE.match(w["text"]):
                hit = next((k for k, rx in self.OCC_RIGHT
                            if abs(w["x1"] - rx) <= 5.0), None)
                if hit:
                    vals[hit] = self._num(w["text"])
                    continue
            if w["x0"] < 200:
                label.append(w["text"])
        return " ".join(label).strip(), vals

    @staticmethod
    def _occ_key(label):
        return re.sub(r"[^a-z]+", " ", label.lower()).strip()

    # -- detail section -------------------------------------------------- #

    def _parse_detail(self, rows, summary):
        subjournals = sorted(summary["sub_journals"], key=len, reverse=True)
        known_codes = set(summary["trans_codes"])
        units, order = {}, []
        u = None            # current UnitRecord
        r = None            # current lease (Resident)
        name_open = False   # still collecting the wrapped resident name
        lease_billing = []  # (unit, resident, printed Total Billing)
        stop = summary["start_row"]

        for i, ws in enumerate(rows):
            if stop is not None and i >= stop:
                break
            txt = re.sub(r"\s+", " ", " ".join(w["text"] for w in ws)).strip()
            if txt.lower().startswith("totals:"):
                continue
            cells, nums, starred, extra = self._split(ws)
            if extra:
                self.flags.append(
                    "unmapped text on a detail row: "
                    + ", ".join("%s@%.0f" % (w["text"], w["x0"])
                                for w in extra))
            unit_txt = cells.get("unit", "").strip()
            status = cells.get("status", "").strip()
            code_txt = cells.get("code", "").strip()

            new_unit = bool(unit_txt) and re.search(r"\d", unit_txt)
            # A status in the Unit/Lease Status column always opens a lease
            # block; on a vacant unit the applicant row is the FIRST block,
            # so this must not be conditioned on an existing lease.
            new_lease = new_unit or (bool(status) and u is not None)

            if new_unit:
                uid = unit_txt
                u = UnitRecord(unit=uid,
                               floor_plan=cells.get("fp", "").strip(),
                               sqft=self._num(cells.get("sqft")),
                               market_rent=nums.get("market"))
                u.onesite_status = status
                u.designation = cells.get("desig", "").strip()
                if self.NONREV_STATUS_RE.match(status):
                    u.apt_status = "NR"
                elif status.lower().startswith("vacant"):
                    u.apt_status = "VL" if "lease" in status.lower() else "VU"
                elif status.lower().startswith("occupied"):
                    u.apt_status = "NA" if "NTV" in status.upper() else "OC"
                else:
                    u.apt_status = "OC"
                    self.flags.append(f"unit {uid}: unrecognised Unit/Lease "
                                      f"status '{status}' - treated as occupied")
                units[uid] = u
                order.append(uid)
                r = None

            if u is None:
                continue

            if new_lease:
                vacant_row = (u.apt_status.startswith("V")
                              and not self.FUTURE_STATUS_RE.match(status))
                if vacant_row:
                    r = None
                    name_open = False
                else:
                    if new_unit:
                        rstat = "N" if u.apt_status == "NA" else "C"
                        if u.apt_status == "NR":
                            rstat = "C"
                    else:
                        rstat = "L"
                        if not self.FUTURE_STATUS_RE.match(status):
                            self.flags.append(
                                f"unit {u.unit}: extra lease block with "
                                f"status '{status}' treated as future/renewal")
                    r = Resident(
                        name=cells.get("name", "").strip(),
                        status=rstat,
                        move_in=self._date(cells.get("mio")),
                        lease_start=self._date(cells.get("lstart")),
                        lease_expires=self._date(cells.get("lend")),
                        deposit=nums.get("dep"),
                        ending_balance=nums.get("balance"),
                    )
                    r.charge_columns = []
                    r.charge_subjournals = []
                    r.balance_starred = "balance" in starred
                    r.lease_status = status
                    u.residents.append(r)
                    if nums.get("billing") is not None:
                        lease_billing.append((u, r, nums["billing"]))
                    name_open = True
            elif r is not None:
                if cells.get("mio"):
                    r.move_out = self._date(cells["mio"])
                if name_open and cells.get("name") and not code_txt:
                    r.name = (r.name + " " + cells["name"]).strip()

            if code_txt and r is not None:
                if not new_lease:
                    name_open = False
                sj = next((j for j in subjournals if code_txt.startswith(j)),
                          "")
                code = code_txt[len(sj):] if sj else code_txt
                if known_codes and code not in known_codes:
                    self.flags.append(
                        f"unit {u.unit}: charge code '{code_txt}' does not "
                        f"resolve against the report's transaction-code "
                        f"summary")
                col = "rent" if nums.get("rent") else (
                    "charges" if nums.get("charges") is not None else None)
                if nums.get("rent") and nums.get("charges"):
                    self.flags.append(
                        f"unit {u.unit}: charge {code} carries an amount in "
                        f"BOTH the Rent and Lease Charges columns")
                amt = (nums.get("rent") or 0) + (nums.get("charges") or 0)
                r.charges.append((code, amt, r.status == "L"))
                r.charge_columns.append(col or "charges")
                r.charge_subjournals.append(sj)

        unit_list = [units[k] for k in order]

        for u in unit_list:
            for r in u.residents:
                if any(self.MTM_CODE_RE.search(c) for c, _, _ in r.charges):
                    r.term_type = "MTM"
            p = u.primary
            if p:
                for pat, label in self.LEASE_TYPE_CODES:
                    if any(pat.search(c) for c, _, _ in p.charges):
                        u.lease_type = label
                if p.subsidy_charge:
                    u.lease_type = "Section 8 Voucher"

        # Bed/bath is not printed anywhere in a OneSite rent roll; the
        # floor-plan code (A1/A2/B1/B2) does not reliably encode it either,
        # so it is left blank for --bedbath / --bedbath-est to fill from a
        # cited source rather than guessed from the code.
        for u in unit_list:
            u.bed_explicit = None
            u.bath_explicit = None
            u.bed_bath_explicit = True

        return unit_list, lease_billing

    # -- reconciliation payload ------------------------------------------ #

    def _checks(self, detail, s):
        units, lease_billing = detail
        checks = {}
        fpt = s["fp_totals"] or {}
        dt = s["detail_totals"] or {}
        occ = s["occupancy"]

        checks["unit_count"] = fpt.get("units")
        checks["total_sqft"] = s["market_sqft"]
        checks["occupied_sqft"] = s["leased_sqft"]
        checks["occupied_count"] = fpt.get("occupied")
        vac = sum(v.get("units", 0) for k, v in occ.items()
                  if k.startswith("vacant"))
        checks["vacant_count"] = vac if occ else None
        checks["total_market_rent"] = dt.get("market")
        checks["total_contract_rent"] = dt.get("rent")
        checks["total_deposits"] = dt.get("dep")
        checks["current_lease_charges"] = dt.get("billing")
        checks["pct_unit_occupancy"] = fpt.get("occupancy_pct")

        def cn(us):
            return [r for u in us for r in u.residents
                    if r.status.upper() in ("C", "N")]

        def col_total(us, col):
            t = 0.0
            for r in cn(us):
                for (c, a, _), cc in zip(r.charges, r.charge_columns):
                    if cc == col:
                        t += a
            return t

        extra = []

        # 1. detail grand-total row --------------------------------------
        if dt:
            extra.append(("detail totals: Rent column",
                          lambda us: col_total(us, "rent"), dt["rent"], 0.01))
            extra.append(("detail totals: Lease Charges/Credits column",
                          lambda us: col_total(us, "charges"), dt["charges"],
                          0.01))
            extra.append((
                "detail totals: Other balance",
                lambda us: sum(r.ending_balance or 0 for u in us
                               for r in u.residents
                               if not getattr(r, "balance_starred", False)),
                dt["balance"], 0.01))

        # 2. per-lease printed Total Billing ------------------------------
        if lease_billing:
            n = len(lease_billing)
            bad = [u.unit for u, r, want in lease_billing
                   if abs(r.total_charges - want) > 0.01]
            checks["_onesite_bad_billing"] = bad
            extra.append((f"per-lease printed Total Billing ({n} leases)",
                          lambda us, n=n, bad=bad: n - len(bad), n, 0.5))

        # 3. floor-plan table --------------------------------------------
        def plan(us, fp):
            return [u for u in us if u.floor_plan == fp]

        def occupied(us):
            return [u for u in us if not u.is_vacant]

        def leased_rent(us):
            return sum((u.primary.rent_charge or 0)
                       for u in occupied(us) if u.primary)

        def avail(us):
            n = 0
            for u in us:
                has_fut = any(r.status.upper() == "L" for r in u.residents)
                if (u.is_vacant or u.on_notice) and not has_fut:
                    n += 1
            return n

        def _plan_checks(label, sel, want):
            g = lambda us: sel(us)                      # noqa: E731
            r2 = self._r2
            return [
                (f"  {label}: # units", lambda us: len(g(us)),
                 want["units"], 0.5),
                (f"  {label}: average SQFT",
                 lambda us: r2(sum(u.sqft or 0 for u in g(us))
                               / max(len(g(us)), 1), 0),
                 want["avg_sqft"], 0.01),
                (f"  {label}: average market rent",
                 lambda us: r2(sum(u.market_rent or 0 for u in g(us))
                               / max(len(g(us)), 1)),
                 want["avg_market"], 0.01),
                (f"  {label}: market amt / SQFT",
                 lambda us: r2(sum(u.market_rent or 0 for u in g(us))
                               / max(sum(u.sqft or 0 for u in g(us)), 1)),
                 want["market_psf"], 0.01),
                (f"  {label}: average leased amt",
                 lambda us: r2(leased_rent(g(us))
                               / max(len(occupied(g(us))), 1)),
                 want["avg_leased"], 0.01),
                (f"  {label}: leased amt / SQFT",
                 lambda us: r2(leased_rent(g(us))
                               / max(sum(u.sqft or 0
                                         for u in occupied(g(us))), 1)),
                 want["leased_psf"], 0.01),
                (f"  {label}: units occupied",
                 lambda us: len(occupied(g(us))), want["occupied"], 0.5),
                (f"  {label}: occupancy %",
                 lambda us: r2(len(occupied(g(us))) / max(len(g(us)), 1) * 100),
                 want["occupancy_pct"], 0.01),
                (f"  {label}: units available",
                 lambda us: avail(g(us)), want["available"], 0.5),
            ]

        for fp, want in sorted(s["floor_plans"].items()):
            extra += _plan_checks(f"plan {fp}",
                                  lambda us, fp=fp: plan(us, fp), want)
        if fpt:
            extra += _plan_checks("plans total/average", lambda us: list(us),
                                  fpt)

        # 4. occupancy and rents summary ----------------------------------
        def bucket(u):
            has_fut = any(r.status.upper() == "L" for r in u.residents)
            if u.apt_status == "NR":
                return "admin down"
            if u.is_vacant:
                return "vacant leased" if has_fut else "vacant not leased"
            if u.on_notice:
                return "occupied ntv leased" if has_fut else "occupied ntv"
            return "occupied no ntv"

        def bucket_units(us, key):
            return [u for u in us if bucket(u) == key]

        def potential(u):
            if u.is_vacant:
                return u.market_rent or 0
            return (u.primary.rent_charge or 0) if u.primary else 0

        for key in sorted(occ):
            want = occ[key]
            if want.get("units") is not None:
                extra.append((f"  occupancy [{key}]: # units",
                              lambda us, k=key: len(bucket_units(us, k)),
                              want["units"], 0.5))
            if want.get("market") is not None:
                extra.append((f"  occupancy [{key}]: market rent",
                              lambda us, k=key: sum(u.market_rent or 0
                                                    for u in bucket_units(us, k)),
                              want["market"], 0.01))
            if want.get("potential") is not None:
                extra.append((f"  occupancy [{key}]: potential rent",
                              lambda us, k=key: sum(potential(u)
                                                    for u in bucket_units(us, k)),
                              want["potential"], 0.01))
        if s["occ_totals"]:
            t = s["occ_totals"]
            if t.get("market") is not None:
                extra.append(("  occupancy totals: market rent",
                              lambda us: sum(u.market_rent or 0 for u in us),
                              t["market"], 0.01))
            if t.get("units") is not None:
                extra.append(("  occupancy totals: # units",
                              lambda us: len(us), t["units"], 0.5))
            if t.get("potential") is not None:
                extra.append(("  occupancy totals: potential rent",
                              lambda us: sum(potential(u) for u in us),
                              t["potential"], 0.01))

        # 5. billing by sub journal ---------------------------------------
        for sj, want in sorted(s["sub_journals"].items()):
            extra.append((
                f"  sub journal {sj}",
                lambda us, sj=sj: sum(
                    a for u in us for r in cn([u])
                    for (c, a, _), j in zip(r.charges, r.charge_subjournals)
                    if j == sj),
                want, 0.01))
        if s["sub_journal_total"] is not None:
            extra.append(("  sub journal total",
                          lambda us: sum(r.total_charges for r in cn(us)),
                          s["sub_journal_total"], 0.01))

        # 6. billing by transaction code ----------------------------------
        for code, want in sorted(s["trans_codes"].items()):
            extra.append((
                f"  trans code {code}",
                lambda us, code=code: sum(a for r in cn(us)
                                          for c, a, _ in r.charges
                                          if c == code),
                want, 0.01))
        if s["trans_code_total"] is not None:
            extra.append(("  trans code total",
                          lambda us: sum(r.total_charges for r in cn(us)),
                          s["trans_code_total"], 0.01))

        checks["extra_checks"] = extra
        return checks


# ----------------------------------------------------------------------------
# Buildium-style "Rent Roll" PDF (Olivos Management / Benbrook, 8/2026)
# ----------------------------------------------------------------------------

def _undouble(tok):
    """Collapse a fake-bold token whose every glyph is printed twice.

    Buildium renders its bold rows by emitting each character TWICE
    ("44663399 WWiilllliiaammss RRooaadd" = "4639 Williams Road",
    "$$3366,,006655..0000" = "$36,065.00"). Applied per TOKEN only after the
    whole ROW has been proved doubled (see `_row_undouble`), because an
    isolated numeric like "4400" is itself a valid pair-doubled string and
    must not be silently halved to "40".
    """
    if len(tok) >= 2 and len(tok) % 2 == 0 and \
            all(tok[i] == tok[i + 1] for i in range(0, len(tok), 2)):
        return tok[::2]
    return tok


def _is_doubled_row(texts):
    """True when every non-trivial token in the row is glyph-doubled."""
    real = [t for t in texts if len(t) >= 2]
    if len(real) < 2 or not any(len(t) >= 6 for t in real):
        return False
    return all(_undouble(t) != t for t in real)


def _row_undouble(texts):
    return [_undouble(t) for t in texts] if _is_doubled_row(texts) else texts


class BuildiumRentRollParser(RentRollParser):
    """Parser for the Buildium-style "Rent Roll" PDF produced by Olivos
    Management (validated on Benbrook Apartments / 4639 Williams Road,
    40 doors, as of 8/5/2026).

    Layout - 792pt landscape, detail on pages 1..n-1, summaries on the last:

        Lease Start | Lease End | Bed/Bath | Rent Cycle | Rent Start |
        Rent | Recurring Charges | Recurring Credits | Total |
        Deposits Held | Prepayments | Balance Due

    Gotchas this parser exists for (read before touching it):

    * **There is NO unit identifier column.** The detail rows are anonymous -
      the only unit-level identity in the whole report is the group header
      ("4639 Williams Road"). Units are therefore numbered in REPORT ORDER
      ("Unit 01".."Unit 40") and the substitution is FLAGged; nothing in the
      output pretends those are the property's real unit numbers. (Verified
      against `page.chars`: there is no white/clipped text and no column left
      of the Lease Start window.)
    * **Fake-bold rows print every glyph twice** - the property group header
      and every totals row. `_row_undouble` fixes them; it is deliberately a
      whole-row test so a real "4400" is never halved.
    * **The Rent column is LEFT-aligned (x0 ~ 336); every other numeric
      column is RIGHT-aligned.** Snapping Rent by right edge fails (its right
      edge moves 362 -> 368 with the digit count, and 385 on the bold totals
      row). Rent is claimed by an x0 window, the rest by right edge.
    * **"Rent Cycle: Monthly" is the BILLING cycle, not a month-to-month
      lease.** MTM is never inferred from it (nothing in this report says
      MTM), and "Rent Start" is the date the current rent schedule began -
      it is a DATE, not a scheduled/market rent figure.
    * **Recurring Charges INCLUDES the Rent.** Other Income is
      (Recurring Charges - Rent), booked as one lump charge because the
      report carries no per-charge detail. The report's own
      "Total for <property>" strip proves both halves.
    * **No per-unit Market Rent and no per-unit Square Feet in the detail** -
      but the last page's "Summary by bed/bath" block prints BOTH per plan.
      Square Feet is exact and uniform (18,000/24 = 750, 14,800/16 = 925), so
      per-unit sqft is allocated from that block - it comes from the rent
      roll, not from the web, and the total ties. Market Rent is NOT
      allocatable: the report prints $13,665 over only 13 of the 40 units
      (the plan "averages" 962.50 / 1,193.00 are averages over that subset,
      not per-unit values), so it is surfaced as a FLAG and left for
      --estimate-market. It is deliberately NOT filed as a tie-out check.
    """

    name = "Buildium"

    # numeric columns: Rent is LEFT-aligned, the rest RIGHT-aligned
    W_RENT_X0 = (330.0, 396.0)
    NUM_EDGES = [("charges", 445.0), ("credits", 504.0), ("total", 564.0),
                 ("deposits", 623.0), ("prepayments", 683.0),
                 ("balance", 758.0)]
    EDGE_TOL = 12.0
    # text columns, by word centre
    W_LEASE_START = (0.0, 95.0)
    W_LEASE_END = (95.0, 152.0)
    W_BEDBATH = (152.0, 212.0)
    W_CYCLE = (212.0, 282.0)
    W_RENT_START = (282.0, 330.0)

    ROW_TOL = 3.0
    MONEY = re.compile(r"^\(?-?\$?[\d,]*\d(?:\.\d{2})?\)?%?$")
    DATE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$")
    BEDBATH = re.compile(
        r"^(\d+(?:\.\d+)?)\s*Bed\s*/\s*(\d+(?:\.\d+)?)\s*Bath$", re.I)

    @staticmethod
    def detect(pdf) -> bool:
        txt = (pdf.pages[0].extract_text() or "")
        head = "\n".join(txt.splitlines()[:12])
        need = ("Lease Start", "Lease End", "Bed/Bath", "Prepayments",
                "Balance Due")
        return "Rent Roll" in head and all(t in head for t in need)

    # -- helpers -------------------------------------------------------------

    @classmethod
    def _rows(cls, page):
        rows = []
        for w in sorted(page.extract_words(), key=lambda z: z["top"]):
            if rows and abs(w["top"] - rows[-1][0]) <= cls.ROW_TOL:
                rows[-1][1].append(w)
            else:
                rows.append((w["top"], [w]))
        for _, ws in rows:
            ws.sort(key=lambda z: z["x0"])
        return rows

    @classmethod
    def _clean(cls, ws):
        """[(x0, x1, text)] with fake-bold doubling undone row-wide."""
        texts = _row_undouble([w["text"] for w in ws])
        return [(w["x0"], w["x1"], t) for w, t in zip(ws, texts)]

    @staticmethod
    def _money(s):
        s = (s or "").strip()
        if not s:
            return None
        neg = s.startswith("(") and s.endswith(")")
        s2 = re.sub(r"[^0-9.\-]", "", s)
        if not s2 or s2 in ("-", "."):
            return None
        v = float(s2)
        return -v if neg else v

    @staticmethod
    def _date(s):
        try:
            return datetime.strptime(s, "%m/%d/%Y").date()
        except (ValueError, TypeError):
            return None

    @classmethod
    def _detail_cells(cls, toks):
        """Bucket one detail row into named cells. Numerics are POSITIONAL:
        Rent by its left edge, everything else by its printed right edge."""
        c = {k: None for k, _ in cls.NUM_EDGES}
        c["rent"] = None
        text = {"lease_start": [], "lease_end": [], "bedbath": [],
                "cycle": [], "rent_start": []}
        unplaced = []
        for x0, x1, t in toks:
            centre = (x0 + x1) / 2.0
            placed = False
            for key, (lo, hi) in (("lease_start", cls.W_LEASE_START),
                                  ("lease_end", cls.W_LEASE_END),
                                  ("bedbath", cls.W_BEDBATH),
                                  ("cycle", cls.W_CYCLE),
                                  ("rent_start", cls.W_RENT_START)):
                if lo <= centre < hi:
                    text[key].append(t)
                    placed = True
                    break
            if placed:
                continue
            if cls.W_RENT_X0[0] <= x0 < cls.W_RENT_X0[1]:
                c["rent"] = t                      # may be the "--" marker
                continue
            hit = min(cls.NUM_EDGES, key=lambda e: abs(e[1] - x1))
            if abs(hit[1] - x1) <= cls.EDGE_TOL:
                c[hit[0]] = t
            else:
                unplaced.append(t)
        for k, v in text.items():
            c[k] = " ".join(v).strip()
        c["_unplaced"] = unplaced
        return c

    @classmethod
    def _snap_row(cls, toks, edges, tol=14.0):
        """Summary tables: numeric tokens snapped to a header row's right
        edges. Returns a list the same length as `edges` (None where the
        report printed nothing)."""
        out = [None] * len(edges)
        for x0, x1, t in toks:
            if not cls.MONEY.match(t):
                continue
            j = min(range(len(edges)), key=lambda k: abs(edges[k] - x1))
            if abs(edges[j] - x1) <= tol:
                out[j] = t
        return out

    # -- second, independent extraction pass ---------------------------------

    @classmethod
    def _reextract(cls, pdf):
        """Independent recount through `page.extract_text()` line splitting
        (the positional word parser above never touches this path): number of
        detail rows, occupied rows and the Rent column sum. Filed as
        "re-extract" checks alongside the report's own printed totals."""
        n = occ = 0
        rent = 0.0
        for page in pdf.pages:
            for raw in (page.extract_text() or "").split("\n"):
                m = re.match(
                    r"^(?:(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}/\d{1,2}/\d{4})"
                    r"\s+)?(\d+(?:\.\d+)?)\s*Bed/(\d+(?:\.\d+)?)\s*Bath\s+"
                    r"(?:Monthly|Weekly|Yearly|Quarterly)?\s*"
                    r"(?:\d{1,2}/\d{1,2}/\d{4})?\s*"
                    r"(--|[\d,]+\.\d{2})\s", raw)
                if not m:
                    continue
                n += 1
                if m.group(5) != "--":
                    occ += 1
                    rent += float(m.group(5).replace(",", ""))
        return n, occ, rent

    # -- main parse ----------------------------------------------------------

    def parse(self, pdf):
        self.flags = []
        self.asof_found = False
        prop, asof = "", None
        units, checks = [], {}
        grand = {}
        plan_summary = {}          # "1/1" -> dict of printed columns
        prop_summary = None
        totals_row = None
        prepayments = []
        seq = 0
        mode = None                # None | "bb" | "prop"
        bb_edges = prop_edges = None

        for pi, page in enumerate(pdf.pages):
            for top, ws in self._rows(page):
                toks = self._clean(ws)
                text = " ".join(t for _, _, t in toks).strip()
                low = text.lower()
                if not text:
                    continue

                # ---- report header -------------------------------------
                m = re.search(r"as of\s+(\d{1,2}/\d{1,2}/\d{4})\s*,\s*(.+?)"
                              r"\s*,\s*(current|all)\b", text, re.I)
                if m:
                    asof = self._date(m.group(1))
                    self.asof_found = asof is not None
                    prop = prop or m.group(2).strip()
                    continue
                if re.match(r"^generated\s+\d", low) or low == "rent roll" \
                        or low.startswith("prepared by") \
                        or re.match(r"^(po box|austin,|amount$)", low):
                    continue

                # ---- summary blocks (last page) ------------------------
                if low.startswith("grand totals"):
                    mode = "grand"
                    continue
                if low.startswith("summary by bed/bath"):
                    mode = "bb"
                    continue
                if low.startswith("summary by property"):
                    mode = "prop"
                    continue
                if mode in ("bb", "prop") and low.startswith(
                        ("bed/bath ", "property ")):
                    # Column header of a summary table. Its nine data columns
                    # are, left to right: No. of Units | Vacant | Occupied |
                    # % Occupied | SF Total | SF Average | Market Rent Total |
                    # Market Rent Average | Avg./Sq.Ft. The right edge of the
                    # LAST word of each caption is the column's right edge
                    # (every value in the block is right-aligned).
                    edges = sorted(x1 for _, x1, t in toks
                                   if t in ("Units", "Vacant", "Occupied",
                                            "Total", "Average",
                                            "Avg./Sq.Ft."))
                    if len(edges) != 9:
                        raise ValueError(
                            f"summary header has {len(edges)} column edges, "
                            "expected 9 - layout changed")
                    if mode == "bb":
                        bb_edges = edges
                    else:
                        prop_edges = edges
                    continue

                if mode == "grand":
                    m = re.match(r"^(market rent|rent|recurring charges|"
                                 r"recurring credits|deposits held|"
                                 r"balance due)\s+\$?([\d,.\-()]+)$", text,
                                 re.I)
                    if m:
                        grand[m.group(1).lower()] = self._money(m.group(2))
                        continue

                if mode in ("bb", "prop") and re.match(
                        r"^totals and averages", low):
                    edges = bb_edges if mode == "bb" else prop_edges
                    row = self._summary_row(self._snap_row(toks, edges))
                    if mode == "bb":
                        plan_summary["_total"] = row
                    else:
                        prop_summary = row
                    continue
                if mode == "bb" and bb_edges:
                    label = " ".join(t for _, _, t in toks[:3])
                    fp = self._fp(label)
                    if fp:
                        plan_summary[fp] = self._summary_row(
                            self._snap_row(toks, bb_edges))
                    continue
                if mode == "prop" and prop_edges:
                    # the property detail row ("4639 Williams Road 40 5 ...")
                    prop_summary = self._summary_row(
                        self._snap_row(toks, prop_edges))
                    continue

                # ---- property group header / totals strip ---------------
                if re.match(r"^total for\s+", low):
                    totals_row = self._detail_cells(toks)
                    continue
                if mode is None and not self.BEDBATH.search(text) and \
                        not any(self.MONEY.match(t) for _, _, t in toks):
                    continue                       # group header line

                # ---- detail row ----------------------------------------
                c = self._detail_cells(toks)
                bb = re.sub(r"\s+", " ", c["bedbath"] or "").strip()
                if not self.BEDBATH.match(bb):
                    continue
                if c["_unplaced"]:
                    raise ValueError(
                        f"unplaced token(s) {c['_unplaced']} on a detail row "
                        "- column windows need review, refusing to guess")
                seq += 1
                unit = f"Unit {seq:02d}"
                rent = None if (c["rent"] or "").strip() in ("--", "", "-") \
                    else self._money(c["rent"])
                charges = self._money(c["charges"]) or 0.0
                credits = self._money(c["credits"]) or 0.0
                dep = self._money(c["deposits"])
                prepay = self._money(c["prepayments"]) or 0.0
                bal = self._money(c["balance"])
                tot = self._money(c["total"])
                if tot is not None and abs((charges - credits) - tot) > 0.005:
                    raise ValueError(
                        f"{unit}: printed Total {tot:,.2f} != Recurring "
                        f"Charges {charges:,.2f} - Credits {credits:,.2f}")
                vacant = rent is None

                u = UnitRecord(unit=unit, floor_plan=self._fp(bb),
                               apt_status="VU" if vacant else "OC",
                               market_rent=None)
                bed, bath = self.BEDBATH.match(bb).groups()
                u.bed_explicit = float(bed)
                u.bath_explicit = float(bath)
                u.bed_bath_explicit = True
                if not vacant:
                    ch = [("RENT", rent, False)]
                    other = round(charges - rent, 2)
                    if abs(other) > 0.004:
                        ch.append(("RECURRING CHARGES - OTHER", other, False))
                    if abs(credits) > 0.004:
                        ch.append(("RECURRING CREDIT", -credits, False))
                        self.flags.append(
                            f"{unit}: recurring credit {credits:,.2f} - the "
                            "report gives no credit code; classified by the "
                            "unknown-negative rule")
                    r = Resident(name="", status="C", charges=ch,
                                 lease_start=self._date(c["lease_start"]),
                                 lease_expires=self._date(c["lease_end"]),
                                 deposit=dep, ending_balance=bal)
                    r.prepayment = prepay
                    u.residents.append(r)
                else:
                    if charges or credits:
                        self.flags.append(
                            f"{unit}: vacant but carries charges "
                            f"{charges:,.2f}/credits {credits:,.2f}")
                    r = Resident(name="", status="X", charges=[],
                                 deposit=dep, ending_balance=bal)
                    r.prepayment = prepay
                    if dep or bal or prepay:
                        u.residents.append(r)
                prepayments.append(prepay)
                units.append(u)

        # ---- square feet, from the report's own Summary by bed/bath -----
        for u in units:
            row = plan_summary.get(u.floor_plan)
            if row and row.get("sf_avg") is not None:
                u.sqft = row["sf_avg"]
        if plan_summary:
            self.flags.append(
                "the detail rows carry NO square footage; per-unit Net Sf is "
                "allocated from the report's own 'Summary by bed/bath' block "
                "(" + ", ".join(
                    f"{fp} = {r['sf_total']:,.0f} sf / {r['units']:,.0f} "
                    f"units = {r['sf_avg']:,.0f} sf"
                    for fp, r in sorted(plan_summary.items())
                    if fp != "_total" and r.get("sf_avg") is not None)
                + "). Each plan's total divides EXACTLY by its unit count, "
                "so the plan is uniform; the parsed total ties to the "
                "report. This is rent-roll data, not a web estimate.")

        self.flags.append(
            "this Buildium report prints NO unit identifiers - the detail "
            f"rows are anonymous. Units are numbered in report order "
            f"(Unit 01..Unit {len(units):02d}); these are NOT the "
            "property's real unit numbers.")
        if grand.get("market rent"):
            n_mr = sum(1 for fp, r in plan_summary.items()
                       if fp != "_total" and r.get("mr_total")
                       and r.get("mr_avg"))
            detail = "; ".join(
                f"{fp}: {r['mr_total']:,.2f} total / {r['mr_avg']:,.2f} avg "
                f"= {round(r['mr_total'] / r['mr_avg']):.0f} of "
                f"{r['units']:,.0f} units"
                for fp, r in sorted(plan_summary.items())
                if fp != "_total" and r.get("mr_avg"))
            self.flags.append(
                "MARKET RENT: the detail has no Market Rent column. The "
                f"report's grand totals print Market rent "
                f"${grand['market rent']:,.2f}, but that is an aggregate "
                f"over only the units that have one set in Buildium "
                f"({detail}) - it cannot be allocated per unit, so it is "
                "NOT used as a tie-out target. Column I is filled by "
                "--estimate-market (house rule 8/2026) and red-flagged.")

        checks.update(self._checks(grand, plan_summary, prop_summary,
                                   totals_row, prepayments, pdf))
        return prop, asof, units, checks

    # -- summary helpers -----------------------------------------------------

    @staticmethod
    def _fp(bedbath):
        m = BuildiumRentRollParser.BEDBATH.match(
            re.sub(r"\s+", " ", (bedbath or "").strip()))
        if not m:
            return ""

        def trim(x):
            f = float(x)
            return str(int(f)) if f == int(f) else str(f)
        return f"{trim(m.group(1))}/{trim(m.group(2))}"

    @classmethod
    def _summary_row(cls, vals):
        keys = ["units", "vacant", "occupied", "pct_occ", "sf_total",
                "sf_avg", "mr_total", "mr_avg", "mr_psf"]
        out = {}
        for k, v in zip(keys, list(vals) + [None] * len(keys)):
            out[k] = cls._money(v) if v is not None else None
        return out

    def _checks(self, grand, plan_summary, prop_summary, totals_row,
                prepayments, pdf):
        checks, src, extra = {}, {}, []
        ps = prop_summary or plan_summary.get("_total") or {}
        if ps.get("units") is not None:
            checks["unit_count"] = ps["units"]
        if ps.get("occupied") is not None:
            checks["occupied_count"] = ps["occupied"]
        if ps.get("vacant") is not None:
            checks["vacant_count"] = ps["vacant"]
        if ps.get("pct_occ") is not None:
            # NOT checks["pct_unit_occupancy"]: reconcile()'s generic check
            # TRUNCATES (Yardi/OneSite behaviour). Buildium rounds half-up.
            extra.append((
                "% unit occupancy (report rounds half-up)",
                lambda us: math.floor(
                    sum(1 for u in us if not u.is_vacant)
                    / max(1, len(us)) * 10000 + 0.5) / 100.0,
                ps["pct_occ"], 0.005))
        if ps.get("sf_total") is not None:
            checks["total_sqft"] = ps["sf_total"]
        if grand.get("rent") is not None:
            checks["total_contract_rent"] = grand["rent"]
        if grand.get("recurring charges") is not None:
            checks["current_lease_charges"] = (
                grand["recurring charges"]
                - (grand.get("recurring credits") or 0.0))
        if grand.get("deposits held") is not None:
            checks["total_deposits"] = grand["deposits held"]

        def _cn(us):
            return [r for u in us for r in u.residents
                    if r.status.upper() in ("C", "N")]

        def _all(us):
            return [r for u in us for r in u.residents]

        if grand.get("recurring charges") is not None:
            extra.append(("Grand total: recurring charges",
                          lambda us: sum(a for r in _cn(us)
                                         for _, a, _ in r.charges if a > 0),
                          grand["recurring charges"], 0.01))
        if grand.get("recurring credits") is not None:
            extra.append(("Grand total: recurring credits",
                          lambda us: -sum(a for r in _cn(us)
                                          for _, a, _ in r.charges if a < 0),
                          grand["recurring credits"], 0.01))
        if grand.get("balance due") is not None:
            extra.append(("Grand total: balance due",
                          lambda us: sum(r.ending_balance or 0
                                         for r in _all(us)),
                          grand["balance due"], 0.01))
        # the property "Total for <name>" strip
        if totals_row:
            tr = {k: self._money(totals_row.get(k))
                  for k in ("rent", "charges", "credits", "total",
                            "deposits", "prepayments", "balance")}
            pairs = (
                ("rent", lambda us: sum(r.rent_charge or 0 for r in _cn(us))),
                ("charges", lambda us: sum(a for r in _cn(us)
                                           for _, a, _ in r.charges if a > 0)),
                ("credits", lambda us: -sum(a for r in _cn(us)
                                            for _, a, _ in r.charges
                                            if a < 0)),
                ("total", lambda us: sum(a for r in _cn(us)
                                         for _, a, _ in r.charges)),
                ("deposits", lambda us: sum(r.deposit or 0
                                            for r in _all(us))),
                ("prepayments", lambda us: sum(getattr(r, "prepayment", 0)
                                               for r in _all(us))),
                ("balance", lambda us: sum(r.ending_balance or 0
                                           for r in _all(us))),
            )
            for key, fn in pairs:
                if tr.get(key) is not None:
                    extra.append((f"Total-for-property row: {key}", fn,
                                  tr[key], 0.01))
        # per-plan blocks
        for fp, row in sorted(plan_summary.items()):
            if fp == "_total":
                continue
            if row.get("units") is not None:
                extra.append((f"  plan {fp}: units",
                              (lambda f: lambda us: sum(
                                  1 for u in us if u.floor_plan == f))(fp),
                              row["units"], 0.5))
            if row.get("vacant") is not None:
                extra.append((f"  plan {fp}: vacant",
                              (lambda f: lambda us: sum(
                                  1 for u in us
                                  if u.floor_plan == f and u.is_vacant))(fp),
                              row["vacant"], 0.5))
            if row.get("occupied") is not None:
                extra.append((f"  plan {fp}: occupied",
                              (lambda f: lambda us: sum(
                                  1 for u in us
                                  if u.floor_plan == f
                                  and not u.is_vacant))(fp),
                              row["occupied"], 0.5))
            if row.get("sf_total") is not None:
                extra.append((f"  plan {fp}: sq ft",
                              (lambda f: lambda us: sum(
                                  u.sqft or 0 for u in us
                                  if u.floor_plan == f))(fp),
                              row["sf_total"], 0.5))
            if row.get("pct_occ") is not None:
                # Buildium ROUNDS its occupancy percentages half-up
                # (19/24 = 79.1666 -> 79.17); Yardi/OneSite truncate. Do not
                # reuse reconcile()'s trunc2 here.
                extra.append((
                    f"  plan {fp}: % occupied",
                    (lambda f: lambda us: (
                        math.floor(sum(1 for u in us if u.floor_plan == f
                                       and not u.is_vacant)
                                   / max(1, sum(1 for u in us
                                                if u.floor_plan == f))
                                   * 10000 + 0.5) / 100.0))(fp),
                    row["pct_occ"], 0.005))
        # second, independent extraction pass
        rx_n, rx_occ, rx_rent = self._reextract(pdf)
        extra.append(("Unit count (2nd pass, extract_text)",
                      lambda us: len(us), rx_n, 0.5))
        extra.append(("Occupied units (2nd pass, extract_text)",
                      lambda us: sum(1 for u in us if not u.is_vacant),
                      rx_occ, 0.5))
        extra.append(("Total contract rent (2nd pass, extract_text)",
                      lambda us: sum(r.rent_charge or 0 for r in _cn(us)),
                      rx_rent, 0.01))
        checks["extra_checks"] = extra
        checks["_src"] = src
        return checks

    def source_note(self, asof):
        d = asof.strftime("%m/%d/%Y") if asof else "unknown date"
        return (
            f"Generated from a Buildium-style rent roll PDF (as of {d}). "
            "Contractual Rent = the report's Rent column; Other Income = "
            "Recurring Charges less Rent (the report prints no per-charge "
            "detail). Net Sf is allocated from the report's own Summary by "
            "bed/bath block. The report prints no unit numbers - Unit NN is "
            "report order. Market Rent is not in the detail; where filled it "
            "is an estimate (highlighted).")


# Detection order matters. OneSite first (its detect() keys on the report's
# own product banner + title). ResManSummary before ResMan: the full-roll
# ResMan detect() ('ResMan' + 'Rent Roll') also matches a Rent Roll Summary,
# so the summary variant must claim its PDFs first. ResMan next: its
# detect() is the strictest of the remaining three.
# SSI410 next (it matches any PDF titled "Rent Roll Report").
# OwnerSheetPdfParser last: its header sniff (unit/tenant/rent/type) is the
# loosest of the three, so it only claims a PDF nothing else recognises.
PARSERS = [BuildiumRentRollParser, OneSiteRentsParser, ResManSummaryParser,
           ResManParser, SSI410Parser, OwnerSheetPdfParser]
XLSX_PARSERS = [AppFolioXlsxParser, YardiRentRollXlsxParser]



# ----------------------------------------------------------------------------
# Excel writer (rediQ-style Rent Roll sheet)
# ----------------------------------------------------------------------------

HEADERS = [
    "Unit No.", "Floor Plan", "Net Sf", "Bed", "Bath", "Lease Type",
    "Renovation Status", "Occupancy Status", "Market Rent",
    "Contractual Rent", "Recurring Concessions", "Net Effective Rent",
    "Supplemental Rent", "Upfront Concessions", "Emp./Other Discounts",
    "Other Income", "Lease Start Date", "Lease Expiration",
    "Lease Term (months)", "MTM", "Move In Date", "Move Out Date",
    "Vac. Notice",
]

COL_WIDTHS = [14.47, 16.70, 10.0, 7.47, 8.0, 15.35, 19.47, 20.18, 15.06,
              19.64, 25.23, 20.23, 21.59, 22.94, 24.12, 16.88, 18.82, 19.06,
              23.29, 8.23, 16.41, 18.06, 15.06]

FMT_TEXT = '\\ @_)'
FMT_SF = '#,##0\\ "sf"_);\\(#,##0\\ "sf"\\)_);_("-"_);\\ @_)'
FMT_NUM = '#,##0_);\\(#,##0\\);_("-"_);\\ @_)'
FMT_CUR = '"$"\\ #,##0_);\\("$"\\ #,##0\\);_("-"_);\\ @_)'
FMT_DATE = '[$-409]mmm\\ d\\,\\ yyyy_);[Red]"must be date";""'
FMT_BED = '#,##0\\ "BR"_);\\(#,##0\\);"Studio"_);\\ @_)'
FMT_BATH = '#,##0.0_);\\(#,##0.0\\);_("-"_);\\ @_)'
FMT_MONTHS = '#,##0\\ "months"_);\\(#,##0\\ "months"\\)_);_("-"_);\\ @_)'

# column-index (1-based) -> number format for data rows
COL_FMTS = {1: FMT_TEXT, 2: FMT_TEXT, 3: FMT_SF, 4: FMT_BED, 5: FMT_BATH,
            6: FMT_TEXT, 7: FMT_TEXT, 8: 'General', 9: FMT_CUR, 10: FMT_CUR,
            11: FMT_CUR, 12: FMT_CUR, 13: FMT_CUR, 14: FMT_CUR, 15: FMT_CUR,
            16: FMT_CUR, 17: FMT_DATE, 18: FMT_DATE, 19: FMT_MONTHS,
            20: FMT_TEXT, 21: FMT_DATE, 22: FMT_DATE, 23: FMT_TEXT}

HEADER_FILL = "1F3864"     # dark navy (theme accent, 50% darker)
TITLE_COLOR = "44546A"     # dark slate blue
ROW_FILL = "F2F2F2"        # light gray band on data rows


def unit_sort_key(u):
    parts = re.findall(r"\d+|\D+", u.unit)
    # mixed schemes ("Unit 01" vs "111 Jasmin") must stay comparable:
    # tag each token so ints sort before strings instead of raising.
    return [(0, int(p), "") if p.isdigit() else (1, 0, p.strip().lower())
            for p in parts if p.strip(" -")]


def find_default_template():
    """rentroll_template.xlsx next to this script, if present."""
    import os
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "rentroll_template.xlsx")
    return p if os.path.exists(p) else None


# ----------------------------------------------------------------------------
# Best-estimate cells (--sqft-est / --bedbath-est)
# ----------------------------------------------------------------------------
# House rule (7/2026): when SF or bed/bath are missing and no source can be
# found, fill the best estimate rather than shipping blanks - but mark it.
# Estimated cells get Excel's standard "bad" pair (light red fill FFC7CE,
# dark red text 9C0006) and one red note line sits under the data, so nobody
# can mistake an estimate for a number the rent roll actually carried.

EST_FILL = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
EST_FONT_COLOR = "9C0006"
EST_NOTE = ("highlighted SF, bed, bath counts are best estimates, not "
            "provided by ownership")
# UnitRecord.estimated key -> Rent Roll column index
EST_COLS = {"fp": 2, "sqft": 3, "bed": 4, "bath": 5, "market": 9}


def _est_note_text(units):
    """The red note under the data names exactly the fields estimated in
    THIS run - one combined line. 'counts' is kept when only SF/bed/bath are
    involved (the wording Dmytro signed off on); a market-rent estimate makes
    it 'values'."""
    keys = set()
    for u in units:
        keys |= set(getattr(u, "estimated", ()) or ())
    if not keys:
        return EST_NOTE
    names = []
    for k, label in (("sqft", "SF"), ("bed", "bed"), ("bath", "bath"),
                     ("fp", "floor plan"), ("market", "market rent")):
        if k in keys:
            names.append(label)
    noun = "counts" if keys <= {"sqft", "bed", "bath", "fp"} else "values"
    return (f"highlighted {', '.join(names)} {noun} are best estimates, not "
            "provided by ownership")


def _add_months(d, n):
    y, m = d.year, d.month + n
    while m < 1:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    day = min(d.day, [31, 29 if y % 4 == 0 and (y % 100 or y % 400 == 0)
                      else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1])
    return d.replace(year=y, month=m, day=day)


def estimate_market_rents(units, asof, window_months=6, min_count=3):
    """Fill a MISSING Market Rent column from the maximum stated contractual
    rents (house rule, Dmytro 8/2026). Per floor plan:

    1. Data points are that plan's OCCUPIED units with a real contractual
       rent. A $0 rent (employee/model unit) is not a market signal.
    2. Recency filter, in order of preference: lease START dates within the
       most recent `window_months` of the as-of date; failing that lease END
       dates in that window; failing that, all occupied units of the plan.
    3. The estimate is the HIGHEST contractual rent occurring at least
       `min_count` times in that set - the highest rent the property can
       actually repeat, not the single best outlier.
    4. Graceful fallback, never silently empty: if nothing reaches
       `min_count`, take the value with the most occurrences (ties -> the
       highest) and FLAG it. A plan whose occupied units carry no rent at
       all is reported and left blank.

    A market rent the source DID provide is never overridden. Returns
    (per-plan derivation lines, flags)."""
    notes, flags = [], []
    if not units:
        return notes, flags
    cutoff = _add_months(asof, -window_months) if asof else None
    plans = []
    for u in units:
        if u.floor_plan not in plans:
            plans.append(u.floor_plan)
    for fp in plans:
        mine = [u for u in units if u.floor_plan == fp]
        if all(u.market_rent is not None for u in mine):
            notes.append(f"  plan {fp or '(none)'}: market rent already "
                         f"provided by the source - left untouched")
            continue
        pts = []
        for u in mine:
            if u.is_vacant:
                continue
            p = u.primary
            rent = p.rent_charge if p else None
            if rent is None or rent <= 0:
                continue
            pts.append((u, p, rent))
        if not pts:
            flags.append(f"market-rent estimate: plan {fp or '(none)'} has "
                         "no occupied unit carrying a rent - column I left "
                         "blank for that plan")
            notes.append(f"  plan {fp or '(none)'}: no rent data - blank")
            continue
        basis, sel = "all occupied units", pts
        if cutoff is not None:
            recent = [t for t in pts
                      if t[1].lease_start and t[1].lease_start >= cutoff]
            if recent:
                basis = (f"lease starts on/after {cutoff:%m/%d/%Y} "
                         f"({window_months}-month window)")
                sel = recent
            else:
                recent = [t for t in pts
                          if t[1].lease_expires
                          and t[1].lease_expires >= cutoff]
                if recent:
                    basis = (f"lease ENDS on/after {cutoff:%m/%d/%Y} "
                             f"(no lease starts in the {window_months}-month "
                             f"window)")
                    sel = recent
        counts = {}
        for _, _, r in sel:
            counts[round(float(r), 2)] = counts.get(round(float(r), 2), 0) + 1
        qualified = [v for v, c in counts.items() if c >= min_count]
        if qualified:
            est = max(qualified)
            thin = False
        else:
            top = max(counts.values())
            est = max(v for v, c in counts.items() if c == top)
            thin = True
            flags.append(
                f"market-rent estimate thin - plan {fp or '(none)'}: no rent "
                f"repeats {min_count}+ times in the basis set "
                f"({len(sel)} data point(s)); took the most-repeated value "
                f"(highest on a tie) = {est:,.2f}")
        n_filled = 0
        for u in mine:
            if u.market_rent is None:
                u.market_rent = est
                u.estimated.add("market")
                n_filled += 1
        notes.append(
            f"  plan {fp or '(none)'}: basis = {basis}; "
            f"{len(sel)} data point(s) "
            + "{" + ", ".join(f"{v:,.0f}x{c}"
                              for v, c in sorted(counts.items(),
                                                 reverse=True)) + "}"
            + f" -> ${est:,.2f}" + (" [THIN]" if thin else "")
            + f"; filled {n_filled} unit(s)")
    return notes, flags


def _mark_estimates(ws, row, u):
    """Apply the red estimate treatment to unit `u`'s estimated cells on the
    Rent Roll sheet row `row`. Keeps the template's font family/size and
    number format; only colour and fill change."""
    for key in getattr(u, "estimated", ()):
        col = EST_COLS.get(key)
        if not col:
            continue
        c = ws.cell(row=row, column=col)
        f = c.font
        c.font = Font(name=f.name, size=f.size, bold=f.bold,
                      italic=f.italic, color=EST_FONT_COLOR)
        c.fill = EST_FILL


def _est_note_font():
    return Font(name="Calibri", size=8, bold=True, color=EST_FONT_COLOR)


def _fmt_bb(x):
    """1.0 -> '1', 1.5 -> '1.5' (floor-plan strings never show a bare .0)."""
    if x is None:
        return ""
    return str(int(x)) if float(x) == int(x) else str(float(x))


def _fill_sqft(units, spec, estimate=False):
    """Apply a '<key>=<sqft>' map. Keys that are not floor plans address
    individual units by name (exact, then case-insensitive substring) and win
    over the floor-plan value. Only ever fills a MISSING value."""
    fpmap = {}
    for pair in spec.split(","):
        fp, v = pair.split("=")
        fpmap[fp.strip()] = float(v)
    filled = unmatched = 0
    fps = {u.floor_plan for u in units}
    unitmap = {k: v for k, v in fpmap.items() if k not in fps}

    def _lookup(u):
        if u.unit in unitmap:                       # exact unit name
            return unitmap[u.unit]
        for k, v in unitmap.items():                # substring on name
            if k.lower() in u.unit.lower():
                return v
        return fpmap.get(u.floor_plan)

    for u in units:
        if u.sqft is not None:
            continue
        v = _lookup(u)
        if v is not None:
            u.sqft = v
            if estimate:
                u.estimated.add("sqft")
            filled += 1
        else:
            unmatched += 1
    flag = "--sqft-est" if estimate else "--sqft"
    print(f"Net Sf filled from {flag} map: {filled} units"
          + (" (marked as estimates)" if estimate else "")
          + (f"; {unmatched} unit(s) left blank (no map entry)"
             if unmatched and not estimate else ""))


def _fill_bedbath(units, spec, estimate=False):
    """Apply a '<key>=<bed>/<bath>' map. Same key resolution as _fill_sqft.

    In estimate mode the two components are handled separately: a component
    is marked red only if it was blank (or disagrees with what the source
    gave), so a parser-derived bed count stays un-highlighted while the bath
    count it could not provide is flagged. A unit with no floor plan at all
    gets one synthesised from the estimate ('1/1') so it joins the Floor Plan
    rollups instead of silently dropping out - that cell is marked too."""
    bbmap = {}
    for pair in spec.split(","):
        fp, v = pair.split("=")
        bed, _, bath = v.strip().partition("/")
        bbmap[fp.strip()] = (float(bed) if bed.strip() else None,
                             float(bath) if bath.strip() else None)
    fps = {u.floor_plan for u in units}
    unitmap = {k: v for k, v in bbmap.items() if k not in fps}

    def _lookup(u):
        if u.unit in unitmap:
            return unitmap[u.unit]
        for k, v in unitmap.items():
            if k.lower() in u.unit.lower():
                return v
        return bbmap.get(u.floor_plan)

    filled = 0
    for u in units:
        hit = _lookup(u)
        if hit is None:
            continue
        new_bed, new_bath = hit
        if not estimate:
            u.bed_explicit, u.bath_explicit = new_bed, new_bath
            u.bed_bath_explicit = True
            filled += 1
            continue
        cur_bed, cur_bath = u.bed_bath
        for key, cur, newv in (("bed", cur_bed, new_bed),
                               ("bath", cur_bath, new_bath)):
            if newv is None:
                continue
            if cur is None or float(cur) != float(newv):
                u.estimated.add(key)
        u.bed_explicit = new_bed if new_bed is not None else cur_bed
        u.bath_explicit = new_bath if new_bath is not None else cur_bath
        u.bed_bath_explicit = True
        if not u.floor_plan.strip() and (new_bed is not None
                                         or new_bath is not None):
            u.floor_plan = f"{_fmt_bb(new_bed)}/{_fmt_bb(new_bath)}"
            u.estimated.add("fp")
        filled += 1
    flag = "--bedbath-est" if estimate else "--bedbath"
    missing = sorted({u.floor_plan for u in units
                      if _lookup(u) is None and u.floor_plan})
    print(f"Bed/Bath filled from {flag} map: {filled} units"
          + (" (marked as estimates)" if estimate else "")
          + (f"; no entry for floor plan(s): {', '.join(missing)}"
             if missing and not estimate else ""))


# ----------------------------------------------------------------------------
# Output normalization (see CLAUDE.md gotcha #1)
# ----------------------------------------------------------------------------
# openpyxl writes strings inline and emits no xl/sharedStrings.xml; DOM-based
# loaders (Power BI web, JS xlsx readers, some import tools) fetch that part
# unconditionally and crash with "getElementsByTagName of null". The rent-roll
# template additionally carries a table part (Table1) whose
# calculatedColumnFormula entries point at sheets that do not exist in the
# deliverable ('Rents, Other Income', 'Bed-Bath Type', 'Data') - Excel repairs
# the file over those. Both are fixed here, exactly as process_t12.py does for
# the T-12 deliverables.


def _purge_broken_names(wb):
    """Remove defined names that reference deleted sheets or #REF!."""
    sheets = set(wb.sheetnames)

    def is_bad(name, ref):
        if not ref or "#REF!" in ref or name.startswith("ExternalData"):
            return True
        for q, u in re.findall(r"'([^']+)'!|([A-Za-z0-9_. ]+)!", ref):
            if (q or u) not in sheets:
                return True
        return False

    for holder in [wb] + list(wb.worksheets):
        dn = getattr(holder, "defined_names", None)
        if dn is None:
            continue
        for n in list(dn):
            try:
                ref = dn[n].value or ""
            except Exception:
                ref = ""
            if is_bad(n, ref):
                del dn[n]


def _normalize_xlsx(path, sheet_names=()):
    """Rewrite an openpyxl-saved package the way Excel writes it: move inline
    strings into xl/sharedStrings.xml, untype empty cells, and clean the
    template's table part (query-table markers + calculated-column formulas
    that reference sheets which are not in this workbook)."""
    import shutil
    import zipfile

    SST_T = ("application/vnd.openxmlformats-officedocument.spreadsheetml"
             ".sharedStrings+xml")
    SST_R = ("http://schemas.openxmlformats.org/officeDocument/2006/"
             "relationships/sharedStrings")
    known = set(sheet_names)
    strings, index = [], {}
    total_refs = 0

    def fix_sheet(xml):
        nonlocal total_refs

        def repl(m):
            nonlocal total_refs
            head, tail, inner = m.group(1), m.group(2), m.group(3) or ""
            t = re.search(r"<t[^>]*>.*?</t>", inner, re.S)
            t_elem = t.group(0) if t else "<t/>"
            if t_elem not in index:
                index[t_elem] = len(strings)
                strings.append(t_elem)
            total_refs += 1
            return f'<c{head} t="s"{tail}><v>{index[t_elem]}</v></c>'

        xml = re.sub(
            r'<c([^>]*?) t="inlineStr"([^>]*)>(?:<is>(.*?)</is>)?</c>',
            repl, xml, flags=re.S)
        # empty cells typed as numbers -> untyped empty cells
        xml = re.sub(r'<c([^>]*?) t="n"([^>]*)></c>', r"<c\1\2/>", xml)
        xml = re.sub(r'<c([^>]*?) t="n"([^>]*)/>', r"<c\1\2/>", xml)
        return xml

    def dangling(formula):
        """True when the formula names a sheet this workbook does not have."""
        for q, u in re.findall(r"'([^']+)'!|([A-Za-z0-9_. ]+)!", formula):
            name = q or u
            if name and name not in known:
                return True
        return False

    def fix_table(xml):
        xml = xml.replace(' tableType="queryTable"', "")
        xml = re.sub(r' queryTableFieldId="\d+"', "", xml)
        xml = re.sub(
            r"<calculatedColumnFormula[^>]*>(.*?)</calculatedColumnFormula>",
            lambda m: "" if dangling(m.group(1)) else m.group(0),
            xml, flags=re.S)
        xml = re.sub(
            r"<calculatedColumnFormula[^>]*/>", "", xml)
        m = re.search(r'<table[^>]* ref="([^"]+)"', xml)
        if m:      # keep the autoFilter range in sync with the table range
            xml = re.sub(r'<autoFilter ref="[^"]+"/>',
                         f'<autoFilter ref="{m.group(1)}"/>', xml)
        return xml

    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        parts = {n: zin.read(n) for n in zin.namelist()}
        for name in list(parts):
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                parts[name] = fix_sheet(
                    parts[name].decode("utf-8")).encode("utf-8")
            elif re.match(r"xl/tables/table\d+\.xml$", name):
                parts[name] = fix_table(
                    parts[name].decode("utf-8")).encode("utf-8")
            elif name == "xl/workbook.xml":
                w = parts[name].decode("utf-8")
                w = re.sub(r"<connections[^/>]*/>", "", w)
                parts[name] = w.encode("utf-8")
        if strings:
            sst = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<sst xmlns="http://schemas.openxmlformats.org/'
                   'spreadsheetml/2006/main" '
                   f'count="{total_refs}" uniqueCount="{len(strings)}">'
                   + "".join(f"<si>{t}</si>" for t in strings) + "</sst>")
            parts["xl/sharedStrings.xml"] = sst.encode("utf-8")
            ct = parts["[Content_Types].xml"].decode("utf-8")
            if "sharedStrings" not in ct:
                ct = ct.replace("</Types>",
                                f'<Override PartName="/xl/sharedStrings.xml"'
                                f' ContentType="{SST_T}"/></Types>')
                parts["[Content_Types].xml"] = ct.encode("utf-8")
            rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
            if "sharedStrings" not in rels:
                rels = rels.replace(
                    "</Relationships>",
                    f'<Relationship Id="rIdSST" Type="{SST_R}" '
                    f'Target="sharedStrings.xml"/></Relationships>')
                parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")
        for name, data in parts.items():
            zout.writestr(name, data)
    shutil.move(tmp, path)


def _save_normalized(wb, path):
    """Single save path for every rent-roll deliverable."""
    _purge_broken_names(wb)
    names = list(wb.sheetnames)
    wb.save(path)
    _normalize_xlsx(path, names)


def write_workbook_from_template(path, template, prop, asof, units,
                                 source_note=""):
    """Fill the pre-styled template (exact 'Rent Roll Processor' formatting):
    write Rent Roll values, floor-plan identity cells, and named ranges.
    All styling and the Floor Plan / Floor Plan Summary formulas already
    live in the template."""
    from openpyxl import load_workbook
    wb = load_workbook(template)
    ws = wb["Rent Roll"]

    ws["A1"] = prop or "Property"
    ws["B2"] = datetime(asof.year, asof.month, asof.day) if asof else None

    def dt(d):
        return datetime(d.year, d.month, d.day) if d else ""

    row = 4
    for u in sorted(units, key=unit_sort_key):
        p = u.primary
        bed, bath = u.bed_bath
        vacant = u.is_vacant
        occ = "Vacant" if vacant else "Occupied"
        mtm = "Yes" if (p and re.search(r"MTM|M-?T-?M", p.term_type or "",
                                        re.I)) else None
        notice = "Yes" if (not vacant and u.on_notice) else None

        if vacant:
            lease_exp = ""
            fut = next((r for r in u.residents
                        if r.status.upper() == "L" and r.lease_expires), None)
            if fut:
                lease_exp = dt(fut.lease_expires)
            elif p and p.lease_expires:
                lease_exp = dt(p.lease_expires)
            values = [u.unit, u.floor_plan, u.sqft, bed, bath, "", "", occ,
                      u.market_rent, "", "", 0, "", "", "", "", "",
                      lease_exp, "", None, "", "", notice]
        else:
            rent = p.rent_charge if p else None
            rconc = p.recurring_concessions if p else None
            uconc = p.upfront_concessions if p else None
            disc = p.discounts if p else None
            other = p.other_income if p else None
            ner = ((rent or 0) + (rconc or 0) + (disc or 0)) \
                if rent is not None else 0
            values = [u.unit, u.floor_plan, u.sqft, bed, bath,
                      u.lease_type or "", "", occ,
                      u.market_rent,
                      rent if rent is not None else "",
                      rconc if rconc is not None else "",
                      ner, "",
                      uconc if uconc is not None else "",
                      disc if disc is not None else "",
                      other if other is not None else "",
                      dt(p.lease_start) if p and p.lease_start else "",
                      dt(p.lease_expires) if p else "",
                      "", mtm, dt(p.move_in) if p else "",
                      dt(p.move_out) if p else "", notice]

        for i, v in enumerate(values, start=1):
            ws.cell(row=row, column=i, value=v)
        _mark_estimates(ws, row, u)
        row += 1

    nrow = row + 1
    if any(getattr(u, "estimated", None) for u in units):
        est = ws.cell(row=nrow, column=1, value=_est_note_text(units))
        est.font = _est_note_font()
        nrow += 1
    if source_note:
        note = ws.cell(row=nrow, column=1, value=source_note)
        note.font = Font(name="Calibri", size=8, italic=True, color="808080")

    # Floor Plan Summary identity cells (A-E); remaining template slots -> ""
    fps_ws = wb["Floor Plan Summary"]
    plans = _floor_plans(units)
    for i, r in enumerate(range(3, 21)):
        if i < len(plans):
            fp = plans[i]
            bed, bath = next((u.bed_bath for u in units
                              if u.floor_plan == fp), (None, None))
            fps_ws[f"A{r}"] = fp
            # Lease Type rolls up only when the whole plan shares one (a plan
            # with a mix of e.g. voucher and market leases has no single
            # answer -- the per-unit values on the Rent Roll tab are the truth)
            lts = {u.lease_type for u in units if u.floor_plan == fp}
            fps_ws[f"B{r}"] = lts.pop() if len(lts) == 1 else ""
            fps_ws[f"C{r}"] = ""
            fps_ws[f"D{r}"] = bed
            fps_ws[f"E{r}"] = bath
        else:
            for col in "ABCDE":
                fps_ws[f"{col}{r}"] = ""

    # Trim unused floor-plan slots on the Floor Plan tab (template has 18)
    # and re-point the Total/Average formulas at the trimmed range.
    n = len(plans)
    if n < 18:
        fp_ws = wb["Floor Plan"]
        fp_ws.delete_rows(6 + n, 18 - n)
        _rewrite_fp_totals(fp_ws, n)

    _define_names(wb, row - 1)
    _save_normalized(wb, path)
    return row - 4


def _rewrite_fp_totals(fp_ws, n):
    """Rewrite the Floor Plan Total/Average rows for n floor-plan rows
    starting at row 6 (styles already in place)."""
    first, last_fp = 6, 5 + n
    tr, ar = last_fp + 1, last_fp + 2
    FP = "'Floor Plan'"
    totals = {
        "A": "Total",
        "E": f"=SUMPRODUCT({FP}!$E${first}:$E${last_fp},"
             f"{FP}!$J${first}:$J${last_fp})",
        "F": f"=SUBTOTAL(109,{FP}!$F${first}:$F${last_fp})",
        "G": f"=SUBTOTAL(109,{FP}!$G${first}:$G${last_fp})",
        "H": f"=SUBTOTAL(109,{FP}!$H${first}:$H${last_fp})",
        "I": f"=SUBTOTAL(109,{FP}!$I${first}:$I${last_fp})",
        "J": f"=SUBTOTAL(109,{FP}!$J${first}:$J${last_fp})",
        "K": f"=SUMPRODUCT({FP}!$K${first}:$K${last_fp},"
             f"{FP}!$J${first}:$J${last_fp})",
        "M": f"=SUMPRODUCT({FP}!$M${first}:$M${last_fp},"
             f"{FP}!$J${first}:$J${last_fp})",
    }
    for col, v in totals.items():
        fp_ws[f"{col}{tr}"] = v
    averages = {
        "A": "Average",
        "E": f"=IF({FP}!$J${tr},{FP}!$E${tr}/{FP}!$J${tr},)",
        "K": f"=IF({FP}!$J${tr},{FP}!$K${tr}/{FP}!$J${tr},)",
        "L": f"=IF(E{ar},K{ar}/E{ar},)",
        "M": f"=IF({FP}!$J${tr},{FP}!$M${tr}/{FP}!$J${tr},)",
        "N": ArrayFormula(
            f"N{ar}",
            f"=IFERROR(SUMPRODUCT({FP}!$M${first}:$M${last_fp}*"
            f"{FP}!$J${first}:$J${last_fp}*({FP}!$E${first}:$E${last_fp}>0))/"
            f"SUMPRODUCT({FP}!$M${first}:$M${last_fp}*"
            f"{FP}!$J${first}:$J${last_fp}*({FP}!$E${first}:$E${last_fp}>0)/"
            f"({FP}!$N${first}:$N${last_fp}*"
            f"({FP}!$M${first}:$M${last_fp}<>0)"
            f"+1*({FP}!$M${first}:$M${last_fp}=0)"
            f"+1*({FP}!$E${first}:$E${last_fp}=0)*"
            f"({FP}!$M${first}:$M${last_fp}<>0)"
            f"+1*({FP}!$E${first}:$E${last_fp}>0)*"
            f"({FP}!$M${first}:$M${last_fp}<>0)*"
            f"({FP}!$J${first}:$J${last_fp}=0))),0)"),
    }
    for col, v in averages.items():
        fp_ws[f"{col}{ar}"] = v


def write_workbook(path, prop, asof, units, source_note=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rent Roll"
    ws.sheet_view.showGridLines = False

    cg = dict(name="Century Gothic", bold=True, color=TITLE_COLOR)
    ws["A1"] = prop or "Property"
    ws["A1"].font = Font(size=14, **cg)
    ws["A1"].fill = PatternFill("solid", fgColor="D9D9D9")
    ws["K1"] = "Rent Roll"
    ws["K1"].font = Font(size=12, **cg)
    ws["A2"] = "Rent Roll as of:"
    ws["A2"].font = Font(size=10, **cg)
    ws["A2"].alignment = Alignment(vertical="center")
    ws["B2"] = datetime(asof.year, asof.month, asof.day) if asof else None
    ws["B2"].font = Font(size=10, **cg)
    ws["B2"].number_format = "mm-dd-yy"
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 17.1
    ws.row_dimensions[3].height = 33.4

    hdr_font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    hdr_align = Alignment(horizontal="center", wrap_text=True)
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.number_format = COL_FMTS.get(i, "General")
        ws.column_dimensions[get_column_letter(i)].width = COL_WIDTHS[i - 1]

    data_font = Font(name="Calibri", size=10)
    data_fill = PatternFill("solid", fgColor=ROW_FILL)
    hair_top = Border(top=Side(style="hair"))

    def dt(d):
        return datetime(d.year, d.month, d.day) if d else ""

    row = 4
    for u in sorted(units, key=unit_sort_key):
        p = u.primary
        bed, bath = u.bed_bath
        vacant = u.is_vacant
        occ = "Vacant" if vacant else "Occupied"
        mtm = "Yes" if (p and re.search(r"MTM|M-?T-?M", p.term_type or "",
                                        re.I)) else None
        notice = "Yes" if (not vacant and u.on_notice) else None

        if vacant:
            lease_exp = ""
            # a future/pending lease may still carry an expiration date
            fut = next((r for r in u.residents
                        if r.status.upper() == "L" and r.lease_expires), None)
            if fut:
                lease_exp = dt(fut.lease_expires)
            elif p and p.lease_expires:
                lease_exp = dt(p.lease_expires)
            values = [u.unit, u.floor_plan, u.sqft, bed, bath,
                      u.lease_type or "", "", occ,
                      u.market_rent, "", "", 0, "", "", "", "", "",
                      lease_exp, "", None, "", "", notice]
        else:
            rent = p.rent_charge if p else None
            rconc = p.recurring_concessions if p else None
            uconc = p.upfront_concessions if p else None
            disc = p.discounts if p else None
            other = p.other_income if p else None
            # Net Effective Rent = contractual rent net of recurring
            # concessions and discounts (stored as negative amounts).
            # Upfront (one-time) concessions are intentionally excluded so
            # they don't understate average in-place rents.
            ner = ((rent or 0) + (rconc or 0) + (disc or 0)) \
                if rent is not None else 0
            values = [u.unit, u.floor_plan, u.sqft, bed, bath,
                      u.lease_type or "", "", occ,
                      u.market_rent,
                      rent if rent is not None else "",
                      rconc if rconc is not None else "",
                      ner,
                      "",
                      uconc if uconc is not None else "",
                      disc if disc is not None else "",
                      other if other is not None else "",
                      dt(p.lease_start) if p and p.lease_start else "",
                      dt(p.lease_expires) if p else "",
                      "", mtm, dt(p.move_in) if p else "",
                      dt(p.move_out) if p else "", notice]

        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = data_font
            c.fill = data_fill
            c.number_format = COL_FMTS.get(i, "General")
            c.border = hair_top
            if i == 1:
                c.alignment = Alignment(horizontal="left")
        _mark_estimates(ws, row, u)
        row += 1

    nrow = row + 1
    if any(getattr(u, "estimated", None) for u in units):
        est = ws.cell(row=nrow, column=1, value=_est_note_text(units))
        est.font = _est_note_font()
        nrow += 1
    if source_note:
        note = ws.cell(row=nrow, column=1, value=source_note)
        note.font = Font(name="Calibri", size=8, italic=True, color="808080")

    last_rr_row = row - 1          # last unit row on the Rent Roll sheet
    _define_names(wb, last_rr_row)
    _add_floor_plan_sheets(wb, units, last_rr_row)

    _save_normalized(wb, path)
    return row - 4  # number of unit rows written


# ----------------------------------------------------------------------------
# Named ranges (rediQ-style) + Floor Plan / Floor Plan Summary sheets
# ----------------------------------------------------------------------------

def _define_names(wb, last):
    """Recreate the named ranges the Floor Plan formulas rely on."""
    names = {
        "rediq_dealname": "'Rent Roll'!$A$1",
        "rediq_rentrollasofdate": "'Rent Roll'!$B$2",
        "rediq_rentroll": "'Rent Roll'!$A$3",
        "rediq_floorplansummary": "'Floor Plan Summary'!$A$1",
        "UnitNo": f"'Rent Roll'!$A$4:$A${last}",
        "FloorPlan": f"'Rent Roll'!$B$4:$B${last}",
        "NetSf": f"'Rent Roll'!$C$4:$C${last}",
        "Bed": f"'Rent Roll'!$D$4:$D${last}",
        "Bath": f"'Rent Roll'!$E$4:$E${last}",
        "RentType": f"'Rent Roll'!$F$4:$F${last}",
        "RenovationString": f"'Rent Roll'!$G$4:$G${last}",
        "OccupancyString": f"'Rent Roll'!$H$4:$H${last}",
        "MarketRent": f"'Rent Roll'!$I$4:$I${last}",
        "CurrentInPlaceRent": f"'Rent Roll'!$J$4:$J${last}",
        "CurrentRecurringConc": f"'Rent Roll'!$K$4:$K${last}",
        "CurrentNetEffectiveRent": f"'Rent Roll'!$L$4:$L${last}",
        "CurrentSuppRent": f"'Rent Roll'!$M$4:$M${last}",
        "CurrentUpfrontConc": f"'Rent Roll'!$N$4:$N${last}",
        "CurrentEmpOtherConc": f"'Rent Roll'!$O$4:$O${last}",
        "CurrentOtherIncome": f"'Rent Roll'!$P$4:$P${last}",
        "CurrentLeaseSignedDate": f"'Rent Roll'!$Q$4:$Q${last}",
        "CurrentLeaseExpirationDate": f"'Rent Roll'!$R$4:$R${last}",
        "CurrentLeaseTerm": f"'Rent Roll'!$S$4:$S${last}",
        "CurrentIsMonthToMonth": f"'Rent Roll'!$T$4:$T${last}",
        "CurrentMoveInDate": f"'Rent Roll'!$U$4:$U${last}",
        "CurrentMoveOutDate": f"'Rent Roll'!$V$4:$V${last}",
        "IsNoticeToVacate": f"'Rent Roll'!$W$4:$W${last}",
    }
    for n, ref in names.items():
        wb.defined_names[n] = DefinedName(n, attr_text=ref)


FPS_GROUPS = [   # (range, title) for Floor Plan Summary row 1
    ("A1:H1", "Unit Info"), ("I1:K1", "Occupancy Status (# Units)"),
    ("L1:N1", "Occupancy Status (%)"), ("O1:O1", "All Units"),
    ("P1:R1", "Currently Occupied Units"),
    ("S1:V1", "In-Place Rent (by Lease Start Date)"),
    ("W1:Z1", "# Leases (by Lease Start Date)"),
]
FPS_HEADERS = ["Floor Plan", "Lease Type", "Renovated", "Bed", "Bath",
               "Net sf", "# Units", "%", "Occupied", "Vacant", "Non-Rev",
               "Occupied", "Vacant", "Non-Rev", "Market Rent", "Market Rent",
               "In-Place Rent", "% of Market Rent", "Recent 2",
               "Last 90 Days", "Last 60 Days", "Last 30 Days", "Recent 2",
               "Last 90 Days", "Last 60 Days", "Last 30 Days"]
FPS_WIDTHS = {"A": 12.64, "B": 12.47, "C": 12.23, "D": 5.7, "E": 6.23,
              "F": 7.29, "G": 8.06, "H": 7.59, "O": 11.47, "P": 11.76,
              "Q": 11.29, "R": 13.88, "S": 8.41, "T": 11.18, "W": 8.41,
              "X": 11.18}
FMT_PCT0 = "0%"
FMT_DOLLAR0 = '"$"#,##0'


def _floor_plans(units):
    """Distinct floor plans, ordered by bed count then name."""
    seen = {}
    for u in units:
        if u.floor_plan and u.floor_plan not in seen:
            bed, bath = u.bed_bath
            seen[u.floor_plan] = (bed if bed is not None else 99,
                                  bath if bath is not None else 99)
    return sorted(seen, key=lambda fp: (seen[fp], fp))


def _add_floor_plan_sheets(wb, units, last):
    hdr_font = Font(name="Century Gothic", size=8, bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=HEADER_FILL)
    band_font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    data_fill = PatternFill("solid", fgColor=ROW_FILL)
    fps = _floor_plans(units)
    n = len(fps)

    # ------------------------------------------------------------------ #
    # Floor Plan Summary                                                 #
    # ------------------------------------------------------------------ #
    ws = wb.create_sheet("Floor Plan Summary")
    for rng, title in FPS_GROUPS:
        if ":" in rng and rng.split(":")[0] != rng.split(":")[1]:
            ws.merge_cells(rng)
        anchor = rng.split(":")[0]
        c = ws[anchor]
        c.value = title
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for col in range(1, 27):        # fill the whole band rows 1-2
        for r in (1, 2):
            cell = ws.cell(row=r, column=col)
            cell.font = hdr_font
            if cell.fill.patternType is None:
                cell.fill = hdr_fill
    for i, h in enumerate(FPS_HEADERS, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
    ws.row_dimensions[1].height = 10.5
    ws.row_dimensions[2].height = 10.5
    for col, w in FPS_WIDTHS.items():
        ws.column_dimensions[col].width = w

    fps_last = 2 + n
    RR = "'Rent Roll'"
    FS = "'Floor Plan Summary'"
    for i, fp in enumerate(fps):
        r = 3 + i
        bed, bath = next((u.bed_bath for u in units if u.floor_plan == fp),
                         (None, None))
        vals = {
            # A-E: identity columns (values; source template pulled these
            # from the excluded 'Bed-Bath Type' tab)
            "A": fp, "B": "", "C": "", "D": bed, "E": bath,
            # F-G: live aggregates from the Rent Roll tab
            "F": f"=AVERAGEIFS({RR}!C:C,{RR}!B:B,$A{r})",
            "G": f"=COUNTIF({RR}!$B:$B,$A{r})",
            "H": f'=IFERROR(G{r}/SUM($G$3:$G${fps_last}),"")',
            "I": f'=IF(A{r}="","",COUNTIFS({RR}!$B:$B,{FS}!$A{r},'
                 f'{RR}!$H:$H,{FS}!I$2))',
            "J": f'=IF(A{r}="","",COUNTIFS({RR}!$B:$B,{FS}!$A{r},'
                 f'{RR}!$H:$H,{FS}!J$2))',
            "K": f'=IF(A{r}="","",COUNTIFS({RR}!$B:$B,{FS}!$A{r},'
                 f'{RR}!$H:$H,{FS}!K$2))',
            "L": f'=IF(A{r}="","",I{r}/G{r})',
            "M": f'=IF(A{r}="","",J{r}/G{r})',
            "N": f'=IF(A{r}="","",K{r}/G{r})',
            "O": f'=IF(A{r}="","",AVERAGEIFS({RR}!I:I,{RR}!B:B,{FS}!A{r}))',
            "P": f'=IF(A{r}="","",AVERAGEIFS({RR}!I:I,{RR}!B:B,{FS}!A{r},'
                 f'{RR}!$H:$H,{FS}!$I$2))',
            "Q": f'=IF(A{r}="","",AVERAGEIFS({RR}!L:L,{RR}!B:B,{FS}!A{r},'
                 f'{RR}!$H:$H,{FS}!$I$2))',
            "R": f'=IF(A{r}="","",Q{r}/P{r})',
            "S": ArrayFormula(
                f"S{r}",
                f'=IFERROR(IF(AVERAGE(LARGE(IF({RR}!$B$4:$B${last}={FS}!$A{r},'
                f'{RR}!$L$4:$L${last}),{{1,2}}))=0,"",'
                f'AVERAGE(LARGE(IF({RR}!$B$4:$B${last}={FS}!$A{r},'
                f'{RR}!$L$4:$L${last}),{{1,2}}))),"")'),
            "T": f'=IFERROR(AVERAGEIFS({RR}!$L$4:$L${last},'
                 f'{RR}!$B$4:$B${last},{FS}!$A{r},{RR}!$Q$4:$Q${last},'
                 f'">="&rediq_rentrollasofdate-90),"")',
            "U": f'=IFERROR(AVERAGEIFS({RR}!$L$4:$L${last},'
                 f'{RR}!$B$4:$B${last},{FS}!$A{r},{RR}!$Q$4:$Q${last},'
                 f'">="&rediq_rentrollasofdate-60),"")',
            "V": f'=IFERROR(AVERAGEIFS({RR}!$L$4:$L${last},'
                 f'{RR}!$B$4:$B${last},{FS}!$A{r},{RR}!$Q$4:$Q${last},'
                 f'">="&rediq_rentrollasofdate-30),"")',
            "X": f'=IF(A{r}="","",COUNTIFS({RR}!$B$4:$B${last},{FS}!$A{r},'
                 f'{RR}!$Q$4:$Q${last},">="&TODAY()-90))',
            "Y": f'=IF(A{r}="","",COUNTIFS({RR}!$B$4:$B${last},{FS}!$A{r},'
                 f'{RR}!$Q$4:$Q${last},">="&TODAY()-60))',
            "Z": f'=IF(A{r}="","",COUNTIFS({RR}!$B$4:$B${last},{FS}!$A{r},'
                 f'{RR}!$Q$4:$Q${last},">="&TODAY()-30))',
        }
        fmts = {"H": FMT_PCT0, "L": FMT_PCT0, "M": FMT_PCT0, "N": FMT_PCT0,
                "R": FMT_PCT0, "O": FMT_DOLLAR0, "P": FMT_DOLLAR0,
                "Q": FMT_DOLLAR0, "S": FMT_DOLLAR0, "T": FMT_DOLLAR0,
                "U": FMT_DOLLAR0, "V": FMT_DOLLAR0}
        for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
            cell = ws[f"{col}{r}"]
            if col in vals:
                cell.value = vals[col]
            cell.font = data_font
            cell.fill = data_fill
            cell.number_format = fmts.get(col, "General")

    # ------------------------------------------------------------------ #
    # Floor Plan                                                         #
    # ------------------------------------------------------------------ #
    fp_ws = wb.create_sheet("Floor Plan", 0)
    fp_ws.sheet_view.showGridLines = False
    title_font = Font(name="Century Gothic", size=14, bold=True,
                      color=TITLE_COLOR)
    fp_ws["A1"] = "=rediq_dealname"
    fp_ws["A1"].font = title_font
    fp_ws["H1"] = "Floor Plan Summary"
    fp_ws["H1"].font = Font(name="Century Gothic", size=12, bold=True,
                            color="808080")
    fp_ws["H1"].alignment = Alignment(horizontal="center")
    row1_fill = PatternFill("solid", fgColor="D9D9D9")
    for col in range(1, 15):
        fp_ws.cell(row=1, column=col).fill = row1_fill
    fp_ws["A2"] = ('="Floor Plans as of:  "'
                   '&TEXT(rediq_rentrollasofdate,"mmmm d, yyyy")')
    fp_ws["A2"].font = Font(name="Century Gothic", size=10, bold=True,
                            color=TITLE_COLOR)
    fp_ws.row_dimensions[1].height = 17.1
    fp_ws.row_dimensions[3].height = 14.7
    for col, w in {"A": 20.88, "B": 13.12, "C": 14.47, "D": 7.59, "E": 10.59,
                   "F": 6.59, "G": 10.29, "K": 10.7, "L": 7.7, "M": 10.7,
                   "N": 7.7}.items():
        fp_ws.column_dimensions[col].width = w

    hdr_fill_navy = PatternFill("solid", fgColor=HEADER_FILL)

    def band(row, texts):
        """Write a navy band row: {col_letter: text}."""
        for col in range(1, 15):
            cell = fp_ws.cell(row=row, column=col)
            cell.fill = hdr_fill_navy
            cell.font = band_font
            cell.alignment = Alignment(horizontal="centerContinuous")
        for col, txt in texts.items():
            fp_ws[f"{col}{row}"] = txt

    band(3, {"A": "FLOOR PLAN INFORMATION", "G": "UNIT STATUS",
             "K": "MONTHLY RENT"})
    band(4, {"K": "Market Rent", "M": "Net Effective Rent"})
    band(5, {"A": "Floor Plan", "B": "Lease Type.", "C": "Renov. Status",
             "D": "Bed", "E": "Net sf", "F": "%", "G": "Occupied",
             "H": "Vacant", "I": "Non-Rev", "J": "Total", "K": "per unit",
             "L": "psf", "M": "per unit", "N": "psf"})
    for col in "ABCGHIJKLMN":   # row-5 labels align like the template
        fp_ws[f"{col}5"].alignment = Alignment(horizontal="left")
    for col in "DEF":
        fp_ws[f"{col}5"].alignment = Alignment(horizontal="right")

    FMT_FP = {
        "A": FMT_TEXT, "B": FMT_TEXT, "C": FMT_TEXT, "D": FMT_BED,
        "E": '#,##0\\ "SF"_);\\(#,##0\\ "SF"\\)_);_("-"_);\\ @_)',
        "F": '#,##0%_);\\(#,##0%\\);"-"_);\\ @_)',
        "G": '#,##0\\ "units"_);\\(#,##0\\ "units"\\)_);_("-"_);\\ @_)',
        "H": '#,##0\\ "units"_);\\(#,##0\\ "units"\\)_);_("-"_);\\ @_)',
        "I": '#,##0\\ "units"_);\\(#,##0\\ "units"\\)_);_("-"_);\\ @_)',
        "J": '#,##0\\ "units"_);\\(#,##0\\ "units"\\)_);_("-"_);\\ @_)',
        "K": FMT_CUR,
        "L": '"$"\\ #,##0.00_);\\("$"\\ #,##0.00\\);_("-"_);\\ @_)',
        "M": FMT_CUR,
        "N": '"$"\\ #,##0.00_);\\("$"\\ #,##0.00\\);_("-"_);\\ @_)',
    }

    first, last_fp = 6, 5 + n
    for i in range(n):
        r = first + i
        sr = 3 + i          # matching Floor Plan Summary row
        vals = {
            "A": f'=IF(AND({FS}!A{sr}<>"",LEFT({FS}!A{sr},5)<>"Total"),'
                 f'{FS}!A{sr},"")',
            "B": f'=IF($A{r}="","",IF(ISBLANK({FS}!B{sr}),"",{FS}!B{sr}))',
            "C": f'=IF($A{r}="","",IF(ISBLANK({FS}!C{sr}),"",{FS}!C{sr}))',
            "D": f'=IF($A{r}="","",IF(ISBLANK({FS}!D{sr}),"",{FS}!D{sr}))',
            "E": f'=IF($A{r}="",0,IF(ISBLANK({FS}!F{sr}),"",{FS}!F{sr}))',
            "F": f'=IF($A{r}="","",IF(ISBLANK({FS}!H{sr}),"",{FS}!H{sr}))',
            "G": f'=IF($A{r}="",0,IF(ISBLANK({FS}!I{sr}),"",{FS}!I{sr}))',
            "H": f'=IF($A{r}="",0,IF(ISBLANK({FS}!J{sr}),"",{FS}!J{sr}))',
            "I": f'=IF($A{r}="",0,IF(ISBLANK({FS}!K{sr}),"",{FS}!K{sr}))',
            "J": f"=SUM(G{r}:I{r})",
            "K": f'=IF($A{r}="",0,IF(ISBLANK({FS}!O{sr}),"",{FS}!O{sr}))',
            "L": f"=IFERROR(K{r}/E{r},0)",
            "M": f'=IF($A{r}="",0,IF(ISBLANK({FS}!Q{sr}),"",{FS}!Q{sr}))',
            "N": f"=IFERROR(M{r}/E{r},0)",
        }
        for col in "ABCDEFGHIJKLMN":
            cell = fp_ws[f"{col}{r}"]
            cell.value = vals[col]
            cell.font = data_font
            cell.fill = data_fill
            cell.number_format = FMT_FP[col]

    # Total / Average rows
    tr, ar = last_fp + 1, last_fp + 2
    FP = "'Floor Plan'"
    top = Border(top=Side(style="medium"))
    totals = {
        "A": "Total",
        "E": f"=SUMPRODUCT({FP}!$E${first}:$E${last_fp},"
             f"{FP}!$J${first}:$J${last_fp})",
        "F": f"=SUBTOTAL(109,{FP}!$F${first}:$F${last_fp})",
        "G": f"=SUBTOTAL(109,{FP}!$G${first}:$G${last_fp})",
        "H": f"=SUBTOTAL(109,{FP}!$H${first}:$H${last_fp})",
        "I": f"=SUBTOTAL(109,{FP}!$I${first}:$I${last_fp})",
        "J": f"=SUBTOTAL(109,{FP}!$J${first}:$J${last_fp})",
        "K": f"=SUMPRODUCT({FP}!$K${first}:$K${last_fp},"
             f"{FP}!$J${first}:$J${last_fp})",
        "M": f"=SUMPRODUCT({FP}!$M${first}:$M${last_fp},"
             f"{FP}!$J${first}:$J${last_fp})",
    }
    for col in "ABCDEFGHIJKLMN":
        cell = fp_ws[f"{col}{tr}"]
        if col in totals:
            cell.value = totals[col]
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.border = top
        cell.number_format = FMT_FP[col]
    averages = {
        "A": "Average",
        "E": f"=IF({FP}!$J${tr},{FP}!$E${tr}/{FP}!$J${tr},)",
        "K": f"=IF({FP}!$J${tr},{FP}!$K${tr}/{FP}!$J${tr},)",
        "L": f"=IF(E{ar},K{ar}/E{ar},)",
        "M": f"=IF({FP}!$J${tr},{FP}!$M${tr}/{FP}!$J${tr},)",
        "N": ArrayFormula(
            f"N{ar}",
            f"=IFERROR(SUMPRODUCT({FP}!$M${first}:$M${last_fp}*"
            f"{FP}!$J${first}:$J${last_fp}*({FP}!$E${first}:$E${last_fp}>0))/"
            f"SUMPRODUCT({FP}!$M${first}:$M${last_fp}*"
            f"{FP}!$J${first}:$J${last_fp}*({FP}!$E${first}:$E${last_fp}>0)/"
            f"({FP}!$N${first}:$N${last_fp}*"
            f"({FP}!$M${first}:$M${last_fp}<>0)"
            f"+1*({FP}!$M${first}:$M${last_fp}=0)"
            f"+1*({FP}!$E${first}:$E${last_fp}=0)*"
            f"({FP}!$M${first}:$M${last_fp}<>0)"
            f"+1*({FP}!$E${first}:$E${last_fp}>0)*"
            f"({FP}!$M${first}:$M${last_fp}<>0)*"
            f"({FP}!$J${first}:$J${last_fp}=0))),0)"),
    }
    for col in "ABCDEFGHIJKLMN":
        cell = fp_ws[f"{col}{ar}"]
        if col in averages:
            cell.value = averages[col]
        cell.font = Font(name="Calibri", size=10, italic=True)
        cell.number_format = FMT_FP[col]


# ----------------------------------------------------------------------------
# Reconciliation
# ----------------------------------------------------------------------------

def reconcile(units, checks):
    lines, ok = [], True
    # Optional per-check provenance labels. Formats that print their own
    # totals leave this empty ("report"); owner-made sheets that print no
    # totals set e.g. {"unit_count": "re-extract"} so the block says plainly
    # what each number was tied to.
    src = checks.get("_src", {})

    def check(label, got, want, tol=0.01, key=None):
        nonlocal ok
        if want is None:
            lines.append(f"  ~ {label}: {got:,.2f} (no report total to check)")
            return
        good = abs(got - want) <= tol
        ok &= good
        mark = "OK " if good else "MISMATCH"
        who = src.get(key, "report")
        lines.append(f"  {mark} {label}: parsed {got:,.2f} vs {who} "
                     f"{want:,.2f}")

    check("Unit count", len(units), checks.get("unit_count"),
          key="unit_count")
    if checks.get("total_sqft") is not None:
        check("Total sq ft", sum(u.sqft or 0 for u in units),
              checks.get("total_sqft"))
    occ = [u for u in units if not u.is_vacant]
    vac = [u for u in units if u.is_vacant]
    check("Occupied units", len(occ), checks.get("occupied_count"), tol=0.5,
          key="occupied_count")
    if checks.get("vacant_count") is not None:
        check("Vacant units", len(vac), checks["vacant_count"], tol=0.5,
              key="vacant_count")
    if checks.get("nonrev_count") is not None:
        # no source format flags non-revenue units today; the report prints
        # the bucket, so tie it out explicitly rather than ignoring it
        check("Non-revenue units",
              sum(1 for u in units if u.apt_status.upper() in ("NR", "MO")),
              checks.get("nonrev_count"), tol=0.5)
    if checks.get("occupied_sqft") is not None:
        check("Occupied sq ft", sum(u.sqft or 0 for u in occ),
              checks["occupied_sqft"])
    if checks.get("vacant_sqft") is not None:
        check("Vacant sq ft", sum(u.sqft or 0 for u in vac),
              checks["vacant_sqft"])
    if checks.get("occupied_market_rent") is not None:
        check("Occupied market rent", sum(u.market_rent or 0 for u in occ),
              checks.get("occupied_market_rent"), key="occupied_market_rent")
    if checks.get("vacant_market_rent") is not None:
        check("Vacant market rent", sum(u.market_rent or 0 for u in vac),
              checks["vacant_market_rent"])
    if checks.get("total_market_rent") is not None:
        check("Total market rent (all units)",
              sum(u.market_rent or 0 for u in units),
              checks.get("total_market_rent"), key="total_market_rent")
    if checks.get("total_contract_rent") is not None:
        check("Total contract rent",
              sum(u.primary.rent_charge or 0 for u in occ if u.primary),
              checks.get("total_contract_rent"), key="total_contract_rent")
    if checks.get("reextract_contract_rent") is not None:
        check("Total contract rent (2nd pass)",
              sum(u.primary.rent_charge or 0 for u in occ if u.primary),
              checks.get("reextract_contract_rent"),
              key="reextract_contract_rent")
    if checks.get("total_deposits") is not None:
        check("Total security deposits",
              sum(r.deposit or 0 for u in units for r in u.residents
                  if r.status.upper() in ("C", "N")),
              checks.get("total_deposits"))
    if checks.get("total_other_deposits") is not None:
        check("Total other deposits",
              sum(r.other_deposit or 0 for u in units for r in u.residents
                  if r.status.upper() in ("C", "N")),
              checks.get("total_other_deposits"))
    fut = [(u, r) for u in units for r in u.residents
           if r.status.upper() == "L"]
    if checks.get("future_count") is not None:
        check("Future residents/applicants", len(fut),
              checks.get("future_count"), tol=0.5)
    if checks.get("future_sqft") is not None:
        check("Future-resident sq ft", sum(u.sqft or 0 for u, _ in fut),
              checks.get("future_sqft"))
    if checks.get("future_market_rent") is not None:
        check("Future-resident market rent",
              sum(u.market_rent or 0 for u, _ in fut),
              checks.get("future_market_rent"))

    def trunc2(x):
        """Reports print occupancy truncated (81.5789% -> 81.57)."""
        return int(x * 100) / 100.0

    if checks.get("pct_unit_occupancy") is not None and units:
        check("% unit occupancy", trunc2(len(occ) / len(units) * 100),
              checks.get("pct_unit_occupancy"))
    tot_sf = sum(u.sqft or 0 for u in units)
    if checks.get("pct_sqft_occupied") is not None and tot_sf:
        check("% sq ft occupied",
              trunc2(sum(u.sqft or 0 for u in occ) / tot_sf * 100),
              checks.get("pct_sqft_occupied"))
    cn_total = sum(r.total_charges for u in units for r in u.residents
                   if r.status.upper() in ("C", "N"))
    check("Current/On-Notice lease charges", cn_total,
          checks.get("current_lease_charges"),
          key="current_lease_charges")


    # --- optional checks (populated by parsers that can read them) ------
    if checks.get("report_total_market_rent") is not None:
        check("Total market rent (grand-total row)",
              sum(u.market_rent or 0 for u in units),
              checks["report_total_market_rent"])
    for label, pk, rk in (
            ("Surety bonds", "parsed_surety_bonds",
             "report_total_surety_bonds"),
            ("Security deposits", "parsed_deposits", "report_total_deposits"),
            ("Resident balances", "parsed_balance", "report_total_balance"),
            ("Applicant balances", "parsed_future_balance",
             "report_future_balance")):
        if checks.get(rk) is not None and checks.get(pk) is not None:
            check(label, checks[pk], checks[rk])

    all_charges = [(d, a) for u in units for r in u.residents
                   for d, a, _ in r.charges]
    if checks.get("total_charges") is not None:
        check("Total charges (debits)",
              sum(a for _, a in all_charges if a > 0), checks["total_charges"])
    if checks.get("total_credits") is not None:
        check("Total credits", -sum(a for _, a in all_charges if a < 0),
              checks["total_credits"])
    for key, sign, title in (("charge_totals", 1, "charge"),
                             ("credit_totals", -1, "credit")):
        want = checks.get(key)
        if not want:
            continue
        got = {}
        for d, a in all_charges:
            if (a < 0) == (sign < 0):
                got[d] = got.get(d, 0.0) + sign * a
        for d in sorted(set(want) | set(got)):
            check(f"  {title}: {d}", got.get(d, 0.0), want.get(d))
    if checks.get("unit_charge_totals"):
        want = checks["unit_charge_totals"]
        bad = [u.unit for u in units
               if u.unit in want
               and abs(sum(r.total_charges for r in u.residents)
                       - want[u.unit]) > 0.01]
        n = len([u for u in units if u.unit in want])
        check(f"Per-unit printed charge totals ({n} units)",
              n - len(bad), n, tol=0.5)
        if bad:
            lines.append(f"        offending units: {', '.join(bad)}")
    if checks.get("floor_plan_totals"):
        for fp, want in sorted(checks["floor_plan_totals"].items()):
            mine = [u for u in units if u.floor_plan == fp]
            check(f"  plan {fp}: units", len(mine), want["units"], tol=0.5)
            check(f"  plan {fp}: market rent",
                  sum(u.market_rent or 0 for u in mine), want["market"])
            check(f"  plan {fp}: sq ft",
                  sum(u.sqft or 0 for u in mine), want["sqft"])
    # Parser-supplied checks: (label, fn(units) -> value, printed value, tol).
    # The value is always recomputed from the reconstructed UnitRecord list,
    # so these tie the OUTPUT to the report, not the parser to itself.
    for label, fn, want, tol in checks.get("extra_checks", []):
        check(label, float(fn(units)), want, tol=tol)
    if checks.get("_onesite_bad_billing"):
        lines.append("        offending leases: "
                     + ", ".join(checks["_onesite_bad_billing"]))
    return ok, "\n".join(lines)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------

def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("pdf", help="source rent roll (PDF or AppFolio xlsx export)")
    ap.add_argument("-o", "--output", help="output .xlsx path")
    ap.add_argument("--property", help="override property name")
    ap.add_argument("--asof",
                    help="as-of date (YYYY-MM-DD or M/D/YYYY). Required for "
                         "sources that print no date of their own (owner-made "
                         "spreadsheets); overrides a parsed date otherwise")
    ap.add_argument("--sqft",
                    help="fill missing Net Sf by floor plan, e.g. "
                         "'1/1.5=1100,2/2.5=1600' (for source formats with "
                         "no sqft; values must come from a cited source). "
                         "A key that is not a floor plan is matched against "
                         "the unit name (exact, then case-insensitive "
                         "substring) and wins over the floor-plan value - "
                         "use this for scattered-site portfolios where one "
                         "floor plan spans properties of different sizes, "
                         "e.g. '3/2.5=1600,Holly=1636,Wren Rowe=1566'")
    ap.add_argument("--bedbath",
                    help="fill Bed/Bath by floor plan for sources that do "
                         "not carry them, e.g. 'F1=1/1,F2=2/1.5'. Values "
                         "must come from a cited source; a bare '2' sets "
                         "beds only and leaves Bath blank.")
    ap.add_argument("--sqft-est",
                    help="same syntax as --sqft, but the cells it fills are "
                         "BEST ESTIMATES: they are highlighted red and a red "
                         "note is added under the data. Keys may be floor "
                         "plans or literal unit names.")
    ap.add_argument("--bedbath-est",
                    help="same syntax as --bedbath, but the cells it fills "
                         "are BEST ESTIMATES (highlighted red + note). Only "
                         "components that were blank (or that disagree with "
                         "the source) are highlighted; a unit with no floor "
                         "plan gets one synthesised so it joins the Floor "
                         "Plan rollups.")
    ap.add_argument("--estimate-market", action="store_true",
                    help="sources that carry NO Market Rent column: estimate "
                         "it per floor plan from the maximum stated "
                         "contractual rents (house rule 8/2026) - the "
                         "highest rent that repeats at least 3 times among "
                         "the plan's occupied units, preferring leases "
                         "started in the last 6 months. Every filled cell is "
                         "highlighted as an estimate; a market rent the "
                         "source DID provide is never overridden.")
    ap.add_argument("--template",
                    help="pre-styled template workbook (default: "
                         "rentroll_template.xlsx next to this script)")
    args = ap.parse_args(argv)

    if args.pdf.lower().endswith((".xlsx", ".xlsm")):
        cls = next((p for p in XLSX_PARSERS if p.detect_xlsx(args.pdf)), None)
        if cls is None:
            sys.exit("ERROR: xlsx input does not match any supported rent "
                     "roll export layout (AppFolio, Yardi/ResMan).")
        parser = cls()
        prop, asof, units, checks = parser.parse(args.pdf)
    else:
        with pdfplumber.open(args.pdf) as pdf:
            parser_cls = next((p for p in PARSERS if p.detect(pdf)), None)
            if parser_cls is None:
                sys.exit("ERROR: no parser recognizes this PDF layout.")
            parser = parser_cls()
            prop, asof, units, checks = parser.parse(pdf)

    if args.property:
        prop = args.property
    if args.asof:
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                asof = datetime.strptime(args.asof, fmt).date()
                break
            except ValueError:
                asof = None
        if asof is None:
            sys.exit(f"ERROR: --asof '{args.asof}' is not a recognised date.")
    # Parsers that explicitly report "I looked and this source prints no
    # date" (asof_found is False) must be given one; formats that normally
    # carry their own date keep the previous lenient behaviour.
    if asof is None and getattr(parser, "asof_found", None) is False:
        sys.exit("ERROR: this rent roll prints no as-of date. Pass one "
                 "explicitly with --asof YYYY-MM-DD (the as-of date drives "
                 "the filename and cell B2 - it must not be guessed).")
    # Anything the parser could not take at face value, surfaced before the
    # reconciliation block so it is never lost in the scroll.
    for f in getattr(parser, "flags", []):
        print(f"FLAG: {f}")
    # Order matters: sourced values first, estimates only fill what is
    # still missing, and --bedbath-est runs last because it can synthesise a
    # floor plan (which must not then be picked up by the sourced --sqft map).
    if args.sqft:
        _fill_sqft(units, args.sqft, estimate=False)
    if args.sqft_est:
        _fill_sqft(units, args.sqft_est, estimate=True)
    if args.bedbath:
        _fill_bedbath(units, args.bedbath, estimate=False)
    if args.bedbath_est:
        _fill_bedbath(units, args.bedbath_est, estimate=True)
    if args.estimate_market:
        if sum(1 for u in units if u.market_rent is None) == 0:
            print("--estimate-market: every unit already carries a market "
                  "rent from the source - nothing estimated.")
        else:
            notes, mflags = estimate_market_rents(units, asof)
            print("Market rent ESTIMATED from the maximum stated "
                  "contractual rents (house rule 8/2026); no market-rent "
                  "reconciliation check is filed - there is no report total "
                  "to tie an estimate to:")
            for ln in notes:
                print(ln)
            for f in mflags:
                print(f"FLAG: {f}")
    n_est = sum(1 for u in units if u.estimated)
    if n_est:
        cells = {}
        for u in units:
            for k in u.estimated:
                cells[k] = cells.get(k, 0) + 1
        print(f"ESTIMATES: {n_est} unit(s) carry best-estimate cells "
              f"({', '.join(f'{k}={v}' for k, v in sorted(cells.items()))}) "
              f"- highlighted red, with the note under the data.")
    if not units:
        sys.exit("ERROR: no units parsed from the PDF.")

    out = args.output
    if not out:
        # naming convention: "RR - Property Name - M-D-YYYY" (no zero padding,
        # e.g. "RR - Harvest Moon - 7-28-2026")
        d = f"{asof.month}-{asof.day}-{asof.year}" if asof else "output"
        out = f"RR - {prop} - {d}.xlsx"

    parser.source_kind = "xlsx export" if args.pdf.lower().endswith(
        (".xlsx", ".xlsm")) else "PDF"
    note = parser.source_note(asof)
    template = args.template or find_default_template()
    if template:
        n = write_workbook_from_template(out, template, prop, asof, units,
                                         source_note=note)
        print(f"Using template: {template}")
    else:
        print("NOTE: rentroll_template.xlsx not found - building formatting "
              "from scratch (close approximation of the template).")
        n = write_workbook(out, prop, asof, units, source_note=note)

    ok, report = reconcile(units, checks)
    print(f"Parsed {len(units)} units ({parser.name}) -> {out} ({n} rows)")
    print("Reconciliation vs report's own totals:")
    print(report)
    if not ok:
        print("WARNING: totals do not tie out — review the output.")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
