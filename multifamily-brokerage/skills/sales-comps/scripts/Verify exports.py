#!/usr/bin/env python3
"""Sanity-check the Excel export and PDF grid before delivery.

Usage: python verify_exports.py <outdir> --selection selection.json
Recalculates the Comp Grid via LibreOffice and confirms the formula-computed
indicated value matches the Python-computed one; checks the PDF opens and
carries the deal name. Exits non-zero on any failure.
"""
import argparse
import glob
import json
import os
import subprocess
import sys
import tempfile

FAILED = []


def check(label, ok, detail=""):
    print(f"  [{'ok' if ok else 'FAIL'}] {label}" + (f" — {detail}" if detail and not ok else ""))
    if not ok:
        FAILED.append(label)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("outdir")
    ap.add_argument("--selection", required=True)
    args = ap.parse_args()
    sel = json.load(open(args.selection))
    name = sel["subject"]["name"]
    xlsx = os.path.join(args.outdir, f"{name} - Sale Comp Export.xlsx")
    uw = os.path.join(args.outdir, f"{name} - Underwriting Sale Data.xlsx")
    pdf = os.path.join(args.outdir, f"{name} - Comparable Sale Grid.pdf")

    check("client xlsx exists", os.path.exists(xlsx), xlsx)
    check("underwriting xlsx exists", os.path.exists(uw), uw)
    check("pdf exists", os.path.exists(pdf), pdf)
    if FAILED:
        sys.exit(1)

    # xlsx structure
    import openpyxl
    wb = openpyxl.load_workbook(xlsx)
    check("sheets: Comp Export + Comp Grid", wb.sheetnames == ["Comp Export", "Comp Grid"],
          str(wb.sheetnames))
    ws = wb["Comp Export"]
    sel_comps_ordered = sorted(
        [c for c in sel["comps"] if c["Rank"] in sel["selected_ranks"]],
        key=lambda c: sel["selected_ranks"].index(c["Rank"]))
    top5 = [ws.cell(r, 1).value for r in range(5, 10)]
    check("grid's 5 comps at the top of the client export",
          top5 == [c["Property Name"] for c in sel_comps_ordered], str(top5))
    client_schema = ["Property Name", "Property Address", "City", "State",
                     "ZIP", "Units", "Year Built", "Avg Unit SF", "Sale Date", "Sold Price",
                     "$/Unit", "Distance (mi)"]
    got = [ws.cell(4, j).value for j in range(1, len(client_schema) + 1)]
    check("client export columns match the agreed schema (no Comp/Rank)",
          got == client_schema,
          str([f"{a}!={b}" for a, b in zip(got, client_schema) if a != b]))
    n_rows = 0
    while ws.cell(5 + n_rows, 1).value:
        n_rows += 1
    check("client list capped (≤20 comps; trimmed to 15 when over)", n_rows <= 20,
          f"{n_rows} rows")
    import statistics as _st
    s_ = sel["subject"]
    cap_ = sel["cap_rate"].get("current_avg_cap")
    adj_ppus = []
    for c in sel_comps_ordered:
        ppu = c["Sold Price"] / c["Unit Count"]
        tot = ((c["Year Built"] - s_["year_built"]) / -100 / 2
               - ((c["Avg Unit SF"] or s_["avg_size"]) / s_["avg_size"] - 1) / 4
               + (-(c.get("CapRateDriftBps", 0) / 10000) / cap_ if cap_ else 0))
        adj_ppus.append(ppu * (1 + tot))
    ind_ = sum(adj_ppus) / len(adj_ppus)
    band = _st.pstdev([c["Sold Price"] / c["Unit Count"] for c in sel_comps_ordered])
    extra_ppus = [ws.cell(r, 11).value for r in range(10, 5 + n_rows)]
    check("every non-grid comp shown is within 1 SD of the grid's indicated $/unit",
          all(abs(p - ind_) <= band + 1 for p in extra_ppus if p is not None),
          f"ind={ind_:,.0f} band={band:,.0f} outliers="
          f"{[f'{p:,.0f}' for p in extra_ppus if p and abs(p-ind_) > band + 1]}")

    uwb = openpyxl.load_workbook(uw)
    uws = uwb["Underwriting Sale Data"]
    uw_schema = ["Property Name", "Property Address", "City", "State", "ZIP", "Unit Count",
                 "Year Built", "Sold Price", "Sold Price/Unit", "Sale Date", "Building SF",
                 "Column1", "Avg Unit SF", "Info Source", "Latitude", "Longitude",
                 "Lat1_Rad", "Lon1_Rad", "Distance (mi.)", "DistancePoints", "AgeSpread",
                 "AgePoints", "DaysSinceSale", "DatePoints", "TotalPoints"]
    got = [uws.cell(1, j).value for j in range(1, len(uw_schema) + 1)]
    check("underwriting columns match the legacy Output Analysis Data schema",
          got == uw_schema, str([f"{a}!={b}" for a, b in zip(got, uw_schema) if a != b]))
    check("underwriting file carries every scored comp",
          uws.max_row - 1 == len(sel["comps"]), f"{uws.max_row-1} vs {len(sel['comps'])}")
    check("Lat1_Rad / Lon1_Rad are live formulas",
          str(uws.cell(2, 17).value).startswith("=O2") and str(uws.cell(2, 18).value).startswith("=P2"),
          f"{uws.cell(2, 17).value} / {uws.cell(2, 18).value}")

    # LibreOffice recalc: convert Comp Grid to csv and compare indicated value
    from shutil import which
    py_rows_ppu = None
    sel_comps = [c for c in sel["comps"] if c["Rank"] in sel["selected_ranks"]]
    if which("soffice"):
        with tempfile.TemporaryDirectory() as td:
            # csv conversion evaluates formulas; sheet 2 via filter options
            r = subprocess.run(
                ["soffice", "--headless", "--convert-to",
                 "csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,false,false,2",
                 xlsx, "--outdir", td], capture_output=True, timeout=300)
            csvs = glob.glob(os.path.join(td, "*.csv"))
            ok = r.returncode == 0 and csvs
            check("LibreOffice recalculates the Comp Grid", bool(ok), r.stderr.decode()[:200])
            if ok:
                import csv as csvmod
                rows = list(csvmod.reader(open(csvs[0])))
                flat = ["|".join(r) for r in rows]
                errs = [x for x in flat if "#REF" in x or "#NAME" in x or "#VALUE" in x
                        or "#DIV" in x or "Err:" in x]
                check("zero formula errors in recalculated grid", not errs, str(errs[:3]))
                ind = next((r for r in rows if r and "Indicated Value/Unit" in r[0]), None)
                got = float(ind[1].replace("$", "").replace(",", "")) if ind and ind[1] else None
                exp_rows = []
                s = sel["subject"]
                cap = sel["cap_rate"].get("current_avg_cap")
                for c in sel_comps:
                    ppu = c["Sold Price"] / c["Unit Count"]
                    tot = ((c["Year Built"] - s["year_built"]) / -100 / 2
                           - ((c["Avg Unit SF"] or s["avg_size"]) / s["avg_size"] - 1) / 4
                           + (-(c.get("CapRateDriftBps", 0) / 10000) / cap if cap else 0))
                    exp_rows.append(ppu * (1 + tot))
                expected = sum(exp_rows) / len(exp_rows)
                check("grid formulas reproduce the Python-computed indicated value",
                      got is not None and abs(got - expected) < 1.5,
                      f"grid={got}, python={expected:.0f}")
    else:
        print("  [skip] LibreOffice not installed — formula recalc not verified")

    # pdf
    try:
        from pypdf import PdfReader
        rd = PdfReader(pdf)
        text = rd.pages[0].extract_text() or ""
        check("PDF is a single page", len(rd.pages) == 1, f"{len(rd.pages)} pages")
        check("PDF carries the deal name and indicated value",
              name in text and "Indicated Value" in text, text[:120])
        check("PDF shows all 5 comps", all(f"Comp {i}" in text for i in range(1, 6)))
    except Exception as e:
        check("PDF opens with pypdf", False, str(e))

    print()
    if FAILED:
        print(f"{len(FAILED)} check(s) FAILED — fix before delivering.")
        sys.exit(1)
    print("All checks passed.")


if __name__ == "__main__":
    main()