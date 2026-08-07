#!/usr/bin/env python3
"""Score, filter, and select sale comps — Python-native, two input workbooks.

Inputs:
  --comps           All Sales Comps workbook (the comps universe, with lat/lon)
  --fannie-freddie  Combined Fannie and Freddie Sales Comps workbook (mostly
                    refinance comps — used ONLY for cap-rate drift analysis)

Scores every comp against the subject, trims $/unit outliers from the best 10,
and writes a selection.json consumed by export_comps.py.

The scoring replicates TMG's original Power Query logic (distance/age/date
point brackets, as since retuned by TMG) and adds a unit-count component so
all four criteria are blended. See references/scoring.md for the rationale.
"""
import argparse
import json
import math
import re
import statistics
import sys
from datetime import date, datetime, timedelta

import openpyxl

# ----------------------------------------------------------------------------
# CONFIG — tune selection behavior here (documented in references/scoring.md)
# ----------------------------------------------------------------------------
CONFIG = {
    "max_distance_mi": 175.0,          # hard filter, same as the original query
    # TMG weighting (Aug 2026): within 1 mi = full, within 3 mi = half,
    # beyond that (incl. 5+) = quarter weight
    "distance_points": [(1, 100), (3, 50), (float("inf"), 25)],
    "age_points": [(10, 40), (20, 30), (30, 20), (40, 10)],       # |yearBuilt - subject|
    # TMG weighting (Aug 2026): sold within 1 yr = full, within 2 yrs = half,
    # older = quarter. Recency is enforced by the hard 3-year cutoff below, not
    # by outsized points (TMG tried 100-pt recency — it dragged in far-away
    # comps; distance should stay the dominant criterion)
    "date_points": [(365, 30), (730, 15), (1095, 8)],             # days since sale
    "max_days_since_sale": 1095,       # hard filter: no comps older than 3 years
    "unit_points": [(0.25, 30), (0.50, 20), (1.00, 10)],          # |units-subj|/subj
    "hard_unit_range": None,           # e.g. (0.4, 10.0) = subj*0.4..subj*10; None = points only
    "shortlist_size": 10,              # "best 10" pool
    "outlier_sd": 1.0,                 # trim comps > this many SDs from mean $/unit of the 10
    "final_count": 5,
    # cap-rate drift (AgencyDrift / Combined Fannie & Freddie):
    "drift_vintage_window": 10,        # subject year built ± this many years
    "drift_min_sample": 8,             # min loans per 12-month window before falling back
}

EARTH_RADIUS_MI = 3959.0


def distance_mi(lat1, lon1, lat2, lon2):
    """Spherical law of cosines — identical to the workbook's Power Query math."""
    la1, lo1, la2, lo2 = map(math.radians, (lat1, lon1, lat2, lon2))
    x = math.sin(la1) * math.sin(la2) + math.cos(la1) * math.cos(la2) * math.cos(lo2 - lo1)
    return math.acos(max(-1.0, min(1.0, x))) * EARTH_RADIUS_MI


def bracket_points(value, brackets):
    for limit, pts in brackets:
        if value < limit:
            return pts
    return 0


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(v.strip()[:10], fmt).date()
            except ValueError:
                continue
    return None


def num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").replace("$", "").strip())
        except ValueError:
            return None
    return None


def read_sheet(path, title=None):
    """Read a sheet into dicts. title=None -> the first sheet; a missing title
    also falls back to the first sheet (the standalone workbooks have one)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if title is None or title not in wb.sheetnames:
        title = wb.sheetnames[0]
    ws = wb[title]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    out = [dict(zip(header, r)) for r in rows]
    wb.close()
    return out


def cap_rate_windows(rows, subject_state, subject_year, cfg, today):
    """Average cap rate in trailing 12-month origination windows, filtered on
    State + Vintage per TMG's instruction; falls back to wider samples when thin."""
    vw = cfg["drift_vintage_window"]

    def sample(state_filter, year_window):
        out = []
        for r in rows:
            cap, orig, built = num(r.get("Cap Rate")), to_date(r.get("Origination")), num(r.get("Built"))
            if cap is None or orig is None or not (0 < cap < 0.25):
                continue
            if state_filter and str(r.get("State", "")).strip().upper() != subject_state:
                continue
            if year_window is not None and built is not None and abs(built - subject_year) > year_window:
                continue
            out.append((orig, cap))
        return out

    attempts = [
        ("state+vintage", True, vw),
        ("state only", True, None),
        ("national+vintage", False, vw),
        ("national", False, None),
    ]
    for label, st, yw in attempts:
        data = sample(st, yw)
        windows = {}
        ok = True
        for w in range(4):  # w=0: last 365d, w=1: 365-730d ago, ...
            lo, hi = today - timedelta(days=365 * (w + 1)), today - timedelta(days=365 * w)
            if w == 0:
                hi = date.max  # fold future-dated originations (data quirks) into "current"
            caps = [c for (d, c) in data if lo < d <= hi]
            if len(caps) < cfg["drift_min_sample"]:
                ok = False
                break
            windows[w] = {"avg_cap": statistics.mean(caps), "n": len(caps)}
        if ok:
            return label, windows
    return None, None


def drift_bps(days_since_sale, windows):
    """Basis points of cap-rate movement since the comp's sale window
    (current-year avg minus the sale-period avg), rounded to 10s like the
    original ROUND(...,-1) formula. Positive = cap rates have risen."""
    if windows is None or days_since_sale is None or days_since_sale <= 365:
        return 0
    w = 1 if days_since_sale <= 730 else (2 if days_since_sale <= 1095 else 3)
    return int(round((windows[0]["avg_cap"] - windows[w]["avg_cap"]) * 10000, -1))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--comps", required=True, help="All Sales Comps .xlsx (comps universe)")
    ap.add_argument("--fannie-freddie", required=True,
                    help="Combined Fannie and Freddie Sales Comps .xlsx (cap-rate drift only)")
    ap.add_argument("--name", required=True)
    ap.add_argument("--address", required=True)
    ap.add_argument("--city", required=True)
    ap.add_argument("--state", required=True)
    ap.add_argument("--zip", required=True)
    ap.add_argument("--year-built", type=int, required=True)
    ap.add_argument("--units", type=int, required=True)
    ap.add_argument("--avg-size", type=float, required=True)
    ap.add_argument("--lat", type=float, required=True)
    ap.add_argument("--lon", type=float, required=True)
    ap.add_argument("--comps-sheet", default=None,
                    help="sheet name in the comps workbook (default: first sheet)")
    ap.add_argument("--outlier-sd", type=float, default=CONFIG["outlier_sd"])
    ap.add_argument("--max-days-since-sale", type=int,
                    default=CONFIG["max_days_since_sale"],
                    help="hard filter: exclude comps sold more than N days ago "
                         "(default 1095 = 3 years; pass a huge number to disable)")
    ap.add_argument("--max-distance", type=float, default=CONFIG["max_distance_mi"],
                    help="hard filter: exclude comps farther than this (miles)")
    ap.add_argument("--out", default="selection.json")
    args = ap.parse_args()
    cfg = dict(CONFIG, outlier_sd=args.outlier_sd, max_distance_mi=args.max_distance,
               max_days_since_sale=args.max_days_since_sale)
    today = date.today()
    subject_state = args.state.strip().upper()

    # ---- comps universe -----------------------------------------------------
    raw = read_sheet(args.comps, args.comps_sheet)
    comps, skipped = [], {"no_coords": 0, "no_price": 0, "too_far": 0, "no_units": 0,
                          "too_old": 0}
    self_sales = []
    subj_addr_num = re.match(r"\S+", args.address.strip())
    subj_addr_num = subj_addr_num.group(0).lower() if subj_addr_num else ""
    for r in raw:
        lat, lon = num(r.get("Latitude")), num(r.get("Longitude"))
        price, units = num(r.get("Sold Price")), num(r.get("Unit Count"))
        if lat is None or lon is None:
            skipped["no_coords"] += 1
            continue
        if not price:
            skipped["no_price"] += 1
            continue
        if not units:
            skipped["no_units"] += 1
            continue
        d = distance_mi(lat, lon, args.lat, args.lon)
        if d >= cfg["max_distance_mi"]:
            skipped["too_far"] += 1
            continue
        # the subject's own past trade must never appear as its own comp —
        # match on street number + ZIP (the workbook's duplicate key), or a
        # near-zero distance with a matching unit count
        addr_num = re.match(r"\S+", str(r.get("Property Address") or "").strip())
        addr_num = addr_num.group(0).lower() if addr_num else "?"
        same_addr = addr_num == subj_addr_num and str(r.get("ZIP") or "").strip() == str(args.zip).strip()
        near_twin = d < 0.05 and abs(units - args.units) <= max(2, 0.1 * args.units)
        if same_addr or near_twin:
            self_sales.append({"Property Name": r.get("Property Name"),
                               "Sale Date": str(r.get("Sale Date"))[:10],
                               "Sold Price": price, "Sold Price/Unit": round(price / units)})
            continue
        sd_pre = to_date(r.get("Sale Date"))
        if cfg["max_days_since_sale"] is not None and (
                sd_pre is None or (today - sd_pre).days > cfg["max_days_since_sale"]):
            skipped["too_old"] += 1
            continue
        if cfg["hard_unit_range"]:
            lo, hi = cfg["hard_unit_range"]
            if not (args.units * lo <= units <= args.units * hi):
                continue
        yb = num(r.get("Year Built"))
        sd = to_date(r.get("Sale Date"))
        ppu = num(r.get("Sold Price/Unit")) or price / units
        bsf = num(r.get("Building SF"))
        avg_sf = num(r.get("Avg Unit SF")) or (round(bsf / units) if bsf else None)
        days = (today - sd).days if sd else None
        age_spread = abs(yb - args.year_built) if yb else None
        unit_dev = abs(units - args.units) / args.units
        pts = {
            "DistancePoints": bracket_points(d, cfg["distance_points"]),
            "AgePoints": bracket_points(age_spread, cfg["age_points"]) if age_spread is not None else 0,
            "DatePoints": bracket_points(days, cfg["date_points"]) if days is not None else 0,
            "UnitPoints": bracket_points(unit_dev, cfg["unit_points"]),
        }
        comps.append({
            "Property Name": r.get("Property Name"),
            "Property Address": r.get("Property Address"),
            "City": r.get("City"), "State": r.get("State"), "ZIP": str(r.get("ZIP") or ""),
            "Unit Count": int(units), "Year Built": int(yb) if yb else None,
            "Sold Price": price, "Sold Price/Unit": round(ppu),
            "Sale Date": sd.isoformat() if sd else None,
            "Building SF": bsf,
            "Column1": r.get("Column1"),
            "Avg Unit SF": avg_sf,
            "Info Source": r.get("Info Source"),
            "Latitude": lat, "Longitude": lon,
            "Lat1_Rad": round(math.radians(lat), 10),
            "Lon1_Rad": round(math.radians(lon), 10),
            "Distance (mi.)": round(d, 4), "AgeSpread": age_spread, "DaysSinceSale": days,
            **pts, "TotalPoints": sum(pts.values()),
            # a comp with no unit-size data would get a bogus size adjustment in
            # the grid (its formula divides by avg SF) — keep it in the table but
            # never in the final 5
            "eligible": bool(avg_sf and yb and sd),
        })

    comps.sort(key=lambda c: (-c["TotalPoints"], c["Distance (mi.)"], c["DaysSinceSale"] or 9e9))
    for i, c in enumerate(comps):
        c["Rank"] = i + 1
    if len(comps) < cfg["final_count"]:
        sys.exit(f"ERROR: only {len(comps)} usable comps after filtering — check the geocode and inputs.")

    # ---- best 10 -> outlier trim -> top 5 ----------------------------------
    eligible = [c for c in comps if c["eligible"]]
    shortlist = eligible[: cfg["shortlist_size"]]
    ppus = [c["Sold Price/Unit"] for c in shortlist]
    mean, sd_ = statistics.mean(ppus), statistics.pstdev(ppus)
    limit = cfg["outlier_sd"] * sd_
    trimmed = [c for c in shortlist if sd_ > 0 and abs(c["Sold Price/Unit"] - mean) > limit]
    trimmed_ranks = {c["Rank"] for c in trimmed}
    survivors = [c for c in shortlist if c["Rank"] not in trimmed_ranks]
    selected = survivors[: cfg["final_count"]]
    backfilled = []
    for c in eligible[cfg["shortlist_size"]:]:
        if len(selected) >= cfg["final_count"]:
            break
        if abs(c["Sold Price/Unit"] - mean) <= limit:
            selected.append(c)
            backfilled.append(c["Rank"])

    # ---- cap-rate drift (Combined Fannie & Freddie, filter State + Vintage) --
    drift_rows = read_sheet(args.fannie_freddie)
    drift_scope, windows = cap_rate_windows(drift_rows, subject_state, args.year_built, cfg, today)
    current_cap = windows[0]["avg_cap"] if windows else None
    for c in comps:
        c["CapRateDriftBps"] = drift_bps(c["DaysSinceSale"], windows)

    result = {
        "generated": today.isoformat(),
        "subject": {
            "name": args.name, "address": args.address, "city": args.city,
            "state": subject_state, "zip": args.zip, "year_built": args.year_built,
            "units": args.units, "avg_size": args.avg_size,
            "lat": args.lat, "lon": args.lon,
        },
        "config": {k: v for k, v in cfg.items()},
        "universe": {"rows_in_tab": len(raw), "scored": len(comps), "skipped": skipped},
        "outlier_trim": {
            "shortlist_ranks": [c["Rank"] for c in shortlist],
            "mean_ppu": round(mean), "sd_ppu": round(sd_),
            "trimmed": [{"Rank": c["Rank"], "Property Name": c["Property Name"],
                         "Sold Price/Unit": c["Sold Price/Unit"]} for c in trimmed],
            "backfilled_ranks": backfilled,
        },
        "cap_rate": {
            "scope": drift_scope,
            "current_avg_cap": current_cap,
            "windows": {f"yr{w}": v for w, v in (windows or {}).items()},
        },
        "subject_self_sales": self_sales,
        "selected_ranks": [c["Rank"] for c in selected],
        "comps": comps,        # every scored comp, rank order (Underwriting Sale Data uses all)
        "all_scored_count": len(comps),
    }
    with open(args.out, "w") as f:
        json.dump(result, f, indent=1, default=str)

    # ---- console summary ----------------------------------------------------
    print(f"Scored {len(comps)} comps (skipped: {skipped}); drift scope: {drift_scope}, "
          f"current avg cap: {current_cap:.2%}" if current_cap else
          f"Scored {len(comps)} comps (skipped: {skipped}); NO cap-rate data — drift=0")
    print(f"Shortlist mean $/unit {mean:,.0f}, SD {sd_:,.0f}; trimmed "
          f"{[c['Property Name'] for c in trimmed] or 'none'}")
    print(f"\n{'#':>3} {'Property':<38} {'Mi':>6} {'Units':>5} {'Built':>5} "
          f"{'Sold':>10} {'$/Unit':>8} {'Pts':>4} {'Drift':>5}")
    for c in selected:
        print(f"{c['Rank']:>3} {str(c['Property Name'])[:38]:<38} {c['Distance (mi.)']:>6.2f} "
              f"{c['Unit Count']:>5} {c['Year Built'] or '?':>5} {c['Sale Date'] or '?':>10} "
              f"{c['Sold Price/Unit']:>8,} {c['TotalPoints']:>4} {c['CapRateDriftBps']:>5}")
    if self_sales:
        print(f"\nNOTE: excluded the subject's own trade(s) from the comp set — mention this "
              f"to the user, it matters: {self_sales}")
    print(f"\nSelection written to {args.out}")


if __name__ == "__main__":
    main()