#!/usr/bin/env python3
"""Build the two deliverables from a selection.json (Python-native, no
external workbook links):

1. "<Deal> - Sales Comp Export.xlsx"
   - 'Comp Export' sheet: every scored comp, the selected best 5 at the top
     (marked Comp 1-5), then the rest by rank, with score components and
     trim/eligibility notes.
   - 'Comp Grid' sheet: the TMG Subject-vs-Comp-1..5 adjustment grid, built
     with live in-sheet formulas (adjustments, indicated value).
2. "<Deal> - Comparable Sale Grid.pdf"
   - TMG-branded portrait page matching the traditional Excel PDF export:
     logo, gold property band, navy "Comparable Sale Analysis" bar, sectioned
     grid (address block / pricing block / adjustments), gold indicated-value
     block, footnotes, and a navy "Map" bar with a coordinate map of the
     subject and comps (drawn offline from the lat/lons — no web tiles).

Brand per the BOV styling guide (tmg-plugins repo): navy #1B3E6F (dark
#16345E), gold #FDB714, pale gold #FDEFD2, ink #333B45; Carlito font; navy
text on gold fills, never white.

Run verify_exports.py afterwards — it recalcs the grid via LibreOffice and
checks the PDF, and must pass before delivery.
"""
import argparse
import json
import math
import os
import tempfile
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1B3E6F"
NAVY_DARK = "16345E"
GOLD = "FDB714"
PALE_GOLD = "FDEFD2"
INK = "333B45"
GREY = "F2F2F2"

HERE = os.path.dirname(os.path.abspath(__file__))
LOGO = os.path.join(HERE, "..", "assets", "tmg_logo_navy.png")

MONEY = '$#,##0'
PCT = '0.0%;(0.0%);"-"'
DATE_FMT = 'm/d/yyyy'

COMP_COLORS = ["#C0392B", "#2E86AB", "#7D3C98", "#1E8449", "#B9770E"]


def load(sel_path):
    with open(sel_path) as f:
        sel = json.load(f)
    comps = sel["comps"]
    selected = [c for c in comps if c["Rank"] in sel["selected_ranks"]]
    selected.sort(key=lambda c: sel["selected_ranks"].index(c["Rank"]))
    return sel, comps, selected


def adj_numbers(sel, selected):
    """Python-side adjustment math — must mirror the Comp Grid formulas."""
    s = sel["subject"]
    cap = sel["cap_rate"].get("current_avg_cap")
    rows = []
    for c in selected:
        ppu = c["Sold Price"] / c["Unit Count"]
        year_adj = (c["Year Built"] - s["year_built"]) / -100 / 2
        size_adj = -((c["Avg Unit SF"] or s["avg_size"]) / s["avg_size"] - 1) / 4
        drift = c.get("CapRateDriftBps", 0)
        drift_adj = -(drift / 10000) / cap if cap else 0.0
        total = year_adj + size_adj + drift_adj
        rows.append({"comp": c, "ppu": ppu, "year_adj": year_adj, "size_adj": size_adj,
                     "drift_bps": drift, "drift_adj": drift_adj, "total_adj": total,
                     "adj_ppu": ppu * (1 + total)})
    ind_ppu = sum(r["adj_ppu"] for r in rows) / len(rows)
    return rows, ind_ppu, ind_ppu * s["units"], cap


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def styles():
    thin = Side(style="thin", color="BFBFBF")
    return {
        "title": Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        "hdr": Font(name="Arial", size=9, bold=True, color="FFFFFF"),
        "base": Font(name="Arial", size=9, color=INK),
        "bold": Font(name="Arial", size=9, bold=True, color=INK),
        "goldtxt": Font(name="Arial", size=9, bold=True, color=NAVY),  # navy on gold
        "navy": PatternFill("solid", fgColor=NAVY),
        "gold": PatternFill("solid", fgColor=GOLD),
        "pale": PatternFill("solid", fgColor=PALE_GOLD),
        "grey": PatternFill("solid", fgColor=GREY),
        "box": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def build_export_sheet(ws, sel, comps, selected, rows, ind_ppu, st):
    import statistics
    s = sel["subject"]
    sel_ranks = set(sel["selected_ranks"])

    ws.merge_cells("A1:L1")
    ws["A1"] = f"{s['name']} — Sale Comp Export"
    ws["A1"].font, ws["A1"].fill = st["title"], st["navy"]
    ws.merge_cells("A2:L2")
    ws["A2"] = (f"Subject: {s['address']}, {s['city']}, {s['state']} {s['zip']}  |  "
                f"{s['units']} units  |  built {s['year_built']}  |  avg {s['avg_size']:.0f} SF  |  "
                f"geocode {s['lat']:.5f}, {s['lon']:.5f}  |  generated {sel['generated']}")
    ws["A2"].font = st["base"]

    # Client-facing schema — no scoring internals, no Comp/Rank columns
    headers = ["Property Name", "Property Address", "City", "State", "ZIP",
               "Units", "Year Built", "Avg Unit SF", "Sale Date", "Sold Price", "$/Unit",
               "Distance (mi)"]
    HRow = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(HRow, j, h)
        c.font, c.fill, c.border = st["hdr"], st["navy"], st["box"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Beyond the grid's 5 comps, only show comps priced consistently with the
    # value we're quoting: within 1 SD (of the grid comps' $/unit) of the
    # grid's indicated value/unit. Then cap the list: if more than 20 rows
    # survive, keep the top 15 by rank.
    grid_ppus = [r_["ppu"] for r_ in rows]
    sd = statistics.pstdev(grid_ppus) or statistics.pstdev(
        [c["Sold Price/Unit"] for c in comps[:20]])
    others = [c for c in comps if c["Rank"] not in sel_ranks
              and abs(c["Sold Price/Unit"] - ind_ppu) <= sd]
    pool = selected + others
    if len(pool) > 20:
        pool = pool[:15]
    screened = len(comps) - len(sel_ranks) - len(others)
    print(f"client export: {len(pool)} comps shown "
          f"({screened} screened out as >1 SD from indicated ${ind_ppu:,.0f}/unit, "
          f"band ±${sd:,.0f})")

    r = HRow
    for c in pool:
        r += 1
        is_sel = c["Rank"] in sel_ranks
        vals = [c["Property Name"], c["Property Address"], c["City"], c["State"], c["ZIP"],
                c["Unit Count"], c["Year Built"], c["Avg Unit SF"],
                datetime.fromisoformat(c["Sale Date"]) if c["Sale Date"] else None,
                c["Sold Price"], c["Sold Price/Unit"], c["Distance (mi.)"]]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.font = st["goldtxt"] if is_sel and j == 1 else st["base"]
            cell.border = st["box"]
            if is_sel:
                cell.fill = st["pale"]
            elif (r - HRow) % 2 == 0:
                cell.fill = st["grey"]
        for col, fmt in [(8, "#,##0"), (9, DATE_FMT), (10, MONEY), (11, MONEY),
                         (12, "0.00")]:
            ws.cell(r, col).number_format = fmt
    widths = [30, 26, 14, 6, 7, 6, 9, 10, 10, 12, 10, 9]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"


UW_COLS = ["Property Name", "Property Address", "City", "State", "ZIP", "Unit Count",
           "Year Built", "Sold Price", "Sold Price/Unit", "Sale Date", "Building SF",
           "Column1", "Avg Unit SF", "Info Source", "Latitude", "Longitude", "Lat1_Rad",
           "Lon1_Rad", "Distance (mi.)", "DistancePoints", "AgeSpread", "AgePoints",
           "DaysSinceSale", "DatePoints", "TotalPoints"]


def build_underwriting_wb(sel, comps, path):
    """Underwriting Sale Data: the legacy Output Analysis Data schema, every
    scored comp in rank order, plain table (feeds the underwriting model)."""
    st = styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting Sale Data"
    for j, h in enumerate(UW_COLS, 1):
        c = ws.cell(1, j, h)
        c.font, c.fill = st["hdr"], st["navy"]
    for i, c in enumerate(comps):
        r = i + 2
        vals = [c["Property Name"], c["Property Address"], c["City"], c["State"], c["ZIP"],
                c["Unit Count"], c["Year Built"], c["Sold Price"], c["Sold Price/Unit"],
                datetime.fromisoformat(c["Sale Date"]) if c["Sale Date"] else None,
                c.get("Building SF"), c.get("Column1"), c["Avg Unit SF"], c["Info Source"],
                c["Latitude"], c["Longitude"],
                f"=O{r}/180*PI()", f"=P{r}/180*PI()",      # Lat1_Rad / Lon1_Rad, live
                c["Distance (mi.)"], c["DistancePoints"], c.get("AgeSpread"),
                c["AgePoints"], c.get("DaysSinceSale"), c["DatePoints"], c["TotalPoints"]]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v).font = st["base"]
        ws.cell(r, 10).number_format = DATE_FMT
        for col, fmt in [(15, "0.000000"), (16, "0.000000"), (17, "0.00000000"),
                         (18, "0.00000000"), (19, "0.00")]:
            ws.cell(r, col).number_format = fmt
    for j, h in enumerate(UW_COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(11, min(28, len(h) + 4))
    ws.freeze_panes = "A2"
    wb.save(path)


def build_grid_sheet(ws, sel, rows, cap, st):
    s = sel["subject"]
    n = len(rows)
    cols = [get_column_letter(3 + i) for i in range(n)]  # C..G comps; B subject

    ws.merge_cells("A1:%s1" % cols[-1])
    ws["A1"] = f"{s['name']} — Comparable Sale Analysis"
    ws["A1"].font, ws["A1"].fill = st["title"], st["navy"]

    labels = ["", "Property Name", "Address", "City", "State", "Distance (mi.)", "Unit Count",
              "Year Built", "Avg Unit SF", "Sale Date", "Sale Price", "Sale Price/Unit",
              "Year Built Adj.", "Unit Size Adj.", "Cap Rate Drift (bps)", "Cap Drift Adj.",
              "Total Adjustments", "Adjusted Sale Price/Unit"]
    base = 3
    for i, lab in enumerate(labels):
        ws.cell(base + i, 1, lab).font = st["bold"]
    ws.cell(base, 2, "Subject")
    for i in range(n):
        ws.cell(base, 3 + i, f"Comp {i+1}")
    for j in range(2, 3 + n):
        c = ws.cell(base, j)
        c.font, c.fill = st["hdr"], st["navy"]
        c.alignment = Alignment(horizontal="center")

    subj_vals = {4: s["name"], 5: s["address"], 6: s["city"], 7: s["state"], 9: s["units"],
                 10: s["year_built"], 11: s["avg_size"]}
    for r_, v in subj_vals.items():
        c = ws.cell(r_, 2, v)
        c.font, c.fill = st["base"], st["pale"]

    for i, row in enumerate(rows):
        c = row["comp"]
        L = cols[i]
        data = {4: c["Property Name"], 5: c["Property Address"], 6: c["City"], 7: c["State"],
                8: c["Distance (mi.)"], 9: c["Unit Count"], 10: c["Year Built"],
                11: c["Avg Unit SF"],
                12: datetime.fromisoformat(c["Sale Date"]) if c["Sale Date"] else None,
                13: c["Sold Price"], 17: row["drift_bps"]}
        for r_, v in data.items():
            ws[f"{L}{r_}"] = v
        # live formulas — the exported grid stays auditable and re-computable
        ws[f"{L}14"] = f"=IFERROR({L}13/{L}9,\"\")"
        ws[f"{L}15"] = f"=({L}10-$B$10)/-100/2"
        ws[f"{L}16"] = f"=-({L}11/$B$11-1)/4"
        ws[f"{L}18"] = f"=IF($B$22=\"\",0,-({L}17/10000)/$B$22)"
        ws[f"{L}19"] = f"=SUM({L}15:{L}16)+{L}18"
        ws[f"{L}20"] = f"={L}14*(1+{L}19)"
        for r_, fmt in [(8, "0.00"), (12, DATE_FMT), (13, MONEY), (14, MONEY),
                        (15, PCT), (16, PCT), (18, PCT), (19, PCT), (20, MONEY)]:
            ws[f"{L}{r_}"].number_format = fmt
        for r_ in range(4, 21):
            ws[f"{L}{r_}"].font = st["base"]

    ws.cell(20, 1, "Adjusted Sale Price/Unit").font = st["bold"]
    ws.cell(22, 1, "Current Avg Cap Rate").font = st["bold"]
    ws["B22"] = cap if cap else ""
    ws["B22"].number_format = "0.00%"
    for r_, formula in [(23, f"=AVERAGE(C20:{cols[-1]}20)"), (24, "=B23*$B$9")]:
        lab = "Subject Indicated Value/Unit" if r_ == 23 else "Subject Indicated Total Value"
        ws.cell(r_, 1, lab).font = st["goldtxt"]
        ws.cell(r_, 1).fill = st["gold"]
        cell = ws.cell(r_, 2, formula)
        cell.number_format = MONEY
        cell.fill, cell.font = st["gold"], st["goldtxt"]
    ws.cell(26, 1, ("Notes: comps scored on distance, vintage, unit count and sale date; "
                    "$/unit outliers >%.1f SD from the best-10 mean trimmed. Cap-rate drift: "
                    "Fannie/Freddie originations, scope = %s." %
                    (sel["config"]["outlier_sd"], sel["cap_rate"]["scope"]))).font = st["base"]
    ws.column_dimensions["A"].width = 26
    for j in range(2, 3 + n):
        ws.column_dimensions[get_column_letter(j)].width = 22


def build_xlsx(sel, comps, selected, rows, ind_ppu, cap, path):
    st = styles()
    wb = openpyxl.Workbook()
    build_export_sheet(wb.active, sel, comps, selected, rows, ind_ppu, st)
    wb.active.title = "Comp Export"
    ws = wb.create_sheet("Comp Grid")
    build_grid_sheet(ws, sel, rows, cap, st)
    wb.save(path)


# ---------------------------------------------------------------------------
# Map. Two renderers:
#   1. _map_tiles — real street basemap (CartoDB Positron via contextily).
#      Needs `pip install contextily` and open internet. Cloud-sandbox package
#      mirrors exclude tile libraries and egress blocks tile servers, so this
#      path activates automatically only in unrestricted environments.
#   2. _map_schematic — offline fallback: to-scale coordinate plot with
#      distance rings. Always works.
# Set TMG_MAP_TILES=off to force the schematic.
# ---------------------------------------------------------------------------
def make_map(sel, selected, path):
    mode = os.environ.get("TMG_MAP_TILES", "auto").lower()
    if mode != "off":
        try:
            return _map_tiles(sel, selected, path)
        except Exception as e:
            print(f"note: tile basemap unavailable ({e.__class__.__name__}: {e}) "
                  "— falling back to schematic map")
    return _map_schematic(sel, selected, path)


def _legend_handles(sel, pts_comps):
    from matplotlib.lines import Line2D
    s = sel["subject"]
    handles = [Line2D([], [], marker="*", ls="", markersize=13, color="#" + GOLD,
                      markeredgecolor="#" + NAVY, label=f"Subject — {s['name']}")]
    for i, c in enumerate(pts_comps):
        handles.append(Line2D([], [], marker="o", ls="", markersize=8,
                              color=COMP_COLORS[i % len(COMP_COLORS)],
                              label=f"Comp {i+1} — {c['Property Name']} ({c['Distance (mi.)']:.1f} mi)"))
    return handles


def _map_tiles(sel, selected, path):
    """Street basemap via contextily (CartoDB Positron — light grey, sits well
    under the navy/gold brand). Raises if contextily or the network is absent;
    make_map() falls back to the schematic."""
    import contextily as ctx
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    R = 6378137.0

    def merc(lat, lon):
        return (math.radians(lon) * R,
                math.log(math.tan(math.pi / 4 + math.radians(lat) / 2)) * R)

    s = sel["subject"]
    x0, y0 = merc(s["lat"], s["lon"])
    k = 1.0 / math.cos(math.radians(s["lat"]))     # mercator meters per true meter
    pts = [(*merc(c["Latitude"], c["Longitude"]), c) for c in selected]
    rmax_m = max(2 * 1609.34 * k,
                 max(math.hypot(x - x0, y - y0) for x, y, _ in pts) * 1.25)

    fig, ax = plt.subplots(figsize=(9.6, 4.1), dpi=150)
    ax.scatter([x0], [y0], marker="*", s=420, color="#" + GOLD,
               edgecolors="#" + NAVY, linewidths=1.4, zorder=5)
    for i, (x, y, c) in enumerate(pts):
        ax.scatter([x], [y], s=150, color=COMP_COLORS[i % len(COMP_COLORS)],
                   edgecolors="white", linewidths=1.2, zorder=5)
        ax.annotate(str(i + 1), (x, y), fontsize=8, fontweight="bold",
                    color="white", ha="center", va="center", zorder=6)
    leg = ax.legend(handles=_legend_handles(sel, [c for _, _, c in pts]),
                    loc="upper left", fontsize=7.5, framealpha=0.95,
                    edgecolor="#B9C2D0", borderpad=0.8)
    leg.set_zorder(7)
    ax.set_xlim(x0 - rmax_m * 1.85, x0 + rmax_m * 1.3)
    ax.set_ylim(y0 - rmax_m * 1.08, y0 + rmax_m * 1.08)
    ax.set_aspect("equal")
    ax.set_xticks([]), ax.set_yticks([])
    ctx.add_basemap(ax, source=ctx.providers.CartoDB.Positron,
                    attribution="© OpenStreetMap contributors © CARTO")
    fig.tight_layout(pad=0.4)
    fig.savefig(path, facecolor="white")
    plt.close(fig)


def _map_schematic(sel, selected, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D

    s = sel["subject"]
    lat0, lon0 = s["lat"], s["lon"]
    kx = math.cos(math.radians(lat0)) * 69.172  # mi per degree lon
    ky = 69.055                                  # mi per degree lat

    pts = [((c["Longitude"] - lon0) * kx, (c["Latitude"] - lat0) * ky, c)
           for c in selected]
    rmax = max(2.0, max(math.hypot(x, y) for x, y, _ in pts) * 1.25)

    fig, ax = plt.subplots(figsize=(9.6, 4.1), dpi=150)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#F7F8FA")
    rings = [r for r in (1, 2, 3, 5, 10, 25, 50) if r <= rmax * 1.05] or [1]
    for r in rings:
        ax.add_patch(plt.Circle((0, 0), r, fill=False, color="#B9C2D0",
                                lw=0.8, ls=(0, (4, 3)), zorder=1))
        ax.annotate(f"{r} mi", (r * 0.7071, r * 0.7071), fontsize=7,
                    color="#8A94A4", ha="left", va="bottom", zorder=2)

    ax.scatter([0], [0], marker="*", s=420, color="#" + GOLD,
               edgecolors="#" + NAVY, linewidths=1.4, zorder=5)
    ax.annotate("  " + s["name"], (0, 0), fontsize=8.5, fontweight="bold",
                color="#" + NAVY, ha="left", va="center", zorder=6)
    for i, (x, y, c) in enumerate(pts):
        col = COMP_COLORS[i % len(COMP_COLORS)]
        ax.scatter([x], [y], s=150, color=col, edgecolors="white",
                   linewidths=1.2, zorder=5)
        ax.annotate(str(i + 1), (x, y), fontsize=8, fontweight="bold",
                    color="white", ha="center", va="center", zorder=6)

    handles = [Line2D([], [], marker="*", ls="", markersize=13, color="#" + GOLD,
                      markeredgecolor="#" + NAVY, label=f"Subject — {s['name']}")]
    for i, (_, _, c) in enumerate(pts):
        handles.append(Line2D([], [], marker="o", ls="", markersize=8,
                              color=COMP_COLORS[i % len(COMP_COLORS)],
                              label=f"Comp {i+1} — {c['Property Name']} ({c['Distance (mi.)']:.1f} mi)"))
    leg = ax.legend(handles=handles, loc="upper left", fontsize=7.5,
                    framealpha=0.95, edgecolor="#B9C2D0", borderpad=0.8)
    leg.set_zorder(7)

    lim = rmax
    ax.set_xlim(-lim * 1.85, lim * 1.3)  # extra room on the left for the legend
    ax.set_ylim(-lim * 1.08, lim * 1.08)
    ax.set_aspect("equal")
    ax.set_xticks([]), ax.set_yticks([])
    for sp in ax.spines.values():
        sp.set_color("#B9C2D0")
    ax.annotate("N", (0.975, 0.94), xycoords="axes fraction", fontsize=10,
                fontweight="bold", color="#" + NAVY, ha="center")
    ax.annotate("▲", (0.975, 0.885), xycoords="axes fraction", fontsize=9,
                color="#" + NAVY, ha="center")
    fig.tight_layout(pad=0.4)
    fig.savefig(path, facecolor="white")
    plt.close(fig)


# ---------------------------------------------------------------------------
# PDF — matches TMG's traditional Excel export layout, BOV brand
# ---------------------------------------------------------------------------
def build_pdf(sel, rows, ind_ppu, ind_total, cap, path):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (Image, KeepTogether, Paragraph,
                                    SimpleDocTemplate, Spacer, Table, TableStyle)

    font, font_b, font_i = "Helvetica", "Helvetica-Bold", "Helvetica-Oblique"
    try:
        pdfmetrics.registerFont(TTFont("Carlito", "/usr/share/fonts/truetype/crosextra/Carlito-Regular.ttf"))
        pdfmetrics.registerFont(TTFont("Carlito-Bold", "/usr/share/fonts/truetype/crosextra/Carlito-Bold.ttf"))
        pdfmetrics.registerFont(TTFont("Carlito-Italic", "/usr/share/fonts/truetype/crosextra/Carlito-Italic.ttf"))
        font, font_b, font_i = "Carlito", "Carlito-Bold", "Carlito-Italic"
    except Exception:
        pass

    navy = colors.HexColor("#" + NAVY)
    navy_dark = colors.HexColor("#" + NAVY_DARK)
    gold = colors.HexColor("#" + GOLD)
    pale = colors.HexColor("#" + PALE_GOLD)
    ink = colors.HexColor("#" + INK)
    zebra = colors.HexColor("#EFF1F5")
    s = sel["subject"]

    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.55 * inch, rightMargin=0.55 * inch,
                            topMargin=0.45 * inch, bottomMargin=0.4 * inch,
                            title=f"{s['name']} - Comparable Sale Grid")
    PW = letter[0] - 1.1 * inch

    note = ParagraphStyle("n", fontName=font, fontSize=6.8, textColor=ink, leading=8.6)
    notes_hdr = ParagraphStyle("nh", fontName=font_b, fontSize=7.5, textColor=ink)

    def money(v):
        return "$%s" % f"{v:,.0f}"

    def pct(v):
        if abs(v) < 0.0005:
            return "0%"
        return f"{v*100:.0f}%" if abs(round(v*100, 1) % 1) < 0.05 else f"{v*100:.1f}%"

    def dt(iso):
        d = date.fromisoformat(iso[:10])
        return f"{d.month}/{d.day}/{d.year}"

    cs = [r["comp"] for r in rows]
    SP = ""  # spacer-row marker

    grid_rows = [
        ["", "Subject"] + [f"Comp {i+1}" for i in range(len(rows))],
        ["Property Name", s["name"]] + [c["Property Name"] for c in cs],
        [SP] * (len(rows) + 2),
        ["Address", s["address"]] + [c["Property Address"] for c in cs],
        ["City", s["city"]] + [c["City"] for c in cs],
        ["State", s["state"]] + [c["State"] for c in cs],
        ["Zip", s["zip"]] + [c["ZIP"] for c in cs],
        ["Distance (mi.)", ""] + [f"{c['Distance (mi.)']:.2f}" for c in cs],
        ["Primary Data Source", ""] + [c["Info Source"] or "" for c in cs],
        ["Sale Price", ""] + [money(c["Sold Price"]) for c in cs],
        ["Sale Date", ""] + [dt(c["Sale Date"]) for c in cs],
        ["Unit Count", f"{s['units']:,}"] + [f"{c['Unit Count']:,}" for c in cs],
        ["Sale Price/Unit", ""] + [money(r["ppu"]) for r in rows],
        [SP] * (len(rows) + 2),
        ["Adjustments", ""] + [""] * len(rows),
        ["Year Built¹", s["year_built"]] + [c["Year Built"] for c in cs],
        ["    Adj.", ""] + [pct(r["year_adj"]) for r in rows],
        ["Average Unit Size²", f"{s['avg_size']:,.0f}"] + [f"{(c['Avg Unit SF'] or 0):,.0f}" for c in cs],
        ["    Adj.", ""] + [pct(r["size_adj"]) for r in rows],
        ["Cap Rate Drift (basis points)³", ""] + [f"{r['drift_bps']:.2f}" for r in rows],
        ["    Adj.", ""] + [pct(r["drift_adj"]) for r in rows],
        [SP] * (len(rows) + 2),
        ["Total Adjustments (%)", ""] + [f"{r['total_adj']*100:.1f}%" for r in rows],
        ["Adj./Unit ($/Unit * Adj %)", ""] + [money(r["adj_ppu"] - r["ppu"]) for r in rows],
        ["Adjusted Sale Price/Unit", ""] + [money(r["adj_ppu"]) for r in rows],
    ]
    IDX_HDR, IDX_NAME, IDX_ADJ_LBL = 0, 1, 14
    spacer_rows = [2, 13, 21]

    widths = [1.55 * inch] + [(PW - 1.55 * inch) / (len(rows) + 1)] * (len(rows) + 1)
    heights = [13] + [11] * (len(grid_rows) - 1)
    for i in spacer_rows:
        heights[i] = 6
    t = Table(grid_rows, colWidths=widths, rowHeights=heights)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), navy_dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), font_b, 8),
        ("FONT", (0, 1), (0, -1), font, 7.8),
        ("FONT", (1, 1), (-1, -1), font, 7.8),
        ("FONT", (1, 1), (1, -1), font_i, 7.8),          # subject column italic
        ("FONT", (0, IDX_ADJ_LBL), (0, IDX_ADJ_LBL), font_b, 7.8),
        ("FONT", (1, IDX_NAME), (-1, IDX_NAME), font_b, 7.8),
        ("TEXTCOLOR", (0, 1), (-1, -1), ink),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (1, 1), (1, -1), pale),           # subject column pale gold
        ("LINEBELOW", (0, 0), (-1, 0), 1, gold),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]
    # bold the Adj. result rows like the reference, zebra the data blocks
    for i in (16, 18, 20, 22, 23, 24):
        style.append(("FONT", (2, i), (-1, i), font_b, 7.8))
    for i in (4, 6, 9, 11, 15, 17, 19):
        style.append(("BACKGROUND", (2, i), (-1, i), zebra))
    for i in spacer_rows:
        style += [("BACKGROUND", (0, i), (-1, i), colors.white),
                  ("LINEABOVE", (0, i), (-1, i), 0, colors.white)]
    t.setStyle(TableStyle(style))

    ind = Table([["Subject Indicated Value/Unit", money(ind_ppu)],
                 ["Subject Indicated Total Value", money(ind_total)]],
                colWidths=[1.9 * inch, 1.0 * inch], rowHeights=[13, 13])
    ind.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), gold),
        ("TEXTCOLOR", (0, 0), (-1, -1), navy),           # navy on gold, never white
        ("FONT", (0, 0), (-1, -1), font_b, 8),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    ind_wrap = Table([[ind, ""]], colWidths=[2.9 * inch, PW - 2.9 * inch])
    ind_wrap.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0),
                                  ("TOPPADDING", (0, 0), (-1, -1), 0),
                                  ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))

    cap_txt = f"{cap*100:.2f}%" if cap else "n/a"
    trimmed = sel["outlier_trim"]["trimmed"]
    trim_txt = ("; trimmed as $/unit outliers: " +
                ", ".join(t["Property Name"] for t in trimmed)) if trimmed else ""
    notes = [
        Paragraph("Notes:", notes_hdr),
        Paragraph("Comparable product was selected based on overall similarity to the subject "
                  "property, comparable market area/linkages, vintage, and date of sale — scored "
                  "on blended distance, vintage, unit count and sale date; sales older than 3 "
                  f"years excluded{trim_txt}.", note),
        Paragraph("Each of the comparables included are given equal weighting for the purposes "
                  "of this analysis.", note),
        Paragraph("¹ Newer product typically trades for a higher $/unit. A 0.5% adjustment "
                  "is applied to account for differences in property age.", note),
        Paragraph("² Larger sized units typically rent for a higher amount & properties "
                  "with a larger average unit size generate higher per unit returns — an "
                  "adjustment is applied for a fourth of the percentage spread between the "
                  "average unit sizes.", note),
        Paragraph("³ Per unit pricing is normalized for historic sales to account for "
                  "differences between past and current cap rates, obtained from available "
                  f"Fannie/Freddie records (current avg cap rate {cap_txt}, scope: "
                  f"{sel['cap_rate']['scope']}).", note),
    ]

    def band(text, bg, fg, fnt, size, h):
        b = Table([[text]], colWidths=[PW], rowHeights=[h])
        b.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), bg),
                               ("TEXTCOLOR", (0, 0), (-1, -1), fg),
                               ("FONT", (0, 0), (-1, -1), fnt, size),
                               ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                               ("LEFTPADDING", (0, 0), (-1, -1), 6)]))
        return b

    story = []
    if os.path.exists(LOGO):
        from PIL import Image as PILImage
        w, h = PILImage.open(LOGO).size
        lw = 2.1 * inch
        logo = Image(LOGO, width=lw, height=lw * h / w)
        logo.hAlign = "LEFT"
        story += [logo, Spacer(1, 8)]
    story.append(band(s["name"], gold, navy, font_b, 12, 20))
    story.append(Spacer(1, 4))
    story.append(band("Comparable Sale Analysis", navy_dark, colors.white, font_b, 8.5, 15))
    story.append(t)
    story.append(Spacer(1, 8))
    story.append(ind_wrap)
    story.append(Spacer(1, 8))
    story += notes

    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tf:
        map_png = tf.name
    try:
        make_map(sel, cs, map_png)
        from PIL import Image as PILImage
        w, h = PILImage.open(map_png).size
        story.append(Spacer(1, 8))
        story.append(KeepTogether([
            band("Map", navy_dark, colors.white, font_b, 8.5, 15),
            Spacer(1, 2),
            Image(map_png, width=PW, height=PW * h / w),
        ]))
    except Exception as e:
        print(f"WARNING: map generation failed ({e}) — PDF built without map")

    doc.build(story)
    if os.path.exists(map_png):
        os.unlink(map_png)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--selection", required=True)
    ap.add_argument("--outdir", default="output")
    args = ap.parse_args()
    sel, comps, selected = load(args.selection)
    rows, ind_ppu, ind_total, cap = adj_numbers(sel, selected)
    os.makedirs(args.outdir, exist_ok=True)
    name = sel["subject"]["name"]
    xlsx = os.path.join(args.outdir, f"{name} - Sale Comp Export.xlsx")
    uw = os.path.join(args.outdir, f"{name} - Underwriting Sale Data.xlsx")
    pdf = os.path.join(args.outdir, f"{name} - Comparable Sale Grid.pdf")
    build_xlsx(sel, comps, selected, rows, ind_ppu, cap, xlsx)
    build_underwriting_wb(sel, comps, uw)
    build_pdf(sel, rows, ind_ppu, ind_total, cap, pdf)
    print(f"Wrote {xlsx} (client-facing)")
    print(f"Wrote {uw} ({len(comps)} scored comps)")
    print(f"Wrote {pdf}")
    print(f"Indicated value: ${ind_ppu:,.0f}/unit  |  ${ind_total:,.0f} total "
          f"({sel['subject']['units']} units)")


if __name__ == "__main__":
    main()