#!/usr/bin/env python3
"""Query the Estimated Loan Terms workbook for current multifamily debt terms.

The workbook (Loan Terms tab) is the source of truth for spreads and structure.
Rates are computed here as: index yield + spread(bps)/10000.

Index yields come from, in priority order:
  1. --yields JSON passed on the command line (fresh web values, in percent)
  2. cached formula results in the workbook (col G, "RATE USED")
  3. the workbook's manual override cells (col F)

Usage:
  python query_terms.py --workbook "Estimated Loan Terms - Multifamily Debt.xlsx" \
      [--program "bridge"] \
      [--yields '{"UST 5 Yr": 4.33, "UST 7 Yr": 4.47, "UST 10 Yr": 4.63, "UST 30 Yr": 5.17, "30-Day Avg SOFR": 3.62}'] \
      [--as-of "08/05/2026"] [--json] [--set-overrides]

  --program        substring filter on program name (case-insensitive); omit for all
  --yields         index yields in PERCENT (4.63 = 4.63%); keys must match the
                   Index names on the Treasury Yields tab
  --set-overrides  also write the supplied yields into the workbook's Manual
                   Override cells (col F) and the as-of column, then save, so
                   the workbook stays current for use in Excel
  --json           machine-readable output
"""
import argparse
import json
import sys

import openpyxl

TREASURY_SHEET = "Treasury Yields"
TERMS_SHEET = "Loan Terms"
T_FIRST = 4          # first index row on Treasury Yields
L_HEADER = 3         # header row on Loan Terms


def load_indices(path, cli_yields):
    """Return {index_name: (rate_decimal, source, as_of)}."""
    wb_v = openpyxl.load_workbook(path, data_only=True)   # cached values
    wb_f = openpyxl.load_workbook(path)                   # formulas/literals
    tv, tf = wb_v[TREASURY_SHEET], wb_f[TREASURY_SHEET]
    out = {}
    r = T_FIRST
    while tf.cell(r, 1).value:
        name = str(tf.cell(r, 1).value).strip()
        as_of = tf.cell(r, 8).value
        if cli_yields and name in cli_yields:
            out[name] = (float(cli_yields[name]) / 100.0, "live (web)", "today")
        else:
            used = tv.cell(r, 7).value        # cached "RATE USED"
            ovr = tf.cell(r, 6).value         # override literal
            if isinstance(used, (int, float)):
                out[name] = (float(used), "workbook cached", as_of)
            elif isinstance(ovr, (int, float)):
                out[name] = (float(ovr), "workbook override", as_of)
            else:
                out[name] = (0.0, "missing", as_of)
        r += 1
    return out


def load_programs(path, indices):
    wb = openpyxl.load_workbook(path)  # literals: spreads & terms are typed values
    ws = wb[TERMS_SHEET]
    headers = [ws.cell(L_HEADER, c).value for c in range(1, 14)]
    rows = []
    r = L_HEADER + 1
    while ws.cell(r, 1).value:
        idx_name = str(ws.cell(r, 2).value).strip()
        idx_rate, src, as_of = indices.get(idx_name, (0.0, "missing", None))
        spread = ws.cell(r, 4).value or 0
        amort = ws.cell(r, 10).value
        rows.append({
            "program": ws.cell(r, 1).value,
            "index": idx_name,
            "index_rate_pct": round(idx_rate * 100, 3),
            "index_source": src,
            "index_as_of": str(as_of) if as_of else None,
            "spread_bps": spread,
            "interest_rate_pct": round((idx_rate + spread / 10000.0) * 100, 3),
            "max_ltv_pct": round((ws.cell(r, 6).value or 0) * 100, 1),
            "min_dscr": ws.cell(r, 7).value,
            "io_years": ws.cell(r, 8).value,
            "term_months": ws.cell(r, 9).value,
            "amortization_months": "Interest Only" if amort == 0 else amort,
            "rate_type": ws.cell(r, 11).value,
            "recourse": ws.cell(r, 12).value,
            "notes": ws.cell(r, 13).value,
        })
        r += 1
    return headers, rows


def set_overrides(path, cli_yields, as_of):
    wb = openpyxl.load_workbook(path)
    ws = wb[TREASURY_SHEET]
    r, changed = T_FIRST, []
    while ws.cell(r, 1).value:
        name = str(ws.cell(r, 1).value).strip()
        if name in cli_yields:
            ws.cell(r, 6).value = float(cli_yields[name]) / 100.0
            if as_of:
                ws.cell(r, 8).value = as_of
            changed.append(name)
        r += 1
    wb.save(path)
    return changed


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", required=True)
    ap.add_argument("--program", default=None)
    ap.add_argument("--yields", default=None, help="JSON dict of index name -> percent")
    ap.add_argument("--as-of", default=None, help="label for the supplied yields, e.g. 08/05/2026")
    ap.add_argument("--set-overrides", action="store_true")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    cli_yields = json.loads(args.yields) if args.yields else {}
    if args.set_overrides and cli_yields:
        changed = set_overrides(args.workbook, cli_yields, args.as_of)
        print(f"Updated workbook overrides for: {', '.join(changed)}", file=sys.stderr)

    indices = load_indices(args.workbook, cli_yields)
    _, rows = load_programs(args.workbook, indices)

    if args.program:
        tokens = args.program.lower().split()
        rows = [x for x in rows
                if all(t in (x["program"] + " " + (x["notes"] or "")).lower()
                       for t in tokens)]
        if not rows:
            print(f"No program matching '{args.program}'. Available:", file=sys.stderr)
            indices2 = load_indices(args.workbook, cli_yields)
            for x in load_programs(args.workbook, indices2)[1]:
                print("  -", x["program"], file=sys.stderr)
            sys.exit(1)

    if args.json:
        print(json.dumps(rows, indent=2, default=str))
    else:
        for x in rows:
            print(f"\n{x['program']}")
            print(f"  Interest Rate : {x['interest_rate_pct']:.2f}%  "
                  f"({x['index']} {x['index_rate_pct']:.2f}% + {x['spread_bps']} bps)"
                  f"  [{x['index_source']}, as of {x['index_as_of']}]")
            print(f"  Max LTV       : {x['max_ltv_pct']:.0f}%")
            print(f"  Min DSCR      : {x['min_dscr']}x")
            print(f"  Interest Only : {x['io_years']} yrs")
            print(f"  Term          : {x['term_months']} months")
            print(f"  Amortization  : {x['amortization_months']}")
            print(f"  Rate Type     : {x['rate_type']} | Recourse: {x['recourse']}")
            print(f"  Notes         : {x['notes']}")


if __name__ == "__main__":
    main()