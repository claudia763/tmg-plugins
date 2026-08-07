#!/usr/bin/env python3
"""loan_terms.py — TMG Estimated Loan Terms: fetch live index rates, build the
Excel workbook from scratch, and render a Loan Terms PDF. Pure Python
(requests + openpyxl + reportlab) — no LibreOffice, no Excel WEBSERVICE, safe
to run headless on a Linux server.

Inputs : loan_terms_config.json (same directory) — lender programs (spreads,
         LTV, DSCR, IO, term, amort) and fallback index rates.
Outputs: "Estimated Loan Terms - Multifamily Debt.xlsx" (4 tabs, dropdown
         lookup) and "Loan Terms.pdf" (one-page landscape terms sheet).

Rate sources are tried in order until every index is filled (the script runs
on a German-IP server, so EU-reachable mirrors back up the US sources):

  UST 5/7/10/30:  treasury.gov XML  ->  FRED CSV  ->  Stooq (PL/EU) CSV
                  ->  Yahoo Finance chart API  ->  finanzen.net (DE, scrape)
                  ->  config fallback
  30-Day SOFR:    NY Fed markets API  ->  FRED CSV  ->  config fallback

(Dukascopy is deliberately absent: its free feeds carry bond *futures prices*,
not yields, and converting price->yield needs curve math we don't want here.)

A 7-Yr gap is linearly interpolated from 5+10 when only it is missing.
Successful fetches are written back into the config as the new fallbacks.

Usage:
  python loan_terms.py                        # fetch, build xlsx + pdf
  python loan_terms.py --program "bridge"     # ...and print a quote card
  python loan_terms.py --offline              # skip web, use fallbacks
  python loan_terms.py --json                 # machine-readable summary
  python loan_terms.py --outdir /tmp/out --no-save
"""
import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path

import requests

UA = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) TMG-loan-terms/2.0"}
TIMEOUT = 12
UST_KEYS = {"UST 5 Yr": "5", "UST 7 Yr": "7", "UST 10 Yr": "10", "UST 30 Yr": "30"}
SOFR_KEY = "30-Day Avg SOFR"

def _sane(v):
    """Accept a plausible percent yield; fix 10x-scaled quotes (e.g. CBOE 46.3)."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return None
    if 15 < v < 100:
        v = v / 10.0
    return v if 0.05 <= v <= 15 else None

def log(msg):
    print(f"  [rates] {msg}", file=sys.stderr)

# ----------------------------------------------------------------- fetchers
def fetch_treasury_gov(sess):
    """Official daily par yield curve, XML month feed (works from EU IPs)."""
    import xml.etree.ElementTree as ET
    out = {}
    today = dt.date.today()
    months = [today.strftime("%Y%m"), (today.replace(day=1) - dt.timedelta(days=1)).strftime("%Y%m")]
    tags = {"BC_5YEAR": "UST 5 Yr", "BC_7YEAR": "UST 7 Yr", "BC_10YEAR": "UST 10 Yr", "BC_30YEAR": "UST 30 Yr"}
    for month in months:
        url = ("https://home.treasury.gov/resource-center/data-chart-center/interest-rates/pages/xml"
               f"?data=daily_treasury_yield_curve&field_tdr_date_value_month={month}")
        r = sess.get(url, timeout=TIMEOUT)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        latest, latest_date = {}, None
        for props in root.iter():
            if props.tag.endswith("}properties"):
                row, row_date = {}, None
                for child in props:
                    name = child.tag.split("}")[-1]
                    if name == "NEW_DATE" and child.text:
                        row_date = child.text[:10]
                    elif name in tags:
                        v = _sane(child.text)
                        if v is not None:
                            row[tags[name]] = v
                if row and row_date and (latest_date is None or row_date > latest_date):
                    latest, latest_date = row, row_date
        if latest:
            for k, v in latest.items():
                out[k] = (v, latest_date)
            return out
    raise RuntimeError("no rows in treasury.gov feed")

def _fred_last(sess, series):
    start = (dt.date.today() - dt.timedelta(days=30)).isoformat()
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series}&cosd={start}"
    r = sess.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    for line in reversed(r.text.strip().splitlines()):
        parts = line.split(",")
        if len(parts) == 2:
            v = _sane(parts[1])
            if v is not None:
                return v, parts[0]
    raise RuntimeError(f"no numeric rows in FRED {series}")

def fetch_fred_ust(sess):
    series = {"UST 5 Yr": "DGS5", "UST 7 Yr": "DGS7", "UST 10 Yr": "DGS10", "UST 30 Yr": "DGS30"}
    return {k: _fred_last(sess, s) for k, s in series.items()}

def fetch_stooq(sess):
    """Stooq (Poland/EU) end-of-day US yield indices. No 7-Yr symbol."""
    syms = {"UST 5 Yr": "5usy.b", "UST 10 Yr": "10usy.b", "UST 30 Yr": "30usy.b"}
    out = {}
    for k, s in syms.items():
        r = sess.get(f"https://stooq.com/q/l/?s={s}&f=sd2t2ohlcv&h&e=csv", timeout=TIMEOUT)
        r.raise_for_status()
        lines = r.text.strip().splitlines()
        if len(lines) >= 2:
            f = lines[1].split(",")           # Symbol,Date,Time,Open,High,Low,Close,Volume
            if len(f) >= 7:
                v = _sane(f[6])
                if v is not None:
                    out[k] = (v, f[1])
    if not out:
        raise RuntimeError("stooq returned no usable rows")
    return out

def fetch_yahoo(sess):
    """Yahoo chart API: ^FVX/^TNX/^TYX are CBOE yield indices (5/10/30-Yr)."""
    syms = {"UST 5 Yr": "%5EFVX", "UST 10 Yr": "%5ETNX", "UST 30 Yr": "%5ETYX"}
    out = {}
    for k, s in syms.items():
        r = sess.get(f"https://query1.finance.yahoo.com/v8/finance/chart/{s}?range=5d&interval=1d",
                     timeout=TIMEOUT)
        r.raise_for_status()
        meta = r.json()["chart"]["result"][0]["meta"]
        v = _sane(meta.get("regularMarketPrice"))
        if v is not None:
            ts = meta.get("regularMarketTime")
            d = dt.datetime.utcfromtimestamp(ts).date().isoformat() if ts else "recent"
            out[k] = (v, d)
    if not out:
        raise RuntimeError("yahoo returned no usable quotes")
    return out

def fetch_finanzen_net(sess):
    """Last-resort German source: scrape finanzen.net's US-Staatsanleihen page.
    Tolerant regex (German decimal commas); values are sanity-checked, so a
    layout change degrades to 'nothing found' rather than a bad number."""
    r = sess.get("https://www.finanzen.net/zinsen/us-staatsanleihen", timeout=TIMEOUT)
    r.raise_for_status()
    text = re.sub(r"<[^>]+>", " ", r.text)
    out = {}
    for key, label in {"UST 5 Yr": "5", "UST 7 Yr": "7", "UST 10 Yr": "10", "UST 30 Yr": "30"}.items():
        m = re.search(rf"\b{label}\s*Jahre?\b(.{{0,120}}?)(\d{{1,2}}[.,]\d{{1,4}})\s*%", text, re.S)
        if m:
            v = _sane(m.group(2).replace(",", "."))
            if v is not None:
                out[key] = (v, dt.date.today().isoformat())
    if not out:
        raise RuntimeError("finanzen.net: no yields recognized (layout change?)")
    return out

def fetch_nyfed_sofr(sess):
    """Official NY Fed SOFR Averages & Index API."""
    r = sess.get("https://markets.newyorkfed.org/api/rates/secured/sofrai/last/1.json",
                 timeout=TIMEOUT)
    r.raise_for_status()
    row = r.json()["refRates"][0]
    v = _sane(row.get("average30day"))
    if v is None:
        raise RuntimeError("NY Fed payload missing average30day")
    return {SOFR_KEY: (v, row.get("effectiveDate", "recent"))}

def fetch_fred_sofr(sess):
    v, d = _fred_last(sess, "SOFR30DAYAVG")
    return {SOFR_KEY: (v, d)}

UST_TREE = [("treasury.gov", fetch_treasury_gov), ("FRED", fetch_fred_ust),
            ("stooq.com (EU)", fetch_stooq), ("Yahoo Finance", fetch_yahoo),
            ("finanzen.net (DE)", fetch_finanzen_net)]
SOFR_TREE = [("NY Fed", fetch_nyfed_sofr), ("FRED", fetch_fred_sofr)]

def resolve_rates(config, offline=False):
    """Walk the source trees; return {index: {pct, as_of, source}} for all indices."""
    rates = {}
    if not offline:
        sess = requests.Session()
        sess.headers.update(UA)
        for wanted, tree in ((set(UST_KEYS), UST_TREE), ({SOFR_KEY}, SOFR_TREE)):
            for name, fn in tree:
                missing = wanted - set(rates)
                if not missing:
                    break
                try:
                    got = fn(sess)
                except Exception as e:
                    log(f"{name}: {type(e).__name__}: {e}")
                    continue
                for k in missing & set(got):
                    v, d = got[k]
                    rates[k] = {"pct": round(v, 4), "as_of": d, "source": name}
                log(f"{name}: filled {sorted(missing & set(got))or '(nothing new)'}")
    # interpolate a lone 7-Yr gap: linear between 5 and 10
    if "UST 7 Yr" not in rates and {"UST 5 Yr", "UST 10 Yr"} <= set(rates):
        y5, y10 = rates["UST 5 Yr"]["pct"], rates["UST 10 Yr"]["pct"]
        rates["UST 7 Yr"] = {"pct": round(y5 + 0.4 * (y10 - y5), 4),
                             "as_of": rates["UST 10 Yr"]["as_of"],
                             "source": "interpolated 5/10-Yr"}
    for k, meta in config["indices"].items():
        if k not in rates:
            rates[k] = {"pct": meta["fallback_pct"], "as_of": meta.get("as_of") or "n/a",
                        "source": "config fallback" if k != "Fixed (no index)" else "n/a"}
    return rates

# ------------------------------------------------------------------- outputs
NAVY, GOLD, LIGHT = "1F3352", "C9A227", "F2F4F8"

def rate_of(p, rates):
    return rates[p["index"]]["pct"] / 100.0 + p["spread_bps"] / 10000.0

def build_workbook(config, rates, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName

    F = lambda **kw: Font(name="Arial", size=10, **kw)
    f_base, f_bold = F(), F(bold=True)
    f_title = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    f_hdr = F(bold=True, color="FFFFFF")
    f_input, f_link = F(color="0000FF"), F(color="008000")
    f_note = Font(name="Arial", size=9, italic=True, color="595959")
    fill_hdr, fill_light = PatternFill("solid", fgColor=NAVY), PatternFill("solid", fgColor=LIGHT)
    fill_yellow = PatternFill("solid", fgColor="FFFF00")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    programs = config["programs"]

    wb = openpyxl.Workbook()

    # ---- Treasury Yields (fetched values, no formulas) ----
    ws = wb.active
    ws.title = "Treasury Yields"
    ws["A1"] = "INDEX RATES — U.S. TREASURY & SOFR"
    ws["A1"].font = f_title; ws["A1"].fill = fill_hdr
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height = 22
    ws["A2"] = ("Values fetched by loan_terms.py at build time. Re-run the script to refresh; "
                "the Loan Terms tab recalculates from this table when opened in Excel.")
    ws["A2"].font = f_note; ws.merge_cells("A2:D2")
    for c, h in enumerate(["Index", "Rate", "Source", "As-Of"], 1):
        cell = ws.cell(3, c, h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
    T0 = 4
    idx_names = list(config["indices"])
    for i, k in enumerate(idx_names):
        r = T0 + i
        ws.cell(r, 1, k).font = f_bold
        rc = ws.cell(r, 2, rates[k]["pct"] / 100.0)
        rc.font = f_input; rc.number_format = "0.00%"
        ws.cell(r, 3, rates[k]["source"]).font = f_note
        ws.cell(r, 4, str(rates[k]["as_of"])).font = f_note
        for c in range(1, 5):
            ws.cell(r, c).border = border
            if i % 2:
                ws.cell(r, c).fill = fill_light
    T1 = T0 + len(idx_names) - 1
    ws.cell(T1 + 2, 1, ("Generated " + dt.datetime.now().strftime("%Y-%m-%d %H:%M") +
                        ". Sources tried in order: treasury.gov, FRED, Stooq, Yahoo, finanzen.net "
                        "(UST); NY Fed, FRED (SOFR).")).font = f_note
    for col, w in zip("ABCD", [18, 10, 22, 12]):
        ws.column_dimensions[col].width = w

    # ---- Loan Terms ----
    wl = wb.create_sheet("Loan Terms")
    wl["A1"] = "ESTIMATED LOAN TERMS — MULTIFAMILY DEBT (LOOKUP TABLE)"
    wl["A1"].font = f_title; wl["A1"].fill = fill_hdr
    wl.merge_cells("A1:M1"); wl.row_dimensions[1].height = 22
    wl["A2"] = ("Blue = assumptions from loan_terms_config.json (edit there and re-run, or edit here "
                "for a one-off). Interest Rate = Index Rate + Spread. Amortization 0 = full-term IO.")
    wl["A2"].font = f_note; wl.merge_cells("A2:M2")
    heads = ["Loan Program", "Index", "Index Rate", "Spread (bps)", "Interest Rate", "Max LTV",
             "Min DSCR", "Interest Only (Yrs)", "Loan Term (Months)", "Amortization (Months)",
             "Rate Type", "Recourse", "Notes"]
    for c, h in enumerate(heads, 1):
        cell = wl.cell(3, c, h); cell.font = f_hdr; cell.fill = fill_hdr; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wl.row_dimensions[3].height = 30
    L0 = 4
    for i, p in enumerate(programs):
        r = L0 + i
        wl.cell(r, 1, p["program"]).font = f_bold
        wl.cell(r, 2, p["index"]).font = f_input
        wl.cell(r, 3, (f"=INDEX('Treasury Yields'!$B${T0}:$B${T1},"
                       f"MATCH(B{r},'Treasury Yields'!$A${T0}:$A${T1},0))")).font = f_link
        wl.cell(r, 4, p["spread_bps"]).font = f_input
        wl.cell(r, 5, f"=C{r}+D{r}/10000").font = f_bold
        wl.cell(r, 6, p["max_ltv"]).font = f_input
        wl.cell(r, 7, p["min_dscr"]).font = f_input
        wl.cell(r, 8, p["io_years"]).font = f_input
        wl.cell(r, 9, p["term_months"]).font = f_input
        wl.cell(r, 10, p["amort_months"]).font = f_input
        wl.cell(r, 11, p["rate_type"]).font = f_input
        wl.cell(r, 12, p["recourse"]).font = f_input
        wl.cell(r, 13, p["notes"]).font = f_note
        for c in range(1, 14):
            wl.cell(r, c).border = border
            if i % 2:
                wl.cell(r, c).fill = fill_light
        for c, fmt in ((3, "0.00%"), (4, "#,##0"), (5, "0.00%"), (6, "0%"), (7, '0.00"x"'),
                       (8, "#,##0"), (9, "#,##0"), (10, "#,##0")):
            wl.cell(r, c).number_format = fmt
        wl.cell(r, 13).alignment = Alignment(wrap_text=True, vertical="top")
    L1 = L0 + len(programs) - 1
    wl.cell(L1 + 2, 1, ("Spreads/terms are broker-maintained estimates — not lender quotes. Index "
                        "rates on the Treasury Yields tab; rate build-up: index + spread.")).font = f_note
    wl.merge_cells(start_row=L1 + 2, start_column=1, end_row=L1 + 2, end_column=13)
    for ci, w in enumerate([26, 16, 10, 12, 10, 9, 9, 11, 11, 13, 10, 20, 46], 1):
        wl.column_dimensions[get_column_letter(ci)].width = w
    wl.freeze_panes = "B4"

    # ---- Loan Assumptions (dropdown) ----
    wa = wb.create_sheet("Loan Assumptions")
    wa["B2"] = "LOAN ASSUMPTIONS — NEW DEBT (FREE & CLEAR)"
    wa["B2"].font = f_title; wa["B2"].fill = fill_hdr
    wa.merge_cells("B2:D2"); wa.row_dimensions[2].height = 24
    wa["B3"] = "Select a loan program from the dropdown (yellow cell)."
    wa["B3"].font = f_note; wa.merge_cells("B3:D3")
    wa["B5"] = "Loan Program"; wa["B5"].font = f_bold
    wa["C5"] = programs[0]["program"]
    wa["C5"].font = f_input; wa["C5"].fill = fill_yellow; wa["C5"].border = border
    wa.merge_cells("C5:D5")
    match = f"MATCH($C$5,'Loan Terms'!$A${L0}:$A${L1},0)"
    lk = lambda col: f"=INDEX('Loan Terms'!${col}${L0}:${col}${L1},{match})"
    rows = [("LTV", lk("F"), "0%"), ("Interest Rate", lk("E"), "0.00%"),
            ("DSCR Requirement (reference)", lk("G"), '0.00"x"'),
            ("Interest Only (Years)", lk("H"), "#,##0"), ("Loan Term (Months)", lk("I"), "#,##0"),
            ("Amortization (Months)",
             f"=IF(INDEX('Loan Terms'!$J${L0}:$J${L1},{match})=0,\"Interest Only\","
             f"INDEX('Loan Terms'!$J${L0}:$J${L1},{match}))", "#,##0"),
            ("", None, None),
            ("Index", lk("B"), None), ("Index Rate", lk("C"), "0.00%"),
            ("Spread (bps)", lk("D"), "#,##0"), ("Rate Type", lk("K"), None),
            ("Recourse", lk("L"), None), ("Notes", lk("M"), None)]
    r = 7
    for label, formula, fmt in rows:
        if not label:
            r += 1
            continue
        lc = wa.cell(r, 2, label); lc.font = f_bold if r < 13 else f_base
        lc.border = border; lc.fill = fill_light
        vc = wa.cell(r, 3, formula); vc.font = f_link; vc.border = border
        vc.alignment = Alignment(horizontal="right", wrap_text=(label == "Notes"))
        if fmt:
            vc.number_format = fmt
        wa.cell(r, 4).border = border
        if label == "Notes":
            wa.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            wa.row_dimensions[r].height = 44
            vc.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
        r += 1
    wa.cell(r + 1, 2, "Estimates only — not a quote. Verify spreads with lenders.").font = f_note
    for col, w in zip("ABCD", [2, 28, 24, 24]):
        wa.column_dimensions[col].width = w
    dv = DataValidation(type="list", formula1="LenderList", allow_blank=False, showDropDown=False)
    wa.add_data_validation(dv); dv.add("C5")

    for name, ref in [("LenderList", f"'Loan Terms'!$A${L0}:$A${L1}"),
                      ("LoanTermsTable", f"'Loan Terms'!$A${L0}:$M${L1}")]:
        try:
            wb.defined_names.add(DefinedName(name, attr_text=ref))
        except AttributeError:
            wb.defined_names[name] = DefinedName(name, attr_text=ref)

    wb.move_sheet("Loan Assumptions", offset=-2)
    wb.move_sheet("Loan Terms", offset=-1)
    wb.save(path)
    return path

def build_pdf(config, rates, path):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle

    navy, gold = colors.HexColor("#1F3352"), colors.HexColor("#C9A227")
    light = colors.HexColor("#F2F4F8")
    programs = config["programs"]
    today = dt.date.today().strftime("%B %d, %Y")

    doc = SimpleDocTemplate(str(path), pagesize=landscape(letter),
                            leftMargin=0.45 * inch, rightMargin=0.45 * inch,
                            topMargin=0.4 * inch, bottomMargin=0.35 * inch)
    title = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=17, textColor=navy)
    sub = ParagraphStyle("s", fontName="Helvetica", fontSize=8.5, textColor=colors.HexColor("#595959"))
    cell = ParagraphStyle("c", fontName="Helvetica", fontSize=7.2, leading=8.6)
    cellb = ParagraphStyle("cb", parent=cell, fontName="Helvetica-Bold")
    head = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=7.4, leading=8.8,
                          textColor=colors.white, alignment=1)

    ust10 = rates["UST 10 Yr"]; sofr = rates[SOFR_KEY]
    header = [Paragraph("ESTIMATED LOAN TERMS — MULTIFAMILY DEBT", title),
              Spacer(1, 3),
              Paragraph(f"As of {today} &nbsp;•&nbsp; 10-Yr UST {ust10['pct']:.2f}% "
                        f"({ust10['source']}, {ust10['as_of']}) &nbsp;•&nbsp; 30-Day SOFR "
                        f"{sofr['pct']:.2f}% ({sofr['source']}, {sofr['as_of']}) &nbsp;•&nbsp; "
                        "Rate = index + spread", sub),
              Spacer(1, 7)]

    cols = ["Loan Program", "Index", "Spread", "Rate", "Max LTV", "Min DSCR", "IO (Yrs)",
            "Term (Mo)", "Amort (Mo)", "Type", "Recourse", "Notes"]
    data = [[Paragraph(c, head) for c in cols]]
    for p in programs:
        amort = "I/O" if p["amort_months"] == 0 else str(p["amort_months"])
        data.append([
            Paragraph(p["program"], cellb), Paragraph(p["index"].replace("Fixed (no index)", "Fixed"), cell),
            Paragraph(f"{p['spread_bps']}", cell), Paragraph(f"{rate_of(p, rates)*100:.2f}%", cellb),
            Paragraph(f"{p['max_ltv']*100:.0f}%", cell), Paragraph(f"{p['min_dscr']:.2f}x", cell),
            Paragraph(str(p["io_years"]), cell), Paragraph(str(p["term_months"]), cell),
            Paragraph(amort, cell), Paragraph(p["rate_type"], cell),
            Paragraph(p["recourse"], cell), Paragraph(p["notes"], cell)])
    widths = [1.5, 0.85, 0.6, 0.55, 0.55, 0.55, 0.5, 0.55, 0.6, 0.55, 1.0, 2.3]
    table = Table(data, colWidths=[w * inch for w in widths], repeatRows=1)
    style = [("BACKGROUND", (0, 0), (-1, 0), navy),
             ("LINEBELOW", (0, 0), (-1, 0), 1.2, gold),
             ("GRID", (0, 1), (-1, -1), 0.4, colors.HexColor("#BFBFBF")),
             ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
             ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5)]
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), light))
    table.setStyle(TableStyle(style))

    footer = Paragraph("Indicative estimates for screening only — not a loan quote, rate lock, or "
                       "commitment. Spreads are broker-maintained placeholders; confirm with lenders. "
                       "HUD rates exclude MIP. Generated by TMG loan-terms-lookup.", sub)
    doc.build(header + [table, Spacer(1, 6), footer])
    return path

def print_terms(config, rates, program_filter=None, as_json=False):
    rows = []
    for p in config["programs"]:
        if program_filter:
            hay = (p["program"] + " " + p["notes"]).lower()
            if not all(t in hay for t in program_filter.lower().split()):
                continue
        rows.append({**p, "interest_rate_pct": round(rate_of(p, rates) * 100, 3),
                     "index_rate_pct": rates[p["index"]]["pct"],
                     "index_source": rates[p["index"]]["source"],
                     "index_as_of": str(rates[p["index"]]["as_of"])})
    if as_json:
        print(json.dumps({"rates": rates, "programs": rows}, indent=2))
        return bool(rows)
    for x in rows:
        amort = "Interest Only" if x["amort_months"] == 0 else f"{x['amort_months']} months"
        print(f"\n{x['program']}")
        print(f"  Interest Rate : {x['interest_rate_pct']:.2f}%  ({x['index']} "
              f"{x['index_rate_pct']:.2f}% + {x['spread_bps']} bps)  "
              f"[{x['index_source']}, as of {x['index_as_of']}]")
        print(f"  Max LTV       : {x['max_ltv']*100:.0f}%   Min DSCR: {x['min_dscr']}x   "
              f"IO: {x['io_years']} yrs")
        print(f"  Term          : {x['term_months']} mo   Amortization: {amort}")
        print(f"  Rate Type     : {x['rate_type']} | Recourse: {x['recourse']}")
        print(f"  Notes         : {x['notes']}")
    return bool(rows)

def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--config", default=None)
    ap.add_argument("--outdir", default=".")
    ap.add_argument("--basename", default="Estimated Loan Terms - Multifamily Debt")
    ap.add_argument("--program", default=None, help="print quote card(s) for matching program(s)")
    ap.add_argument("--offline", action="store_true", help="skip web fetches, use config fallbacks")
    ap.add_argument("--no-save", action="store_true", help="don't write fetched rates back to config")
    ap.add_argument("--no-files", action="store_true", help="skip xlsx/pdf output (query only)")
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()

    cfg_path = Path(args.config) if args.config else Path(__file__).parent / "loan_terms_config.json"
    config = json.loads(cfg_path.read_text(encoding="utf-8"))
    rates = resolve_rates(config, offline=args.offline)
    for k, v in rates.items():
        log(f"{k:<17} {v['pct']:.3f}%  <- {v['source']} ({v['as_of']})")

    fetched = {k: v for k, v in rates.items() if v["source"] not in ("config fallback", "n/a")
               and not v["source"].startswith("interpolated")}
    if fetched and not args.no_save:
        for k, v in fetched.items():
            config["indices"][k]["fallback_pct"] = v["pct"]
            config["indices"][k]["as_of"] = str(v["as_of"])
        cfg_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
        log(f"config fallbacks updated ({len(fetched)} indices)")

    if not args.no_files:
        outdir = Path(args.outdir); outdir.mkdir(parents=True, exist_ok=True)
        xlsx = build_workbook(config, rates, outdir / f"{args.basename}.xlsx")
        pdf = build_pdf(config, rates, outdir / "Loan Terms.pdf")
        log(f"wrote {xlsx}"); log(f"wrote {pdf}")

    ok = print_terms(config, rates, args.program, args.json)
    if args.program and not ok:
        print(f"No program matching '{args.program}'. Programs: "
              + "; ".join(p["program"] for p in config["programs"]), file=sys.stderr)
        sys.exit(1)

if __name__ == "__main__":
    main()
