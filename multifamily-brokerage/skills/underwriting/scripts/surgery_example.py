#!/usr/bin/env python3
"""Zip-level surgery: write all underwriting inputs into a pristine copy of the
original model, preserving Power Query / connections / queryTables byte-for-byte."""
import zipfile, re, shutil, datetime, warnings
from collections import defaultdict
from lxml import etree
import openpyxl
warnings.filterwarnings('ignore')

ORIG = '/root/.claude/uploads/2de58f0f-5997-594d-88ea-a35aa307b914/070090a7-Initials__Property_Name__Date__TEST.xlsx'
OUT = '/home/claude/preserved.xlsx'
NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
E = lambda tag: f'{{{NS}}}{tag}'
EPOCH = datetime.datetime(1899, 12, 30)

def col_to_num(col):
    n = 0
    for ch in col: n = n*26 + ord(ch) - 64
    return n
def num_to_col(n):
    s = ''
    while n: n, r = divmod(n-1, 26); s = chr(65+r) + s
    return s
def split_ref(ref):
    m = re.match(r'([A-Z]+)(\d+)', ref)
    return m.group(1), int(m.group(2))

class Sheet:
    def __init__(self, data):
        self.tree = etree.fromstring(data)
        sd = self.tree.find(E('sheetData'))
        self.sheetdata = sd
        self.rows = {int(r.get('r')): r for r in sd.findall(E('row'))}
    def _get_row(self, rn):
        if rn in self.rows: return self.rows[rn]
        row = etree.SubElement(self.sheetdata, E('row')); row.set('r', str(rn))
        self.rows[rn] = row
        return row
    def _sort(self):
        rows = sorted(self.sheetdata.findall(E('row')), key=lambda r: int(r.get('r')))
        for r in rows:
            self.sheetdata.remove(r); self.sheetdata.append(r)
            cells = sorted(r.findall(E('c')), key=lambda c: col_to_num(split_ref(c.get('r'))[0]))
            for c in cells: r.remove(c); r.append(c)
    def _get_cell(self, ref, create=True):
        col, rn = split_ref(ref)
        row = self._get_row(rn) if create else self.rows.get(rn)
        if row is None: return None
        for c in row.findall(E('c')):
            if c.get('r') == ref: return c
        if not create: return None
        c = etree.SubElement(row, E('c')); c.set('r', ref)
        return c
    def clear(self, ref, drop_cell=False):
        c = self._get_cell(ref, create=False)
        if c is None: return
        if drop_cell:
            c.getparent().remove(c); return
        for child in list(c): c.remove(child)
        if 't' in c.attrib: del c.attrib['t']
    def set(self, ref, value, style=None, formula=None, array_ref=None):
        c = self._get_cell(ref)
        for child in list(c): c.remove(child)
        if 't' in c.attrib: del c.attrib['t']
        if style is not None and c.get('s') is None:
            c.set('s', str(style))
        elif style is not None and c.get('s') in (None, '0'):
            c.set('s', str(style))
        if formula is not None:
            f = etree.SubElement(c, E('f'))
            if array_ref:
                f.set('t', 'array'); f.set('ref', array_ref)
            f.text = formula
            return
        if value is None: return
        if isinstance(value, str):
            c.set('t', 'inlineStr')
            is_ = etree.SubElement(c, E('is'))
            t = etree.SubElement(is_, E('t'))
            t.text = value
            if value != value.strip(): t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
        elif isinstance(value, bool):
            c.set('t', 'b')
            etree.SubElement(c, E('v')).text = '1' if value else '0'
        elif isinstance(value, datetime.datetime):
            serial = (value - EPOCH).days + (value - EPOCH).seconds/86400
            etree.SubElement(c, E('v')).text = repr(serial)
        else:
            etree.SubElement(c, E('v')).text = repr(value) if isinstance(value, float) else str(value)
    def clear_range(self, c1, r1, c2, r2, keep_formulas=False):
        for rn in range(r1, r2+1):
            row = self.rows.get(rn)
            if row is None: continue
            for c in list(row.findall(E('c'))):
                col = col_to_num(split_ref(c.get('r'))[0])
                if col_to_num(c1) <= col <= col_to_num(c2):
                    if keep_formulas and c.find(E('f')) is not None: continue
                    for child in list(c): c.remove(child)
                    if 't' in c.attrib: del c.attrib['t']
    def out(self):
        self._sort()
        return etree.tostring(self.tree, xml_declaration=True, encoding='UTF-8', standalone=True)

# ------------------------------------------------------------------ load parts
z = zipfile.ZipFile(ORIG)
parts = {}
SHEETS = {'Master':'xl/worksheets/sheet9.xml','FinalRR':'xl/worksheets/sheet11.xml',
 'FPlan':'xl/worksheets/sheet6.xml','RRdate':'xl/worksheets/sheet10.xml',
 'T12':'xl/worksheets/sheet12.xml','TRL':'xl/worksheets/sheet1.xml',
 'TPD':'xl/worksheets/sheet2.xml','RawRents':'xl/worksheets/sheet28.xml',
 'RentComp':'xl/worksheets/sheet20.xml','VA':'xl/worksheets/sheet14.xml',
 'Factors':'xl/worksheets/sheet15.xml','Assum':'xl/worksheets/sheet16.xml',
 'UWFC':'xl/worksheets/sheet17.xml','PDFFC':'xl/worksheets/sheet22.xml'}
sh = {k: Sheet(z.read(p)) for k, p in SHEETS.items()}

# ------------------------------------------------------------------ styles: append xfs
styles = z.read('xl/styles.xml').decode('utf8')
# add custom numFmt mmm-yy (id 300)
if '<numFmts' in styles:
    styles = re.sub(r'(<numFmts count=")(\d+)(">)',
                    lambda m: f'{m.group(1)}{int(m.group(2))+1}{m.group(3)}<numFmt numFmtId="300" formatCode="mmm\\-yy"/>',
                    styles, count=1)
else:
    raise SystemExit('no numFmts block')
m = re.search(r'<cellXfs count="(\d+)">', styles)
base = int(m.group(1))
XF_DATE, XF_MMMYY, XF_PCT1 = base, base+1, base+2
new_xfs = ('<xf numFmtId="14" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
           '<xf numFmtId="300" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
           '<xf numFmtId="167" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>')
styles = styles.replace(m.group(0), f'<cellXfs count="{base+3}">', 1)
# insert before </cellXfs>
styles = styles.replace('</cellXfs>', new_xfs + '</cellXfs>', 1)
parts['xl/styles.xml'] = styles.encode('utf8')

# ------------------------------------------------------------------ source data (same derivation as phase A)
rr = openpyxl.load_workbook('rr.xlsx', data_only=True)
rrs = rr['Rent Roll']
rr_rows = []
for r in range(4, rrs.max_row+1):
    unit = rrs.cell(row=r, column=1).value
    if unit in (None, '', 0): continue
    rr_rows.append([rrs.cell(row=r, column=c).value for c in range(1, 24)])
fps = rr['Floor Plan Summary']
fp_rows = [[fps.cell(row=r, column=c).value for c in range(1, 27)]
           for r in range(3, 20) if fps.cell(row=r, column=1).value not in (None, '')]

t12 = openpyxl.load_workbook('t12.xlsx', data_only=True)
st = t12['Trailing 12 Month Statement']
agg = defaultdict(lambda: [0.0]*12)
for r in range(2, st.max_row+1):
    code = st.cell(row=r, column=1).value
    if not code: continue
    vals = [st.cell(row=r, column=c).value for c in range(3, 15)]
    if all(v in (None, '') for v in vals): continue
    for i, v in enumerate(vals):
        if isinstance(v, (int, float)): agg[code][i] += v
agg['ro'] = [a+b for a, b in zip(agg['ro'], agg.get('rt', [0.0]*12))]

rd = openpyxl.load_workbook('rentdata.xlsx', data_only=True)
uld = rd['Unit-Level Data']
SUBJECT = 'The Flats at Shadowglen'
lease_rows = []
for r in range(4, uld.max_row+1):
    name = uld.cell(row=r, column=2).value
    if name in (None, '') or name == SUBJECT: continue
    lease_rows.append([uld.cell(row=r, column=c).value for c in range(2, 22)])
cut = datetime.datetime(2025, 7, 13)
recent = [x for x in lease_rows if (isinstance(x[10], datetime.datetime) and x[10] >= cut) or x[11] in (True, 'True', 'TRUE')]
by_prop = defaultdict(list)
for rec in recent: by_prop[rec[0]].append(rec)
final_leases = []
for prop in by_prop:
    lst = sorted(by_prop[prop], key=lambda x: (x[10] or datetime.datetime(2100,1,1)), reverse=True)
    final_leases.extend(lst[:138])
assert len(final_leases) == 860, len(final_leases)

rc = rd['Rent Comps']
def addr_split(full):
    if not full or full == '--': return ('--', None, None, None)
    p = [x.strip() for x in full.split(',')]
    sz = p[2].split() if len(p) > 2 else []
    return (p[0], p[1] if len(p) > 1 else None, sz[0] if sz else None,
            int(sz[1]) if len(sz) > 1 and sz[1].isdigit() else None)
prop_rows = []
for r in range(4, 15):
    name = rc.cell(row=r, column=2).value
    if name in (None, ''): continue
    prop_rows.append((name, rc.cell(row=r, column=3).value,
                      [rc.cell(row=r, column=c).value for c in range(4, 27)],
                      [rc.cell(row=r, column=c).value for c in range(27, 41)]))

# ------------------------------------------------------------------ 1. Master
M = sh['Master']
for ref, v in [('B1','Flats at Shadowglen'), ('B2','12500 Shadowglen Trace'), ('B3','Manor'),
               ('C3','TX'), ('D3',78653), ('B4','https://travis.prodigycad.com/property-detail/786700'),
               ('B5','786700'), ('B6','Travis'), ('B7','Manor ISD'), ('B9',38875783),
               ('B11',0), ('B12',1.0814), ('B13',0.375845), ('B14',0.1034), ('B15',0.118023),
               ('B16',0.772), ('B17',0.1), ('B18',2020), ('B19',530474)]:
    M.set(ref, v)
M.set('B20', datetime.datetime(2026, 4, 30), style=XF_DATE)

# ------------------------------------------------------------------ 2. FinalRR
F = sh['FinalRR']
F.clear_range('C', 9, 'Y', 658)
F.clear_range('BA', 9, 'BZ', 26)          # removes array anchors + stale zeros
F.set('C3', 'Rent Roll as of:  April 30, 2026')
DATE_COLS_RR = {19, 20, 23, 24}           # S,T,W,X within C..Y => offsets 16,17,20,21 -> cols 19,20,23,24
for i, data in enumerate(rr_rows):
    rn = 9 + i
    for j, v in enumerate(data):
        if v in (None, ''): continue
        colnum = 3 + j
        style = XF_DATE if (colnum in DATE_COLS_RR and isinstance(v, datetime.datetime)) else None
        F.set(f'{num_to_col(colnum)}{rn}', v, style=style)
for i, data in enumerate(fp_rows):
    rn = 9 + i
    for j, v in enumerate(data):
        if v in (None, ''): continue
        F.set(f'{num_to_col(53+j)}{rn}', v)

# ------------------------------------------------------------------ 3. Floor plan sheet + RR_asof_date
P = sh['FPlan']
P.clear_range('A', 2, 'Z', 19)
for i, data in enumerate(fp_rows):
    rn = 2 + i
    for j, v in enumerate(data):
        if v in (None, ''): continue
        P.set(f'{num_to_col(1+j)}{rn}', v)
sh['RRdate'].set('A2', datetime.datetime(2026, 4, 30), style=XF_DATE)

# ------------------------------------------------------------------ 4. Final_T_12
T = sh['T12']
months = [datetime.datetime(2025,2,1), datetime.datetime(2025,3,1), datetime.datetime(2025,4,1),
          datetime.datetime(2025,5,1), datetime.datetime(2025,6,1), datetime.datetime(2025,7,1),
          datetime.datetime(2025,10,1), datetime.datetime(2025,11,1), datetime.datetime(2025,12,1),
          datetime.datetime(2026,1,1), datetime.datetime(2026,2,1), datetime.datetime(2026,3,1)]
for i, mth in enumerate(months):
    T.set(f'{num_to_col(3+i)}2', mth, style=XF_MMMYY)
codemap = {'r':4,'ll':5,'v':6,'nr':7,'bd':8,'rw':9,'ro':10,'oi':11,'cs':14,'rm':15,
           'ad':16,'m':17,'pr':18,'w':19,'tr':20,'e':21,'o':22,'mf':23,'i':24,'tx':25}
for code, rn in codemap.items():
    vals = agg.get(code, [0.0]*12)
    for i, v in enumerate(vals):
        T.set(f'{num_to_col(3+i)}{rn}', round(v, 2))
    T.set(f'O{rn}', round(sum(vals), 2))

# ------------------------------------------------------------------ 5. TableRecentLeases
L = sh['TRL']
L.clear_range('A', 2, 'V', 458)
for i, rec in enumerate(final_leases):
    rn = 2 + i
    for j, v in enumerate(rec):
        if v in (None, ''): continue
        style = XF_DATE if (j in (9, 10) and isinstance(v, datetime.datetime)) else None
        L.set(f'{num_to_col(1+j)}{rn}', v, style=style)
    L.set(f'U{rn}', 1)
    L.set(f'V{rn}', rec[0])

# ------------------------------------------------------------------ 6. TablePropertyData
D = sh['TPD']
D.clear_range('A', 2, 'AQ', 14)
for i, (name, addr, v, eff) in enumerate(prop_rows):
    rn = 2 + i
    street, city, state, zc = addr_split(addr)
    D.set(f'A{rn}', name); D.set(f'B{rn}', street)
    if city:  D.set(f'C{rn}', ' ' + city)
    if state: D.set(f'D{rn}', state)
    if zc:    D.set(f'E{rn}', zc)
    for j in range(9):
        if v[j] not in (None, ''): D.set(f'{num_to_col(6+j)}{rn}', v[j])
    for j in range(12):
        if v[9+j] not in (None, ''): D.set(f'{num_to_col(15+j)}{rn}', v[9+j])
    for j in range(2):
        if v[21+j] not in (None, ''): D.set(f'{num_to_col(27+j)}{rn}', v[21+j])
    for j in range(12):
        if eff[j] not in (None, ''): D.set(f'{num_to_col(29+j)}{rn}', eff[j])
    for j in range(2):
        if eff[12+j] not in (None, ''): D.set(f'{num_to_col(41+j)}{rn}', eff[12+j])
    D.set(f'AQ{rn}', name)

# ------------------------------------------------------------------ 7. Raw Rents
R = sh['RawRents']
for c1, c2 in [('W','AA'), ('AE','AH')]:
    R.clear_range(c1, 4, c2, 999)
R.clear_range('AL', 4, 'BS', 16)
for i, rec in enumerate(final_leases):
    rn = 4 + i
    R.set(f'W{rn}', rec[0]); R.set(f'X{rn}', rec[5] if rec[5] is not None else 0)
    if rec[6] is not None: R.set(f'Y{rn}', rec[6])
    if rec[8] is not None: R.set(f'Z{rn}', rec[8])
    R.set(f'AA{rn}', 1)
    for off, idx in [(31,13),(32,14),(33,15),(34,16)]:
        if rec[idx] is not None: R.set(f'{num_to_col(off)}{rn}', rec[idx])
colmap = {38:'street',39:'name',40:'units',46:'quality',49:'city',50:'state',51:'zip',
          52:'avgsf',54:'avg',64:'leased',66:'yr'}
for i, (name, addr, v, eff) in enumerate(prop_rows):
    rn = 4 + i
    street, city, state, zc = addr_split(addr)
    data = {'street':street,'name':name,'quality':v[2] or 0,'yr':v[3] or 0,'units':v[4] or 0,
            'avgsf':v[6] or 0,'leased':v[7] or 0,'avg':v[21] or 0,
            'city':city or 0,'state':state or 0,'zip':zc or 0}
    if name == 'Comp Average':
        data = {k: 0 for k in data}; data.update({'street':'--','name':'Comp Average'})
    for cn, key in colmap.items():
        R.set(f'{num_to_col(cn)}{rn}', data[key])

# ------------------------------------------------------------------ 8. Rent comps / Value-Add / Factors selections
RC = sh['RentComp']
RC.clear_range('AK', 4, 'AK', 103)
for rn in [4, 5, 7, 9, 10]: RC.set(f'AK{rn}', 'x')
VA = sh['VA']
VA.clear_range('C', 16, 'C', 52)
for rn in [28, 33, 35, 47, 50, 51]: VA.set(f'C{rn}', 'x')
FX = sh['Factors']
FX.set('G24', 'Lease-Up Risk / Below-Market Occupancy (81% vs 94% comp avg)')
FX.set('I24', 50); FX.set('J24', 'x')
FX.set('G25', 'MUD District Tax Burden (Wilbarger Creek MUD No 1, +0.772/$100)')
FX.set('I25', 25); FX.set('J25', 'x')

# ------------------------------------------------------------------ 9. Assumptions
A = sh['Assum']
for ref, v in [('G17',0.01),('G18',0.05),('G19',0.10),('G20',0.03),('G21',0.025),
               ('G37',0.03),('G39',1.0),('G43',0.025),('G46',0.025),
               ('F48',0.25),('G48',0.25),('H48',0.25),('G50',32000000),
               ('G61',0.70),('G64',3),('G65',120)]:
    A.set(ref, v)
A.clear('G62')   # blank -> avg agency quote

# ------------------------------------------------------------------ 10. UW - F&C G69 header typo
U = sh['UWFC']
U.set('G69', None, formula='Final_T_12[date2]', array_ref='G69')

# ------------------------------------------------------------------ 11. PDF sheet: drop black-box CF, pct format on E35
PD = sh['PDFFC']
for cf in PD.tree.findall(E('conditionalFormatting')):
    if cf.get('sqref') == 'B52:J77':
        cf.getparent().remove(cf)
e35 = PD._get_cell('E35')
e35.set('s', str(XF_PCT1))

# ------------------------------------------------------------------ 12. tables + workbook calcPr
t1 = z.read('xl/tables/table1.xml').decode('utf8')
t1 = t1.replace('ref="A1:V458"', 'ref="A1:V861"')
parts['xl/tables/table1.xml'] = t1.encode('utf8')
t2 = z.read('xl/tables/table2.xml').decode('utf8')
t2 = t2.replace('ref="A1:AQ14"', 'ref="A1:AQ10"')
parts['xl/tables/table2.xml'] = t2.encode('utf8')
t4 = z.read('xl/tables/table4.xml').decode('utf8')
t4 = t4.replace('ref="A1:Z19"', 'ref="A1:Z11"')
parts['xl/tables/table4.xml'] = t4.encode('utf8')
wbxml = z.read('xl/workbook.xml').decode('utf8')
wbxml = wbxml.replace('<calcPr calcId="191029"/>', '<calcPr calcId="191029" fullCalcOnLoad="1"/>')
parts['xl/workbook.xml'] = wbxml.encode('utf8')

for key, part in SHEETS.items():
    parts[part] = sh[key].out()

# ------------------------------------------------------------------ write zip: copy everything else byte-for-byte
with zipfile.ZipFile(OUT, 'w', zipfile.ZIP_DEFLATED) as zo:
    for item in z.infolist():
        data = parts.get(item.filename, None)
        if data is None:
            zo.writestr(item, z.read(item.filename))
        else:
            zo.writestr(item, data)
print('wrote', OUT)
import os
print('size:', os.path.getsize(OUT))
