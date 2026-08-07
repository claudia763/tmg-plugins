#!/usr/bin/env python3
"""Read the recalculated TMG model's key output cells and diff new errors vs the template.

Usage: python3 read_model.py <recalced.xlsx> [<pristine template.xlsx>]
"""
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl

CELLS = [
    ("Assumptions", "F5", "Project IRR"),
    ("Assumptions", "F7", "Avg Cash-on-Cash"),
    ("Assumptions", "F6", "Equity Multiple"),
    ("Assumptions", "I8", "T-3 DSCR"),
    ("Assumptions", "G48", "Target IRR"),
    ("Assumptions", "F50", "Price (Low)"),
    ("Assumptions", "G50", "Price (STRIKE)"),
    ("Assumptions", "H50", "Price (High)"),
    ("Assumptions", "G58", "Terminal Cap"),
    ("Assumptions", "G62", "Loan Program"),
    ("Assumptions", "G63", "Recourse"),
    ("Assumptions", "G64", "LTV"),
    ("Assumptions", "G65", "Interest Rate"),
    ("Assumptions", "I5", "Loan Amount"),
    ("Assumptions", "I6", "Equity Required"),
    ("Assumptions", "I9", "Debt Yield Yr1"),
    ("Master", "G33", "Total Units"),
    ("Master", "H33", "Total SF"),
    ("Master", "B22", "Occupancy"),
    ("Master", "B24", "Total Tax Rate"),
    ("Master", "B25", "Taxes @ CAD value"),
    ("Factors", "N16", "Avg 5-yr rent growth"),
    ("Factors", "N17", "Mkt occ at reversion"),
    ("Agency Loan-Sale Comps", "Z40", "Sale-comp cap base"),
    ("UW - F&C", "AK38", "Year-1 NOI"),
    ("UW - F&C", "AB13", "Occupancy (bridge test)"),
    ("UW - F&C", "AC8", "T-3 vacancy %"),
    ("UW - F&C", "AC9", "T-3 concessions %"),
    ("UW - F&C", "AC10", "T-3 bad debt %"),
]

ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!")


def errset(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value in ERRS:
                    out.add(f"{ws.title}!{c.coordinate}={c.value}")
    return out


def main():
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    print(f"{'CELL':<34}{'LABEL':<26}VALUE")
    for sheet, cell, label in CELLS:
        if sheet not in wb.sheetnames:
            print(f"  {sheet}!{cell:<28}{label:<26}<sheet missing>")
            continue
        v = wb[sheet][cell].value
        print(f"  {sheet+'!'+cell:<32}{label:<26}{v!r}")

    if len(sys.argv) > 2:
        new = errset(sys.argv[1])
        base = errset(sys.argv[2])
        added = sorted(new - base)
        print(f"\nerror cells: template {len(base)}, deliverable {len(new)}, "
              f"NEW {len(added)}")
        for e in added[:40]:
            print("   NEW " + e)


if __name__ == "__main__":
    main()
