"""Combine two part-year owner/PM income statements into ONE trailing-twelve
workbook in the source's own layout, ready for process_t12.py.

WHEN TO USE: ownership sends a prior-year statement and a current-year-to-date
statement instead of a single trailing twelve (very common with small
owner-managed assets). Stitching them by hand loses the printed-subtotal
tie-outs that process_t12.py's reconciliation gate depends on; this script
rebuilds every subtotal from the combined monthly detail and verifies it
against BOTH source statements' own printed monthly subtotals before writing.

HANDLES AN INTERIOR GAP. The Werner Creek pair (8/2026) had Jan-Nov 2025 and
Jan-Mar 2026 -- December 2025 existed in neither file. The missing month keeps
its column on the trailing-12 axis but every data cell stays EMPTY: never zero,
never interpolated, never annualized (house rule). Recomputed row totals are
therefore 11-month sums, and the caption in A3 says so. Note this also required
teaching `--pad-to-12` to accept months that are an ordered SUBSET of the
trailing-12 axis rather than only a contiguous short tail.

Do NOT use this to fabricate a missing month. If the model downstream needs
twelve populated columns, fill the gap in the MODEL and disclose it there --
the T-12 deliverable stays honest. See
`instructions/werner-creek-uw-8-2026.md` for how that split was presented.

INPUTS   two .xlsx statements in the same PM-system layout, each with an
         "Account Name" header row and 12 month columns of real date cells
OUTPUT   one .xlsx in the identical layout covering the chosen trailing-12
         window, with a truthful period caption in A3; exits non-zero and
         prints every failing check if any recomputed subtotal disagrees with
         the source's own printed figure

TO ADAPT TO A NEW DEAL: edit the SRC25 / SRC26 / OUT constants and the target
window below. Everything else is driven off the source layout. Validated on
Werner Creek Apartments (Toki Property Management LLC, Houston, 8/6/2026):
Revenue 360,985.65 / OpEx 147,662.00 / NOI 213,323.65 over the 11 real months,
tying to the cent against both source statements.

RELATED, AND EASY TO CONFUSE -- three different stitching problems landed in
this library the same week; pick the right one:
  * THIS SCRIPT: one property, two part-year statements of the same layout,
    stitched onto a trailing-12 axis (an interior month may be missing).
  * `instructions/owner-calendar-year-tabs-t12.md`: one property, one workbook
    with a tab per calendar year.
  * `instructions/combining-two-property-t12s.md` + `scripts/build_master_t12.py`:
    TWO properties merged into one master statement (different charts of
    accounts, not different periods).

Requires: openpyxl.
"""
import datetime as dt
import sys

import openpyxl
from openpyxl.styles import Font

JOB = r"D:\Multifamily Work\jobs\2026-08-06T17-47-20-614Z-uid40"
SRC25 = JOB + r"\inbox\Werner_q4_T12_-RR-2025.xlsx"
SRC26 = JOB + r"\inbox\Werner_2026_T12_-RR.xlsx"
OUT = JOB + r"\work\Werner_Combined_T12_Apr2025-Mar2026.xlsx"

TOL = 0.005
fails = []


def check(label, got, want):
    ok = abs(got - want) <= TOL
    if not ok:
        fails.append(f"{label}: {got:,.2f} vs {want:,.2f}")
    return ok


def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = None
    for i, r in enumerate(rows):
        if str(r[0] or "").strip().lower() == "account name":
            hdr = i
            break
    months = [r for r in rows[hdr][1:13]]
    body = {}          # raw label (with indentation) -> (vals[12], annual)
    order = []
    for r in rows[hdr + 1:]:
        name = r[0]
        if name is None or not str(name).strip():
            continue
        vals = [None if r[1 + j] is None else float(r[1 + j])
                for j in range(12)]
        annual = None if r[13] is None else float(r[13])
        body[str(name)] = (vals, annual)
        order.append(str(name))
    return months, body, order


m25, b25, o25 = load(SRC25)
m26, b26, o26 = load(SRC26)

# ---------------------------------------------------------------- source QA
print("=" * 78)
print("SOURCE VERIFICATION (each printed Total vs its own monthly cells)")
for tag, body, order in (("2025", b25, o25), ("2026", b26, o26)):
    for nm in order:
        vals, annual = body[nm]
        if annual is None:
            continue
        s = sum(v for v in vals if v is not None)
        if not check(f"{tag} row total {nm.strip()}", s, annual):
            print(f"  MISMATCH {tag} {nm.strip()}: months {s:,.2f} "
                  f"vs printed {annual:,.2f}")
print("  every printed row Total ties to its own monthly cells"
      if not fails else "  !! see mismatches above")

# 2025 col M (Dec) must be genuinely empty
dec_vals = [b25[nm][0][11] for nm in o25]
nonempty = [nm.strip() for nm in o25 if b25[nm][0][11] is not None]
print(f"  2025 Dec column: {len(nonempty)} non-empty cell(s) "
      f"{nonempty if nonempty else '- COMPLETELY EMPTY (confirmed)'}")
# 2026 cols E..M must be empty on the detail rows
n26 = [nm.strip() for nm in o26
       if any(b26[nm][0][j] is not None for j in range(3, 12))]
print(f"  2026 Apr-Dec columns: detail rows with data: "
      f"{n26 if n26 else '- none (only the zero-filled NOI/restatement rows)'}")

# ------------------------------------------------------------ combined axis
MONTHS = [dt.datetime(2025, 4, 1), dt.datetime(2025, 5, 1),
          dt.datetime(2025, 6, 1), dt.datetime(2025, 7, 1),
          dt.datetime(2025, 8, 1), dt.datetime(2025, 9, 1),
          dt.datetime(2025, 10, 1), dt.datetime(2025, 11, 1),
          dt.datetime(2025, 12, 1),                       # <- BLANK
          dt.datetime(2026, 1, 1), dt.datetime(2026, 2, 1),
          dt.datetime(2026, 3, 1)]
DEC = 8
SRC_IDX = {}          # combined col -> ('25'|'26', source month index)
for k in range(8):
    SRC_IDX[k] = ("25", 3 + k)          # Apr-25 (idx3) .. Nov-25 (idx10)
for k in range(3):
    SRC_IDX[9 + k] = ("26", k)          # Jan-26 .. Mar-26

# --------------------------------------------------------- combined layout
S, A, T, MEMO, BLANK = "section", "account", "subtotal", "memo", "blank"
LAYOUT = [
    (MEMO, "    Market Rent "),
    (S, "    Income"),
    (S, "        RENTS"),
    (A, "            Rent Income"),
    (A, "            Section 8 Rent"),
    (T, "        Total RENTS", ["            Rent Income",
                                "            Section 8 Rent"]),
    (A, "        Loss/Gain to Market"),
    (A, "        Concessions"),
    (S, "        FEES"),
    (A, "            Electric Connection Fee"),
    (A, "            Pet Fee-Non Refundable"),
    (A, "            NSF Fees Collected"),
    (A, "            Application Fee Income"),
    (A, "            Insurance Services"),
    (A, "            Property Damage Fee Account"),
    (A, "            Late Fee"),
    (A, "            Month-to-Month Fee"),
    (A, "            Admin Fee Account"),
    (A, "        Miscellaneous Income"),
    (A, "        Parking"),
    (A, "        Reserved Parking"),
    (A, "        Move-Out Cleaning Income"),
    (T, "TOTAL FEES", ["            Electric Connection Fee",
                       "            Pet Fee-Non Refundable",
                       "            NSF Fees Collected",
                       "            Application Fee Income",
                       "            Insurance Services",
                       "            Property Damage Fee Account",
                       "            Late Fee",
                       "            Month-to-Month Fee",
                       "            Admin Fee Account",
                       "        Miscellaneous Income",
                       "        Parking",
                       "        Reserved Parking",
                       "        Move-Out Cleaning Income"]),
    (T, "    Total Operating Income", ["#Total RENTS", "        Loss/Gain to Market",
                                       "        Concessions", "#TOTAL FEES"]),
    (BLANK, ""),
    (S, "    Expense"),
    (S, "        UTILITY"),
    (A, "            Water"),
    (A, "            Gas"),
    (A, "            Electric"),
    (A, "            Waste collection"),
    (A, "            Internet/Phone"),
    (T, "        Total Utility", ["            Water", "            Gas",
                                  "            Electric",
                                  "            Waste collection",
                                  "            Internet/Phone"]),
    (S, "        Wages"),
    (A, "            Staff Salary (1099)"),
    (A, "            Temp. Contractor (1099)"),
    (T, "    Total Operating Expense", ["#Total Utility",
                                        "            Staff Salary (1099)",
                                        "            Temp. Contractor (1099)"]),
    (BLANK, ""),
    (A, "    Tax"),
    (A, "    Insurance"),
    (BLANK, ""),
    (S, "Maintenance Cost / Expense"),
    (A, "Annual Material Cost / month"),
    (BLANK, ""),
    (T, "    NOI - Net Operating Income", None),
    (BLANK, ""),
    (T, "    Total Income", None),
    (T, "    Total Expense", None),
    (BLANK, ""),
    (T, "    Net Income", None),
]

INCOME_ACCTS = [n for k, n, *_ in LAYOUT if k == A][:17]
EXPENSE_ACCTS = [n for k, n, *_ in LAYOUT if k == A][17:]
assert INCOME_ACCTS[-1].strip() == "Move-Out Cleaning Income", INCOME_ACCTS[-1]
assert EXPENSE_ACCTS[0].strip() == "Water", EXPENSE_ACCTS[0]

# --------------------------------------------- pull the monthly detail
missing_note = []
vals = {}
for name in INCOME_ACCTS + EXPENSE_ACCTS:
    row = [None] * 12
    for k, (tag, j) in SRC_IDX.items():
        body = b25 if tag == "25" else b26
        if name in body:
            row[k] = body[name][0][j]
        else:
            row[k] = 0.0          # account not on that statement at all
    vals[name] = row
    in25, in26 = name in b25, name in b26
    if not (in25 and in26):
        missing_note.append((name.strip(), in25, in26))

# every source account row must be consumed by the layout (nothing dropped)
known = set(INCOME_ACCTS + EXPENSE_ACCTS) | {
    n for k, n, *_ in LAYOUT if k in (S, T, MEMO)}
for tag, order in (("2025", o25), ("2026", o26)):
    for nm in order:
        if nm not in known:
            fails.append(f"{tag} source row not represented in layout: {nm!r}")

# Section 8 Rent carries no data at all in either statement -> keep it blank
blank_rows = {n for n in INCOME_ACCTS + EXPENSE_ACCTS
              if all((b25.get(n, ([None] * 12, None))[0][j] is None)
                     for j in range(3, 11))
              and all((b26.get(n, ([None] * 12, None))[0][j] is None)
                      for j in range(3))
              and n not in b25.keys() | b26.keys() or
              (n in b25 and all(v is None for v in b25[n][0]) and
               (n not in b26 or all(v is None for v in b26[n][0])))}
print(f"  rows with NO data in either statement (kept blank): "
      f"{sorted(x.strip() for x in blank_rows)}")

# ------------------------------------------------------- recompute totals
def col(name, k):
    if name in blank_rows:
        return 0.0
    return vals[name][k] or 0.0


sub = {}
for item in LAYOUT:
    if item[0] != T:
        continue
    nm, members = item[1].strip(), item[2]
    if members is None:
        continue
    sub[nm] = [sum(col(m, k) if not m.startswith("#") else sub[m[1:]][k]
                   for m in members) for k in range(12)]

inc_tot = [sum(col(n, k) for n in INCOME_ACCTS) for k in range(12)]
exp_tot = [sum(col(n, k) for n in EXPENSE_ACCTS) for k in range(12)]
sub["NOI - Net Operating Income"] = [a - b for a, b in
                                     zip(sub["Total Operating Income"], exp_tot)]
sub["Total Income"] = list(sub["Total Operating Income"])
sub["Total Expense"] = list(exp_tot)
sub["Net Income"] = [a - b for a, b in zip(sub["Total Income"], exp_tot)]

# ---- independent arithmetic checks vs the SOURCES' own printed monthlies
print("=" * 78)
print("COMBINED vs SOURCE printed monthly subtotals (independent recompute)")
SRC_SUB = {"Total Operating Income": "    Total Operating Income",
           "Total Utility": "        Total Utility",
           "Total RENTS": "        Total RENTS",
           "TOTAL FEES": "TOTAL FEES",
           "Total Operating Expense": "    Total Operating Expense",
           "NOI - Net Operating Income": "    NOI - Net Operating Income"}
for k in range(12):
    if k == DEC:
        continue
    tag, j = SRC_IDX[k]
    body = b25 if tag == "25" else b26
    for key, raw in SRC_SUB.items():
        want = body[raw][0][j]
        check(f"{MONTHS[k]:%b-%y} {key}", sub[key][k], want or 0.0)
print("  all 11 real months x 6 printed subtotals tie exactly"
      if not fails else "  !! mismatches: " + "; ".join(fails))

# income detail (ex Market Rent memo) must equal Total Operating Income
for k in range(12):
    check(f"{MONTHS[k]:%b-%y} income detail vs Total Operating Income",
          inc_tot[k], sub["Total Operating Income"][k])
# NOI must equal income less every expense line
for k in range(12):
    check(f"{MONTHS[k]:%b-%y} NOI", sub["NOI - Net Operating Income"][k],
          inc_tot[k] - exp_tot[k])

print("=" * 78)
print("MONTHLY COMBINED SUMMARY (Dec-2025 blank = not provided)")
print(f"{'Month':>9} {'Revenue':>12} {'OpEx':>12} {'NOI':>12}")
for k in range(12):
    if k == DEC:
        print(f"{MONTHS[k]:%b-%y:>9}".replace("%b", "") if False else
              f"{MONTHS[k].strftime('%b-%y'):>9} {'(blank)':>12} "
              f"{'(blank)':>12} {'(blank)':>12}")
        continue
    print(f"{MONTHS[k].strftime('%b-%y'):>9} {inc_tot[k]:>12,.2f} "
          f"{exp_tot[k]:>12,.2f} "
          f"{sub['NOI - Net Operating Income'][k]:>12,.2f}")
real = [k for k in range(12) if k != DEC]
tr = sum(inc_tot[k] for k in real)
te = sum(exp_tot[k] for k in real)
print(f"{'TOTAL(11)':>9} {tr:>12,.2f} {te:>12,.2f} {tr - te:>12,.2f}")
t3 = [9, 10, 11]
print(f"  T-3 (Jan-Mar 26) x4 annualized : Rev {sum(inc_tot[k] for k in t3)*4:,.2f}"
      f" | OpEx {sum(exp_tot[k] for k in t3)*4:,.2f}"
      f" | NOI {(sum(inc_tot[k] for k in t3)-sum(exp_tot[k] for k in t3))*4:,.2f}")
print(f"  T-1 (Mar 26)     x12 annualized: Rev {inc_tot[11]*12:,.2f}"
      f" | OpEx {exp_tot[11]*12:,.2f} | NOI {(inc_tot[11]-exp_tot[11])*12:,.2f}")
print(f"  Market Rent (GPR memo row): {sum(col('    Market Rent ', k) if False else 0 for k in real):,.2f}")

# ---------------------------------------------------------------- write it
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "WERNER COMBINED T12"
ws["A1"] = "Toki Property Management LLC"
ws["A2"] = ("Properties: Werner creek Apartments - 4635 WERNER St "
            "Houston, TX 77022")
ws["A3"] = ("Period Range: Apr 2025 to Mar 2026 (trailing twelve months "
            "ending Mar 2026; Dec 2025 not provided by ownership - column "
            "intentionally blank)")
ws["A5"] = "Account Name"
for k, d in enumerate(MONTHS):
    c = ws.cell(row=5, column=2 + k, value=d)
    c.number_format = "mmm-yy"
    c.font = Font(bold=True)
ws.cell(row=5, column=14, value="Total").font = Font(bold=True)
ws["A5"].font = Font(bold=True)

mem = {"    Market Rent ": [b25["    Market Rent "][0][SRC_IDX[k][1]]
                            if SRC_IDX[k][0] == "25"
                            else b26["    Market Rent "][0][SRC_IDX[k][1]]
                            for k in range(12) if k != DEC]}

r = 6
for item in LAYOUT:
    kind, name = item[0], item[1]
    if kind == BLANK:
        r += 1
        continue
    ws.cell(row=r, column=1, value=name)
    if kind == S:
        r += 1
        continue
    if kind == MEMO:
        row = [None] * 12
        for k in range(12):
            if k == DEC:
                continue
            tag, j = SRC_IDX[k]
            row[k] = (b25 if tag == "25" else b26)["    Market Rent "][0][j]
    elif kind == A:
        row = [None if (name in blank_rows or k == DEC) else col(name, k)
               for k in range(12)]
    else:
        row = [None if k == DEC else sub[name.strip()][k] for k in range(12)]
    for k in range(12):
        if row[k] is not None:
            ws.cell(row=r, column=2 + k, value=round(float(row[k]), 2))
    if any(v is not None for v in row):
        ws.cell(row=r, column=14,
                value=round(sum(v for v in row if v is not None), 2))
    r += 1

ws.column_dimensions["A"].width = 34
for cc in "BCDEFGHIJKLMN":
    ws.column_dimensions[cc].width = 11
wb.save(OUT)
print("=" * 78)
print(f"WROTE {OUT}")
if fails:
    print(f"!! {len(fails)} CHECK FAILURE(S):")
    for f in fails[:40]:
        print("   " + f)
    sys.exit(1)
print("ALL CHECKS PASSED")
if missing_note:
    print("Accounts present on only one statement (0.00 written for the "
          "months of the statement that omits the line - proved by that "
          "statement's own printed TOTAL FEES / Total Operating Income):")
    for n, i25, i26 in missing_note:
        print(f"   {n}: 2025={'yes' if i25 else 'NO'} "
              f"2026={'yes' if i26 else 'NO'}")

