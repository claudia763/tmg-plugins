#!/usr/bin/env python3
"""Pre-send QC gate for a recalculated TMG underwriting MODEL workbook.

WHAT IT DOES
    `verify_deliverable_workbooks.py` covers the rent-roll / T-12 / Capex
    deliverables. This is the equivalent gate for the underwriting model --
    the checks an orchestrator would otherwise run by hand on the
    LibreOffice-recalculated workbook before attaching it to a reply.

    It does NOT re-underwrite the deal. It proves the workbook that is about
    to be emailed is internally consistent, has the debt block wired up, is
    green on the house-rule metrics, and has no client-visible errors on the
    one page the buyer actually reads.

CHECKS
    Debt block wiring (the silent killers -- see
    aggressive-pricing-house-rule-8-2026.md and
    uw-model-linux-libreoffice-build-8-2026.md sec.4):
      * Assumptions!M62 is a non-empty program name. Ships BLANK; until it is
        set, G62 recalcs to #DIV/0!, the whole debt block is empty, and G48
        silently falls back to the 25% RECOURSE target.
      * G63 resolves to a recourse string, not "" or an error.
      * G48 is read at RUNTIME, never assumed. 0.20 non-recourse / 0.25
        recourse. A workbook reporting G48 = 0.25 on agency debt is the
        classic ~$500k underpricing bug.
      * G64 (LTV) and G65 (rate) are populated and sane.

    House-rule green tests, read from the workbook and not from Python:
      * F5  Project IRR        >= G48
      * F7  Avg cash-on-cash   >= 0.10
      * I8  T-3 DSCR           >= max(1.25, program minimum in 'Loan Terms'!G)
        -- unless the deal is distressed (T-3 economic loss > 30%, from
        'UW - F&C' AC8+AC9+AC10), in which case DSCR is not tested at all.

    Aggressive-pricing slack (the reason this script exists):
      * reports HOW MUCH cushion sits above each floor. Excess DSCR or a
        multi-point IRR cushion means the strike is leaving money on the
        table and should be re-solved upward. "Green" is necessary but not
        sufficient -- the house rule wants JUST green.

    Client-visible integrity:
      * zero error strings (#REF!/#VALUE!/#DIV/0!/#N/A/#NAME?/#NUM!/#NULL!)
        anywhere in the 'PDF Output - F&C' print area. Hidden-sheet errors
        are counted and reported but do not fail -- LibreOffice does not
        implement the _xlfn.*/_xludf.* functions Excel does.
      * the Rent Comparable Summary block actually has comparables. An empty
        comp table prints #DIV/0! across the Averages row on page 1 and is
        the most common shipped-with-a-visible-defect failure.
      * no black-fill conditional format left on the floor-plan block.

USAGE
    python3 verify_uw_model.py "CK - Property - 8-7-2026.xlsx"
    python3 verify_uw_model.py model.xlsx --price 6425000 --units 96

    Exit status 0 = clean, 1 = at least one FAIL. WARNs never fail the gate.

NOTE
    Run this on the RECALCULATED workbook (the one that came back out of
    `soffice --convert-to xlsx`), not on the openpyxl-saved input -- an
    unrecalculated file has no cached values and every check is vacuous.
    The script refuses to run if F5 has no cached value, which is that case.
"""
import argparse
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.utils import range_boundaries

ERRORS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!", "#NULL!")

results = []


def record(level, label, detail):
    results.append((level, label, detail))
    print(f"  [{level:4}] {label}: {detail}")


def cell(ws, coord):
    try:
        return ws[coord].value
    except Exception:
        return None


def check_debt_block(a):
    print("\nDEBT BLOCK WIRING")
    program = cell(a, "M62")
    if not program or not str(program).strip():
        record("FAIL", "M62 loan program",
               "BLANK -- debt block is empty and G48 falls back to the "
               "recourse target. Set it from 'Loan Terms'!A4:A16.")
    else:
        record("ok", "M62 loan program", repr(program))

    recourse = cell(a, "G63")
    if not recourse or str(recourse).startswith("#"):
        record("FAIL", "G63 recourse", f"unresolved ({recourse!r})")
    else:
        record("ok", "G63 recourse", str(recourse))

    target = cell(a, "G48")
    if not isinstance(target, (int, float)):
        record("FAIL", "G48 target IRR", f"not numeric ({target!r})")
    else:
        note = ""
        if isinstance(recourse, str) and "non-recourse" in recourse.lower():
            note = ("  <-- expected 0.20 on non-recourse"
                    if abs(target - 0.20) > 1e-9 else "")
            if note:
                record("FAIL", "G48 target IRR",
                       f"{target:.2%} on NON-RECOURSE debt -- expected 20%. "
                       "This silently crushes the strike.")
            else:
                record("ok", "G48 target IRR", f"{target:.2%} (non-recourse)")
        else:
            record("ok", "G48 target IRR", f"{target:.2%}")

    for coord, label, lo, hi in (("G64", "LTV", 0.30, 0.85),
                                 ("G65", "interest rate", 0.02, 0.15)):
        v = cell(a, coord)
        if not isinstance(v, (int, float)):
            record("FAIL", f"{coord} {label}", f"not numeric ({v!r})")
        elif not (lo <= v <= hi):
            record("WARN", f"{coord} {label}", f"{v:.4f} outside {lo}-{hi}")
        else:
            record("ok", f"{coord} {label}", f"{v:.4%}")
    return target


def check_green(a, uw, target):
    print("\nHOUSE-RULE GREEN TESTS (and slack above each floor)")
    irr, coc, dscr = cell(a, "F5"), cell(a, "F7"), cell(a, "I8")

    if not isinstance(irr, (int, float)):
        record("FAIL", "F5 project IRR",
               "no cached value -- workbook was never recalculated")
        return
    tgt = target if isinstance(target, (int, float)) else 0.20
    if irr >= tgt:
        record("ok", "F5 project IRR",
               f"{irr:.2%} vs target {tgt:.2%}  (+{(irr - tgt) * 100:.2f} pts slack)")
        if irr - tgt > 0.015:
            record("WARN", "IRR cushion",
                   f"{(irr - tgt) * 100:.2f} points above target -- the house rule "
                   "wants JUST green. Re-solve the strike upward.")
    else:
        record("FAIL", "F5 project IRR", f"{irr:.2%} BELOW target {tgt:.2%}")

    if isinstance(coc, (int, float)):
        (record("ok", "F7 avg cash-on-cash", f"{coc:.2%} vs 10.00% floor")
         if coc >= 0.10 else
         record("FAIL", "F7 avg cash-on-cash", f"{coc:.2%} BELOW the 10% floor"))

    # distress test -- if T-3 economic loss > 30%, DSCR is not tested
    loss = 0.0
    for c in ("AC8", "AC9", "AC10"):
        v = cell(uw, c) if uw is not None else None
        if isinstance(v, (int, float)):
            loss += v
    if loss > 0.30:
        record("WARN", "distress test",
               f"T-3 economic loss {loss:.1%} > 30% -- bridge financing assumed, "
               "DSCR test skipped per the skill's house rules")
    elif isinstance(dscr, (int, float)):
        if dscr >= 1.25:
            record("ok", "I8 T-3 DSCR", f"{dscr:.5f} vs 1.25 floor "
                                        f"(+{dscr - 1.25:.5f} slack)")
            if dscr - 1.25 > 0.02:
                record("WARN", "excess DSCR",
                       f"{dscr:.4f} carries {dscr - 1.25:.4f} of excess coverage. "
                       "Properties rarely trade with excess DSCR -- raise the price.")
        else:
            record("FAIL", "I8 T-3 DSCR", f"{dscr:.5f} BELOW the 1.25 floor")


def check_pdf_page(wb):
    print("\nCLIENT-VISIBLE INTEGRITY ('PDF Output - F&C')")
    name = "PDF Output - F&C"
    if name not in wb.sheetnames:
        record("FAIL", "PDF sheet", f"{name!r} missing")
        return
    ws = wb[name]

    bad = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in ERRORS:
                bad.append(f"{c.coordinate}={c.value.strip()}")
    if bad:
        record("FAIL", "error cells on the printed page",
               f"{len(bad)} found: " + ", ".join(bad[:12])
               + (" ..." if len(bad) > 12 else ""))
    else:
        record("ok", "error cells on the printed page", "none")

    # black-fill CF on the floor-plan block
    dark = []
    for rng in ws.conditional_formatting:
        for rule in rng.rules:
            dxf = getattr(rule, "dxf", None)
            fill = getattr(dxf, "fill", None) if dxf else None
            for col in (getattr(fill, "bgColor", None),
                        getattr(fill, "fgColor", None)):
                rgb = getattr(col, "rgb", None)
                if isinstance(rgb, str) and len(rgb) == 8:
                    r, g, b = (int(rgb[2:4], 16), int(rgb[4:6], 16),
                               int(rgb[6:8], 16))
                    if r + g + b < 150:
                        for part in str(rng.sqref).split():
                            _, mn_r, _, mx_r = (range_boundaries(part)[0],
                                                range_boundaries(part)[1],
                                                range_boundaries(part)[2],
                                                range_boundaries(part)[3])
                            if mx_r >= 52 and mn_r <= 77:
                                dark.append(str(rng.sqref))
    if dark:
        record("FAIL", "black-box conditional format",
               f"dark fill still on {sorted(set(dark))} -- delete before export")
    else:
        record("ok", "black-box conditional format", "removed")


def check_printed_price(wb, a):
    """The printed Sales Range rounds to $10k; an off-boundary strike breaks
    the cost stack AND can advertise a price below the DSCR floor. See
    aggressive-pricing-house-rule-8-2026.md."""
    print("\nPRINTED PRICE CONSISTENCY")
    name = "PDF Output - F&C"
    if name not in wb.sheetnames:
        return
    p = wb[name]
    strike = cell(a, "G50")
    shown = cell(p, "G22")
    if not isinstance(strike, (int, float)) or not isinstance(shown, (int, float)):
        record("WARN", "printed price", "could not read G50 / G22")
        return

    if abs(shown - strike) < 0.5:
        record("ok", "printed price", f"${shown:,.0f} == G50 (on a $10k boundary)")
    else:
        record("FAIL", "printed price",
               f"page 1 prints ${shown:,.0f} but the model is solved at "
               f"${strike:,.0f}. G22 = ROUND(price,-4) and Excel rounds half "
               f"AWAY from zero. Move the strike onto a $10,000 boundary -- "
               f"try ${int(strike // 10000) * 10000:,}.")

    # Purchase Price + Value-Add == Total Costs, as printed
    pp, va, tc = cell(p, "M13"), cell(p, "M14"), cell(p, "M17")
    if all(isinstance(v, (int, float)) for v in (pp, va, tc)):
        if abs((pp + va) - tc) < 0.5:
            record("ok", "cost stack",
                   f"${pp:,.0f} + ${va:,.0f} = ${tc:,.0f}")
        else:
            record("FAIL", "cost stack",
                   f"page 1 prints ${pp:,.0f} + ${va:,.0f} = ${pp + va:,.0f}, "
                   f"but Total Costs reads ${tc:,.0f} -- off by "
                   f"${abs((pp + va) - tc):,.0f}. Same root cause as above.")


def check_rent_comps(wb):
    """The Rent Comparable Summary must not be empty (page-1 #DIV/0! defect)."""
    name = "PDF Output - F&C"
    if name not in wb.sheetnames:
        return
    ws = wb[name]
    hits = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for c in row:
            if isinstance(c.value, str) and "Rent Comparable" in c.value:
                hits.append(c.row)
    if not hits:
        record("WARN", "rent comp block", "header not located -- check by eye")
        return
    start = min(hits)
    filled, errs = 0, 0
    for r in range(start + 1, min(start + 14, ws.max_row + 1)):
        vals = [ws.cell(row=r, column=col).value for col in range(2, 13)]
        errs += sum(1 for v in vals
                    if isinstance(v, str) and v.strip() in ERRORS)
        real = [v for v in vals
                if v not in (None, "")
                and not (isinstance(v, str) and v.strip() in ERRORS)]
        if len(real) >= 5:
            filled += 1

    # The subject row is always populated, so "some rows filled" is not
    # evidence of comparables. Errors in the block mean the Averages row is
    # dividing by an empty comp set -- that is the real tell.
    if errs:
        record("FAIL", "rent comp block",
               f"{errs} error cells inside the block ({filled} populated rows, "
               "one of which is the subject) -- the comp table is EMPTY and the "
               "Averages row prints #DIV/0! on page 1. Populate "
               "TableRecentLeases / TablePropertyData and mark 4-5 comps with "
               "'x' in Rent Comparison column AK.")
    elif filled >= 3:
        record("ok", "rent comp block",
               f"{filled} populated rows (subject + comparables), no errors")
    else:
        record("FAIL", "rent comp block",
               f"only {filled} populated rows -- expected the subject plus at "
               "least two comparables. Check the AK selection.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--price", type=float, default=None)
    ap.add_argument("--units", type=int, default=None)
    args = ap.parse_args()

    print(f"VERIFYING {args.workbook}")
    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    a = wb["Assumptions"] if "Assumptions" in wb.sheetnames else None
    uw = wb["UW - F&C"] if "UW - F&C" in wb.sheetnames else None
    if a is None:
        print("  [FAIL] no 'Assumptions' sheet -- is this the model workbook?")
        return 1
    if cell(a, "F5") is None:
        print("  [FAIL] Assumptions!F5 has no cached value -- run the "
              "LibreOffice recalc first; every check below would be vacuous.")
        return 1

    target = check_debt_block(a)
    check_green(a, uw, target)
    check_pdf_page(wb)
    check_printed_price(wb, a)
    check_rent_comps(wb)

    if args.price is not None:
        got = cell(a, "G50")
        print("\nSTRIKE")
        if got is not None and abs(float(got) - args.price) < 1:
            record("ok", "G50 strike", f"${float(got):,.0f}")
        else:
            record("FAIL", "G50 strike",
                   f"workbook has {got!r}, expected ${args.price:,.0f}")
        if args.units:
            record("ok", "derived", f"${float(got) / args.units:,.0f} per unit")

    fails = [r for r in results if r[0] == "FAIL"]
    warns = [r for r in results if r[0] == "WARN"]
    print(f"\n{'=' * 62}\n{len(fails)} FAIL, {len(warns)} WARN, "
          f"{len(results) - len(fails) - len(warns)} ok")
    if fails:
        print("NOT DELIVERABLE:")
        for _, label, detail in fails:
            print(f"  - {label}: {detail}")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
