#!/usr/bin/env python3
"""OwnerSimpleRowRentRollXlsxParser - reference copy (TMG library addition)

Added 8/2026 on Aldine Apartments (Goldenwrist Capital LLC, 4826 Aldine Mail
Rte Rd, Houston TX, 96 doors). Companion write-up:
`instructions/owner-simple-row-rent-roll-xlsx.md`.

This file is a DOCUMENTED REFERENCE COPY of a parser that belongs inside
`process_rent_roll.py`; it is not importable on its own (it subclasses
`RentRollParser` and uses that module's `UnitRecord` / `Resident` / helpers).

To install it into a toolkit copy of `process_rent_roll.py`:

1. Paste the class below immediately BEFORE `class OwnerSheetXlsxParser`.

2. Register it immediately BEFORE `OwnerSheetXlsxParser` in `XLSX_PARSERS`:

       XLSX_PARSERS = [OwnerBlockRentRollXlsxParser, AppFolioUnitTypeXlsxParser,
                       AppFolioXlsxParser, YardiRentRollXlsxParser,
                       OwnerSimpleRowRentRollXlsxParser, OwnerSheetXlsxParser]

3. Tighten `OwnerSheetXlsxParser.detect_xlsx` so the two dialects are
   mutually exclusive rather than merely order-dependent (they print the SAME
   nine header captions). Add to that class:

       STATUS_RE = re.compile(
           r"^(current|occupied|evict(ion|ed)?|notice(\s*-\s*\w+)?|"
           r"vacant(\s*-\s*\w+)?|vac)$", re.I)

   and replace its `detect_xlsx` with one that returns True only when at
   least one data row's Tenants cell matches `STATUS_RE`.

4. Apply the three supporting deltas below (non-revenue units), which are
   what let a door be neither occupied nor vacant. All three are additive:
   `non_revenue` defaults False, so no existing format changes behaviour.

   (a) `UnitRecord` gains two fields:

           non_revenue: bool = False
           non_revenue_use: str = ""

   (b) `UnitRecord.is_vacant` and `UnitRecord.on_notice` both return False
       first thing when `self.non_revenue` - a non-revenue door is not a
       vacancy and is not on notice.

   (c) Both writers stamp the template's own literal:

           occ = ("Non-Rev" if u.non_revenue
                  else "Vacant" if vacant else "Occupied")

       `rentroll_template.xlsx` already counts "Non-Rev" in Floor Plan
       Summary col K ("Occupancy Status (# Units) / Non-Rev") and shows it on
       the Floor Plan tab - no template change is needed.

   (d) `reconcile()` splits the three buckets and asserts the identity:

           nrv = [u for u in units if u.non_revenue]
           occ = [u for u in units if not u.is_vacant and not u.non_revenue]
           vac = [u for u in units if u.is_vacant]
           ...
           if nrv:
               check("Occupied + vacant + non-revenue = unit count",
                     len(occ) + len(vac) + len(nrv), len(units), tol=0.5)

5. Optional but recommended: the Comments tab. `_add_comments_tab(wb, prop,
   notes)` writes a plain-black "Comments" sheet (mirroring the T-12 house
   rule of 8/6/2026), created only when there is at least one note; both
   rent-roll writers take a `comments=()` argument and `main()` builds the
   list from `parser.comments` plus a repeatable `--note` flag. This parser
   uses it for the non-revenue roster, the MTM ruling, the INFERRED as-of
   date and the companion-statement market-rent comparison.

Regression-check detection before shipping:

    python scripts/parser_detection_regression.py --toolkit ./toolkit \
        --files "./sources/*.xlsx" "./toolkit/*.xlsx" \
        --expect OwnerSimpleRowRentRollXlsxParser="./sources/<new roll>.xlsx" \
        --expect OwnerSheetXlsxParser="./sources/<a Werner-style roll>.xlsx"

Aldine 8/2026: 16 reconciliation checks, all tying to an independent
re-extraction of the sheet XML (96 units; 85 occupied / 4 vacant / 7
non-revenue; 69,600 sf; 104,160 market rent; 88,839 contract rent; per-plan
counts, sq ft and market rent for both plans).
"""

import re
import sys
from datetime import datetime, date

from process_rent_roll import RentRollParser, UnitRecord, Resident


class OwnerSimpleRowRentRollXlsxParser(RentRollParser):
    """Owner/PM-maintained ONE-ROW-PER-UNIT rent roll (.xlsx) whose "Tenants"
    column carries a FACILITY LABEL or nothing at all - never a status word.

    Validated on Aldine Apartments (Goldenwrist Capital LLC, 4826 Aldine Mail
    Rte Rd, Houston TX, 96 doors, 8/2026), sheet `RR` of a workbook whose
    first sheet is the T-12.

        (idx) | COMPLEX | APT #NUMBER | FLOOR TYPE | Sq. FEET | Tenants |
        Move In | Lease Expires | ADV. RENT | CURRENT RENT
        1 | Aldine Apartment | A01 | 1/1.00 | 650 | Office | Office | Office
          | 995 |
        2 | Aldine Apartment | A02 | 1/1.00 | 650 |  | 2024-12-01
          | 2026-11-30 | 995 | 950

    SAME HEADER ROW as `OwnerSheetXlsxParser` (Werner Creek) - and a DIFFERENT
    dialect. In the Werner sheet the "Tenants" column is a STATUS column
    ("Current" / "Vacant-Unrented"); here it is blank on every leased door and
    carries a facility USE ("Office", "Shop", "Storage") on the doors that are
    not apartments at all. Reading this sheet with the Werner parser gets
    every answer wrong: its openpyxl pass reads a blank status as occupied
    while its re-extraction pass reads the same blank as vacant, it reads
    "Office"/"Shop"/"Storage" as resident names on occupied doors, and it has
    no non-revenue concept at all.

    The two are told apart by the Tenants column itself and the split is
    exhaustive, so neither can steal the other's file:

      * this parser REQUIRES that no data row's Tenants cell is a Werner
        status word, and that at least three rows pair a BLANK Tenants cell
        with a numeric CURRENT RENT (the signature of "the roll does not
        print who lives here");
      * `OwnerSheetXlsxParser`'s sniff is looser (four header captions), so
        this parser is registered BEFORE it - the same precedent as
        `ResManSummaryParser` sitting before `ResManParser`.

    Layout rules:

    * **Occupancy is read from CURRENT RENT, never from the dates.** The sheet
      prints no status column, so a door with a contractual rent is occupied
      and a door without one is vacant. This is the trap in the format:
      Aldine A94 carries Move In 6/6/2025 and Lease Expires 5/31/2026 but no
      CURRENT RENT - it is a vacated/notice door whose dates were left behind,
      and any rule of the form "has dates => occupied" mis-states it. The
      dates are still carried through to the deliverable verbatim.
    * **A facility label in Tenants is a NON-REVENUE door, not a vacancy.**
      Office / Shop / Storage / Laundry / Model / Down units are doors
      ownership has taken out of the rentable pool: `UnitRecord.non_revenue`
      is set, Occupancy Status prints "Non-Rev", and the unit lands in neither
      the occupied nor the vacant bucket. Their ADV. RENT is still carried
      (the source's own market-rent total, and the companion statement's
      MARKET RENT TOTAL, both include them). A facility row that DOES print a
      current rent is treated as a revenue door instead and FLAGged - a door
      collecting rent is not non-revenue.
    * **"FLOOR TYPE" is bed/bath ("1/1.00"), not a plan name.** Plans are
      named `{bed}x{bath}` and only carry a size suffix (`1x1-650`) when one
      bed/bath count spans more than one Sq. FEET value in the same roll, so
      the plan roll-ups are always valid without inventing plan names.
      Bed, bath and sqft all come from the source: nothing here is estimated.
    * **ADV. RENT is market/asking rent, CURRENT RENT the in-place rent.**
      Both verbatim; an occupied door with no CURRENT RENT is never
      back-filled from ADV. RENT. No charge detail exists, so Other Income,
      concessions and deposits stay blank.
    * **No as-of date** (`asof_found = False`, `--asof` mandatory) and no
      totals row of any kind, so - like both owner-sheet parsers - `parse`
      ALWAYS re-derives the door count, the three occupancy buckets, total
      sqft, total market rent and total contract rent through an independent
      second pass over the raw sheet XML (`_reextract`) and files them as
      "re-extract" checks.
    * **Companion-statement cross-check.** This dialect ships the T-12 and the
      rent roll in ONE workbook. If a sibling sheet prints a
      "MARKET RENT TOTAL" row, the roll's total ADV. RENT is compared against
      it month by month: an exact match becomes a real reconciliation check
      naming the month(s), and no match becomes a FLAG naming the statement's
      values and the difference (never a silent pass).
    """

    name = "OwnerSimpleRow-xlsx"
    asof_found = False

    # All nine captions, in one row. Deliberately stricter than
    # OwnerSheetXlsxParser's four.
    HEADER_KEYS = ("complex", "apt #number", "floor type", "sq. feet",
                   "tenants", "move in", "lease expires", "adv. rent",
                   "current rent")

    # The Werner dialect's status vocabulary. One of these anywhere in the
    # Tenants column means the sheet is that dialect, not this one.
    STATUS_RE = re.compile(
        r"^(current|occupied|evict(ion|ed)?|notice(\s*-\s*\w+)?|"
        r"vacant(\s*-\s*\w+)?|vac|vacant-rented|vacant-unrented)$", re.I)

    # A Tenants cell naming a USE rather than a resident.
    FACILITY_RE = re.compile(
        r"^(office|shop|storage|store\s*room|workshop|warehouse|laundry|"
        r"leasing|model|clubhouse|club\s*house|maintenance|maint\.?|"
        r"mechanical|utility|guest|security|courtesy\s*officer|down|"
        r"non.?rev\w*|vacant\s*-\s*(office|shop|storage))\b", re.I)

    # Explicit month-to-month markers the owner may write into the Lease
    # Expires cell. "M2M" is the one Aldine uses; it must be here, because
    # the house rule allows MTM ONLY from an explicit marker and an
    # unrecognised marker would silently become "no lease expiration".
    MTM_RE = re.compile(
        r"^\s*(?:mtm|mtom|m-?t-?m|m\s*-?\s*2\s*-?\s*m|"
        r"month\s*-?\s*to\s*-?\s*month)\s*$", re.I)
    NON_UNIT = {"total", "totals", "grand total", "sum", "subtotal",
                "apt #number", "apt", "unit", "complex"}

    MIN_BLANK_TENANT_ROWS = 3

    def __init__(self):
        self.flags = []
        self.comments = []

    # -- small helpers (kept local so the two owner-sheet parsers stay
    #    independent of each other) -------------------------------------
    @staticmethod
    def _norm(v):
        return re.sub(r"\s+", " ", str(v if v is not None else "")).strip()

    @classmethod
    def _key(cls, v):
        return cls._norm(v).lower()

    @staticmethod
    def _num(v):
        if v is None:
            return None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return float(v)
        s = str(v).strip().replace(",", "").replace("$", "")
        if not s:
            return None
        neg = s.startswith("(") and s.endswith(")")
        try:
            f = float(s.strip("()"))
        except ValueError:
            return None
        return -f if neg else f

    @staticmethod
    def _date(v):
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        s = str(v or "").strip()
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    @staticmethod
    def _bb(v):
        """'1/1.00' -> (1, 1); '2/1.50' -> (2, 1.5); '' -> (None, None)."""
        m = re.match(r"^\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)",
                     str(v if v is not None else ""))
        if not m:
            return None, None

        def trim(x):
            f = float(x)
            return int(f) if f == int(f) else f
        return trim(m.group(1)), trim(m.group(2))

    @classmethod
    def _find_header(cls, rows):
        for i, r in enumerate(rows[:40]):
            vals = [cls._key(v) for v in r]
            if all(k in vals for k in cls.HEADER_KEYS):
                return i, {v: j for j, v in enumerate(vals) if v}
        return None, {}

    @classmethod
    def _sheets(cls, path):
        """[(sheet title, rows)] for every sheet, plus the index of the first
        one carrying the header row (or None)."""
        from openpyxl import load_workbook as _lw
        wb = _lw(path, read_only=True, data_only=True)
        out, hit = [], None
        for name in wb.sheetnames:
            rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
            out.append((name, rows))
            if hit is None and cls._find_header(rows)[0] is not None:
                hit = len(out) - 1
        return out, hit

    # -- detection ------------------------------------------------------
    @classmethod
    def _classify_tenant_column(cls, rows, hdr_i, cols):
        """-> (n_status_words, n_blank_with_rent) over the data rows."""
        jt, ju, jc = (cols.get("tenants"), cols.get("apt #number"),
                      cols.get("current rent"))
        n_status = n_blank_rent = 0
        for r in rows[hdr_i + 1:]:
            def cell(j):
                return r[j] if j is not None and j < len(r) else None
            unit = cls._norm(cell(ju))
            if not unit or unit.lower() in cls.NON_UNIT:
                continue
            t = cls._norm(cell(jt))
            if cls.STATUS_RE.match(t):
                n_status += 1
            elif not t and cls._num(cell(jc)) is not None:
                n_blank_rent += 1
        return n_status, n_blank_rent

    @staticmethod
    def detect_xlsx(path):
        cls = OwnerSimpleRowRentRollXlsxParser
        try:
            sheets, hit = cls._sheets(path)
        except Exception:
            return False
        if hit is None:
            return False
        _, rows = sheets[hit]
        hdr_i, cols = cls._find_header(rows)
        n_status, n_blank_rent = cls._classify_tenant_column(rows, hdr_i, cols)
        # Any Werner-style status word disqualifies this parser outright, and
        # the blank-tenant-with-rent signature must actually be present.
        return n_status == 0 and n_blank_rent >= cls.MIN_BLANK_TENANT_ROWS

    # -- independent second pass ----------------------------------------
    @classmethod
    def _reextract(cls, path):
        """Re-derive every total from the worksheet XML inside the .xlsx zip.

        Shares no code with the openpyxl pass in `parse` - not the cell
        reader, not the row assembler - so the reconciliation block is tying
        the output to the file rather than the parser to itself. The sheet
        prints no totals row at all, which is why this exists.
        """
        import zipfile
        import xml.etree.ElementTree as ET
        NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        REL = ("{http://schemas.openxmlformats.org/officeDocument/2006/"
               "relationships}")

        def col_of(ref):
            return re.match(r"[A-Za-z]+", ref or "").group(0)

        with zipfile.ZipFile(path) as z:
            names = z.namelist()
            shared = []
            if "xl/sharedStrings.xml" in names:
                sst = ET.fromstring(z.read("xl/sharedStrings.xml"))
                for si in sst.findall(f"{NS}si"):
                    shared.append("".join(t.text or ""
                                          for t in si.iter(f"{NS}t")))
            targets = []
            if "xl/workbook.xml" in names and \
                    "xl/_rels/workbook.xml.rels" in names:
                rels = {r.get("Id"): r.get("Target") for r in
                        ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))}
                wbx = ET.fromstring(z.read("xl/workbook.xml"))
                sheets_el = wbx.find(f"{NS}sheets")
                for sh in (list(sheets_el) if sheets_el is not None else []):
                    tgt = rels.get(sh.get(f"{REL}id"), "")
                    if tgt:
                        tgt = tgt.lstrip("/")
                        if not tgt.startswith("xl/"):
                            tgt = "xl/" + tgt
                        targets.append(tgt)
            if not targets:
                targets = sorted(n for n in names
                                 if n.startswith("xl/worksheets/sheet"))

            for tgt in targets:
                if tgt not in names:
                    continue
                grid = []
                for row in ET.fromstring(z.read(tgt)).iter(f"{NS}row"):
                    cells = {}
                    for c in row.findall(f"{NS}c"):
                        t = c.get("t")
                        if t == "inlineStr":
                            is_ = c.find(f"{NS}is")
                            val = "".join(x.text or ""
                                          for x in is_.iter(f"{NS}t")) \
                                if is_ is not None else None
                        else:
                            v = c.find(f"{NS}v")
                            val = v.text if v is not None else None
                            if val is not None and t == "s":
                                try:
                                    val = shared[int(val)]
                                except (ValueError, IndexError):
                                    pass
                        cells[col_of(c.get("r"))] = val
                    grid.append(cells)

                hdr, cols = None, {}
                for i, cells in enumerate(grid[:40]):
                    labels = {cls._key(v): k for k, v in cells.items() if v}
                    if all(k in labels for k in cls.HEADER_KEYS):
                        hdr, cols = i, labels
                        break
                if hdr is None:
                    continue

                out = {"units": 0, "occupied": 0, "vacant": 0, "nonrev": 0,
                       "sqft": 0.0, "market": 0.0, "contract": 0.0, "mtm": 0,
                       "plans": {}}
                for cells in grid[hdr + 1:]:
                    unit = cls._norm(cells.get(cols.get("apt #number")))
                    if not unit or unit.lower() in cls.NON_UNIT:
                        continue
                    ten = cls._norm(cells.get(cols.get("tenants")))
                    rent = cls._num(cells.get(cols.get("current rent")))
                    sf = cls._num(cells.get(cols.get("sq. feet"))) or 0.0
                    out["units"] += 1
                    out["sqft"] += sf
                    out["market"] += cls._num(
                        cells.get(cols.get("adv. rent"))) or 0.0
                    if cls.MTM_RE.match(
                            cls._norm(cells.get(cols.get("lease expires")))):
                        out["mtm"] += 1
                    bed, bath = cls._bb(cells.get(cols.get("floor type")))
                    key = (bed, bath, sf)
                    p = out["plans"].setdefault(
                        key, {"units": 0, "sqft": 0.0, "market": 0.0})
                    p["units"] += 1
                    p["sqft"] += sf
                    p["market"] += cls._num(
                        cells.get(cols.get("adv. rent"))) or 0.0
                    if cls.FACILITY_RE.match(ten) and rent is None:
                        out["nonrev"] += 1
                    elif rent is None:
                        out["vacant"] += 1
                    else:
                        out["occupied"] += 1
                        out["contract"] += rent
                return out
        return None

    # -- companion T-12 in the same workbook ----------------------------
    MARKET_ROW_RE = re.compile(r"^market\s+rent\s+total$", re.I)

    @classmethod
    def _companion_market_rent(cls, sheets, roll_idx):
        """[(sheet, [values])] for every 'MARKET RENT TOTAL' row found on a
        sheet other than the rent roll."""
        found = []
        for i, (name, rows) in enumerate(sheets):
            if i == roll_idx:
                continue
            for r in rows:
                if not r:
                    continue
                if cls.MARKET_ROW_RE.match(cls._norm(r[0])):
                    vals = [cls._num(v) for v in r[1:]]
                    found.append((name, [v for v in vals if v]))
        return found

    # -- parse ----------------------------------------------------------
    def parse(self, path):
        self.flags, self.comments = [], []
        sheets, hit = self._sheets(path)
        if hit is None:
            sys.exit("ERROR: owner one-row-per-unit rent roll header row "
                     "(COMPLEX / APT #NUMBER / FLOOR TYPE / Sq. FEET / "
                     "Tenants / Move In / Lease Expires / ADV. RENT / "
                     "CURRENT RENT) not found in any sheet of this workbook.")
        sheet, rows = sheets[hit]
        hdr_i, cols = self._find_header(rows)

        def cell(r, key):
            j = cols.get(key)
            return r[j] if j is not None and j < len(r) else None

        # ---- property name: title band above the header, else COMPLEX
        prop = ""
        band = [self._norm(v) for r in rows[:hdr_i] for v in r
                if self._norm(v)]
        for txt in band:
            m = re.match(r"^propert(?:y|ies)\s*:\s*(.+)$", txt, re.I)
            if m:
                prop = m.group(1)
                break
        if not prop:
            for txt in band:
                if re.search(r"apartments?|apts?\b|villas?|homes?|creek|"
                             r"place|park|manor|court|estates?", txt, re.I):
                    prop = txt
                    break
        if not prop and band:
            prop = band[-1]
        prop = re.split(r"\s+-\s+(?=\d)", prop, maxsplit=1)[0].strip()

        raw = []          # (unit, bed, bath, sqft, tenant, mi, exp, adv, cur)
        for r in rows[hdr_i + 1:]:
            unit = self._norm(cell(r, "apt #number"))
            if not unit or unit.lower() in self.NON_UNIT:
                continue
            bed, bath = self._bb(cell(r, "floor type"))
            raw.append({
                "unit": unit, "bed": bed, "bath": bath,
                "ft": self._norm(cell(r, "floor type")),
                "sqft": self._num(cell(r, "sq. feet")),
                "tenant": self._norm(cell(r, "tenants")),
                "mi": self._date(cell(r, "move in")),
                "exp_raw": self._norm(cell(r, "lease expires")),
                "exp": self._date(cell(r, "lease expires")),
                "adv": self._num(cell(r, "adv. rent")),
                "cur": self._num(cell(r, "current rent"))})

        # ---- floor-plan naming: "{bed}x{bath}", size-suffixed only when one
        #      bed/bath count spans more than one size in this roll.
        sizes = {}
        for d in raw:
            if d["bed"] is not None:
                sizes.setdefault((d["bed"], d["bath"]), set()).add(d["sqft"])

        def trim(x):
            return str(int(x)) if float(x) == int(float(x)) else str(x)

        def plan(d):
            if d["bed"] is None or d["bath"] is None:
                return ""
            code = f"{trim(d['bed'])}x{trim(d['bath'])}"
            if len(sizes.get((d["bed"], d["bath"]), ())) > 1 and d["sqft"]:
                code += f"-{int(round(d['sqft']))}"
            return code

        multi = [k for k, v in sizes.items() if len(v) > 1]
        if multi:
            self.flags.append(
                "bed/bath count(s) " + ", ".join(f"{trim(b)}/{trim(ba)}"
                                                 for b, ba in multi)
                + " span more than one Sq. FEET value - those floor plans are "
                  "named with the size suffix so the plan roll-ups stay valid")

        units, nonrev, vacant, occupied = [], [], [], []
        mtm_units, no_exp, named = [], [], []
        for d in raw:
            u = UnitRecord(
                unit=d["unit"], floor_plan=plan(d), sqft=d["sqft"],
                market_rent=d["adv"],
                bed_explicit=d["bed"], bath_explicit=d["bath"],
                bed_bath_explicit=True)
            if d["bed"] is None:
                self.flags.append(
                    f"unit {d['unit']}: FLOOR TYPE '{d['ft']}' is not a "
                    "bed/bath - Bed, Bath and Floor Plan left blank")

            is_fac = bool(self.FACILITY_RE.match(d["tenant"]))
            if is_fac and d["cur"] is not None:
                self.flags.append(
                    f"unit {d['unit']}: Tenants reads '{d['tenant']}' (a "
                    f"facility label) but the sheet also prints a CURRENT "
                    f"RENT of ${d['cur']:,.2f} - a door collecting rent is "
                    f"not a non-revenue door, so it is carried as OCCUPIED. "
                    f"Ownership should confirm.")
                is_fac = False

            if is_fac:
                u.apt_status = "NR"
                u.non_revenue = True
                u.non_revenue_use = d["tenant"]
                u.lease_type = f"Non-Revenue ({d['tenant']})"
                nonrev.append(d["unit"])
                units.append(u)
                continue

            is_mtm = bool(self.MTM_RE.match(d["exp_raw"]))
            exp = None if is_mtm else d["exp"]
            if d["exp_raw"] and not is_mtm and exp is None:
                self.flags.append(
                    f"unit {d['unit']}: Lease Expires reads '{d['exp_raw']}' "
                    "- not a date and not an MTM marker; left blank")

            if d["cur"] is None:
                # No contractual rent -> vacant. The dates (if any) are still
                # written to the deliverable; they are NOT read as occupancy.
                u.apt_status = "VU"
                vacant.append(d["unit"])
                if d["mi"] or exp:
                    self.flags.append(
                        f"unit {d['unit']}: VACANT (no CURRENT RENT) but the "
                        f"sheet still prints "
                        + " and ".join(
                            x for x in (
                                f"Move In {d['mi']:%m/%d/%Y}" if d["mi"]
                                else "",
                                f"Lease Expires {exp:%m/%d/%Y}" if exp
                                else "") if x)
                        + " - a vacated/notice door whose dates were left "
                          "behind. Occupancy in this dialect comes from "
                          "CURRENT RENT, never from the dates.")
                units.append(u)
                continue

            if d["tenant"]:
                named.append(d["unit"])
            u.apt_status = "OC"
            res = Resident(
                name=d["tenant"] if d["tenant"] else "",
                status="C",
                charges=[("RENT", d["cur"], False)],
                move_in=d["mi"],
                lease_start=d["mi"],     # the sheet prints one date only
                lease_expires=exp,
                term_type="MTM" if is_mtm else "")
            u.residents.append(res)
            occupied.append(d["unit"])
            if is_mtm:
                mtm_units.append(d["unit"])
            if exp is None and not is_mtm:
                no_exp.append(d["unit"])
            if d["cur"] == 0:
                self.flags.append(
                    f"unit {d['unit']}: CURRENT RENT is printed as 0 - "
                    "counted as occupied at $0 (employee/model rent?), not "
                    "as vacant. Ownership should confirm.")
            units.append(u)

        dupes = sorted({u.unit for u in units
                        if [x.unit for x in units].count(u.unit) > 1})
        if dupes:
            self.flags.append("duplicate unit id(s): " + ", ".join(dupes))
        if nonrev:
            byuse = {}
            for d in raw:
                if d["unit"] in nonrev:
                    byuse.setdefault(d["tenant"], []).append(d["unit"])
            txt = "; ".join(f"{k}: {', '.join(v)}"
                            for k, v in sorted(byuse.items()))
            self.flags.append(
                f"{len(nonrev)} NON-REVENUE door(s) - the Tenants column "
                f"names a facility use instead of a resident ({txt}). They "
                f"are marked Occupancy Status 'Non-Rev': NOT vacant (they do "
                f"not belong in the vacancy calculation) and NOT occupied. "
                f"Their ADV. RENT is still carried, because the sheet's own "
                f"market-rent total includes them.")
            self.comments.append(
                f"Non-revenue units ({len(nonrev)} of {len(units)}): {txt}. "
                f"The rent roll's Tenants column names a facility use rather "
                f"than a resident on these doors and prints no CURRENT RENT. "
                f"They are reported as Occupancy Status 'Non-Rev' - they are "
                f"not vacancies and are excluded from both the occupied and "
                f"the vacant counts. Their ADV. RENT (market rent) is "
                f"carried, since the source's own market-rent total includes "
                f"them.")
        if named:
            self.flags.append(
                "the Tenants column carries text that is neither blank nor a "
                "recognised facility label on: " + ", ".join(named)
                + " - read as a resident name")
        if no_exp:
            self.flags.append(
                "occupied with NO Lease Expires printed and no MTM marker: "
                + ", ".join(no_exp)
                + " - Lease Expiration left blank and MTM left blank (MTM is "
                  "never inferred from a missing or expired date)")
        if mtm_units:
            self.flags.append(
                "month-to-month per an explicit marker in Lease Expires: "
                + ", ".join(mtm_units)
                + " - MTM = Yes and Lease Expiration left blank (no date "
                  "exists). This is the source STATING a month-to-month "
                  "term, not an inference from an expired date.")
            self.comments.append(
                "Month-to-month: " + ", ".join(mtm_units)
                + ". The rent roll writes an explicit month-to-month marker "
                  "in the Lease Expires cell for "
                + ("these doors" if len(mtm_units) > 1 else "this door")
                + ", so MTM = Yes and Lease Expiration is left blank. MTM is "
                  "never inferred from a missing or expired lease date.")
        self.flags.append(
            "the workbook prints NO as-of date - the date in the filename and "
            "in cell B2 came from --asof and is an INFERRED date, not a "
            "sourced one")
        self.comments.append(
            "AS-OF DATE IS INFERRED. This rent roll prints no as-of date "
            "anywhere; the date shown in cell B2 and in the file name was "
            "supplied to the processor (--asof) and derived from the lease "
            "dates on the roll itself. It is not a date stated by ownership.")

        # ---- checks -------------------------------------------------------
        checks, src = {}, {}
        rx = self._reextract(path)
        if not rx:
            sys.exit("ERROR: the independent re-extraction pass found no rent "
                     "roll sheet - refusing to ship an unreconciled workbook.")
        checks["unit_count"] = rx["units"]
        checks["occupied_count"] = rx["occupied"]
        checks["vacant_count"] = rx["vacant"]
        checks["nonrev_count"] = rx["nonrev"]
        checks["total_sqft"] = rx["sqft"]
        checks["total_market_rent"] = rx["market"]
        checks["total_contract_rent"] = rx["contract"]
        for k in ("unit_count", "occupied_count", "vacant_count",
                  "total_sqft", "total_market_rent", "total_contract_rent"):
            src[k] = "re-extract"
        checks["_src"] = src

        extra = [
            ("MTM units (explicit marker)",
             lambda us: sum(1 for u in us for r in u.residents
                            if re.search(r"MTM", r.term_type or "", re.I)),
             rx["mtm"], 0.5),
        ]
        # per-plan counts / sqft / market rent, from the independent pass
        for (bed, bath, sf), want in sorted(
                rx["plans"].items(), key=lambda kv: (str(kv[0][0]),
                                                     str(kv[0][1]), kv[0][2])):
            code = plan({"bed": bed, "bath": bath, "sqft": sf})
            extra += [
                (f"  plan {code}: units",
                 lambda us, c=code: sum(1 for u in us if u.floor_plan == c),
                 want["units"], 0.5),
                (f"  plan {code}: sq ft",
                 lambda us, c=code: sum(u.sqft or 0 for u in us
                                        if u.floor_plan == c),
                 want["sqft"], 0.01),
                (f"  plan {code}: market rent",
                 lambda us, c=code: sum(u.market_rent or 0 for u in us
                                        if u.floor_plan == c),
                 want["market"], 0.01),
            ]

        # ---- companion statement in the same workbook ---------------------
        comp = self._companion_market_rent(sheets, hit)
        total_adv = rx["market"]
        for name, vals in comp:
            match = [v for v in vals if abs(v - total_adv) <= 0.01]
            uniq = sorted(set(round(v, 2) for v in vals))
            if match:
                extra.append(
                    (f"Total ADV. RENT vs sheet '{name}' MARKET RENT TOTAL "
                     f"({len(match)} of {len(vals)} reported month(s))",
                     lambda us: sum(u.market_rent or 0 for u in us),
                     match[0], 0.01))
                self.comments.append(
                    f"Total market rent (ADV. RENT) of {total_adv:,.0f} ties "
                    f"exactly to the companion operating statement's "
                    f"\"MARKET RENT TOTAL\" line on sheet '{name}' in "
                    f"{len(match)} of its {len(vals)} reported months.")
            else:
                closest = min(uniq, key=lambda v: abs(v - total_adv))
                self.flags.append(
                    f"companion statement sheet '{name}' prints MARKET RENT "
                    f"TOTAL of {', '.join(f'{v:,.0f}' for v in uniq)} in the "
                    f"months it reports, none of which equals the rent "
                    f"roll's total ADV. RENT of {total_adv:,.0f} (closest "
                    f"{closest:,.0f}, difference {total_adv - closest:+,.0f}) "
                    f"- the roll and the statement disagree about market "
                    f"rent; the rent roll is the authority for the "
                    f"deliverable.")
                self.comments.append(
                    f"The companion operating statement in this workbook "
                    f"(sheet '{name}') prints a \"MARKET RENT TOTAL\" of "
                    f"{', '.join(f'{v:,.0f}' for v in uniq)} for the months "
                    f"it reports, versus {total_adv:,.0f} of ADV. RENT on the "
                    f"rent roll ({total_adv - closest:+,.0f} vs the closest). "
                    f"The rent roll is the authority for the rent-roll "
                    f"deliverable.")
        checks["extra_checks"] = extra
        return prop, None, units, checks

    def source_note(self, asof):
        d = asof.strftime("%m/%d/%Y") if asof else "unknown date"
        return (f"Generated from an owner/PM-prepared one-row-per-unit rent "
                f"roll spreadsheet (xlsx) as of {d}; the workbook prints no "
                f"as-of date, so it was supplied with --asof (an INFERRED "
                f"date - see the Comments tab). Market Rent = the sheet's "
                f"ADV. RENT column; Contractual Rent = its CURRENT RENT "
                f"column. Bed/Bath come from FLOOR TYPE and Net Sf from "
                f"Sq. FEET - both from the source, neither estimated - and "
                f"Floor Plan is bed x bath. Occupancy is read from CURRENT "
                f"RENT: the sheet prints no status column, so a door with a "
                f"contractual rent is occupied and a door without one is "
                f"vacant regardless of any lease dates left on the row. A "
                f"Tenants cell naming a facility use (Office / Shop / "
                f"Storage) is a NON-REVENUE door, not a vacancy, and not a "
                f"resident name. MTM = Yes only where the sheet writes an "
                f"explicit MTM marker; it is never inferred. The sheet "
                f"prints one 'Move In' date per unit, used for both Move In "
                f"Date and Lease Start Date. It carries no charge detail, so "
                f"concessions, other income and deposits are intentionally "
                f"blank.")
