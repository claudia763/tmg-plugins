#!/usr/bin/env python3
"""
apply_comp_corrections.py -- apply primary-source corrections to TMG's sale-comp
universe before `select_comps.py` runs, so a verified grid stays reproducible.

WHY THIS EXISTS
    `sales-comps/SKILL.md` says "edit constants there rather than hand-picking
    comps; keep the pipeline reproducible." But
    `instructions/sales-comps-from-an-address-only-8-2026.md` section 6b requires
    web-verifying every selected comp, and roughly half come back with a defect:
    a price that disagrees with the recorded deed, a wrong sale date, a wrong
    address, or a product-type mismatch.

    Hand-editing the exported grid would break reproducibility. This script keeps
    the correction as DATA: a reviewed CSV of primary-source findings, applied to
    the universe workbook, with every change logged. Re-running the pipeline on
    the corrected workbook gives a "Verified" grid alongside the "As-Run" one,
    which is exactly the two-grid delivery section 6b asks for.

INPUT
    --universe    the "All Sales Comps.xlsx" produced by extract_all_sale_comps.py
    --corrections a CSV, one row per comp to correct or drop:

        property_name,sale_date_match,action,field,new_value,source
          property_name    matched case-insensitively against "Property Name"
          sale_date_match  YYYY-MM-DD of the row to touch; blank = any row with
                           that name (use when the name is unique)
          action           "set" or "exclude"
          field            column to overwrite when action=set
          new_value        the verified value
          source           URL / record id justifying the change (logged, required)

    Excluded rows are dropped from the universe entirely. "Sold Price/Unit" is
    NOT auto-recomputed -- state it explicitly if the verified per-unit differs,
    because the database's stated column and price/units genuinely disagree on
    ~31% of rows and which one is right must be a human judgment.

OUTPUT
    --out  a corrected universe workbook in the identical schema, plus a printed
           change log. Nothing is changed silently.

    python apply_comp_corrections.py --universe "All Sales Comps.xlsx" \
        --corrections corrections.csv --out "All Sales Comps (verified).xlsx"
"""
import argparse
import csv
import sys

import openpyxl


def norm(v):
    return str(v or "").strip().lower()


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("--universe", required=True)
    ap.add_argument("--corrections", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.universe)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else ""
              for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header)}

    rules = list(csv.DictReader(open(args.corrections, newline="")))
    for r in rules:
        if not r.get("source", "").strip():
            sys.exit(f"ERROR: correction for {r.get('property_name')!r} has no "
                     f"source — every change must cite a primary record")
        if r["action"] == "set" and r["field"] not in col:
            sys.exit(f"ERROR: unknown column {r['field']!r}")

    drop_rows, changes, unmatched = set(), [], []
    for r in rules:
        name, want_date = norm(r["property_name"]), r.get("sale_date_match", "").strip()
        hit = 0
        for row in range(2, ws.max_row + 1):
            if norm(ws.cell(row, col["Property Name"]).value) != name:
                continue
            if want_date:
                sd = ws.cell(row, col["Sale Date"]).value
                sd = sd.date().isoformat() if hasattr(sd, "date") else str(sd)[:10]
                if sd != want_date:
                    continue
            hit += 1
            if r["action"] == "exclude":
                drop_rows.add(row)
                changes.append(f"  EXCLUDE  {r['property_name']}  ({r['source']})")
            else:
                cell = ws.cell(row, col[r["field"]])
                old, new = cell.value, r["new_value"]
                try:
                    new = float(new) if "." in new else int(new)
                except ValueError:
                    pass
                cell.value = new
                changes.append(f"  SET      {r['property_name']}: "
                               f"{r['field']} {old!r} -> {new!r}  ({r['source']})")
        if not hit:
            unmatched.append(r["property_name"])

    for row in sorted(drop_rows, reverse=True):
        ws.delete_rows(row)

    wb.save(args.out)
    print(f"applied {len(changes)} correction(s) to {args.universe}")
    for c in changes:
        print(c)
    if unmatched:
        print("WARNING: no row matched -> " + ", ".join(sorted(set(unmatched))))
    print(f"wrote {args.out} ({ws.max_row - 1} comps)")


if __name__ == "__main__":
    main()
