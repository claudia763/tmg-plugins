#!/usr/bin/env python3
"""Neutralise the model's =WEBSERVICE() FRED pulls so LibreOffice can recalc headless,
and set the yellow manual-override curve for THIS deal.

WHY: 'Treasury Yields'!D4:D8 are `=IFERROR(_xludf.webservice(C4),"")` live pulls from
fred.stlouisfed.org. LibreOffice evaluates WEBSERVICE during calculateAll and blocks on
the socket -- soffice.bin sits at 0% CPU indefinitely and the driving macro never runs.
(Shady Oaks 8/7/2026: two runs killed after 10+ minutes each, 14 s CPU total.)

The tab is DESIGNED for this: E4:E8 parse the CSV response, and
G4:G8 = IF(ISNUMBER(E),E,F) fall back to the yellow manual-override column F whenever the
pull yields a non-number. Writing "" into D4:D8 therefore takes the documented fallback
path rather than defeating the model -- the rates used become exactly column F, which we
set from the loan-terms lookup and stamp with an as-of date in H.

RATES ARE PER-DEAL. Earlier revisions of this script hardcoded one job's curve, which
silently carried a stale 8/6/2026 Shady Oaks curve into the next deal and forced an
agent to fork a local copy. Pass the curve on the command line instead; the hardcoded
DEFAULTS below are only a last-resort fallback and print a warning when used.

Row layout is fixed in the 8/2026 template and asserted on the column-A labels:
  row 4 UST 5 Yr | row 5 UST 7 Yr | row 6 UST 10 Yr | row 7 UST 30 Yr | row 8 30-Day Avg SOFR
(Note: column A holds the labels; B is the FRED series id and C the request URL.)

INPUT   a TMG model .xlsx
OUTPUT  the same workbook with D4:D8 literal "" and F/H refreshed
Usage:
    python3 fix_treasury_yields.py in.xlsx out.xlsx \
        --asof 08/05/2026 --ust5 4.33 --ust7 4.47 --ust10 4.63 --ust30 5.17 --sofr30 3.622

Rates may be given as percents (4.33) or decimals (0.0433) -- anything >= 0.5 is read
as a percent and divided by 100. Omit a rate to leave the template's existing override
in place for that tenor.
"""
import argparse
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl

ROWS = {
    "ust5": (4, "UST 5 Yr"),
    "ust7": (5, "UST 7 Yr"),
    "ust10": (6, "UST 10 Yr"),
    "ust30": (7, "UST 30 Yr"),
    "sofr30": (8, "30-Day Avg SOFR"),
}

# Last-resort fallback only (8/5/2026 curve, Pointe at Garden Oaks). Passing --asof
# plus explicit rates is always preferred; using these emits a warning.
DEFAULTS = {"ust5": 4.33, "ust7": 4.47, "ust10": 4.63, "ust30": 5.17, "sofr30": 3.622}
DEFAULT_ASOF = "08/05/2026"


def as_decimal(v):
    """Accept 4.33 or 0.0433; anything >= 0.5 is treated as a percent."""
    if v is None:
        return None
    # round(..., 8) keeps 5.17 -> 0.0517 rather than 0.051699999999999996
    return round(v / 100.0, 8) if v >= 0.5 else v


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("src")
    p.add_argument("dst")
    p.add_argument("--asof", help="as-of date stamped into column H, e.g. 08/05/2026")
    for k in ROWS:
        p.add_argument(f"--{k}", type=float, help=f"{ROWS[k][1]} yield")
    a = p.parse_args(argv)

    rates = {k: getattr(a, k) for k in ROWS}
    asof = a.asof
    if all(v is None for v in rates.values()) and asof is None:
        print(f"WARNING: no rates given; falling back to the hardcoded {DEFAULT_ASOF} "
              f"curve. Pass --asof and explicit rates for this deal.", file=sys.stderr)
        rates, asof = dict(DEFAULTS), DEFAULT_ASOF
    if asof is None:
        p.error("--asof is required whenever any rate is supplied")

    wb = openpyxl.load_workbook(a.src)
    ws = wb["Treasury Yields"]

    killed = 0
    for key, (row, label) in ROWS.items():
        actual = ws.cell(row, 1).value
        assert actual == label, f"row {row} is {actual!r}, expected {label!r}"
        ws.cell(row, 4).value = ""             # D: kill the WEBSERVICE call
        killed += 1
        rate = as_decimal(rates[key])
        if rate is not None:
            ws.cell(row, 6).value = rate       # F: manual override
            ws.cell(row, 8).value = asof       # H: as-of
        print(f"  {label:16s} rate used -> {ws.cell(row, 6).value}  "
              f"as-of {ws.cell(row, 8).value}")

    left = sum(
        1 for w in wb.worksheets for r in w.iter_rows() for c in r
        if isinstance(c.value, str) and "webservice" in c.value.lower()
        and c.value.startswith("=")
    )
    wb.save(a.dst)
    print(f"cleared {killed} WEBSERVICE formula cells; {left} WEBSERVICE formulas remain")
    print(f"wrote {a.dst}")


if __name__ == "__main__":
    main()
