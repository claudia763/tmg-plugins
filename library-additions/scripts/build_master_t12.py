#!/usr/bin/env python3
"""Combine two QuickBooks-Online P&L exports into ONE master T-12 source.

WHAT IT DOES
    Reads two QuickBooks Online "Profit and Loss" .xlsx exports (two entities,
    two charts of accounts, possibly two different reporting spans) and writes
    a single combined statement in the SAME QuickBooks-Online xlsx layout, so
    that `process_t12.py`'s registered `parse_t12_quickbooks_xlsx` reads the
    master exactly like a native QuickBooks export - no hand-built workbook, no
    one-off processing (toolkit iron rule #4/#5).

INPUTS  (see SOURCES below)
    Westlake East  - "V Westlake LLC" P&L, Oct 2025 - Jul 2026 (10 months)
    Westlake West  - "Westlake West"  P&L, Aug 2025 - Jul 2026 (12 months)

OUTPUT
    "Westlake Master T-12 Jul 26.xlsx" - A1 entity, A2 "Profit and Loss",
    A3 period caption, row 5 month headers + Total, Income section,
    "Total for Income", "Gross Profit", Expenses section, "Total for
    Expenses", "Net Operating Income", trailing basis line.

--------------------------------------------------------------------------
RULE 1 - THE OVERLAPPING-PERIOD RULE (the judgment call)
--------------------------------------------------------------------------
    A combined statement may only cover the months BOTH properties' books
    cover. Westlake West has 12 months (Aug-25 - Jul-26); Westlake East's
    ownership books (V Westlake LLC) begin 10/1/2025, so East has 10
    (Oct-25 - Jul-26). A combined Aug-25 or Sep-25 column would be West-only,
    which would make the "12-month" total NOT a combined run rate - it would
    silently understate two months of the combined asset and overstate
    nothing, i.e. an invented number.

    So the MASTER covers the 10 shared months, Oct 2025 - Jul 2026. East is
    NOT zero-filled, West's Aug-25/Sep-25 are NOT carried, and nothing is
    annualized or grossed up. The master is then processed with
    `--allow-partial --pad-to-12`, which displays it on the full Aug-25 -
    Jul-26 trailing axis with Aug-25 and Sep-25 GENUINELY BLANK (never zeros)
    plus a Comments-tab note naming the gap - the standing house rule for a
    short statement.

--------------------------------------------------------------------------
RULE 2 - MERGE BY ACCOUNT NAME, NEVER BY GL NUMBER
--------------------------------------------------------------------------
    Two QuickBooks files are two independent charts of accounts and they
    REUSE the same numbers for different things:

        50400   East = "Software"        West = "Bank Fees"
        50610   East = "Electricity"     West = "HVAC Repair"
        50620   East = "Water & sewer"   West = (unused; water is 50720)
        50640   East = "Trash Removal"   West = "General Repairs"
        50650   East = "Landscaping & Pest Control"  West = "Shipping Container"

    Merging on the number would cross-post electricity into HVAC repair and
    trash into general repairs. Every merge below is declared explicitly by
    NAME, and the script refuses to run if any source account is not
    explicitly merged, carried through, or dropped (see `verify_coverage`).

    Where the two charts genuinely disagree about granularity the lines are
    KEPT SEPARATE rather than forced together: East reports one
    "50650 Landscaping & Pest Control" line while West splits
    "50550 Pest Control" and "50750 Landscaping". Keeping all three carries
    every source number through unchanged; merging them would require
    splitting East's line, which is an invented number.

--------------------------------------------------------------------------
RULE 3 - PROVE IT
--------------------------------------------------------------------------
    Nothing is written until every check passes:
      * every source account is merged / carried / dropped exactly once;
      * every merged line equals the sum of its named components, month by
        month and in total;
      * the master's Total for Income / Gross Profit / Total for Expenses /
        Net Operating Income equal East + West over the shared months,
        computed from the merged detail;
      * the master's control rows are compared to each source's own PRINTED
        control rows and any difference is reported in full (Westlake East's
        printed "Total for Expenses" omits its own "Make Ready Cleaning"
        line, $1,185.00 - the monthly detail wins, house rule).
"""

# ---------------------------------------------------------------------------
# HOW TO REUSE THIS SCRIPT ON A NEW DEAL
# ---------------------------------------------------------------------------
# Everything deal-specific lives in the CONFIGURATION block below - SOURCES,
# OUT_PATH, OUT_ENTITY, DROP, INCOME_MAP, EXPENSE_MAP. Edit those; the reader,
# the checks and the writer are generic.
#
#   1. Run `process_t12.py` on EACH source first and let it print the account
#      list. Copy the account names verbatim into INCOME_MAP / EXPENSE_MAP -
#      spelling, GL prefix, case and typos included ("50930 Quicbkooks" is a
#      real account name). `verify_coverage()` aborts on a name that does not
#      exist and on any source account you forgot, so a stale map cannot
#      silently drop money.
#   2. Merge by NAME. Never by GL number - see the module docstring, RULE 2.
#   3. Keep genuinely different granularity SEPARATE (one property reports
#      "Landscaping & Pest Control" as one line while the other splits them):
#      carrying all three lines preserves every source number exactly, while
#      forcing them together would require splitting a line, i.e. inventing a
#      number.
#   4. Do not "fix" anything on the way through. Suspect lines (a possible
#      duplicate insurance booking, an owner's margin note about double
#      billing) are carried unchanged and reported to the Comments tab via
#      `process_t12.py --header-note`; ownership confirms, not the agent.
#   5. Process the master with `--allow-partial --pad-to-12` when the shared
#      period is short - see RULE 1.
#
# Companion notes: instructions/combining-two-property-t12s.md and
# instructions/quickbooks-online-xlsx-t12-parser.md.


from __future__ import annotations

import os
import re
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# CONFIGURATION - this is the only block you edit per deal
# ---------------------------------------------------------------------------

HERE = os.path.dirname(os.path.abspath(__file__))
JOB = os.environ.get("T12_JOB_DIR", os.path.dirname(HERE))

SOURCES = {
    "East": os.path.join(JOB, "inbox", "Westlake East P_L T10.xlsx"),
    "West": os.path.join(JOB, "inbox", "WLW T12 Jul 26.xlsx"),
}
OUT_PATH = os.path.join(HERE, "Westlake Master T-12 Jul 26.xlsx")
OUT_ENTITY = "Westlake East + Westlake West (Combined)"

# Rows whose only value is 0/blank: QuickBooks parent-account headers with no
# postings of their own. Structure, not money.
DROP = {("East", "50600 Utilities"), ("East", "50620 Water")}

# (master line name, [(source, source account name), ...])
# ORDER IS THE OUTPUT ORDER.
INCOME_MAP = [
    ("Total Net Collections", [("East", "Total Net Collections"),
                               ("West", "Total Net Collections")]),
    ("Application Fees",      [("East", "42000 Application fees"),
                               ("West", "42000 application fees")]),
]

EXPENSE_MAP = [
    ("Payroll Expenses",                 [("East", "50100 Payroll Labor"),
                                          ("West", "50000 Payroll Expenses")]),
    ("50200 Advertising and Marketing",  [("West",
                                           "50200 Advertising and Marketing")]),
    ("Property Management",              [("East", "Property Management"),
                                          ("West", "Property Management")]),
    ("Software",                         [("East", "50400 Software"),
                                          ("West", "50325 - Software")]),
    ("Quickbooks",                       [("East", "50930 Quicbkooks"),
                                          ("West", "50327 Quickbooks")]),
    ("50400 Bank Fees",                  [("West", "50400 Bank Fees")]),
    ("Repairs & Maintenance",            [("East",
                                           "50500 Repairs & Maintenance"),
                                          ("West",
                                           "50630 Repairs & maintenance")]),
    ("50640 General Repairs",            [("West", "50640 General Repairs")]),
    ("50610 HVAC Repair",                [("West", "50610 HVAC Repair")]),
    ("50650 Shipping Container",         [("West", "50650 Shipping Container")]),
    ("50550 Pest Control",               [("West", "50550 Pest Control")]),
    ("50750 Landscaping",                [("West", "50750 Landscaping")]),
    ("50650 Landscaping & Pest Control", [("East",
                                           "50650 Landscaping & Pest Control")]),
    ("Electricity",                      [("East", "50610 Electricity"),
                                          ("West", "50710 Electricity")]),
    ("Water & Sewer",                    [("East", "50620 Water & sewer"),
                                          ("West", "50720 Water")]),
    ("Waste Removal",                    [("East", "50640 Trash Removal"),
                                          ("West", "50740 Waste Removal")]),
    ("Make Ready Labor",                 [("East", "50710 Make Ready Labor"),
                                          ("West", "50840 make ready labor")]),
    ("Make Ready Cleaning",              [("East", "Make Ready Cleaning"),
                                          ("West", "50830 Make Ready Cleaning")]),
    ("50820 Make Ready Materials",       [("West",
                                           "50820 Make Ready Materials")]),
    ("50910 Legal Expenses",             [("West", "50910 Legal Expenses")]),
    ("General business expenses",        [("East", "General business expenses")]),
    # BOTH of East's insurance lines are kept (ownership has not confirmed the
    # second one is a duplicate, so it is never dropped - house rule) and they
    # merge into the one combined Insurance line.
    ("Insurance",                        [("East", "50800 Insurance"),
                                          ("East", "Insurance"),
                                          ("West", "52000 Insurance")]),
    ("Property Tax",                     [("East", "51000 Property Tax"),
                                          ("West", "51000 Property Tax")]),
]

TOL = 0.005


def normalize(path):
    """Iron rule #5: any deliverable .xlsx goes through the toolkit's
    normalized save path (`_normalize_xlsx`: inline strings -> sharedStrings,
    broken defined names purged). Raw openpyxl output crashes Excel/JS
    loaders. The master source workbook IS a deliverable, so it gets the same
    treatment as the processed T-12s."""
    sys.path.insert(0, os.path.join(JOB, "toolkit"))
    try:
        from process_t12 import _normalize_xlsx
    except Exception as exc:                       # pragma: no cover
        sys.exit(f"ERROR: cannot import the toolkit's _normalize_xlsx "
                 f"({exc}); refusing to ship a non-normalized workbook.")
    _normalize_xlsx(path)

# ---------------------------------------------------------------------------
# QuickBooks-Online xlsx reader (same structural rules as
# process_t12.parse_t12_quickbooks_xlsx - kept standalone so this script has
# no dependency on the toolkit's import path)
# ---------------------------------------------------------------------------

MONTH_RE = re.compile(r"^([A-Za-z]{3,9})\.?[-/. ]+(\d{4})$")
TOTAL_FOR_RE = re.compile(r"^Total\s+for\s+(.+)$", re.I)
CONTROL_RE = re.compile(r"^(gross profit|net operating income|net income"
                        r"|net other income|net revenue)$", re.I)
SECTION_RE = re.compile(r"^(income|revenues?|expenses?|cost of goods sold"
                        r"|other income|other expenses?)$", re.I)
FOOTER_RE = re.compile(r"^(cash|accrual)\s+basis\b|^page \d+\b", re.I)


def txt(v):
    return re.sub(r"\s+", " ", str(v if v is not None else "")).strip()


def num(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace(",", "").replace("$", "")
        if not s:
            return None
        neg = s.startswith("(") and s.endswith(")")
        try:
            f = float(s.strip("()"))
        except ValueError:
            return None
        return -f if neg else f
    if isinstance(v, datetime):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def month_of(v):
    if isinstance(v, datetime):
        return datetime(v.year, v.month, 1)
    m = MONTH_RE.match(txt(v))
    if not m:
        return None
    for cand in (m.group(1), m.group(1)[:3]):
        for fmt in ("%B", "%b"):
            try:
                return datetime.strptime(f"{cand.title()} {m.group(2)}",
                                         f"{fmt} %Y")
            except ValueError:
                pass
    return None


def read_qbo(path):
    """-> dict(entity, months, accounts{name: {month: value|None}},
              order[names], controls{label: {month: value}}, printed_total)."""
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    hdr_i = mcols = months = tot_col = None
    for i, r in enumerate(rows[:15]):
        if not r:
            continue
        got = [(j, month_of(v)) for j, v in enumerate(r) if j >= 1]
        got = [(j, d) for j, d in got if d]
        if len(got) < 2:
            continue
        hdr_i, mcols, months = i, [j for j, _ in got], [d for _, d in got]
        for j, v in enumerate(r):
            if j > max(mcols) and txt(v).lower() == "total":
                tot_col = j
                break
        break
    if hdr_i is None or tot_col is None:
        sys.exit(f"ERROR: {os.path.basename(path)} is not a QuickBooks-Online "
                 f"P&L export (no month header row / no Total column).")

    accounts, order, controls, printed_tot = {}, [], {}, {}
    for r in rows[hdr_i + 1:]:
        name = txt(r[0] if r else None)
        if not name:
            continue                      # blank label -> not a statement row
        if FOOTER_RE.match(name):
            break                         # basis/timestamp line ends the block
        # only the month columns and the Total column are ever read, so stray
        # marginal cells (notes, a bare 7500 in K40) cannot enter the numbers
        vals = [num(r[j]) if j < len(r) else None for j in mcols]
        annual = num(r[tot_col]) if tot_col < len(r) else None
        has = any(v is not None for v in vals)
        low = name.lower()

        if TOTAL_FOR_RE.match(name) or CONTROL_RE.match(low):
            key = low
            controls[key] = {m: (v or 0.0) for m, v in zip(months, vals)}
            printed_tot[key] = annual
            continue
        if not has and (annual is None or abs(annual) <= TOL):
            continue                      # section head / empty parent account
        if name in accounts:
            sys.exit(f"ERROR: {os.path.basename(path)} lists {name!r} twice.")
        accounts[name] = dict(zip(months, vals))
        printed_tot[name] = annual
        order.append(name)

    return {"entity": txt(rows[0][0]) if rows and rows[0] else "",
            "caption": txt(rows[2][0]) if len(rows) > 2 and rows[2] else "",
            "months": months, "accounts": accounts, "order": order,
            "controls": controls, "printed": printed_tot,
            "path": path}


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------

def verify_coverage(src):
    """Every source account must be merged, carried through, or dropped."""
    used = set()
    for _, comps in INCOME_MAP + EXPENSE_MAP:
        for s, a in comps:
            if a not in src[s]["accounts"]:
                sys.exit(f"ERROR: {s} has no account named {a!r} - the merge "
                         f"map is stale. Source accounts: "
                         f"{sorted(src[s]['accounts'])}")
            if (s, a) in used:
                sys.exit(f"ERROR: {s} {a!r} is used twice in the merge map.")
            used.add((s, a))
    missing = []
    for s in src:
        for a in src[s]["order"]:
            if (s, a) not in used and (s, a) not in DROP:
                missing.append((s, a))
    if missing:
        sys.exit("ERROR: these source accounts are neither merged, carried "
                 "through nor dropped - refusing to silently lose money:\n  "
                 + "\n  ".join(f"{s}: {a}" for s, a in missing))
    for s, a in DROP:
        tot = sum(v or 0.0 for v in src[s]["accounts"].get(a, {}).values())
        if abs(tot) > TOL:
            sys.exit(f"ERROR: {s} {a!r} is on the DROP list but carries "
                     f"{tot:,.2f} - it is not an empty parent row.")
    print(f"  coverage: {len(used)} source account line(s) merged/carried, "
          f"{len(DROP)} empty parent row(s) dropped, none unaccounted for.")


def shared_months(src):
    sets = [set((d.year, d.month) for d in s["months"]) for s in src.values()]
    keep = sorted(set.intersection(*sets))
    return [datetime(y, m, 1) for y, m in keep]


def merge_line(src, comps, months):
    """-> [value|None per month]. None only when EVERY component is blank in
    that month (QuickBooks prints nothing when nothing posted)."""
    out = []
    for m in months:
        parts = [src[s]["accounts"][a].get(m) for s, a in comps]
        out.append(None if all(p is None for p in parts)
                   else sum(p or 0.0 for p in parts))
    return out


def main():
    src = {k: read_qbo(p) for k, p in SOURCES.items()}
    for k, s in src.items():
        print(f"{k}: {s['entity']!r} | {len(s['months'])} months "
              f"{s['months'][0]:%b %Y}-{s['months'][-1]:%b %Y} | "
              f"{len(s['order'])} accounts | caption {s['caption']!r}")

    verify_coverage(src)

    months = shared_months(src)
    print(f"  shared period: {len(months)} months "
          f"{months[0]:%b %Y} - {months[-1]:%b %Y} "
          f"(the overlapping-period rule - see module docstring)")
    for k, s in src.items():
        skipped = [f"{d:%b %Y}" for d in s["months"]
                   if (d.year, d.month) not in
                   {(m.year, m.month) for m in months}]
        if skipped:
            print(f"  {k}: {len(skipped)} month(s) NOT carried into the "
                  f"master ({', '.join(skipped)}) - the other property's "
                  f"books do not cover them.")

    income = [(nm, merge_line(src, comps, months)) for nm, comps in INCOME_MAP]
    expense = [(nm, merge_line(src, comps, months))
               for nm, comps in EXPENSE_MAP]

    # ---- every merged line = the sum of its named components ---------------
    for group, mapping in ((income, INCOME_MAP), (expense, EXPENSE_MAP)):
        for (nm, vals), (_, comps) in zip(group, mapping):
            want = sum(sum(v or 0.0
                           for m, v in src[s]["accounts"][a].items()
                           if (m.year, m.month) in
                           {(x.year, x.month) for x in months})
                       for s, a in comps)
            got = sum(v or 0.0 for v in vals)
            if abs(got - want) > TOL:
                sys.exit(f"ERROR: merged line {nm!r} sums to {got:,.2f} but "
                         f"its components sum to {want:,.2f}")
    print(f"  line check: {len(income) + len(expense)} merged/carried line(s) "
          f"each equal the sum of their named components - all tie.")

    # ---- control rows, computed from the merged detail ---------------------
    def col(rows, m):
        return sum((dict(zip(months, v)).get(m) or 0.0) for _, v in rows)

    tot_inc = {m: col(income, m) for m in months}
    tot_exp = {m: col(expense, m) for m in months}
    gross = dict(tot_inc)
    noi = {m: tot_inc[m] - tot_exp[m] for m in months}

    # ---- ... and proved against East + West over the shared months ---------
    def src_detail(s, side, m):
        names = [a for nm, comps in (INCOME_MAP if side == "inc"
                                     else EXPENSE_MAP)
                 for ss, a in comps if ss == s]
        return sum(src[s]["accounts"][a].get(m) or 0.0 for a in names)

    print("  control-row proof (master = East + West, month by month):")
    bad = 0
    for label, mine, side in (("Total for Income", tot_inc, "inc"),
                              ("Total for Expenses", tot_exp, "exp")):
        for m in months:
            want = sum(src_detail(s, side, m) for s in src)
            if abs(mine[m] - want) > TOL:
                bad += 1
                print(f"    MISMATCH {label} {m:%b %Y}: master {mine[m]:,.2f} "
                      f"vs East+West {want:,.2f}")
        a = sum(mine.values())
        b = sum(sum(src_detail(s, side, m) for s in src) for m in months)
        assert abs(a - b) <= TOL, f"{label}: {a} vs {b}"
        print(f"    OK  {label} ANNUAL: master {a:,.2f} = "
              f"East {sum(src_detail('East', side, m) for m in months):,.2f} "
              f"+ West {sum(src_detail('West', side, m) for m in months):,.2f}")
    assert bad == 0, "control rows do not equal East + West"
    a = sum(gross.values())
    print(f"    OK  Gross Profit ANNUAL: {a:,.2f} (no Cost of Goods Sold on "
          f"either statement, so Gross Profit = Total for Income)")
    print(f"    OK  Net Operating Income ANNUAL: {sum(noi.values()):,.2f} "
          f"= Total for Income - Total for Expenses")

    # ---- and compared to what each SOURCE printed (never silent) -----------
    print("  vs each source's own PRINTED control rows over the shared "
          "months:")
    for s in src:
        for key, mine_side in (("total for income", "inc"),
                               ("total for expenses", "exp")):
            ctrl = src[s]["controls"].get(key)
            if not ctrl:
                continue
            printed = sum(v for m, v in ctrl.items()
                          if (m.year, m.month) in
                          {(x.year, x.month) for x in months})
            detail = sum(src_detail(s, mine_side, m) for m in months)
            flag = "OK " if abs(printed - detail) <= 0.05 else "VARIANCE"
            print(f"    {flag} {s} printed '{key.title()}' {printed:,.2f} vs "
                  f"its own detail {detail:,.2f} "
                  f"(diff {detail - printed:+,.2f})")

    write_master(months, income, expense, tot_inc, gross, tot_exp, noi)


def write_master(months, income, expense, tot_inc, gross, tot_exp, noi):
    """Write the combined statement in the QuickBooks-Online xlsx layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = OUT_ENTITY
    ws["A2"] = "Profit and Loss"
    last_day = (datetime(months[-1].year + (months[-1].month == 12),
                         months[-1].month % 12 + 1, 1) - timedelta(days=1))
    ws["A3"] = (f"{months[0]:%B} 1, {months[0]:%Y}-{last_day:%B} "
                f"{last_day.day}, {last_day:%Y}")
    for i, d in enumerate(months):
        ws.cell(row=5, column=2 + i, value=f"{d:%b %Y}")
    tot_col = 2 + len(months)
    ws.cell(row=5, column=tot_col, value="Total")

    def put(r, name, vals=None):
        ws.cell(row=r, column=1, value=name)
        if vals is not None:
            for i, v in enumerate(vals):
                if v is not None:          # a QuickBooks blank stays blank
                    ws.cell(row=r, column=2 + i, value=v)
            ws.cell(row=r, column=tot_col,
                    value=sum(v or 0.0 for v in vals))

    r = 6
    put(r, "Income")
    r += 1
    for nm, vals in income:
        put(r, nm, vals)
        r += 1
    put(r, "Total for Income", [tot_inc[m] for m in months])
    r += 1
    put(r, "Gross Profit", [gross[m] for m in months])
    r += 1
    put(r, "Expenses")
    r += 1
    for nm, vals in expense:
        put(r, nm, vals)
        r += 1
    put(r, "Total for Expenses", [tot_exp[m] for m in months])
    r += 1
    put(r, "Net Operating Income", [noi[m] for m in months])
    r += 2
    ws.cell(row=r, column=1,
            value=f"Cash Basis {datetime.now():%A, %B %d, %Y %I:%M %p} "
                  f"(combined from {', '.join(os.path.basename(p) for p in SOURCES.values())})")
    ws.column_dimensions["A"].width = 38
    wb.save(OUT_PATH)
    normalize(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  {len(months)} months, {len(income)} income line(s), "
          f"{len(expense)} expense line(s); Total Revenue "
          f"{sum(tot_inc.values()):,.2f}, Total Operating Expense "
          f"{sum(tot_exp.values()):,.2f}, NOI {sum(noi.values()):,.2f}")


if __name__ == "__main__":
    main()
