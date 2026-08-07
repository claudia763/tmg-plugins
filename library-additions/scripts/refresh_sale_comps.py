#!/usr/bin/env python3
"""Load this deal's sale comps into the TMG model's Sale Comparable Summary.

THE DEFECT THIS FIXES (it ships silently on every deal): `PDF Output - F&C` rows
118-147 print the client-facing "Sale Comparable Summary". That block is a pure mirror
of `Comparable Grid`!C7:I35, which is formula-driven: `Comparable Grid`!N3:Z52 INDEX
into the table **Output_Analysis_Data__2** (it lives on the `Auto Sales` sheet,
A1:Y286) whenever the mode switch `Comparable Grid`!M1 reads "Automatic" -- which is how
the template ships. And the template ships that table **preloaded with a previous deal's
CMA export**. On Shady Oaks (Houston, 8/2026) it printed five Irving/Dallas comps and a
"Subject Indicated Total Value" of $2,030,830 against a real income value of $1,020,000.

Nothing about that is loud. There is no error cell, no #REF!, no blank. The print-area
error gate passes with 0 errors, the model is ALL GREEN, and the page still shows the
wrong city. **Look at the rendered PDF -- the gate will not catch this.** Sale comps are
the third stale tab, alongside `Agency Region` and `YardiProjections` (see
`uw-model-linux-libreoffice-build-8-2026.md` s.5).

SOURCE: the `Automatic CMA Analysis.xlsx` produced by the `sales-comps` run for THIS
deal -- its `Output Analysis Data` sheet is the same table in the same order, so the
delivered comp workbook and the model agree by construction. Point --cma at it.

THE COLUMN OFFSET, WHICH IS THE ONE REAL TRAP: the model's table carries an extra empty
column `Column1` at L that the CMA export does not have. CMA A..K map straight across to
model A..K; CMA L..X map to model **M..Y**, shifted one right. Copying column-for-column
silently drops Avg Unit SF into the Column1 slot and shears every field after it -- with
no error, just quietly wrong comps. This script asserts the two headers line up under
the offset before it writes anything, so a template revision that moves a column fails
loudly instead of producing a plausible wrong page.

SELECTION: `Comparable Grid`!L3:L52 is the "Include (x)" column. Grid row r reads table
row r-2, i.e. `Auto Sales` sheet row **r-1**. Pass --marks as grid rows. Only the first
five marks print (the helper chain is x/xx/xxx/xxxx/xxxxx).

AFTER RUNNING: recalculate (a plain `soffice --convert-to xlsx`), then re-export the PDF
and LOOK at the Sale Comparable Summary page. Also replace the map picture -- it is a
static image with no formula behind it, so it still shows the old deal's geography even
once the grid is right; see `comp-map-generation.md` and `swap_workbook_image.py`.

Usage:
  python3 refresh_sale_comps.py <in.xlsx> <out.xlsx> --cma <Automatic CMA Analysis.xlsx>
                                [--marks 3 4 5 6 8] [--count 50]
"""
import argparse
import warnings

warnings.filterwarnings("ignore")
import openpyxl

SHEET = "Auto Sales"            # hosts Output_Analysis_Data__2
GRID = "Comparable Grid"
INCLUDE_COL = 12                # 'Comparable Grid'!L
COLUMN1 = 12                    # 'Auto Sales'!L, the extra empty column
TABLE_LAST_ROW = 286            # Output_Analysis_Data__2 = A1:Y286
N_COLS = 24


def model_col(cma_col):
    """CMA A..K -> model A..K; CMA L..X -> model M..Y (skip the extra 'Column1')."""
    return cma_col if cma_col < COLUMN1 else cma_col + 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("book")
    ap.add_argument("out")
    ap.add_argument("--cma", required=True,
                    help="Automatic CMA Analysis.xlsx for THIS deal")
    ap.add_argument("--marks", type=int, nargs="*", default=[3, 4, 5, 6, 7],
                    help="'Comparable Grid' rows to mark 'x' (grid row r = sheet row r-1)")
    ap.add_argument("--count", type=int, default=None,
                    help="comps to copy; default = every populated CMA row")
    a = ap.parse_args()

    cma = openpyxl.load_workbook(a.cma, data_only=True)["Output Analysis Data"]
    n = a.count or sum(1 for r in range(2, cma.max_row + 1)
                       if cma.cell(r, 1).value not in (None, ""))

    wb = openpyxl.load_workbook(a.book)
    auto, grid = wb[SHEET], wb[GRID]

    mode = grid["M1"].value
    if mode != "Automatic":
        print(f"  WARNING: 'Comparable Grid'!M1 = {mode!r}, not 'Automatic' — this "
              f"script feeds the Automatic path; the grid may be reading 'Sale Comps'")

    for c in range(1, N_COLS + 1):
        want, got = cma.cell(1, c).value, auto.cell(1, model_col(c)).value
        if want is not None and want != got:
            raise SystemExit(
                f"header mismatch at CMA col {c}: {want!r} vs model "
                f"col {model_col(c)} {got!r} — the column offset has moved, fix it "
                f"before trusting this script")

    for i in range(n):
        for c in range(1, N_COLS + 1):
            auto.cell(2 + i, model_col(c)).value = cma.cell(2 + i, c).value
        auto.cell(2 + i, COLUMN1).value = None
    print(f"  {SHEET}!A2:Y{1 + n}: wrote {n} comps from {a.cma}")

    cleared = 0
    for r in range(2 + n, TABLE_LAST_ROW + 1):
        if any(auto.cell(r, c).value not in (None, "") for c in range(1, 26)):
            cleared += 1
        for c in range(1, 26):
            auto.cell(r, c).value = None
    print(f"  {SHEET} rows {2 + n}-{TABLE_LAST_ROW}: cleared {cleared} stale rows")

    for r in range(3, 53):
        grid.cell(r, INCLUDE_COL).value = None
    for r in a.marks:
        grid.cell(r, INCLUDE_COL).value = "x"
    print(f"  {GRID}!L3:L52 marked {a.marks}:")
    for r in a.marks:
        print(f"     L{r} -> {SHEET} row {r - 1}: {auto.cell(r - 1, 1).value}")
    if len(a.marks) > 5:
        print(f"  NOTE: {len(a.marks)} marks but only the first 5 print")

    wb.save(a.out)
    print(f"  wrote {a.out}  — now recalc, re-export the PDF, and LOOK at the page")


if __name__ == "__main__":
    main()
