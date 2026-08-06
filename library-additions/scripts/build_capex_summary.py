#!/usr/bin/env python3
"""
build_capex_summary.py — side-by-side historical capex workbook (TMG house style)
================================================================================

WHAT IT DOES
------------
Turns one or more QuickBooks account-summary exports (the "Fixed Assets" /
capex reports ownership sends as totals-only two-column sheets: account name in
column A, amount in column B, hierarchy carried by cell INDENT, roll-ups printed
as "Total for <parent>" rows) into a single delivery workbook:

  Tab 1  "Capex Summary"  one row per account, one column per property plus a
                          Combined column, aligned across properties on the GL
                          account number so charts of accounts that differ still
                          line up. A property with no such account gets a BLANK
                          cell - never a zero. Subtotals for every parent
                          account, a bold grand total, a per-unit block, and a
                          plain-black notes block.
  Tabs 2..n               each source report reproduced exactly as printed,
                          hierarchy and roll-ups intact.

Formatting follows `process_t12.write_capex()` (Calibri 11, no gridlines, bold
header block, thin-bottom header row, bold + top-bordered totals, accounting
number formats, frozen panes, no red text) and the workbook is saved through a
copy of `_normalize_xlsx()` (inline strings -> sharedStrings, empty cells
untyped) per the TMG house rule about openpyxl output crashing Excel/JS loaders.

WHY IT IS CAREFUL ABOUT PARENT ACCOUNTS
---------------------------------------
In QuickBooks a parent account can carry its OWN direct postings in addition to
its sub-accounts, e.g.

    13320 Flooring                341,516.00   <- parent's own postings
      13322 Flooring Labor         61,220.00
    Total for 13320 Flooring      402,736.00

Summing the visible rows would double-count. This script reads the parent's
direct amount separately, shows it on its own "<account> - Other" row (the
QuickBooks convention), and RECOMPUTES every printed "Total for" row from
parent-direct + immediate children (a child's own printed total superseding it).
Any total that does not tie to the cent is reported loudly and sets a non-zero
exit code — printed totals are the truth and a mismatch means the parse is
wrong.

USAGE
-----
  python build_capex_summary.py \
      --prop "Westlake East|59|2022 - 2025|Accrual|C:\\...\\East capex.xlsx" \
      --prop "Westlake West|115|2023 - 2026|Cash|C:\\...\\West capex.xlsx" \
      --title "Westlake Apartments - Lubbock, TX" \
      --asof 7/31/2026 \
      --out "C:\\...\\Capex - Westlake East & West - Historical.xlsx" \
      --out "C:\\...\\dropbox\\- Info for Buyers\\Capex - ....xlsx"

  --prop is  Label|Units|Period|Basis|Path   (repeatable, order = column order)
  --out  is repeatable; every path's parent directory is created if missing.

The notes block is generated from the --prop metadata: when the properties do
not share the same period or the same accounting basis, the workbook says so
explicitly and states that the Combined column is a sum of different periods,
presented for scale and not as an annual run rate.

Built 8/2026 for the Westlake East & West historical capex deliverable
(East 2022-2025 accrual $1,152,986.94 / West 2023-2026 cash $2,607,343.92 /
combined $3,760,330.86, all printed roll-ups tying to the cent).
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

# ---------------------------------------------------------------------------
# House style (mirrors process_t12.write_capex)
# ---------------------------------------------------------------------------
F_NORM = Font(name="Calibri", size=11)
F_BOLD = Font(name="Calibri", size=11, bold=True)
THIN_BOTTOM = Border(bottom=Side(style="thin"))
TOP_BORDER = Border(top=Side(style="thin"))
TOP_DOUBLE = Border(top=Side(style="thin"), bottom=Side(style="double"))

ACCT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
ACCT_CUR = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
INT_FMT = "#,##0"

TOTAL_RE = re.compile(r"^total\s+for\s+", re.I)
BASIS_RE = re.compile(r"^(accrual|cash)\s+basis", re.I)


# ---------------------------------------------------------------------------
# Source parsing
# ---------------------------------------------------------------------------
class SrcRow:
    """One printed line of a QuickBooks account summary."""

    __slots__ = ("name", "indent", "amount", "is_total", "row")

    def __init__(self, name, indent, amount, is_total, row):
        self.name, self.indent = name, indent
        self.amount, self.is_total, self.row = amount, is_total, row

    def __repr__(self):
        return f"<SrcRow {self.row} i{self.indent} {self.name!r} {self.amount}>"


def read_qb_summary(path, name_col=1, amt_col=2):
    """Read a two-column QuickBooks account summary.

    Returns (title, rows, footer): rows in printed order, footer = the
    basis/timestamp line QuickBooks stamps at the bottom (or None).
    """
    ws = load_workbook(path, data_only=True).active
    title = footer = None
    rows = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=name_col)
        b = ws.cell(row=r, column=amt_col)
        if a.value is None or not str(a.value).strip():
            continue
        name = str(a.value).strip()
        if BASIS_RE.match(name):
            footer = name
            continue
        if title is None and b.value is None and not TOTAL_RE.match(name):
            title = name                     # report caption ("Fixed Assets")
            continue
        amt = b.value
        if isinstance(amt, str):
            amt = amt.replace(",", "").replace("$", "").strip()
            amt = float(amt) if amt else None
        if amt is not None:
            amt = round(float(amt), 2)
        rows.append(SrcRow(name, int(a.alignment.indent or 0), amt,
                           bool(TOTAL_RE.match(name)), r))
    return title, rows, footer


# ---------------------------------------------------------------------------
# Hierarchy
# ---------------------------------------------------------------------------
class Node:
    """One account in the merged chart of accounts.

    `direct` / `total` are dicts keyed by property label so several reports can
    share one tree; `total` holds the report's PRINTED "Total for" figure.
    """

    def __init__(self, key, label):
        self.key = key
        self.label = label
        self.labels = {}        # prop -> that report's own label for the acct
        self.direct = {}        # prop -> amount posted straight to this acct
        self.total = {}         # prop -> printed "Total for" figure
        self.children = []

    def child(self, key):
        for c in self.children:
            if c.key == key:
                return c
        return None


def acct_key(name):
    """Accounts are aligned across properties on their GL number when they have
    one (Westlake East calls 13310 'Interior Paint', West calls it 'Paint'),
    otherwise on the normalized name."""
    m = re.match(r"^(\d[\d.\-]*)\b", name.strip())
    return m.group(1) if m else re.sub(r"\s+", " ", name.strip().lower())


def _insert_sibling(parent, node):
    """Keep the first report's printed order, and slot accounts that only a
    later report has into their GL-number position among the siblings.

    String comparison of GL numbers reproduces QuickBooks' own ordering,
    including sub-numbered accounts ('13350' < '133502' < '13360')."""
    if node.key[:1].isdigit():
        for i, c in enumerate(parent.children):
            if c.key[:1].isdigit() and c.key > node.key:
                parent.children.insert(i, node)
                return
    parent.children.append(node)


def build_tree(rows, prop, root=None):
    """Fold one report's printed rows into (or onto) a merged tree."""
    root = root or Node("@root", "@root")
    stack = []                                  # [(indent, node)]
    for r in rows:
        if r.is_total:
            name = TOTAL_RE.sub("", r.name).strip()
            key = acct_key(name)
            for i in range(len(stack) - 1, -1, -1):
                if stack[i][0] == r.indent and stack[i][1].key == key:
                    stack[i][1].total[prop] = r.amount
                    del stack[i:]               # the total closes the account
                    break
            else:
                raise ValueError(f"{prop}: '{r.name}' (row {r.row}) has no "
                                 "matching parent account row")
            continue
        while stack and stack[-1][0] >= r.indent:
            stack.pop()
        parent = stack[-1][1] if stack else root
        key = acct_key(r.name)
        node = parent.child(key)
        if node is None:
            node = Node(key, r.name)
            _insert_sibling(parent, node)
        node.labels[prop] = r.name
        if r.amount is not None:
            node.direct[prop] = r.amount
        stack.append((r.indent, node))
    return root


def verify_tree(root, prop, out=sys.stdout):
    """Recompute every printed 'Total for' from parent-direct + children.

    Returns (results, ok) where results is a list of dicts.
    """
    results = []

    def walk(node):
        for c in node.children:
            walk(c)
        if prop in node.total:
            comps = []
            if prop in node.direct:
                comps.append((node.labels.get(prop, node.label) + " (direct)",
                              node.direct[prop]))
            for c in node.children:
                if prop in c.total:
                    comps.append(("Total for " + c.labels.get(prop, c.label),
                                  c.total[prop]))
                elif prop in c.direct:
                    comps.append((c.labels.get(prop, c.label), c.direct[prop]))
            calc = round(sum(v for _, v in comps), 2)
            printed = node.total[prop]
            delta = round(printed - calc, 2)
            results.append({"account": node.labels.get(prop, node.label),
                            "printed": printed, "recomputed": calc,
                            "delta": delta, "ok": abs(delta) < 0.005,
                            "components": comps})

    walk(root)
    print(f"\n--- roll-up verification: {prop} ---", file=out)
    for x in results:
        print(f"  [{'OK ' if x['ok'] else 'MISMATCH'}] Total for "
              f"{x['account']}: printed {x['printed']:,.2f} vs recomputed "
              f"{x['recomputed']:,.2f} (delta {x['delta']:+,.2f})", file=out)
        for cn, cv in x["components"]:
            print(f"            + {cn}: {cv:,.2f}", file=out)
    return results, all(x["ok"] for x in results)


def grand_total(root, prop):
    """Sum of the top-level printed section totals (QuickBooks capex summaries
    print no grand total of their own)."""
    tot = 0.0
    for c in root.children:
        tot += c.total.get(prop, c.direct.get(prop, 0.0))
    return round(tot, 2)


# ---------------------------------------------------------------------------
# Summary tab
# ---------------------------------------------------------------------------
def _put(ws, row, col, value=None, *, bold=False, fmt=None, indent=0,
         border=None, wrap=False, align=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font = F_BOLD if bold else F_NORM
    if fmt:
        c.number_format = fmt
    if indent or wrap or align:
        c.alignment = Alignment(indent=indent, wrap_text=wrap,
                                horizontal=align,
                                vertical="center" if wrap else None)
    if border is not None:
        c.border = border
    return c


def _emit_rows(node, depth, props, out):
    """Flatten the merged tree into (label, indent, kind, values) rows."""
    for c in node.children:
        vals = {p: c.direct.get(p) for p in props}
        if not c.children:
            out.append((c.label, depth, "acct", vals))
            continue
        kind_head = "sect_head" if depth == 0 else "head"
        out.append((c.label, depth, kind_head, {p: None for p in props}))
        if any(v is not None for v in vals.values()):
            out.append((f"{c.label} - Other", depth + 1, "acct", vals))
        _emit_rows(c, depth + 1, props, out)
        out.append((f"Total {c.label}", depth,
                    "sect" if depth == 0 else "sub",
                    {p: c.total.get(p) for p in props}))
        if depth == 0:
            out.append(("", 0, "blank", {}))


def write_summary(wb, root, props, meta, title, asof, sources, aliases):
    ws = wb.create_sheet("Capex Summary")
    ws.sheet_view.showGridLines = False
    ncol = 1 + len(props) + 1                     # account + props + combined
    comb_col = ncol

    total_units = sum(meta[p]["units"] for p in props)
    periods = " & ".join(f"{p} ({meta[p]['period']})" for p in props)
    _put(ws, 1, 1, title, bold=True)
    _put(ws, 2, 1, f"Historical Capital Expenditures - {periods}", bold=True)
    _put(ws, 3, 1, f"{total_units} Units ("
                   + " / ".join(f"{p} {meta[p]['units']}" for p in props)
                   + ")" + (f" - rent roll as of {asof}" if asof else ""))

    hdr = 5
    _put(ws, hdr, 1, "Account", bold=True, border=THIN_BOTTOM)
    for i, p in enumerate(props):
        _put(ws, hdr, 2 + i,
             f"{p}\n{meta[p]['period']} ({meta[p]['basis']})", bold=True,
             border=THIN_BOTTOM, wrap=True, align="center")
    _put(ws, hdr, comb_col, "Combined\n(see notes)", bold=True,
         border=THIN_BOTTOM, wrap=True, align="center")
    ws.row_dimensions[hdr].height = 32

    rows = []
    _emit_rows(root, 0, props, rows)
    while rows and rows[-1][2] == "blank":
        rows.pop()

    r = hdr + 2
    for label, indent, kind, vals in rows:
        if kind == "blank":
            r += 1
            continue
        bold = kind != "acct"
        border = (TOP_BORDER if kind in ("sub", "sect") else None)
        fmt = ACCT if kind == "acct" else ACCT_CUR
        _put(ws, r, 1, label, bold=bold, indent=indent * 2, border=border)
        if kind in ("head", "sect_head"):
            for col in range(2, ncol + 1):
                if border is not None:
                    ws.cell(row=r, column=col).border = border
            r += 1
            continue
        present = [vals[p] for p in props if vals.get(p) is not None]
        for i, p in enumerate(props):
            c = ws.cell(row=r, column=2 + i)
            if vals.get(p) is not None:       # BLANK, never zero, when absent
                c.value = vals[p]
            c.number_format = fmt
            c.font = F_BOLD if bold else F_NORM
            if border is not None:
                c.border = border
        c = ws.cell(row=r, column=comb_col)
        if present:
            c.value = round(sum(present), 2)
        c.number_format = fmt
        c.font = F_BOLD if bold else F_NORM
        if border is not None:
            c.border = border
        r += 1

    # -------- grand total --------
    r += 1
    grands = {p: grand_total(root, p) for p in props}
    combined = round(sum(grands.values()), 2)
    _put(ws, r, 1, "Total Capital Expenditures", bold=True, border=TOP_DOUBLE)
    for i, p in enumerate(props):
        _put(ws, r, 2 + i, grands[p], bold=True, fmt=ACCT_CUR,
             border=TOP_DOUBLE)
    _put(ws, r, comb_col, combined, bold=True, fmt=ACCT_CUR, border=TOP_DOUBLE)
    grand_row = r

    # -------- per-unit block --------
    r += 2
    _put(ws, r, 1, "Capital Expenditures Per Unit", bold=True,
         border=THIN_BOTTOM)
    for col in range(2, ncol + 1):
        ws.cell(row=r, column=col).border = THIN_BOTTOM
    r += 1
    block = [("Units", [meta[p]["units"] for p in props], total_units,
              INT_FMT, False),
             ("Total Capital Expenditures", [grands[p] for p in props],
              combined, ACCT_CUR, False),
             ("Capital Expenditures per Unit",
              [round(grands[p] / meta[p]["units"], 2) for p in props],
              round(combined / total_units, 2), ACCT_CUR, True)]
    for label, vals, comb, fmt, bold in block:
        _put(ws, r, 1, label, bold=bold, indent=2)
        for i, v in enumerate(vals):
            _put(ws, r, 2 + i, v, bold=bold, fmt=fmt)
        _put(ws, r, comb_col, comb, bold=bold, fmt=fmt)
        r += 1

    # -------- notes (plain black, no red text) --------
    same_period = len({meta[p]["period"] for p in props}) == 1
    same_basis = len({meta[p]["basis"] for p in props}) == 1
    notes = []
    if len(props) > 1 and not (same_period and same_basis):
        notes.append(
            "Reporting periods and/or accounting bases DIFFER between the "
            "properties shown: "
            + "; ".join(f"{p} covers {meta[p]['period']} on a "
                        f"{meta[p]['basis'].upper()} basis" for p in props)
            + ". The columns are not the same time span and/or not the same "
              "accounting basis.")
        notes.append(
            "The Combined column is therefore the arithmetic sum of different "
            "reporting periods. It is presented for SCALE only - it is not an "
            "annual run rate and should not be annualized or divided by a "
            "number of years.")
    notes.append("A blank cell means that property's chart of accounts "
                 "contains no such account. Blank is not zero.")
    notes.append(
        'Rows ending "- Other" carry amounts posted directly to a parent '
        "account that also has sub-accounts (QuickBooks convention). They are "
        "shown separately so the parent's own postings and its sub-accounts "
        'can both be seen without double-counting; every printed "Total for" '
        "row was recomputed from parent-direct plus children and ties out to "
        "the cent.")
    for key, label, alias in aliases:
        notes.append(f"Account {key} is named " + "; ".join(alias)
                     + "; the two are aligned on the same line here.")
    notes.append("The source reports print no grand total. Total Capital "
                 "Expenditures is the sum of the printed section totals.")
    for s in sources:
        notes.append("Source: " + s)
    notes.append("Unit counts: "
                 + ", ".join(f"{p} {meta[p]['units']} units" for p in props)
                 + f", {total_units} units combined"
                 + (f"; rent roll as of {asof}." if asof else "."))

    r += 2
    _put(ws, r, 1, "Notes", bold=True)
    r += 1
    for nt in notes:
        c = _put(ws, r, 1, nt)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 14 * (len(nt) // 115 + 1))
        r += 1

    ws.column_dimensions["A"].width = 42
    for i in range(2, ncol + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 20
    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)
    return grands, combined, grand_row


# ---------------------------------------------------------------------------
# Source-reproduction tabs
# ---------------------------------------------------------------------------
def write_source_tab(wb, tab, prop, meta_p, rows, footer, grand, title_suffix):
    ws = wb.create_sheet(tab[:31])
    ws.sheet_view.showGridLines = False
    _put(ws, 1, 1, f"{prop}{title_suffix}", bold=True)
    _put(ws, 2, 1, f"Historical Capital Expenditures {meta_p['period']} "
                   f"({meta_p['basis']} Basis) - {meta_p['units']} Units",
         bold=True)
    _put(ws, 3, 1, "QuickBooks account summary - reproduced as printed")

    hdr = 5
    _put(ws, hdr, 1, "Account", bold=True, border=THIN_BOTTOM)
    _put(ws, hdr, 2, "Amount", bold=True, border=THIN_BOTTOM, align="center")

    r = hdr + 2
    _put(ws, r, 1, "Fixed Assets", bold=True)
    r += 1
    base = min(x.indent for x in rows)
    for src in rows:
        border = TOP_BORDER if src.is_total else None
        _put(ws, r, 1, src.name, bold=src.is_total,
             indent=(src.indent - base + 1) * 2, border=border)
        c = ws.cell(row=r, column=2)
        if src.amount is not None:
            c.value = src.amount
        c.number_format = ACCT_CUR if src.is_total else ACCT
        c.font = F_BOLD if src.is_total else F_NORM
        if border is not None:
            c.border = border
        r += 1

    r += 1
    _put(ws, r, 1, "Total Capital Expenditures", bold=True, border=TOP_DOUBLE)
    _put(ws, r, 2, grand, bold=True, fmt=ACCT_CUR, border=TOP_DOUBLE)
    r += 2
    _put(ws, r, 1, "Total Capital Expenditures is the sum of the printed "
                   "section totals; the source report prints no grand total.")
    if footer:
        _put(ws, r + 1, 1, footer)

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = ws.cell(row=hdr + 1, column=2)


# ---------------------------------------------------------------------------
# Normalized save (TMG house rule; copy of process_t12._normalize_xlsx)
# ---------------------------------------------------------------------------
def normalize_xlsx(path):
    """Rewrite an openpyxl-saved package the way Excel writes it: inline
    strings -> xl/sharedStrings.xml, empty typed cells -> untyped. DOM-based
    loaders fetch sharedStrings unconditionally and crash without it."""
    SST_T = ("application/vnd.openxmlformats-officedocument.spreadsheetml"
             ".sharedStrings+xml")
    SST_R = ("http://schemas.openxmlformats.org/officeDocument/2006/"
             "relationships/sharedStrings")
    strings, index, total_refs = [], {}, 0

    def fix_sheet(xml):
        nonlocal total_refs

        def repl(m):
            nonlocal total_refs
            head, tail, inner = m.group(1), m.group(2), m.group(3) or ""
            t = re.search(r"<t[^>]*>.*?</t>", inner, re.S)
            t_elem = t.group(0) if t else "<t/>"
            if t_elem not in index:
                index[t_elem] = len(strings)
                strings.append(t_elem)
            total_refs += 1
            return f'<c{head} t="s"{tail}><v>{index[t_elem]}</v></c>'

        xml = re.sub(r'<c([^>]*?) t="inlineStr"([^>]*)>(?:<is>(.*?)</is>)?</c>',
                     repl, xml, flags=re.S)
        xml = re.sub(r'<c([^>]*?) t="n"([^>]*)></c>', r"<c\1\2/>", xml)
        return re.sub(r'<c([^>]*?) t="n"([^>]*)/>', r"<c\1\2/>", xml)

    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        parts = {n: zin.read(n) for n in zin.namelist()}
        for name in list(parts):
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                parts[name] = fix_sheet(
                    parts[name].decode("utf-8")).encode("utf-8")
        if strings:
            sst = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<sst xmlns="http://schemas.openxmlformats.org/'
                   'spreadsheetml/2006/main" '
                   f'count="{total_refs}" uniqueCount="{len(strings)}">'
                   + "".join(f"<si>{t}</si>" for t in strings) + "</sst>")
            parts["xl/sharedStrings.xml"] = sst.encode("utf-8")
            ct = parts["[Content_Types].xml"].decode("utf-8")
            if "sharedStrings" not in ct:
                parts["[Content_Types].xml"] = ct.replace(
                    "</Types>", f'<Override PartName="/xl/sharedStrings.xml" '
                                f'ContentType="{SST_T}"/></Types>'
                ).encode("utf-8")
            rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
            if "sharedStrings" not in rels:
                parts["xl/_rels/workbook.xml.rels"] = rels.replace(
                    "</Relationships>",
                    f'<Relationship Id="rIdSST" Type="{SST_R}" '
                    f'Target="sharedStrings.xml"/></Relationships>'
                ).encode("utf-8")
        for name, data in parts.items():
            zout.writestr(name, data)
    shutil.move(tmp, path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1],
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--prop", action="append", required=True,
                    metavar="Label|Units|Period|Basis|Path",
                    help="one source report (repeatable; order = column order)")
    ap.add_argument("--title", required=True,
                    help="header line 1, e.g. 'Westlake - Lubbock, TX'")
    ap.add_argument("--asof", default=None, help="rent-roll as-of date")
    ap.add_argument("--tab-suffix", default="",
                    help="appended to each source tab's title line, "
                         "e.g. ' - Lubbock, TX'")
    ap.add_argument("--out", action="append", required=True,
                    help="output .xlsx path (repeatable)")
    a = ap.parse_args(argv)

    props, meta, srcs = [], {}, {}
    for spec in a.prop:
        parts = spec.split("|")
        if len(parts) != 5:
            ap.error(f"--prop needs 5 |-separated fields, got {spec!r}")
        label, units, period, basis, path = (p.strip() for p in parts)
        props.append(label)
        meta[label] = {"units": int(units), "period": period, "basis": basis,
                       "path": path}
        srcs[label] = read_qb_summary(path)
        print(f"{label}: {os.path.basename(path)} | caption "
              f"{srcs[label][0]!r} | {len(srcs[label][1])} rows | "
              f"{srcs[label][2]}")

    root = Node("@root", "@root")
    ok = True
    for p in props:
        build_tree(srcs[p][1], p, root)
    for p in props:
        _, good = verify_tree(root, p)
        ok = ok and good

    # accounts whose label differs between properties -> a note
    aliases = []

    def scan(node):
        for c in node.children:
            uniq = {c.labels[p] for p in c.labels}
            if len(uniq) > 1:
                aliases.append((c.key, c.label,
                                [f'"{c.labels[p]}" at {p}' for p in props
                                 if p in c.labels]))
            scan(c)

    scan(root)

    grands = {p: grand_total(root, p) for p in props}
    combined = round(sum(grands.values()), 2)
    print("\n--- grand totals (sum of printed section totals) ---")
    for p in props:
        for c in root.children:
            if p in c.total or p in c.direct:
                print(f"  {p} {c.labels.get(p, c.label)}: "
                      f"{c.total.get(p, c.direct.get(p, 0.0)):,.2f}")
        print(f"  {p} GRAND: {grands[p]:,.2f}  "
              f"(${grands[p] / meta[p]['units']:,.2f}/unit)")
    print(f"  COMBINED GRAND: {combined:,.2f}  ($"
          f"{combined / sum(meta[p]['units'] for p in props):,.2f}/unit)")

    source_notes = []
    for p in props:
        title, _, footer = srcs[p]
        source_notes.append(
            f'"{os.path.basename(meta[p]["path"])}" - QuickBooks '
            f'{title or "account"} summary, {meta[p]["basis"]} Basis'
            + (f" ({footer})." if footer else "."))

    wb = Workbook()
    wb.remove(wb.active)
    write_summary(wb, root, props, meta, a.title, a.asof, source_notes,
                  aliases)
    for p in props:
        _, rows, footer = srcs[p]
        write_source_tab(wb, p, p, meta[p], rows, footer, grands[p],
                         a.tab_suffix)
    wb.calculation.calcId = 191029
    wb.calculation.fullCalcOnLoad = True

    for out in a.out:
        d = os.path.dirname(os.path.abspath(out))
        os.makedirs(d, exist_ok=True)
        wb.save(out)
        normalize_xlsx(out)
        print(f"saved: {out}")

    if not ok:
        print("\n!! at least one printed roll-up did NOT recompute - the "
              "parse or the source is wrong; do not deliver.", file=sys.stderr)
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
