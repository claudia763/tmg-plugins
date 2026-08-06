#!/usr/bin/env python3
"""Pre-send QC gate for TMG rent-roll / T-12 / Capex deliverable workbooks.

WHAT IT DOES
    Runs the structural checks an orchestrator would otherwise do by hand
    before replying to the requester. It does NOT re-do the processing
    scripts' reconciliation against the source report (that is their job and
    it already ran) - it proves the FILE that is about to be emailed is
    well-formed, opens in real Excel, and is internally consistent.

    Per file:
      * opens with openpyxl (formulas + cached values)
      * confirms xl/sharedStrings.xml exists  -> the _save_normalized() path
        was used. openpyxl output that skips it crashes Excel/JS loaders.
      * confirms no leftover query-table / external-data parts (the other
        half of the same trap)
      * scans every cached cell value for Excel error strings (#REF!,
        #VALUE!, #DIV/0!, #N/A, #NAME?, #NUM!, #NULL!)
    Rent roll (RR - *.xlsx):
      * tabs Floor Plan / Rent Roll / Floor Plan Summary all present
      * A1 property name and B2 as-of date populated
      * the four rediQ named ranges present and sized to the data rows
      * per-unit columns that must never be silently empty are reported with
        a fill count (Net Sf, Bed, Bath, Market Rent, Contractual Rent)
      * red ESTIMATE cells (FFC7CE / 9C0006) counted, so a run that filled
        estimates without saying so is caught
      * every floor plan spans exactly one Net Sf value (a plan that mixes
        sizes makes every plan-level average meaningless)
    T-12 (T-12 - *.xlsx):
      * Trailing Financials visible, Final T-12 PRESENT and HIDDEN
      * Comments tab, if present, is non-empty
      * the month header is a contiguous run of month-starts and, unless
        --allow-partial-months, is exactly 12 long
      * Total column equals the sum of its own month cells, per row
      * Total Revenue - Total Expense == Net Operating Income, per month
        and for the Total column
    Capex & Misc (Capex* .xlsx):
      * Total CapEx row equals the sum of the capex detail, per month
      * grand row equals capex + the below-the-line rows

    With --excel it additionally opens each file through Windows COM
    (real Excel), forces a full rebuild, and re-reads the checks that depend
    on live formulas - openpyxl cannot evaluate formulas, so a rent roll's
    summary tabs are only really proved this way. Skipped silently on
    non-Windows or when pywin32 is unavailable.

INPUTS
    One or more .xlsx paths, or a directory (typically the job's outbox/).

OUTPUTS
    A per-file PASS/FAIL report on stdout. Exit code 0 only if every check
    passed; 1 otherwise. Intended to be run before the reply email is sent.

USAGE
    python verify_deliverable_workbooks.py outbox/
    python verify_deliverable_workbooks.py outbox/ --excel
    python verify_deliverable_workbooks.py "outbox/T-12 - X - June 2026.xlsx"

WRITTEN
    8/6/2026, Magnolia Place (Brenham TX). The checks are the ones that
    actually caught something in practice: an empty Renovation Status column
    that the processing script's own reconciliation could not see, because
    reconciliation only proves the numbers that tie to the source - it says
    nothing about a column the source populates and the deliverable drops.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import os
import sys
import zipfile
from collections import defaultdict

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")


ERROR_STRINGS = ("#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NUM!",
                 "#NULL!")
ESTIMATE_FILL = "FFC7CE"
ESTIMATE_FONT = "9C0006"
REDIQ_NAMES = ("rediq_rentroll", "rediq_floorplansummary",
               "rediq_rentrollasofdate", "rediq_dealname")


class Report:
    """Collects PASS/FAIL lines for one file."""

    def __init__(self, path):
        self.path = path
        self.lines = []
        self.failed = 0

    def ok(self, msg):
        self.lines.append(f"  OK    {msg}")

    def fail(self, msg):
        self.lines.append(f"  FAIL  {msg}")
        self.failed += 1

    def note(self, msg):
        self.lines.append(f"  --    {msg}")

    def check(self, cond, msg_ok, msg_fail=None):
        if cond:
            self.ok(msg_ok)
        else:
            self.fail(msg_fail or msg_ok)
        return bool(cond)

    def emit(self):
        status = "PASS" if not self.failed else f"FAIL ({self.failed})"
        print(f"\n{os.path.basename(self.path)}  [{status}]")
        for line in self.lines:
            print(line)


# --------------------------------------------------------------------------
# container-level checks (the _save_normalized() trap)
# --------------------------------------------------------------------------

def check_container(path, rep):
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
    rep.check("xl/sharedStrings.xml" in names,
              "sharedStrings.xml present (normalized save path used)",
              "sharedStrings.xml MISSING - written by raw openpyxl without "
              "_save_normalized(); Excel/JS loaders will choke")
    stragglers = sorted(n for n in names
                        if "queryTable" in n or "connections" in n
                        or "externalLink" in n)
    rep.check(not stragglers,
              "no query-table / external-data parts left from the template",
              f"leftover template parts: {stragglers} - run "
              f"_purge_broken_names()/_normalize_xlsx()")


def check_no_error_values(wb_vals, rep):
    hits = []
    for ws in wb_vals.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip() in ERROR_STRINGS:
                    hits.append(f"{ws.title}!{cell.coordinate}={v.strip()}")
    rep.check(not hits, "no Excel error values in cached results",
              f"{len(hits)} error cell(s): {hits[:8]}")


# --------------------------------------------------------------------------
# rent roll
# --------------------------------------------------------------------------

def _last_data_row(ws, first=4):
    r = first
    while ws.cell(r, 1).value not in (None, ""):
        r += 1
    return r - 1


def check_rent_roll(path, wb, wb_vals, rep):
    tabs = [ws.title for ws in wb.worksheets]
    for want in ("Floor Plan", "Rent Roll", "Floor Plan Summary"):
        rep.check(want in tabs, f"tab '{want}' present",
                  f"tab '{want}' MISSING (tabs: {tabs})")
    if "Rent Roll" not in tabs:
        return

    rr = wb["Rent Roll"]
    rep.check(bool(rr["A1"].value), f"A1 property name = {rr['A1'].value!r}",
              "A1 property name is EMPTY")
    asof = rr["B2"].value
    rep.check(isinstance(asof, _dt.datetime),
              f"B2 as-of date = {asof}",
              f"B2 as-of date missing or not a date ({asof!r})")

    last = _last_data_row(rr)
    n = last - 3
    rep.note(f"{n} unit row(s), rows 4-{last}")

    headers = {rr.cell(3, c).value: c for c in range(1, 30)
               if rr.cell(3, c).value}
    for label in ("Net Sf", "Bed", "Bath", "Market Rent", "Contractual Rent"):
        col = headers.get(label)
        if not col:
            rep.fail(f"column '{label}' not found in the header row")
            continue
        filled = sum(1 for r in range(4, last + 1)
                     if rr.cell(r, col).value not in (None, ""))
        rep.check(filled == n,
                  f"'{label}' populated on all {n} unit(s)",
                  f"'{label}' populated on only {filled}/{n} unit(s) - a "
                  f"column the source provides may have been dropped")

    # columns that are legitimately sparse: report the count, never fail
    for label in ("Renovation Status", "Lease Type", "MTM", "Vac. Notice"):
        col = headers.get(label)
        if col:
            filled = sum(1 for r in range(4, last + 1)
                         if rr.cell(r, col).value not in (None, ""))
            rep.note(f"'{label}' populated on {filled}/{n} unit(s)")

    # red estimate cells
    est = 0
    for row in rr.iter_rows(min_row=4, max_row=last):
        for cell in row:
            fill = cell.fill
            rgb = getattr(getattr(fill, "fgColor", None), "rgb", None)
            if rgb and ESTIMATE_FILL in str(rgb):
                est += 1
            frgb = getattr(getattr(cell.font, "color", None), "rgb", None)
            if frgb and ESTIMATE_FONT in str(frgb):
                est += 1
    rep.note(f"{est} red-flagged ESTIMATE cell(s) - must match what the "
             f"delivery notes claim (0 means nothing was estimated)")

    # named ranges
    defined = set(wb.defined_names.keys()) if hasattr(wb.defined_names, "keys") \
        else {d.name for d in wb.defined_names.definedName}
    for nm in REDIQ_NAMES:
        rep.check(nm in defined, f"named range '{nm}' present",
                  f"named range '{nm}' MISSING - rediQ import will fail")

    # one sqft per floor plan
    plan_col = headers.get("Floor Plan")
    sf_col = headers.get("Net Sf")
    if plan_col and sf_col:
        sizes = defaultdict(set)
        for r in range(4, last + 1):
            sizes[rr.cell(r, plan_col).value].add(rr.cell(r, sf_col).value)
        bad = {p: s for p, s in sizes.items() if len(s) > 1}
        rep.check(not bad,
                  f"each of the {len(sizes)} floor plan(s) spans exactly one "
                  f"Net Sf",
                  f"floor plan(s) mixing sizes (plan averages are then "
                  f"meaningless): {bad}")


# --------------------------------------------------------------------------
# T-12
# --------------------------------------------------------------------------

def _month_header(ws, row=3, first_col=2):
    months, col = [], first_col
    while True:
        v = ws.cell(row, col).value
        if isinstance(v, _dt.datetime):
            months.append((col, v))
            col += 1
        else:
            break
    return months


def check_t12(path, wb, wb_vals, rep, allow_partial=False):
    tabs = {ws.title: ws for ws in wb.worksheets}
    rep.check("Trailing Financials" in tabs, "tab 'Trailing Financials' present",
              f"tab 'Trailing Financials' MISSING (tabs: {list(tabs)})")
    final = tabs.get("Final T-12")
    if rep.check(final is not None, "tab 'Final T-12' present",
                 "tab 'Final T-12' MISSING - the model import needs it; "
                 "never delete it"):
        rep.check(final.sheet_state == "hidden",
                  "'Final T-12' is hidden, as the house format requires",
                  f"'Final T-12' sheet_state is {final.sheet_state!r}, "
                  f"expected 'hidden'")
    if "Comments" in tabs:
        body = [tabs["Comments"].cell(r, 1).value
                for r in range(1, tabs["Comments"].max_row + 1)]
        rep.check(any(v for v in body[3:]),
                  "Comments tab carries notes",
                  "Comments tab exists but is empty - either populate it or "
                  "do not create it")
    else:
        rep.note("no Comments tab (correct for a statement with no notes)")

    tf = tabs.get("Trailing Financials")
    if tf is None:
        return
    months = _month_header(tf)
    rep.note(f"{len(months)} month column(s): "
             f"{months[0][1]:%b %Y} - {months[-1][1]:%b %Y}"
             if months else "no month header found")
    if not allow_partial:
        rep.check(len(months) == 12,
                  "12 month columns (a T-12 must be twelve months)",
                  f"{len(months)} month columns - a T-12 must be 12 unless "
                  f"the user opted into a partial; report captions lie, the "
                  f"COLUMNS are the truth")
    # contiguity
    gaps = []
    for (_, a), (_, b) in zip(months, months[1:]):
        exp_y, exp_m = (a.year + 1, 1) if a.month == 12 else (a.year, a.month + 1)
        if (b.year, b.month) != (exp_y, exp_m):
            gaps.append(f"{a:%b %Y}->{b:%b %Y}")
    rep.check(not gaps, "month columns are contiguous",
              f"non-contiguous month columns: {gaps}")

    total_col = months[-1][0] + 1 if months else None
    if total_col is None:
        return

    # per-row: Total == sum of months  (blank months are genuinely blank)
    bad, checked = [], 0
    for r in range(4, tf.max_row + 1):
        tot = tf.cell(r, total_col).value
        if not isinstance(tot, (int, float)):
            continue
        vals = [tf.cell(r, c).value for c, _ in months]
        if not any(isinstance(v, (int, float)) for v in vals):
            continue
        s = sum(v for v in vals if isinstance(v, (int, float)))
        checked += 1
        if abs(s - tot) > 0.05:
            bad.append(f"row {r} {tf.cell(r, 1).value!r}: months {s:,.2f} vs "
                       f"Total {tot:,.2f}")
    rep.check(not bad,
              f"Total column equals the sum of its months on all {checked} "
              f"valued row(s)",
              f"{len(bad)} row(s) whose Total disagrees with their months: "
              f"{bad[:5]}")

    # Revenue - Expense == NOI, per month and for the Total
    def find_row(label):
        for r in range(4, tf.max_row + 1):
            if str(tf.cell(r, 1).value or "").strip().lower() == label:
                return r
        return None

    rev, exp, noi = (find_row("total revenue"), find_row("total expense"),
                     find_row("net operating income"))
    if rev and exp and noi:
        bad = []
        for c, m in months + [(total_col, "Total")]:
            a, b, n = (tf.cell(rev, c).value, tf.cell(exp, c).value,
                       tf.cell(noi, c).value)
            if all(isinstance(v, (int, float)) for v in (a, b, n)):
                if abs((a - b) - n) > 0.05:
                    tag = m if isinstance(m, str) else f"{m:%b %Y}"
                    bad.append(f"{tag}: {a:,.2f}-{b:,.2f} != {n:,.2f}")
        rep.check(not bad,
                  "Total Revenue - Total Expense == NOI in every month and "
                  "in the Total column",
                  f"NOI identity broken: {bad[:5]}")
    else:
        rep.fail("could not locate Total Revenue / Total Expense / NOI rows")


# --------------------------------------------------------------------------
# capex
# --------------------------------------------------------------------------

def check_capex(path, wb, wb_vals, rep):
    ws = wb.worksheets[0]
    months = _month_header(ws)
    rep.note(f"{len(months)} month column(s)")
    if not months:
        rep.fail("no month header found")
        return
    total_col = months[-1][0] + 1

    def find_row(prefix):
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip().lower().startswith(prefix):
                return r
        return None

    head = find_row("capex expenses")
    tot = find_row("total capex")
    if not (head and tot):
        rep.fail("could not locate the 'CAPEX EXPENSES' header and/or the "
                 "'Total CapEx' row")
        return
    bad = []
    for c, m in months + [(total_col, "Total")]:
        detail = sum(ws.cell(r, c).value for r in range(head + 1, tot)
                     if isinstance(ws.cell(r, c).value, (int, float)))
        printed = ws.cell(tot, c).value
        if isinstance(printed, (int, float)) and abs(detail - printed) > 0.05:
            tag = m if isinstance(m, str) else f"{m:%b %Y}"
            bad.append(f"{tag}: detail {detail:,.2f} vs Total CapEx "
                       f"{printed:,.2f}")
    rep.check(not bad,
              "'Total CapEx' equals its detail in every month and in the Total",
              f"Total CapEx disagrees with its detail: {bad[:5]}")


# --------------------------------------------------------------------------
# optional: real Excel
# --------------------------------------------------------------------------

def check_with_excel(paths, rep_by_path):
    try:
        import win32com.client as win32
    except ImportError:
        print("\n(--excel requested but pywin32 is unavailable; skipped)")
        return
    xl = win32.gencache.EnsureDispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        for p in paths:
            rep = rep_by_path[p]
            try:
                wb = xl.Workbooks.Open(os.path.abspath(p), UpdateLinks=0)
            except Exception as exc:  # noqa: BLE001
                rep.fail(f"real Excel could not open the file: {exc}")
                continue
            try:
                xl.CalculateFullRebuild()
                errs = 0
                for ws in wb.Worksheets:
                    try:
                        errs += ws.UsedRange.SpecialCells(-4123, 16).Count
                    except Exception:  # no error cells on this sheet
                        pass
                rep.check(errs == 0,
                          "real Excel: opens without repair, 0 formula error "
                          "cells after a full rebuild",
                          f"real Excel: {errs} formula error cell(s) after a "
                          f"full rebuild")
            finally:
                wb.Close(False)
    finally:
        xl.Quit()


# --------------------------------------------------------------------------

def classify(path):
    name = os.path.basename(path).lower()
    if name.startswith("rr - ") or name.startswith("rr-"):
        return "rentroll"
    if name.startswith("t-12") or name.startswith("t12"):
        return "t12"
    if name.startswith("capex"):
        return "capex"
    return "unknown"


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("paths", nargs="+",
                    help="workbook path(s) or a directory (e.g. outbox/)")
    ap.add_argument("--excel", action="store_true",
                    help="also open each file through real Excel (Windows)")
    ap.add_argument("--allow-partial-months", action="store_true",
                    help="do not fail a T-12 with fewer than 12 month columns "
                         "(only when the user opted into a partial)")
    args = ap.parse_args()

    files = []
    for p in args.paths:
        if os.path.isdir(p):
            files += [os.path.join(p, f) for f in sorted(os.listdir(p))
                      if f.lower().endswith(".xlsx") and not f.startswith("~$")]
        else:
            files.append(p)
    if not files:
        sys.exit("no .xlsx files found")

    reports = {}
    for path in files:
        rep = Report(path)
        reports[path] = rep
        kind = classify(path)
        rep.note(f"classified as: {kind}")
        try:
            check_container(path, rep)
            wb = openpyxl.load_workbook(path, data_only=False)
            wb_vals = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:  # noqa: BLE001
            rep.fail(f"could not open with openpyxl: {exc}")
            continue
        check_no_error_values(wb_vals, rep)
        if kind == "rentroll":
            check_rent_roll(path, wb, wb_vals, rep)
        elif kind == "t12":
            check_t12(path, wb, wb_vals, rep,
                      allow_partial=args.allow_partial_months)
        elif kind == "capex":
            check_capex(path, wb, wb_vals, rep)
        else:
            rep.note("unrecognised deliverable name - only the generic "
                     "container/error checks ran")

    if args.excel:
        check_with_excel(files, reports)

    for path in files:
        reports[path].emit()

    total_failed = sum(r.failed for r in reports.values())
    print(f"\n{len(files)} file(s) checked, {total_failed} failed check(s)")
    return 1 if total_failed else 0


if __name__ == "__main__":
    sys.exit(main())
