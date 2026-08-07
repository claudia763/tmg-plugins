#!/usr/bin/env python3
"""Extract the comps universe from TMG's Automatic CMA Analysis workbook into a
standalone "All Sales Comps.xlsx" that the `sales-comps` skill can read.

WHY THIS EXISTS
    `sales-comps/SKILL.md` asks the user to attach an `All Sales Comps.xlsx`.
    **No such standalone file exists in TMG Dropbox.** The comps universe lives as
    the `All Sale Comps` TAB (note: singular "Sale") inside

        TMG Dropbox/Databases/- Sales Comps/Automatic CMA Analysis.xlsx

    alongside `Combined Fannie and Freddie Sales Comps.xlsx`, which the skill wants
    as-is for cap-rate drift. So on a job where the broker just names an address and
    attaches nothing, do NOT ask them for the workbook — mount Dropbox and run this.

THE ONE TRAP: the tab carries a junk column `Column1` at position L (it holds
    source-file names like "11.xlsx", left over from the Power Query append). The
    skill's reader maps columns by HEADER NAME so it would survive, but downstream
    consumers that map positionally will shear every field after it — the same
    off-by-one-column hazard documented in `refresh_sale_comps.py`. This script drops
    `Column1` so the delivered universe has a clean 15-column schema.

OUTPUT SCHEMA (what select_comps.py reads):
    Property Name | Property Address | City | State | ZIP | Unit Count | Year Built |
    Sold Price | Sold Price/Unit | Sale Date | Building SF | Avg Unit SF |
    Info Source | Latitude | Longitude

USAGE
    python3 extract_all_sale_comps.py \
        "/path/to/TMG Dropbox/Databases/- Sales Comps/Automatic CMA Analysis.xlsx" \
        "All Sales Comps.xlsx"

    # then, per sales-comps/SKILL.md step 3:
    python3 salescomps/scripts/select_comps.py --comps "All Sales Comps.xlsx" ...

The run prints the row count and the sale-date span so you can confirm the database
is current before quoting anything off it (as of 8/7/2026: 3,990 rows spanning
2021-06-01 -> 2026-06-01).
"""
import argparse
import datetime
import sys

import openpyxl

SRC_SHEET = "All Sale Comps"   # singular "Sale" — the tab name, not the file name
DROP_COLUMNS = {"Column1"}     # Power Query append artifact; see module docstring
REQUIRED = ["Property Name", "Unit Count", "Sold Price", "Sale Date",
            "Latitude", "Longitude"]


def extract(src, dest, sheet=SRC_SHEET):
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"ERROR: no sheet {sheet!r} in {src}\n  sheets: {wb.sheetnames}")
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    keep = [i for i, h in enumerate(header) if h not in DROP_COLUMNS]
    kept_names = [header[i] for i in keep]

    missing = [c for c in REQUIRED if c not in kept_names]
    if missing:
        sys.exit(f"ERROR: source tab is missing expected columns: {missing}\n"
                 f"  found: {kept_names}\n"
                 "  The CMA workbook schema changed — fix the mapping rather than "
                 "shipping a sheared universe.")

    out = openpyxl.Workbook()
    o = out.active
    o.title = sheet
    o.append(kept_names)

    date_i = kept_names.index("Sale Date")
    n, dates = 0, []
    for r in rows:
        if all(v is None for v in r):
            continue
        vals = [r[i] for i in keep]
        o.append(vals)
        n += 1
        if isinstance(vals[date_i], datetime.datetime):
            dates.append(vals[date_i])
    out.save(dest)
    wb.close()

    dropped = [h for h in header if h in DROP_COLUMNS]
    print(f"Wrote {dest}: {n} comps, {len(kept_names)} columns"
          + (f" (dropped {dropped})" if dropped else ""))
    if dates:
        print(f"Sale dates span {min(dates).date()} -> {max(dates).date()}")
        stale = (datetime.date.today() - max(dates).date()).days
        if stale > 120:
            print(f"WARNING: newest sale is {stale} days old — the comps database "
                  "may not have been refreshed. Say so when you quote off it.")
    return n


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("cma", help='"Automatic CMA Analysis.xlsx"')
    ap.add_argument("dest", nargs="?", default="All Sales Comps.xlsx")
    ap.add_argument("--sheet", default=SRC_SHEET)
    extract(*[getattr(ap.parse_args(), k) for k in ("cma", "dest", "sheet")])


if __name__ == "__main__":
    main()
