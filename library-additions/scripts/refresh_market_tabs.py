#!/usr/bin/env python3
"""Refresh the TMG model's two PRIOR-DEAL market tabs on Linux (no Excel/COM).

Linux counterpart of library-additions/scripts/refresh_agency_region.py, which needs
pywin32 + desktop Excel. Both tabs below are Power Query tables whose source is not
available on a job machine, so they NEVER refresh -- the template ships loaded with
whatever metro the previous deal used, and left alone they silently drive client-facing
numbers:

1. 'Agency Region' (table Region_Comps) ships with a DALLAS-FORT WORTH extract. It feeds
   'Agency Loan-Sale Comps'!Z40 = AGGREGATE(1,3,Z16:Z37) -- the average cap rate over the
   first 22 rows -- which is the BASE of the terminal cap in Assumptions!G58. It also
   prints as the PDF's "Agency Loan-Sale Comparables (Geographic Region)" page, which
   would otherwise show Dallas properties under this deal's Houston search criteria.
   Source: the sales-comps skill's "Automatic CMA Analysis.xlsx" AgencyDrift sheet, which
   carries the SAME 26-column schema plus two trailing columns we drop.

2. 'YardiProjections' ships with a LITTLE ROCK forecast. Rows 5/6/7 feed
   Factors!N16 (avg 5-yr rent growth) and Factors!N17 (market occupancy at reversion).

INPUTS   --model, --cma, --region (substring of AgencyDrift "TMG Region")
OUTPUT   the model saved with both tabs repopulated; prints the resulting Z40 average
         so it can be sanity-checked against the CMA's own state/vintage average.

Usage:
  python3 refresh_market_tabs.py --model in.xlsx --out out.xlsx \
      --cma "Automatic CMA Analysis.xlsx" --region Houston
"""
import argparse
import datetime as dt
import warnings

warnings.filterwarnings("ignore")
import openpyxl

# Yardi "3 - Mount Houston (Houston - East)", Forecast Trends revised 8/4/2026, p.1.
# Columns B..O; the band structure (4 history Q, 4 forecast Q, 5 annual Q4, 1 ten-year)
# matches the template's own row-1 banding exactly.
YARDI_LABEL = "Mount Houston (Houston - East)"
YARDI_PERIODS = [(2025, 3), (2025, 4), (2026, 1), (2026, 2), (2026, 3), (2026, 4),
                 (2027, 1), (2027, 2), (2027, 4), (2028, 4), (2029, 4), (2030, 4),
                 (2031, 4), (2036, 4)]
YARDI_RENT = [1214, 1216, 1212, 1209, 1207, 1206, 1209, 1213,
              1218, 1233, 1263, 1304, 1348, 1583]
YARDI_YOY = [0.018, 0.016, 0.007, -0.001, -0.006, -0.008, -0.002, 0.003,
             0.009, 0.012, 0.025, 0.032, 0.034, 0.031]
YARDI_OCC = [0.917, 0.914, 0.915, 0.914, 0.913, 0.911, 0.911, 0.911,
             0.911, 0.910, 0.909, 0.910, 0.909, 0.909]

N_COLS = 26          # Agency Region A..Z
ORIGINATION_COL = 7  # 1-based


def refresh_agency(model, cma_path, region, years_back, year_min, year_max,
                   units_min, units_max):
    cma = openpyxl.load_workbook(cma_path, data_only=True)
    ad = cma["AgencyDrift"]
    hdr = [ad.cell(1, c).value for c in range(1, N_COLS + 1)]

    rows = []
    for r in range(2, ad.max_row + 1):
        vals = [ad.cell(r, c).value for c in range(1, N_COLS + 1)]
        reg, orig, built, units = vals[4], vals[6], vals[10], vals[11]
        if not reg or region.lower() not in str(reg).lower():
            continue
        if not isinstance(orig, dt.datetime):
            continue
        try:
            built, units = int(built), int(units)
        except (TypeError, ValueError):
            continue
        if not (year_min <= built <= year_max and units_min <= units <= units_max):
            continue
        rows.append((orig, vals))

    rows.sort(key=lambda x: x[0], reverse=True)
    cutoff = max(o for o, _ in rows) - dt.timedelta(days=365 * years_back) if rows else None
    rows = [r for r in rows if r[0] >= cutoff]

    ws = model["Agency Region"]
    for r in range(2, ws.max_row + 1):
        for c in range(1, N_COLS + 1):
            ws.cell(r, c).value = None
    for i, (_o, vals) in enumerate(rows):
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(2 + i, c)
            cell.value = v
            if c == ORIGINATION_COL and isinstance(v, dt.datetime):
                cell.number_format = "mm/dd/yyyy"

    last = max(2, 1 + len(rows))
    ws.tables["Region_Comps"].ref = f"A1:Z{last}"
    ws.tables["Region_Comps"].autoFilter.ref = f"A1:Z{last}"

    caps = [v[13] for _o, v in rows[:22] if isinstance(v[13], (int, float)) and v[13]]
    avg = sum(caps) / len(caps) if caps else None
    print(f"  Agency Region: {len(rows)} {region} rows written "
          f"(built {year_min}-{year_max}, {units_min}-{units_max} units, "
          f"{years_back}y back)")
    print(f"  -> Z40 will average the first 22 rows: n={len(caps)}, "
          f"avg cap {avg:.6f}" if avg else "  -> no cap rates found")
    assert hdr == [model["Agency Region"].cell(1, c).value for c in range(1, N_COLS + 1)], \
        "AgencyDrift and Agency Region headers differ"
    return avg


def refresh_yardi(model):
    ws = model["YardiProjections"]
    ws["A4"] = YARDI_LABEL
    for i, (yr, q) in enumerate(YARDI_PERIODS):
        col = 2 + i
        ws.cell(2, col).value = yr
        ws.cell(3, col).value = q
        ws.cell(5, col).value = YARDI_RENT[i]
        ws.cell(6, col).value = YARDI_YOY[i]
        ws.cell(7, col).value = YARDI_OCC[i]
    n16 = sum(YARDI_YOY[8:13]) / 5
    n17 = sum(YARDI_OCC[8:13]) / 5
    print(f"  YardiProjections: {YARDI_LABEL}, 14 columns written")
    print(f"  -> Factors!N16 avg(J6:N6)={n16:.5f} -> MROUND 0.0025 = "
          f"{round(n16 / 0.0025) * 0.0025:.4f}")
    print(f"  -> Factors!N17 avg(J7:N7)={n17:.5f} -> MROUND 0.0025 = "
          f"{round(n17 / 0.0025) * 0.0025:.4f}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--model", required=True)
    p.add_argument("--out", required=True)
    p.add_argument("--cma", required=True)
    p.add_argument("--region", default="Houston")
    p.add_argument("--years-back", type=int, default=3)
    p.add_argument("--year-min", type=int, default=1950)
    p.add_argument("--year-max", type=int, default=2000)
    p.add_argument("--units-min", type=int, default=0)
    p.add_argument("--units-max", type=int, default=500)
    a = p.parse_args()

    wb = openpyxl.load_workbook(a.model)
    refresh_agency(wb, a.cma, a.region, a.years_back, a.year_min, a.year_max,
                   a.units_min, a.units_max)
    refresh_yardi(wb)
    wb.save(a.out)
    print(f"  wrote {a.out}")


if __name__ == "__main__":
    main()
