#!/usr/bin/env python3
"""
verify_ssi410_rentroll.py - independent tie-out of a finished TMG rent-roll
workbook against the SSI410 "Rent Roll Report" PDF it was built from.

WHY THIS EXISTS
---------------
`process_rent_roll.py`'s SSI410Parser harvests report totals by caption regex.
Some SSI410 installs print the summary blocks with different captions, so those
checks silently register as informational ("~ ... (no report total to check)")
instead of as hard OK/FAIL checks. The toolkit's iron rule is that every run
shows a reconciliation block that ties out - when the built-in block degrades to
"~", run this script to get the real tie-out. Known caption variants that dodge
the built-in checks (seen at Harvest Moon, Oak Leaf Management, 6/1/2026):

  * occupancy block captioned "Unit Analysis" (not "Occupancy Status"), and it
    prints Description / Units / Percent only - there is NO sqft column, so the
    parser's 3-group `Occupied\\s+(\\d+)\\s+([\\d,]+)\\s+([\\d,.]+)` cannot match.
    It is also laid out in newspaper columns, so each row arrives glued to the
    income-code legend: "MISCI AMENITY FEE L Leased OC Occupied Occupied 72 96.00".
    Hence the end-of-line anchors used below.
  * charge summary captioned "(Current, On-Notice, Transfer Out residents only)"
    rather than "Current/On-Notice".

This script reads the PDF's own printed grand totals - it does NOT re-use the
parser - so it is a genuinely independent second path, not the parse checking
itself.

INPUTS
------
  python verify_ssi410_rentroll.py "<source.pdf>" "<RR - Property - M-D-YYYY.xlsx>"

OUTPUT
------
A reconciliation block on stdout, one line per check, and exit code 0 if every
check ties or 1 if any check fails. Tolerances are $0.01 on money and exact on
counts - NEVER widen them. A failure means the parse or the mapping is wrong.

Workbook column assumptions (TMG rentroll_template.xlsx, Rent Roll tab):
  header row 3, data from row 4; C Net Sf | H Occupancy Status |
  I Market Rent | J Contractual Rent | P Other Income.

Requires: pdfplumber, openpyxl
"""

import re
import sys

import openpyxl
import pdfplumber

MONEY_TOL = 0.01

# --- PDF grand-total captions -------------------------------------------------
RE_HEADER_UNITS = re.compile(r"([\d,]+)\s+Apts,\s+([\d,]+)\s+Sq\.\s*Ft\.")
RE_OCCUPIED = re.compile(r"Occupied\s+(\d+)\s+[\d.]+\s*$")
RE_VACANT = re.compile(r"Vacant\s+(\d+)\s+[\d.]+\s*$")
RE_TOTAL_UNITS = re.compile(r"Total Units\s+(\d+)\s+[\d.]+\s*$")
RE_GRAND_BANNER = re.compile(r"^Grand Total\s*:")
RE_CODE_LINE = re.compile(r"^([A-Za-z][\w/-]{2,7})\s+([A-Z][A-Z /]+?)\s+([\d,]+\.\d\d)\s*$")


def money(text):
    return float(text.replace(",", ""))


def read_pdf(path):
    """Pull the SSI410 report's own printed totals. Returns a dict."""
    with pdfplumber.open(path) as pdf:
        lines = []
        for page in pdf.pages:
            lines.extend((page.extract_text() or "").split("\n"))

    out = {}

    for line in lines:
        m = RE_HEADER_UNITS.search(line)
        if m:
            out["units"] = int(m.group(1).replace(",", ""))
            out["sqft"] = int(m.group(2).replace(",", ""))
            break

    for line in lines:
        for key, rx in (("occupied", RE_OCCUPIED), ("vacant", RE_VACANT),
                        ("total_units", RE_TOTAL_UNITS)):
            m = rx.search(line)
            if m and key not in out:
                out[key] = int(m.group(1))

    # Grand Total strip: the banner, then a two-row column caption, then values.
    for i, line in enumerate(lines):
        if RE_GRAND_BANNER.match(line):
            for probe in lines[i + 1:i + 6]:
                nums = re.findall(r"[\d,]+\.\d\d", probe)
                if len(nums) >= 4:
                    (out["market_rent"], out["actual_lease_rent"],
                     out["gross_possible"], out["potential_charges"]) = \
                        [money(n) for n in nums[:4]]
                    break
            break

    # Grand Summary of Actual Charges by Income Code - take the LAST such block
    # (per-page blocks repeat; the final one is the grand summary).
    codes, block = {}, {}
    for line in lines:
        m = RE_CODE_LINE.match(line.strip())
        if m:
            block[m.group(1)] = money(m.group(3))
        elif block:
            codes = block
            block = {}
    if block:
        codes = block
    out["codes"] = codes
    return out


def read_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["Rent Roll"]
    rows, r = [], 4
    while ws.cell(r, 1).value not in (None, ""):
        rows.append({
            "unit": ws.cell(r, 1).value,
            "sqft": ws.cell(r, 3).value,
            "status": ws.cell(r, 8).value,
            "market": ws.cell(r, 9).value,
            "contract": ws.cell(r, 10).value,
            "other": ws.cell(r, 16).value,
        })
        r += 1
    return rows


def num(v):
    return v if isinstance(v, (int, float)) else 0.0


def main(pdf_path, xlsx_path):
    rep = read_pdf(pdf_path)
    rows = read_workbook(xlsx_path)

    base_rent = rep["codes"].get("RENT")
    other_codes = sum(v for k, v in rep["codes"].items() if k != "RENT")
    vac_market = sum(num(x["market"]) for x in rows if x["status"] == "Vacant")

    checks = [
        ("Unit count", len(rows), rep.get("units"), 0),
        ("Total units (Unit Analysis)", len(rows), rep.get("total_units"), 0),
        ("Total Net Sf", sum(num(x["sqft"]) for x in rows), rep.get("sqft"), 0),
        ("Occupied units", sum(1 for x in rows if x["status"] == "Occupied"),
         rep.get("occupied"), 0),
        ("Vacant units", sum(1 for x in rows if x["status"] == "Vacant"),
         rep.get("vacant"), 0),
        ("Total Market Rent", sum(num(x["market"]) for x in rows),
         rep.get("market_rent"), MONEY_TOL),
        ("Contractual Rent vs BASE RENT code",
         sum(num(x["contract"]) for x in rows), base_rent, MONEY_TOL),
        ("Other Income vs non-RENT codes",
         sum(num(x["other"]) for x in rows), other_codes, MONEY_TOL),
        ("Total lease charges (rent + other)",
         sum(num(x["contract"]) + num(x["other"]) for x in rows),
         rep.get("actual_lease_rent"), MONEY_TOL),
        ("Vacant units at market rent",
         vac_market,
         (rep["gross_possible"] - rep["potential_charges"])
         if "gross_possible" in rep else None, MONEY_TOL),
    ]

    print("Independent reconciliation vs the SSI410 report's own printed totals:")
    failed = 0
    for label, got, want, tol in checks:
        if want is None:
            print("  ~   %-38s %s (report prints no total)" % (label, fmt(got)))
            continue
        ok = abs(float(got) - float(want)) <= tol
        failed += 0 if ok else 1
        print("  %-3s %-38s parsed %s vs report %s"
              % ("OK" if ok else "FAIL", label, fmt(got), fmt(want)))

    print("\n%d of %d checks tie out." %
          (sum(1 for c in checks if c[2] is not None) - failed,
           sum(1 for c in checks if c[2] is not None)))
    return 1 if failed else 0


def fmt(v):
    if isinstance(v, float):
        return "{:,.2f}".format(v)
    return "{:,}".format(v)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    sys.exit(main(sys.argv[1], sys.argv[2]))
