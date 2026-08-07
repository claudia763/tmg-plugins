#!/usr/bin/env python3
"""parse_t12_toki_xlsx - the OWNER INDENTED-TREE T-12 dialect, updated 8/2026

Reference copy of the registered T-12 parser for the owner/PM income
statement whose body is an indented account tree under an "Account Name"
header row with REAL DATE month cells. Companion write-up:
`instructions/owner-indented-tree-t12-xlsx.md`.

Originally added for Werner Creek Apartments (Goldenwrist Investments LLC /
Toki Property Management LLC, Houston TX, 8/2026). This version EXTENDS it to
the second structural variant of the same dialect, found on Aldine Apartments
(Goldenwrist Capital LLC, 8/2026) - see the write-up for the four changes:

  1. memo-row detection generalised to PARENT rows (`_is_parent_row`);
  2. a grand-EXPENSE roll-up branch that can see accounts printed at the
     OUTER indent level;
  3. below-NOI restatement matching moved AHEAD of the account branch and
     widened to a bare "Net" caption;
  4. the Comments-tab note about "Total Operating Expense" is now chosen by
     comparing the printed row to the canonical total, instead of always
     asserting the Werner-specific partial-roll-up story.

Plus one mapping-engine one-liner, in `SECTION_ALLOWED`:

    (re.compile(r"payroll|benefits|wages?\b", re.I), {"pr"}),

  Owner books head the payroll block "Wages" rather than "Payroll". Without
  it the section carries no constraint at all and "Temp. Contractor (1099)"
  lands in Repair & Maintenance, because the `a/?c` alternative in
  KEYWORD_RULES matches the "ac" inside "Contractor".

This file is NOT importable on its own: it is the body of functions that live
inside `process_t12.py` and use that module's `Line`, `datetime`, `re` and
`sys`. To install, replace the same-named functions in a toolkit copy of
`process_t12.py`, add the SECTION_ALLOWED line above, and confirm
`T12_XLSX_PARSERS` still reads:

    T12_XLSX_PARSERS = [
        (_is_owner_calendar_year_tabs, parse_t12_owner_calendar_year_tabs),
        (_is_resman_trailing_xlsx, parse_t12_resman_trailing_xlsx),
        (_is_toki_xlsx, parse_t12_toki_xlsx),
        (_is_appfolio_xlsx, parse_t12_appfolio_xlsx),
        (_is_onesite_xlsx, parse_t12_onesite_xlsx),
        (_is_yardi_xlsx, parse_t12_yardi_xlsx),
    ]

Aldine 8/2026 result: 38 printed row totals and 10 structural checks all tie;
Total Revenue 1,046,863.71 / Total Operating Expense 483,777.53 / NOI
563,086.18 over 11 real months (Dec-2025 blank).
"""

import re
import sys
from datetime import datetime

from process_t12 import Line


# ----------------------------------------------------------------------------
# Toki Property Management LLC owner/PM income statement XLSX
# ----------------------------------------------------------------------------
# Added for Werner Creek Apartments, Houston TX, 8/2026 (owner Goldenwrist
# Investments LLC, PM Toki Property Management LLC). The generic
# parse_t12_xlsx header sniff cannot read this dialect: the month headers are
# REAL DATE CELLS (2025-04-01 ...), not "June-2025" text, so no 12-month
# header row is ever found.
#
#   row 1  PM company            row 2  "Properties: <name> - <address>"
#   row 3  "Period Range: ..."   row 5  "Account Name" | 12 date cells | Total
#   body   leading-space indentation (4 per level) in col A
#
# Layout gotchas this parser handles explicitly:
#
# - MEMO ROW ABOVE THE LEDGER. "Market Rent" (gross potential rent) is printed
#   before the "Income" section and is NOT part of the statement's own Total
#   Operating Income. Counting it as revenue would double-count rent, so any
#   account row that appears before the first section header is treated as a
#   memo row - and that judgement is PROVED, not assumed: the revenue grand row
#   is verified against the detail with the memo rows excluded, and if the
#   statement actually did include them the parser says so and includes them.
# - THE PRINTED "Total Operating Expense" IS NOT THE OPERATING EXPENSE TOTAL.
#   It rolls up UTILITY + Wages only; Tax, Insurance and the Maintenance
#   section sit BELOW it and outside it. The canonical Total Operating Expenses
#   is therefore taken as printed Total Operating Income less printed NOI -
#   which the statement itself proves month by month - and that figure is then
#   checked against the sum of every expense account. (Same reasoning as the
#   QuickBooks COGS case: it is arithmetic on printed rows, not a derived row
#   checked against itself.)
# - ROLL-UP CAPTIONS DO NOT MATCH THEIR SECTION'S INDENT. "TOTAL FEES" is
#   printed at indent 0 while the FEES accounts sit at 8/12, so roll-ups are
#   resolved by NAME first ("Total <X>" -> the still-open section <X>, all of
#   its not-yet-consumed rows) and only fall back to "every unconsumed row
#   indented deeper than me" when no section matches that name.
# - RESTATEMENT ROWS BELOW NOI. "Total Income" / "Total Expense" / "Net Income"
#   repeat the operating totals. They are verified month-by-month against their
#   operating counterparts and then dropped - left in they would double-count
#   in RawData (write_workbook codes "Total Income" as 'rev') and land in the
#   Capex & Misc workbook spanning both sides of the ledger.
# - EMPTY MONTH COLUMNS. A dated month column whose every data cell is blank
#   was not reported at all (Werner: Dec-2025). It is dropped from the parse -
#   never read as zero - and named loudly; --pad-to-12 then puts it back on the
#   axis as a genuinely blank column with a Comments-tab note.
#
# Verification layers (all abort unless --trust-monthly): every printed row
# Total vs its own monthly cells; every roll-up vs the detail it consumes, per
# month; the revenue grand vs all unconsumed income detail; NOI vs revenue less
# every expense account; and the below-NOI restatement rows vs the operating
# totals.

_TOKI_DROP = re.compile(
    r"^total\s+(operating\s+)?(income|revenue|expense)s?$"
    r"|net operating income|^net income$|\bNOI\b", re.I)


def _toki_month(v):
    """A Toki month header: a real date cell -> first of that month."""
    if v is None or isinstance(v, str):
        return None
    if hasattr(v, "year") and hasattr(v, "month"):
        try:
            return datetime(v.year, v.month, 1)
        except (TypeError, ValueError):
            return None
    return None


def _toki_header(rows):
    """-> (hdr_row_index, [month col idxs], [months], total col idx) | None."""
    for i, r in enumerate(rows[:15]):
        if not r or str(r[0] or "").strip().lower() != "account name":
            continue
        got = [(j, _toki_month(v)) for j, v in enumerate(r) if j]
        got = [(j, d) for j, d in got if d]
        if len(got) < 3 or got[0][0] != 1:
            continue
        tot = None
        for j, v in enumerate(r):
            if j > got[-1][0] and \
                    str(v or "").strip().upper() in ("TOTAL", "TOTALS"):
                tot = j
        return i, [j for j, _ in got], [d for _, d in got], tot
    return None


def _is_toki_xlsx(path):
    """True for the Toki/owner income statement: an 'Account Name' header row
    with REAL DATE month cells starting at col B and a Total column, under a
    'Properties: ...' line, over a leading-space indented account tree.

    Deliberately narrow: AppFolio also captions col A 'Account Name' and also
    prints a 'Properties:' preamble line, but its month headers are TEXT
    ('Jul 2025'), so the date-cell test cannot collide with it. OneSite uses
    date headers too but keys on '<GL> - <Name>' account labels, which this
    statement has none of."""
    from openpyxl import load_workbook as _lw
    if not str(path).lower().endswith((".xlsx", ".xlsm")):
        return False
    try:
        wb = _lw(path, read_only=True, data_only=True)
    except Exception:
        return False
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True, max_row=200)]
    hd = _toki_header(rows)
    if not hd or hd[3] is None:
        return False
    if not any(re.match(r"^\s*propert(y|ies)\s*:", str(r[0] or ""), re.I)
               for r in rows[:hd[0]]):
        return False
    return sum(1 for r in rows[hd[0] + 1:]
               if r and isinstance(r[0], str) and r[0].startswith(" ")) >= 5


def parse_t12_toki_xlsx(path, trust_monthly=False, allow_partial=False):
    """Toki/owner income statement xlsx -> (property, months, [Line], meta)."""
    from openpyxl import load_workbook as _lw

    def _num(v):
        if v is None:
            return None
        if isinstance(v, str):
            s = v.strip().replace(",", "").replace("$", "")
            if not s:
                return None
            neg = s.startswith("(") and s.endswith(")")
            try:
                x = float(s.strip("()"))
            except ValueError:
                return None
            return -x if neg else x
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    wb = _lw(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr_i, mcols, months_all, tot_col = _toki_header(rows)

    prop = ""
    for r in rows[:hdr_i]:
        a = re.sub(r"\s+", " ", str(r[0] or "")).strip()
        m = re.match(r"^propert(?:y|ies)\s*:\s*(.+)$", a, re.I)
        if m:
            prop = re.split(r"\s+-\s+", m.group(1))[0].strip()
    if not prop:
        prop = re.sub(r"\s+", " ", str((rows[0] or [""])[0] or "")).strip()

    body = []
    for r in rows[hdr_i + 1:]:
        raw = r[0] if r else None
        if raw is None or not str(raw).strip():
            continue
        raw = str(raw)
        name = re.sub(r"\s+", " ", raw).strip()
        if name.lower() == "account name":
            continue
        body.append({
            "name": name,
            "indent": len(raw) - len(raw.lstrip(" ")),
            "vals": [_num(r[j]) if j < len(r) else None for j in mcols],
            "annual": (_num(r[tot_col])
                       if tot_col is not None and tot_col < len(r) else None)})

    # ---- month columns the statement leaves genuinely EMPTY --------------
    real = [k for k in range(len(mcols))
            if any(b["vals"][k] is not None for b in body)]
    blank_months = [months_all[k] for k in range(len(mcols)) if k not in real]
    months = [months_all[k] for k in real]
    nm = len(months)
    notes = []
    if blank_months:
        txt = ", ".join(f"{d:%b %Y}" for d in blank_months)
        print(f"!! EMPTY MONTH COLUMN(S): {txt} - the column header is on the "
              f"statement but every data cell in it is BLANK. Ownership did "
              f"not report {'those months' if len(blank_months) > 1 else 'that month'}"
              f"; the column is dropped from the parse and never read as "
              f"zero. Use --pad-to-12 to keep it on a full trailing-12 axis "
              f"as a genuinely blank column.")
        notes.append(f"{txt} carried no data at all on the statement "
                     f"(blank cells, not zeros) - not reported by ownership.")

    def mv(vals):
        return [vals[k] or 0.0 for k in real]

    def emask(vals):
        return [vals[k] is None for k in real]

    def diff(a, b):
        return max((abs(x - y) for x, y in zip(a, b)), default=0.0)

    lines_out, items, consumed, stack = [], [], set(), []
    sect_at, checks, fails = {}, [], []
    side, seen_section = "inc", False
    rev_v = exp_v = noi_v = printed_exp_grand = None
    n_rowchk = 0

    # A valued row whose IMMEDIATELY FOLLOWING row is a valued, deeper-indented
    # account is a PARENT row, not a leaf: the money below it is the real
    # ledger line and the parent is a memo/reference figure printed above it
    # (Aldine 8/2026: "MARKET RENT TOTAL" with "Rent Income" nested under it).
    # This is only a CANDIDATE - like the Werner-style memo row printed above
    # the first section header, it is excluded from revenue only if the
    # printed revenue grand row then ties month by month, and put back if not.
    def _is_parent_row(i):
        b0 = body[i]
        if not any(b0["vals"][k] is not None for k in real):
            return False
        if re.match(r"^total\b", b0["name"], re.I):
            return False
        nxt = body[i + 1] if i + 1 < len(body) else None
        if nxt is None or nxt["indent"] <= b0["indent"]:
            return False
        if re.match(r"^total\b", nxt["name"], re.I):
            return False
        return any(nxt["vals"][k] is not None for k in real)

    for bi, b in enumerate(body):
        name, indent = b["name"], b["indent"]
        vals, annual = b["vals"], b["annual"]
        has = any(vals[k] is not None for k in real)
        is_noi = bool(re.search(r"\bNOI\b|net operating income", name, re.I))
        is_tot = is_noi or bool(re.match(r"^total\b", name, re.I)) or \
            bool(re.fullmatch(r"net income", name, re.I))

        if not has and not is_tot:
            while stack and stack[-1][0] >= indent:
                stack.pop()
            stack.append((indent, name))
            sect_at.setdefault(name.lower(), len(items))
            seen_section = True
            ln = Line("section", name, None, name, side)
            items.append({"kind": "section", "name": name, "indent": indent,
                          "vals": None, "line": ln})
            lines_out.append(ln)
            if re.fullmatch(r"(operating\s+)?(income|revenue)s?", name, re.I):
                side = "inc"
            elif re.search(r"expense", name, re.I) and \
                    not re.search(r"income|revenue", name, re.I):
                side = "exp"
            continue

        v = mv(vals)
        if annual is not None:
            n_rowchk += 1
            if abs(sum(v) - annual) > 0.05:
                fails.append(f"ROW-CHECK {name}: months sum {sum(v):,.2f} vs "
                             f"printed Total {annual:,.2f} "
                             f"(variance {sum(v) - annual:+,.2f})")

        # ---- anything printed AFTER the statement's own NOI row ----------
        # In this dialect the rows below NOI restate the operating totals
        # ("Total Income" / "Total Expense" / "Net"). They are matched here
        # BEFORE the account branch, because the caption is not always a
        # "Total ..." one - Aldine's final row is captioned just "Net" and
        # would otherwise be read as a 563,086.18 expense account.
        if noi_v is not None:
            idx = len(items)
            items.append({"kind": "subtotal", "name": name, "indent": indent,
                          "vals": v, "side": side, "memo": False,
                          "line": None})
            tgt, tname = None, ""
            if re.fullmatch(r"total\s+(operating\s+)?(income|revenue)s?",
                            name, re.I):
                tgt, tname = rev_v, "printed Total Operating Income"
            elif re.fullmatch(r"total\s+(operating\s+)?expenses?",
                              name, re.I):
                tgt, tname = exp_v, "every expense account"
            elif re.fullmatch(r"net(\s+(income|profit|loss|"
                              r"operating\s+income))?", name, re.I):
                tgt, tname = noi_v, "printed NOI"
            if tgt is None:
                fails.append(f"BELOW-NOI row not recognised as a restatement "
                             f"of an operating total: {name} - it carries "
                             f"values and would otherwise be counted in "
                             f"operations")
            elif diff(v, tgt) > 0.05:
                fails.append(f"RESTATEMENT {name} vs {tname}: worst month "
                             f"variance {diff(v, tgt):+,.2f}")
            else:
                checks.append(f"  OK  restatement row '{name}' = {tname}, "
                              f"all {nm} months (dropped: it repeats an "
                              f"operating total)")
            consumed.add(idx)
            continue

        if not is_tot:
            while stack and stack[-1][0] > indent:
                stack.pop()
            sec = stack[-1][1] if stack else name
            memo = (not seen_section) or _is_parent_row(bi)
            ln = Line("account", name, v, sec, side)
            ln.empty = emask(vals)
            items.append({"kind": "account", "name": name, "indent": indent,
                          "vals": v, "side": side, "memo": memo, "line": ln})
            if not memo:
                lines_out.append(ln)
            continue

        # ---------------- roll-up / grand / restatement rows --------------
        idx = len(items)
        items.append({"kind": "subtotal", "name": name, "indent": indent,
                      "vals": v, "side": side, "memo": False, "line": None})

        if is_noi:
            exp_v = [sum(it["vals"][k] for it in items
                         if it["kind"] == "account" and it["side"] == "exp"
                         and not it["memo"]) for k in range(nm)]
            noi_v = v
            want = [a - b2 for a, b2 in zip(rev_v or [0.0] * nm, exp_v)]
            if diff(v, want) > 0.05:
                fails.append(f"NOI vs revenue less every expense account: "
                             f"worst month variance {diff(v, want):+,.2f}")
            else:
                checks.append(f"  OK  printed NOI = Total Operating Income "
                              f"less every expense account, all {nm} months")
            lines_out.append(Line("subtotal", name, v, "", ""))
            continue

        if side == "inc" and re.fullmatch(
                r"total\s+(operating\s+)?(income|revenue)s?", name, re.I):
            def rollup(with_memo):
                return [i for i in range(idx) if i not in consumed
                        and items[i]["kind"] in ("account", "subtotal")
                        and (with_memo or not items[i]["memo"])]
            best = None
            for with_memo in (False, True):
                mem = rollup(with_memo)
                tot = [sum(items[i]["vals"][k] for i in mem)
                       for k in range(nm)]
                if diff(tot, v) <= 0.05:
                    best = (with_memo, mem)
                    break
            memo_names = [items[i]["name"] for i in range(idx)
                          if items[i].get("memo")]
            if best is None:
                mem = rollup(False)
                tot = [sum(items[i]["vals"][k] for i in mem)
                       for k in range(nm)]
                fails.append(f"REVENUE GRAND {name} vs its own detail: worst "
                             f"month variance {diff(tot, v):+,.2f}")
            else:
                with_memo, mem = best
                for i in mem:
                    consumed.add(i)
                checks.append(
                    f"  OK  printed '{name}' = its income detail "
                    f"({len(mem)} rows), all {nm} months")
                if memo_names and not with_memo:
                    print("!! MEMO ROW(S) EXCLUDED FROM REVENUE: "
                          + "; ".join(memo_names)
                          + " - printed above the income section, or as a "
                            "PARENT of the ledger line nested under it, and "
                            "proved month-by-month to be OUTSIDE the "
                            "statement's own Total Operating Income "
                            "(gross-potential-rent reference). Counting it "
                            "as revenue would double-count rent.")
                elif memo_names and with_memo:
                    print("   memo-position row(s) ARE inside the printed "
                          "revenue total and were kept: "
                          + "; ".join(memo_names))
                    # Put them back into the output in their printed position.
                    # They were held out of lines_out while they were only
                    # memo CANDIDATES; the statement has now proved they are
                    # real revenue, so they must not vanish from the T-12.
                    for i in range(idx - 1, -1, -1):
                        if not items[i].get("memo"):
                            continue
                        items[i]["memo"] = False
                        mln = items[i].get("line")
                        if mln is None or mln in lines_out:
                            continue
                        nxt = next((items[j]["line"]
                                    for j in range(i + 1, len(items))
                                    if items[j].get("line") is not None
                                    and items[j]["line"] in lines_out), None)
                        if nxt is None:
                            lines_out.append(mln)
                        else:
                            lines_out.insert(lines_out.index(nxt), mln)
            rev_v = v
            side = "exp"
            lines_out.append(Line("subtotal", name, v, "", "inc"))
            continue

        # ---- grand EXPENSE row -------------------------------------------
        # "Total Operating Expense" is not reliably an indent subtree in this
        # dialect. Two readings are seen, and both are proved month by month
        # rather than assumed:
        #   (1) the indent subtree - Werner Creek 8/2026, where the printed
        #       row is a PARTIAL roll-up (UTILITY + Wages) and the Tax /
        #       Insurance / Maintenance blocks are printed BELOW it;
        #   (2) every unconsumed expense-side row printed above it - Aldine
        #       8/2026, where all expense blocks precede the row AND two real
        #       maintenance accounts ("Annual Material Cost / month", "Other
        #       annual maintenance / month") are printed at the OUTER indent
        #       level (indent 0), which reading (1) structurally cannot see.
        # The first reading that ties wins; if neither ties the run aborts.
        if side == "exp" and re.fullmatch(
                r"total\s+(operating\s+)?expenses?", name, re.I):
            cands = [
                ("every unconsumed row indented deeper than it",
                 [i for i in range(idx) if i not in consumed
                  and items[i]["kind"] in ("account", "subtotal")
                  and items[i]["indent"] > indent and not items[i]["memo"]]),
                ("every unconsumed expense-side row printed above it "
                 "(incl. accounts at the outer indent level)",
                 [i for i in range(idx) if i not in consumed
                  and items[i]["kind"] in ("account", "subtotal")
                  and items[i].get("side") == "exp"
                  and not items[i]["memo"]]),
            ]
            picked = None
            for how, mem in cands:
                tot = [sum(items[i]["vals"][k] for i in mem)
                       for k in range(nm)]
                if diff(tot, v) <= 0.05:
                    picked = (how, mem)
                    break
            if picked is None:
                how, mem = cands[0]
                tot = [sum(items[i]["vals"][k] for i in mem)
                       for k in range(nm)]
                fails.append(f"EXPENSE GRAND {name} vs {how}: worst month "
                             f"variance {diff(tot, v):+,.2f}")
            else:
                how, mem = picked
                outer = [items[i]["name"] for i in mem
                         if items[i]["indent"] <= indent]
                checks.append(f"  OK  printed '{name}' = {how} "
                              f"({len(mem)} rows), all {nm} months")
                if outer:
                    print("   expense account(s) printed at the OUTER indent "
                          "level and proved to be INSIDE the statement's own "
                          "Total Operating Expense: " + "; ".join(outer))
                for i in mem:
                    consumed.add(i)
            printed_exp_grand = list(v)
            lines_out.append(Line("subtotal", name, v, "", side))
            continue

        m = re.match(r"^total\s+(.+)$", name, re.I)
        tail = m.group(1).strip().lower() if m else ""
        start = sect_at.get(tail)
        if start is not None:
            mem = [i for i in range(start + 1, idx) if i not in consumed
                   and items[i]["kind"] in ("account", "subtotal")]
            how = f"section '{tail}'"
        else:
            mem = [i for i in range(idx) if i not in consumed
                   and items[i]["kind"] in ("account", "subtotal")
                   and items[i]["indent"] > indent and not items[i]["memo"]]
            how = "every unconsumed row indented deeper than it"
        tot = [sum(items[i]["vals"][k] for i in mem) for k in range(nm)]
        if diff(tot, v) > 0.05:
            fails.append(f"ROLL-UP {name} vs {how}: worst month variance "
                         f"{diff(tot, v):+,.2f}")
        else:
            checks.append(f"  OK  '{name}' = {how} ({len(mem)} rows), "
                          f"all {nm} months")
        for i in mem:
            consumed.add(i)
        lines_out.append(Line("subtotal", name, v, "", side))

    if rev_v is None or noi_v is None:
        sys.exit("ERROR: Toki statement: no printed Total Operating Income / "
                 "NOI row found - cannot tie out.")

    # The printed "Total Operating Expense" rolls up only part of the expense
    # side, so the canonical operating-expense total is taken from the two
    # printed grand rows (Total Operating Income less NOI) and then proved
    # against the sum of every expense account.
    exp_grand = [a - b2 for a, b2 in zip(rev_v, noi_v)]
    if diff(exp_grand, exp_v) > 0.05:
        fails.append(f"canonical Total Operating Expenses (printed Total "
                     f"Operating Income less printed NOI) vs the sum of every "
                     f"expense account: worst month variance "
                     f"{diff(exp_grand, exp_v):+,.2f}")
    else:
        checks.append(f"  OK  Total Operating Expenses (printed Total "
                      f"Operating Income less printed NOI) = the sum of every "
                      f"expense account, all {nm} months")

    print(f"Toki statement verification ({n_rowchk} printed row totals, "
          f"{len(checks)} structural checks):")
    print(f"  OK  every printed row Total ties to its own monthly cells "
          f"({n_rowchk} rows)" if not any(f.startswith("ROW-CHECK")
                                          for f in fails)
          else "  see ROW-CHECK failures below")
    for c in checks:
        print(c)
    for f in fails:
        print("  FAIL " + f)
    if fails:
        if not trust_monthly:
            sys.exit("ERROR: Toki statement does not tie out - aborting. The "
                     "parse or the statement is wrong; never widen the "
                     "tolerance. (--trust-monthly treats the monthly detail "
                     "as source of truth.)")
        print("  --trust-monthly: monthly detail wins over the printed rows.")

    lines_out = [l for l in lines_out
                 if not (l.kind == "subtotal" and _TOKI_DROP.search(l.name))]
    lines_out.append(Line("subtotal", "Total Revenue", list(rev_v), "", "inc"))
    lines_out.append(Line("subtotal", "Total Operating Expenses",
                          list(exp_grand), "", "exp"))
    lines_out.append(Line("subtotal", "Total Net Operating Income",
                          list(noi_v), "", ""))

    memo_rows = [it for it in items if it.get("memo")]
    note_line = []
    for it in memo_rows:
        tot = sum(it["vals"])
        print(f"   MEMO (not revenue): {it['name']} = "
              f"{tot:,.2f} over {nm} month(s)")
        note_line.append(
            f"{it['name']} (gross potential rent) totals {tot:,.2f} over the "
            f"{nm} reported months. The statement prints it as a memo row - "
            f"above the income section, or as the parent of the ledger line "
            f"nested under it - and EXCLUDES it from its own Total Operating "
            f"Income (verified month by month), so it is carried as a memo "
            f"reference only and is NOT counted as revenue; including it "
            f"would double-count rent against Rent Income.")
    if printed_exp_grand is not None and diff(printed_exp_grand,
                                              exp_grand) <= 0.05:
        note_line.append(
            "The statement's own \"Total Operating Expense\" row equals the "
            "sum of every expense line on the statement, month by month, and "
            "equals its printed Total Operating Income less its printed NOI. "
            "Expense accounts printed at the outer indent level rather than "
            "nested under a section head are included in Total Operating "
            "Expense - that is proved against the printed row, not assumed.")
    else:
        note_line.append(
            "The statement's own \"Total Operating Expense\" row rolls up "
            "UTILITY + Wages only; Tax, Insurance and the Maintenance section "
            "sit outside it. Total Operating Expenses here is the statement's "
            "printed Total Operating Income less its printed NOI, proved "
            "equal to the sum of every expense line, month by month.")
    meta = {"notes": notes, "note_line": " ".join(note_line),
            "exempt_full_year": False}
    return prop, months, lines_out, meta
