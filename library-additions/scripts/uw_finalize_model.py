#!/usr/bin/env python3
"""Final presentation pass + gate checks on the TMG underwriting model.

Does two things openpyxl has no direct API for, then runs the checks the
underwriting SKILL and aggressive-pricing-house-rule-8-2026.md require:

1. Removes the "black box" conditional format on 'PDF Output - F&C'!B52:J77
   (empty floor-plan rows render as solid black fill on the exported PDF).
   openpyxl has no delete API, so the CF list is rebuilt without that range.
2. Sets the final strike in Assumptions!G50.

CHECKS (run with --check on a RECALCULATED workbook):
  - green tests: F5 >= G48, F7 >= 0.10, I8 >= 1.25
  - the $10,000-boundary trap: 'PDF Output - F&C'!G22 is ROUND(...,-4), so an
    off-boundary strike prints a headline price the deal cannot support AND
    breaks the cost stack on page 1. Assert G22 == G50.
  - the cost stack ties: M13 + value-add capital == M17
  - zero formula errors inside the PDF print area

Usage:
  python3 finalize_model.py prep <in.xlsx> <out.xlsx> [--price 3100000]
  python3 finalize_model.py check <recalced.xlsx>
"""
import argparse
import warnings

warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList

BLACK_BOX = "B52:J77"
PDF = "PDF Output - F&C"
ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def prep(src, out, price):
    wb = openpyxl.load_workbook(src)
    ws = wb[PDF]
    new = ConditionalFormattingList()
    dropped = 0
    for rng in ws.conditional_formatting:
        if str(rng.sqref) == BLACK_BOX:
            dropped += 1
            continue
        for rule in rng.rules:
            new.add(str(rng.sqref), rule)
    ws.conditional_formatting = new
    print(f"{PDF}: removed {dropped} black-box conditional format(s) on {BLACK_BOX}")
    # 'Rent Summary'!H7:M7 average the four selected rent comps. With no rent
    # comp set loaded they are AVERAGE() over an empty range -> #DIV/0!, which
    # prints on the client-facing Rent Comparable Summary (PDF row 44). Wrap in
    # IFERROR so the row prints blank instead of an error, WITHOUT inventing a
    # comp: the formula still populates itself the moment comps are loaded.
    rs = wb["Rent Summary"]
    wrapped = 0
    for col in range(8, 14):                       # H..M
        c = rs.cell(7, col)
        f = getattr(c.value, "text", c.value)
        if isinstance(f, str) and f.startswith("=AVERAGE(") and "IFERROR" not in f:
            c.value = f'=IFERROR({f[1:]},"")'
            wrapped += 1
    print(f"Rent Summary!H7:M7: wrapped {wrapped} AVERAGE() in IFERROR "
          f"(rent comps not loaded — row prints blank, not #DIV/0!)")
    # PDF row 44 mostly mirrors 'Rent Summary'!7, but J44 averages the comp
    # rows on the page itself (=AVERAGE(J40:J43)) and needs the same guard.
    for col in range(2, 16):
        c = ws.cell(44, col)
        f = getattr(c.value, "text", c.value)
        if isinstance(f, str) and f.startswith("=AVERAGE(") and "IFERROR" not in f:
            c.value = f'=IFERROR({f[1:]},"")'
            print(f"{PDF}!{c.coordinate}: wrapped AVERAGE() in IFERROR")

    if price:
        wb["Assumptions"]["G50"] = price
        print(f"Assumptions!G50 = {price:,}")
    wb.save(out)
    print(f"wrote {out}")


def check(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    a, pdf = wb["Assumptions"], wb[PDF]
    irr, coc, dscr = a["F5"].value, a["F7"].value, a["I8"].value
    tgt, price = a["G48"].value, a["G50"].value
    va = wb["Value-Add"]["H93"].value or 0
    ok = True

    def line(name, got, need, passed):
        nonlocal ok
        ok = ok and passed
        print(f"  [{'ok' if passed else 'FAIL'}] {name}: {got} (need {need})")

    print("GREEN TESTS")
    line("Project IRR F5", f"{irr:.2%}", f">= {tgt:.0%}", irr >= tgt)
    line("Avg cash-on-cash F7", f"{coc:.2%}", ">= 10%", coc >= 0.10)
    line("T-3 DSCR I8", f"{dscr:.4f}", ">= 1.25", dscr >= 1.25)

    print("PRINTED-PAGE INTEGRITY")
    g22 = pdf["G22"].value
    line("strike sits on a $10,000 boundary (G22 == G50)",
         f"G22={g22:,.0f} G50={price:,.0f}", "equal", g22 == price)
    m13, m17 = pdf["M13"].value, pdf["M17"].value
    tie = abs((m13 or 0) + (va or 0) - (m17 or 0)) < 1.0
    line("cost stack ties (M13 + value-add == M17)",
         f"{m13:,.0f} + {va:,.0f} = {(m13 or 0)+(va or 0):,.0f} vs {m17:,.0f}",
         "equal", tie)

    bad = []
    for row in pdf.iter_rows(min_row=1, max_row=333, min_col=2, max_col=15):
        for c in row:
            if isinstance(c.value, str) and c.value in ERRS:
                bad.append(f"{c.coordinate}={c.value}")
    line("zero formula errors in the PDF print area",
         f"{len(bad)} found" + (f" -> {bad[:8]}" if bad else ""), "0", not bad)

    print(f"\n{'ALL GREEN' if ok else 'NOT GREEN — fix before delivery'}")
    return 0 if ok else 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("mode", choices=["prep", "check"])
    ap.add_argument("book")
    ap.add_argument("out", nargs="?")
    ap.add_argument("--price", type=int)
    a = ap.parse_args()
    if a.mode == "prep":
        prep(a.book, a.out, a.price)
    else:
        raise SystemExit(check(a.book))


if __name__ == "__main__":
    main()
