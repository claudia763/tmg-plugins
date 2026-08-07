#!/usr/bin/env python3
"""LibreOffice-headless driver for the TMG underwriting model (Linux, no Excel/COM).

Linux counterpart of library-additions/scripts/excel_model_recalc.py (which needs
pywin32 + desktop Excel). Drives LibreOffice through a Basic macro installed into a
throwaway profile -- the same trick the underwriting skill's export_pdf.py uses --
so we can force a full recalculation, read the resulting values, write input cells,
and export the single-sheet PDF without ever leaving the shell.

Subcommands
  prep   <in.xlsx> <out.xlsx>
         Pre-recalc cleanup required by underwriting/references/technical-notes.md
         SS1-2: delete every sheet-scoped defined name pointing at an external
         workbook ("[1]Raw Rents"!...), and replace direct external-link formula
         cells with their cached values read from the source with data_only=True.
  recalc <in.xlsx> [--out out.xlsx] [--set Sheet!A1=value ...] [--read Sheet!A1 ...]
         Optionally write input cells, hard-recalculate (calculateAll), print the
         requested cells as TSV, and save.
  read   <in.xlsx> --read Sheet!A1 ...
         Recalculate and print cells without saving.
  pdf    <in.xlsx> <out.pdf> [--sheet "PDF Output - F&C"] [--range B1:O333]
         Recalculate then export only that sheet range.

Values written with --set are parsed as float when possible, else as string.
Cells are read as (value, string) pairs so text and numbers both come back.

Inputs  : a TMG model .xlsx (openpyxl-edited is fine)
Outputs : TSV on stdout (sheet, cell, numeric value, display string), plus the
          saved .xlsx / .pdf when requested.

Timing on the 8/2026 template (67 sheets, ~106k formulas, 23 MB): a cold open +
calculateAll + save runs about 6-10 minutes, so budget accordingly and tune the
deal in Python first (underwriting SKILL Phase A) rather than in the spreadsheet.
"""
from __future__ import annotations

import argparse
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from xml.sax.saxutils import escape

SOFFICE = "soffice"
EXT_NAME = re.compile(r"""(?<![\w"\[])'?\[\d+\][^!"\[\]]*'?!""")


# --------------------------------------------------------------------------- prep
def prep(src: str, dst: str) -> None:
    import warnings

    warnings.filterwarnings("ignore")
    import openpyxl

    wb = openpyxl.load_workbook(src)
    vals = openpyxl.load_workbook(src, data_only=True)

    killed = 0
    for ws in wb.worksheets:
        for name in list(ws.defined_names):
            dn = ws.defined_names[name]
            if isinstance(dn.value, str) and EXT_NAME.search(dn.value):
                del ws.defined_names[name]
                killed += 1
    print(f"prep: deleted {killed} external sheet-scoped defined names")

    patched = 0
    for ws in wb.worksheets:
        vs = vals[ws.title]
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if isinstance(f, str) and f.startswith("=") and EXT_NAME.search(f):
                    cached = vs[c.coordinate].value
                    c.value = 0 if cached is None else cached
                    patched += 1
    print(f"prep: replaced {patched} external-link formula cells with cached values")

    wb.save(dst)
    print(f"prep: wrote {dst}")


# ------------------------------------------------------------------------- macro
def _basic(sets, reads, save_to, pdf_to, pdf_sheet, pdf_range, dump):
    lines = [
        "Sub Drive()",
        "  Dim oDoc, oSheet, oCell, oRange, s As String",
        "  oDoc = ThisComponent",
    ]
    for sheet, cell, value in sets:
        lines.append(f'  oSheet = oDoc.Sheets.getByName("{escape(sheet)}")')
        lines.append(f'  oCell = oSheet.getCellRangeByName("{cell}")')
        if isinstance(value, float):
            lines.append(f"  oCell.setValue({value!r})")
        elif value == "":
            lines.append("  oCell.setString(\"\")")
        elif value.startswith("="):
            lines.append(f'  oCell.setFormula("{escape(value)}")')
        else:
            lines.append(f'  oCell.setString("{escape(value)}")')
    lines.append("  oDoc.calculateAll()")

    lines.append('  s = ""')
    for sheet, cell in reads:
        lines.append(f'  oSheet = oDoc.Sheets.getByName("{escape(sheet)}")')
        lines.append(f'  oCell = oSheet.getCellRangeByName("{cell}")')
        lines.append(
            f'  s = s & "{escape(sheet)}" & Chr(9) & "{cell}" & Chr(9) '
            "& oCell.getValue() & Chr(9) & oCell.getString() & Chr(10)"
        )
    lines += [
        "  Dim iFile As Integer",
        "  iFile = FreeFile",
        f'  Open "{dump}" For Output As #iFile',
        "  Print #iFile, s",
        "  Close #iFile",
    ]

    if save_to:
        lines += [
            "  Dim aSave(0) As New com.sun.star.beans.PropertyValue",
            '  aSave(0).Name = "FilterName"',
            '  aSave(0).Value = "Calc MS Excel 2007 XML"',
            f'  oDoc.storeToURL("{Path(save_to).resolve().as_uri()}", aSave())',
        ]
    if pdf_to:
        lines += [
            "  Dim aFD(0) As New com.sun.star.beans.PropertyValue",
            "  Dim aPdf(1) As New com.sun.star.beans.PropertyValue",
            f'  oSheet = oDoc.Sheets.getByName("{escape(pdf_sheet)}")',
            "  oDoc.CurrentController.setActiveSheet(oSheet)",
            f'  oRange = oSheet.getCellRangeByName("{pdf_range}")',
            '  aFD(0).Name = "Selection"',
            "  aFD(0).Value = oRange",
            '  aPdf(0).Name = "FilterName"',
            '  aPdf(0).Value = "calc_pdf_Export"',
            '  aPdf(1).Name = "FilterData"',
            "  aPdf(1).Value = aFD()",
            f'  oDoc.storeToURL("{Path(pdf_to).resolve().as_uri()}", aPdf())',
        ]

    lines += ["  oDoc.close(False)", "End Sub"]
    body = "\n".join(lines)
    return (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">\n'
        '<script:module xmlns:script="http://openoffice.org/2000/script" '
        'script:name="Module1" script:language="StarBasic">\n' + body + "\n</script:module>"
    )


def drive(book, sets, reads, save_to, pdf_to, pdf_sheet, pdf_range, timeout):
    book = Path(book).resolve()
    profile = Path(tempfile.mkdtemp(prefix="lo_uw_profile_"))
    dump = profile / "cells.tsv"
    url = profile.as_uri()
    subprocess.run(
        [SOFFICE, "--headless", "--terminate_after_init", f"-env:UserInstallation={url}"],
        capture_output=True, timeout=180,
    )
    macro_dir = profile / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        sys.exit("LibreOffice profile bootstrap failed")
    (macro_dir / "Module1.xba").write_text(
        _basic(sets, reads, save_to, pdf_to, pdf_sheet, pdf_range, str(dump))
    )
    r = subprocess.run(
        [SOFFICE, "--headless", "--norestore", f"-env:UserInstallation={url}", str(book),
         "vnd.sun.star.script:Standard.Module1.Drive?language=Basic&location=application"],
        capture_output=True, timeout=timeout,
    )
    out = dump.read_text() if dump.exists() else ""
    if not out and reads:
        sys.exit(f"macro produced no output (rc={r.returncode}): {r.stderr[-2000:]!r}")
    return out


def _cellref(s):
    sheet, _, cell = s.rpartition("!")
    return sheet.strip("'"), cell


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("cmd", choices=["prep", "recalc", "read", "pdf"])
    ap.add_argument("book")
    ap.add_argument("target", nargs="?")
    ap.add_argument("--out")
    ap.add_argument("--set", action="append", default=[])
    ap.add_argument("--read", action="append", default=[])
    ap.add_argument("--sheet", default="PDF Output - F&C")
    ap.add_argument("--range", dest="rng", default="B1:O333")
    ap.add_argument("--timeout", type=int, default=3000)
    a = ap.parse_args()

    if a.cmd == "prep":
        prep(a.book, a.target or a.out)
        return

    sets = []
    for s in a.set:
        ref, _, val = s.partition("=")
        sheet, cell = _cellref(ref)
        try:
            sets.append((sheet, cell, float(val.replace(",", ""))))
        except ValueError:
            sets.append((sheet, cell, val))
    reads = [_cellref(r) for r in a.read]

    save_to = a.out if a.cmd == "recalc" else None
    pdf_to = a.target if a.cmd == "pdf" else None
    print(drive(a.book, sets, reads, save_to, pdf_to, a.sheet, a.rng, a.timeout), end="")
    if save_to:
        print(f"saved {save_to}")
    if pdf_to:
        p = Path(pdf_to)
        print(f"pdf {'OK ' + str(p.stat().st_size) + ' bytes' if p.exists() else 'FAILED'}")


if __name__ == "__main__":
    main()
